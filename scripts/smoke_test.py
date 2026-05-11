from __future__ import annotations

import json
import os
import sys
from pathlib import Path

from fastapi import HTTPException
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

os.environ.setdefault('UPLOAD_ASSISTANT_DB', str(ROOT / 'data' / 'smoke_test.db'))

from backend.app import main
from scripts import windows_executor


def assert_true(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def main_smoke() -> None:
    if main.DB_PATH.exists():
        main.DB_PATH.unlink()
    main.init_db()
    previous_settings = {item.key: item.value for item in main.list_settings()}
    main.save_settings([main.AppSetting(key='enable_external_collection', value='false')])
    main.save_settings([
        main.AppSetting(key='temu_shop_account', value='My Smoke Shop'),
        main.AppSetting(key='temu_site', value='美国'),
        main.AppSetting(key='temu_product_template', value='My Product Template'),
        main.AppSetting(key='temu_size_category', value='My Size Category'),
        main.AppSetting(key='temu_size_template', value='My Size Template'),
        main.AppSetting(key='temu_warehouse_template', value='My Warehouse'),
        main.AppSetting(key='temu_logistics_template', value='My Logistics'),
    ])
    command = main.upload_rpa_command(r'D:\tmp\smoke.xlsx')
    assert_true('--shop-account' in command and command[command.index('--shop-account') + 1] == 'My Smoke Shop', 'shop account not passed to RPA')
    assert_true('--site' in command and command[command.index('--site') + 1] == '美国', 'site not passed to RPA')
    assert_true('--product-template' in command and command[command.index('--product-template') + 1] == 'My Product Template', 'product template not passed to RPA')
    assert_true('--size-template' in command and command[command.index('--size-template') + 1] == 'My Size Template', 'size template not passed to RPA')
    assert_true('--warehouse-template' in command and command[command.index('--warehouse-template') + 1] == 'My Warehouse', 'warehouse template not passed to RPA')
    assert_true('--logistics-template' in command and command[command.index('--logistics-template') + 1] == 'My Logistics', 'logistics template not passed to RPA')

    dashboard = main.dashboard()
    assert_true('product_count' in dashboard, 'dashboard missing product_count')
    status = main.system_status()
    assert_true(status['database_exists'], 'system status database check failed')
    assert_true(status['data_dir'].endswith('data'), 'system status data dir missing')
    assert_true(all(item['exists'] and item['writable'] for item in status['directory_checks']), 'runtime directory check failed')
    assert_true((main.DATA_DIR / 'logs').exists(), 'logs directory was not initialized')
    assert_true((main.DATA_DIR / 'collection_requests').exists(), 'collection request directory was not initialized')
    assert_true((main.DATA_DIR / 'collection_results').exists(), 'collection result directory was not initialized')
    assert_true('collection_preflight' in status, 'system status missing collection preflight')
    upload_preflight = main.upload_preflight()
    assert_true('items' in upload_preflight, 'upload preflight missing item diagnostics')
    assert_true('blocked_count' in upload_preflight['items'], 'upload preflight missing blocked count')
    settings = {item.key: item.value for item in main.list_settings()}
    assert_true('onebound_key' in settings and 'onebound_secret' in settings, 'onebound settings missing')
    health = main.health()
    assert_true(health['ok'] is True and health['app'] == 'upload-assistant', 'health endpoint failed')
    rows = main.normalize_onebound_items(
        [{'title': 'Onebound Smoke Candidate', 'num_iid': '123456', 'price': '12.5', 'sales': '88', 'pic_url': 'https://img.example/a.jpg'}],
        '1688',
    )
    assert_true(rows[0]['url'].endswith('/123456.html'), 'onebound item url normalization failed')

    try:
        main.create_collection_task(
            main.CollectionTaskPayload(
                keyword='smoke test shorts',
                source='1688',
                mode='simulate',
                collector='local',
                target_count=2,
                min_price=10,
                max_price=30,
                blacklist='',
            )
        )
    except HTTPException as exc:
        assert_true(exc.status_code == 410, 'simulate collection was not removed')
    else:
        raise AssertionError('simulate collection unexpectedly succeeded')

    candidate = main.create_collection_item(
        main.CollectionItemPayload(
            title='Smoke Test Shorts Imported Candidate',
            source='外部文件',
            source_url='https://example.com/smoke-candidate',
            image_url='https://img.example.com/smoke.jpg',
            price=26.8,
            sales=500,
            image_count=12,
        )
    )
    candidates = [candidate]

    imported_products = main.import_collection_items(main.ImportCollectionPayload(ids=[candidates[0].id]))
    assert_true(len(imported_products) == 1, 'candidate import did not create product')
    assert_true(imported_products[0].weight_g == 350, 'imported product should keep default product weight')
    assert_true(imported_products[0].status == 'pending_main_image', 'failed external image download should keep product pending main image')
    local_collection_dir = main.IMAGE_DIR / 'smoke_collection'
    local_collection_dir.mkdir(parents=True, exist_ok=True)
    local_collection_image = local_collection_dir / 'main.jpg'
    local_collection_image.write_bytes(b'smoke-local-image')
    local_candidate = main.create_collection_item(
        main.CollectionItemPayload(
            title='Smoke Test Local Image Candidate',
            source='外部文件',
            source_url='https://example.com/smoke-local-candidate',
            image_url='/images/smoke_collection/main.jpg',
            price=21.5,
            sales=12,
            image_count=1,
        )
    )
    local_products = main.import_collection_items(main.ImportCollectionPayload(ids=[local_candidate.id]))
    assert_true(len(local_products) == 1, 'local image candidate import did not create product')
    assert_true(local_products[0].main_image.startswith('/images/'), 'local collection image was not copied to product image')
    assert_true(main.local_image_path_from_url(local_products[0].main_image) is not None, 'copied product image file does not exist')
    detail_payload = {
        'item': {
            'title': 'Smoke Detail Denim Shorts',
            'price': '19.8',
            'props_name': '0:0:颜色:Blue;0:1:颜色:Black;1:0:尺码:S;1:1:尺码:M',
            'prop_imgs': {'prop_img': [{'properties': '0:0', 'url': 'https://img.example.com/blue-prop.jpg'}]},
            'item_imgs': [
                {'url': 'https://img.example.com/detail-main-1.jpg'},
                {'url': 'https://img.example.com/detail-main-2.jpg'},
            ],
            'skus': {'sku': [
                {'sku_id': 'sku-blue-s', 'properties': '0:0;1:0', 'properties_name': '0:0:颜色:Blue;1:0:尺码:S'},
                {'sku_id': 'sku-blue-m', 'properties': '0:0;1:1', 'properties_name': '0:0:颜色:Blue;1:1:尺码:M'},
            ]},
        }
    }
    with main.connect() as db:
        main.apply_detail_to_product(db, imported_products[0].id, 'https://detail.1688.com/offer/123456.html', detail_payload)
        db.commit()
    enriched_processing = main.get_processing_item(imported_products[0].id)
    assert_true(enriched_processing.detail_status == 'completed', 'detail snapshot status not exposed')
    assert_true('蓝色' in enriched_processing.color and 'S' in enriched_processing.size, 'detail color/size did not map to processing item')
    assert_true(any(option.kind == 'detail_sku_image' for option in enriched_processing.image_options), 'detail sku images not available for image processing')
    delete_imported_candidate = main.delete_collection_items(main.CollectionBulkPayload(ids=[candidates[0].id]))
    assert_true(delete_imported_candidate['deleted_count'] == 1, 'imported collection item was not deleted')
    imported_product_still_exists = main.connect().execute('SELECT id FROM products WHERE id = ?', (imported_products[0].id,)).fetchone()
    assert_true(imported_product_still_exists is not None, 'deleting collection item removed product unexpectedly')

    try:
        main.generate_missing_image_placeholders()
    except HTTPException as exc:
        assert_true(exc.status_code == 410, 'placeholder generation was not removed')
    else:
        raise AssertionError('placeholder generation unexpectedly succeeded')

    with main.connect() as db:
        db.execute('DELETE FROM products WHERE id != ?', (imported_products[0].id,))
        db.execute('DELETE FROM processing_overrides WHERE product_id != ?', (imported_products[0].id,))
        db.execute('DELETE FROM product_color_image_assignments WHERE product_id != ?', (imported_products[0].id,))
        db.commit()
    product = main.product_from_row(main.connect().execute('SELECT * FROM products WHERE id = ?', (imported_products[0].id,)).fetchone())
    processing_before_image = main.get_processing_item(product.id)
    assert_true(processing_before_image.image_options, 'processing item missing selectable image options')
    local_source_image = next((option.url for option in processing_before_image.image_options if option.url.startswith('/images/')), '')
    if local_source_image:
        local_image_task = main.process_product_image(main.ProcessImagePayload(product_id=product.id, source_image=local_source_image))
        assert_true(local_image_task.status == 'completed' and local_image_task.provider == 'local', 'local image processing did not complete')
        assert_true(main.local_image_path_from_url(local_image_task.local_path) is not None, 'local processed image file does not exist')
    selected_source_image = next((option.url for option in processing_before_image.image_options if option.url.startswith('http')), '')
    assert_true(selected_source_image, 'processing item missing public image option')
    try:
        main.process_product_image(main.ProcessImagePayload(product_id=product.id, source_image=selected_source_image))
    except HTTPException as exc:
        assert_true(exc.status_code == 400 and '图生图 API 未配置' in str(exc.detail), 'image processing did not block missing API key')
    else:
        tasks = main.list_product_image_tasks(product.id)
        assert_true(tasks and tasks[0].task_id, 'image processing did not create a provider task')
    generated = main.generate_processing_title(main.GenerateTitlePayload(product_id=product.id))
    assert_true(generated.english_title, 'title generation returned empty title')

    default_processing = main.get_processing_item(product.id).model_copy(
        update={
            'color': 'pending',
            'size': 'pending',
            'sku_code': f'{product.skc}-COLOR-SIZE',
        }
    )
    default_diagnostics = main.upload_item_diagnostics([default_processing])[0]
    assert_true(not default_diagnostics['ready'], 'default pending SKU item should not be upload-ready')
    assert_true('SKU Code 使用默认占位' in default_diagnostics['issues'], 'default SKU placeholder did not block upload')

    with main.connect() as db:
        db.execute(
            'UPDATE products SET main_image = ?, updated_at = ? WHERE id = ?',
            ('https://img.example.com/smoke-main.jpg', main.now_text(), product.id),
        )
        db.commit()

    processing = main.update_processing_item(
        product.id,
        main.ProcessingItemPayload(
            english_title='Smoke Test Product for Temu',
            color='Blue',
            size='S-XL',
            sku_code=f'{product.skc}-BLUE-SXL',
            declared_price=29.99,
            weight_g=420,
            length_cm=20,
            width_cm=15,
            height_cm=4,
            source_url='https://example.com/smoke',
            stock=888,
            ship_days=7,
        ),
    )
    assert_true(processing.english_title == 'Smoke Test Product for Temu', 'processing override did not persist')
    unquoted_product = main.product_from_row(main.connect().execute('SELECT * FROM products WHERE id = ?', (product.id,)).fetchone())
    assert_true(unquoted_product.platform_quote_price == 0 and unquoted_product.gross_margin == 0, 'unquoted product should not show profit margin')
    assert_true(unquoted_product.weight_g == 420, 'processing weight should sync to product library weight')
    main.save_manual_color_image_assignment(main.ManualColorImageAssignmentPayload(product_id=product.id, color='Blue', slot_index=0, image_url='https://img.example.com/blue-a.jpg'))
    main.save_manual_color_image_assignment(main.ManualColorImageAssignmentPayload(product_id=product.id, color='Blue', slot_index=1, image_url='https://img.example.com/blue-b.jpg'))
    main.save_manual_color_image_assignment(main.ManualColorImageAssignmentPayload(product_id=product.id, color='Blue', slot_index=2, image_url='https://img.example.com/blue-c.jpg'))
    main.save_manual_color_image_assignment(main.ManualColorImageAssignmentPayload(product_id=product.id, color='Black', slot_index=0, image_url='https://img.example.com/black-a.jpg'))
    main.save_manual_color_image_assignment(main.ManualColorImageAssignmentPayload(product_id=product.id, color='Black', slot_index=1, image_url='https://img.example.com/black-b.jpg'))
    main.save_manual_color_image_assignment(main.ManualColorImageAssignmentPayload(product_id=product.id, color='Black', slot_index=2, image_url='https://img.example.com/black-c.jpg'))
    main.save_manual_detail_image_assignment(main.ManualDetailImageAssignmentPayload(product_id=product.id, slot_index=0, image_url='https://img.example.com/detail-a.jpg'))
    main.save_manual_detail_image_assignment(main.ManualDetailImageAssignmentPayload(product_id=product.id, slot_index=1, image_url='https://img.example.com/detail-b.jpg'))
    processing = main.get_processing_item(product.id)
    upload_rows = main.miaoshou_rows([processing])
    color_values = {row[5] for row in upload_rows}
    sku_codes = [row[10] for row in upload_rows]
    assert_true({'蓝色', '黑色'}.issubset(color_values), 'miaoshou rows did not expand all colors')
    assert_true(len(sku_codes) == len(set(sku_codes)), 'miaoshou SKU codes are duplicated')
    first_row = upload_rows[0]
    assert_true('detail-a.jpg' in first_row[2] and 'detail-b.jpg' in first_row[2], 'detail images should be written to product description')
    assert_true(all('detail-a.jpg' not in row[18] and 'detail-b.jpg' not in row[18] for row in upload_rows), 'detail images leaked into carousel column')
    blue_rows = [row for row in upload_rows if row[5] == '蓝色']
    black_rows = [row for row in upload_rows if row[5] == '黑色']
    assert_true(blue_rows and all('blue-' in row[18] and 'black-' not in row[18] for row in blue_rows), 'blue carousel column should use blue images only')
    assert_true(black_rows and all('black-' in row[18] and 'blue-' not in row[18] for row in black_rows), 'black carousel column should use black images only')
    main.save_settings([
        main.AppSetting(key='upload_image_source', value='cos'),
        main.AppSetting(key='cos_region', value='ap-guangzhou'),
        main.AppSetting(key='cos_bucket', value='smoke-bucket-123'),
        main.AppSetting(key='cos_prefix', value='temu/smoke'),
    ])
    rpa_root = main.configured_rpa_sku_image_dir()
    product_dir = rpa_root / main.safe_path_segment(processing.skc, f'product_{processing.product_id}')
    for folder, count in [('detail', 2), (main.safe_path_segment('蓝色', 'Color'), 3), (main.safe_path_segment('黑色', 'Color'), 3)]:
        target_dir = product_dir / folder
        target_dir.mkdir(parents=True, exist_ok=True)
        for index in range(1, count + 1):
            (target_dir / f'{index}.jpg').write_bytes(b'smoke')
    cos_rows = main.miaoshou_rows([processing], image_url_resolver=main.cos_export_url_resolver([processing]))
    cos_first_row = cos_rows[0]
    cos_blue_rows = [row for row in cos_rows if row[5] == '蓝色']
    assert_true('https://smoke-bucket-123.cos.ap-guangzhou.myqcloud.com/temu/smoke/' in cos_first_row[2], 'detail description should use COS URLs in COS mode')
    assert_true(all('img.example.com/detail-' not in row[18] for row in cos_rows), 'detail images should not leak into carousel in COS mode')
    assert_true(cos_blue_rows and all('https://smoke-bucket-123.cos.ap-guangzhou.myqcloud.com/temu/smoke/' in row[18] for row in cos_blue_rows), 'carousel should use COS URLs in COS mode')
    assert_true(all('img.example.com/blue-' not in row[18] for row in cos_blue_rows), 'blue source URLs should be replaced by COS URLs')
    assert_true(all('img.example.com' not in str(cell) for row in cos_rows for cell in [row[2], row[18], row[19]]), 'COS export image columns should not contain source image URLs')
    main.save_settings([
        main.AppSetting(key='cos_secret_id', value='smoke-secret-id'),
        main.AppSetting(key='cos_secret_key', value='smoke-secret-key'),
    ])
    uploaded_keys = []
    original_build_cos_client = main.build_cos_client
    try:
        class FakeCosClient:
            def put_object(self, Bucket, Body, Key):
                assert_true(Bucket == 'smoke-bucket-123', 'COS upload bucket mismatch')
                Body.read()
                uploaded_keys.append(Key)

        main.build_cos_client = lambda region, secret_id, secret_key: FakeCosClient()
        sync_result = main.sync_rpa_images_to_cos([processing])
        assert_true(sync_result['uploaded_count'] >= 8, 'COS sync did not upload prepared RPA images')
        assert_true(any(key.endswith('/detail/1.jpg') for key in uploaded_keys), 'COS sync did not upload detail image')
        assert_true(any('/蓝色/' in key and key.endswith('/1.jpg') for key in uploaded_keys), 'COS sync did not upload color image')
    finally:
        main.build_cos_client = original_build_cos_client
    main.save_settings([main.AppSetting(key='upload_image_source', value='')])
    processing.size = 'XL,XS,M,XXL,S,L'
    size_order_rows = main.miaoshou_rows([processing])
    first_color = size_order_rows[0][5]
    ordered_sizes = [row[7] for row in size_order_rows if row[5] == first_color]
    assert_true(ordered_sizes == ['XS', 'S', 'M', 'L', 'XL', 'XXL'], 'miaoshou size rows are not sorted XS to XXL')

    color_mapping = main.spec_mapping_colors()
    assert_true('黑色' in color_mapping['standard_colors'], 'standard color table missing black')
    main.save_spec_alias(main.SpecAliasPayload(kind='color', alias='midnight blue smoke', target='深蓝色'))
    main.save_spec_alias(main.SpecAliasPayload(kind='size', alias='5XL Smoke', target='XXL'))
    assert_true(main.clean_color_label('midnight blue smoke') == '深蓝色', 'custom color alias did not resolve')
    assert_true(main.normalize_upload_size('5XL Smoke') == 'XXL', 'custom size alias did not resolve')
    assert_true(main.delete_spec_alias(main.list_spec_aliases('color')[0].id)['deleted'], 'custom spec alias delete failed')

    freight_path = main.DATA_DIR / 'smoke_freight.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['头程渠道', '重量段（g）', '价格/kg', '附加费'])
    ws.append(['空运', 10000000, 69, 4])
    ws2 = wb.create_sheet('Sheet2')
    ws2.append(['尾程渠道', '重量-英制', '重量（g）', 'zone1', 'zone5'])
    ws2.append(['Temu', '1LB', 454, 38.7, 40.1])
    wb.save(freight_path)
    freight_rules = main.parse_freight_rules(freight_path)
    no_header_freight_path = main.DATA_DIR / 'smoke_freight_no_header.xlsx'
    wb_no_header = Workbook()
    ws_no_header = wb_no_header.active
    ws_no_header.title = 'Sheet1'
    ws_no_header.append(['空运', 10000000, 69, 4])
    ws2_no_header = wb_no_header.create_sheet('Sheet2')
    ws2_no_header.append(['Temu', '1LB', 454, 38.7, 39.2, 39.5, 39.8, 40.1])
    wb_no_header.save(no_header_freight_path)
    no_header_rules = main.parse_freight_rules(no_header_freight_path)
    assert_true(len(no_header_rules['first_mile']) == 1, 'no-header first-mile rules were not parsed')
    assert_true(len(no_header_rules['last_mile']) == 1, 'no-header last-mile rules were not parsed')
    assert_true(no_header_rules['last_mile'][0]['zones']['zone5'] == 40.1, 'no-header zone mapping failed')
    with main.connect() as db:
        main.save_setting_value(db, main.FREIGHT_RULES_SETTING_KEY, json.dumps(freight_rules, ensure_ascii=False))
        main.save_setting_value(db, main.DEFAULT_FREIGHT_ZONE_KEY, 'zone5')
        db.commit()
    assert_true(main.calculate_first_mile(420) == 33, 'first-mile freight calculation failed')
    assert_true(main.calculate_last_mile(420) == 40.1, 'last-mile freight calculation failed')

    manual_rules = main.save_freight_rules(main.FreightRulesPayload(
        default_zone='zone5',
        warehouse_fee=12,
        first_mile=[main.FreightFirstMileRule(channel='Air', max_weight_g=10000000, price_per_kg=70, fixed_fee=5)],
        last_mile=[main.FreightLastMileRule(channel='Temu', max_weight_g=500, zones={'zone5': 31.7})],
    ))
    assert_true(manual_rules['saved'], 'manual freight template was not saved')
    assert_true(main.calculate_first_mile(420) == 35, 'manual first-mile freight calculation failed')
    assert_true(main.calculate_last_mile(420) == 31.7, 'manual last-mile freight calculation failed')
    assert_true(main.default_warehouse_fee() == 12, 'manual warehouse fee was not saved')

    refreshed_processing = main.update_processing_item(
        product.id,
        main.ProcessingItemPayload(
            english_title=processing.english_title,
            color=processing.color,
            size=processing.size,
            sku_code=processing.sku_code,
            declared_price=processing.declared_price,
            weight_g=420,
            length_cm=processing.length_cm,
            width_cm=processing.width_cm,
            height_cm=processing.height_cm,
            source_url=processing.source_url,
            stock=processing.stock,
            ship_days=processing.ship_days,
        ),
    )
    freight_synced_product = main.product_from_row(main.connect().execute('SELECT * FROM products WHERE id = ?', (product.id,)).fetchone())
    assert_true(refreshed_processing.weight_g == 420, 'processing weight refresh failed')
    assert_true(freight_synced_product.first_mile > 0, 'processing weight sync should recalculate first-mile freight')
    assert_true(freight_synced_product.last_mile > 0, 'processing weight sync should recalculate last-mile freight')

    quote_result = main.import_product_quotes_text(main.QuoteImportPayload(text=f'{product.skc}\t120\t420'))
    assert_true(quote_result.updated_count == 1, 'quote import did not update product')
    quoted_product = main.connect().execute('SELECT * FROM products WHERE id = ?', (product.id,)).fetchone()
    assert_true(float(quoted_product['platform_quote_price']) == 120, 'platform quote price was not saved')
    assert_true(int(quoted_product['weight_g']) == 420, 'quote import weight was not saved')
    assert_true(float(quoted_product['first_mile']) == 35, 'quote import did not apply first-mile freight')
    assert_true(float(quoted_product['warehouse_fee']) == 12, 'quote import did not apply warehouse fee')
    assert_true(float(quoted_product['last_mile']) == 31.7, 'quote import did not apply last-mile freight')
    assert_true(float(quoted_product['estimated_profit']) > 0, 'quote import did not recalculate profit')

    temu_quote_path = main.DATA_DIR / 'smoke_temu_quote.xlsx'
    quote_wb = Workbook()
    quote_ws = quote_wb.active
    quote_ws.title = '表格信息'
    quote_ws.append(['SKC ID', 'SKC 货号', '站点', 'SKU ID', 'SKU 货号', '原申报价格', '调整后申报价格'])
    quote_ws.append(['58179930234', '', '美国站', '51674533649', f'{product.skc}-Blue-L', '¥ 299.50', '¥ 114.00'])
    quote_wb.save(temu_quote_path)
    temu_quote_rows = main.parse_quote_rows_from_file(temu_quote_path)
    temu_quote_result = main.apply_quote_rows(temu_quote_rows)
    assert_true(temu_quote_result.updated_count == 1, 'Temu quote export should match product by SKU 货号 prefix')
    temu_quoted_product = main.connect().execute('SELECT * FROM products WHERE id = ?', (product.id,)).fetchone()
    assert_true(float(temu_quoted_product['platform_quote_price']) == 114, 'Temu adjusted declared price was not imported')

    export = main.export_miaoshou()
    assert_true(Path(export['path']).exists(), 'export file does not exist')

    upload_task = main.create_upload_task()
    assert_true(upload_task.total_count >= 1, 'upload task has no items')
    assert_true(upload_task.failed_count == 0, f'upload task has failed items: {upload_task.failed_count}')

    main.save_settings([main.AppSetting(key='rpa_script_dir', value=str(ROOT / 'missing-rpa-script-dir'))])
    real_task = main.run_upload_task()
    assert_true(real_task.status == 'queued_for_executor', 'formal upload should queue for Windows executor instead of running on server')
    assert_true('enable_real_rpa is false' not in real_task.run_log, 'formal upload still blocked by removed safety switch')
    assert_true('Windows executor' in real_task.run_log, 'formal upload did not explain Windows executor queue')
    with main.connect() as db:
        db.execute('UPDATE upload_tasks SET executor_id = ? WHERE id = ?', ('smoke-executor', real_task.id))
        db.commit()
    reported = main.executor_report_task(
        real_task.id,
        main.ExecutorReportPayload(
            executor_id='smoke-executor',
            status='rpa_failed',
            success_count=0,
            failed_count=1,
            run_log='保存失败：Smoke validation error',
            stdout='错误：保存失败，截图 D:\\tmp\\smoke_save.png',
        ),
    )
    assert_true(reported.failure_stage == 'save', 'executor failure stage was not inferred')
    assert_true('保存失败' in reported.failure_reason, 'executor failure reason was not stored')
    assert_true(reported.evidence_path.endswith('smoke_save.png'), 'executor evidence path was not stored')
    executor_work_dir = main.DATA_DIR / 'executor-smoke'
    windows_executor.write_executor_status(executor_work_dir, 'smoke-executor', 'idle', 'smoke ready')
    main.save_settings([main.AppSetting(key='executor_work_dir', value=str(executor_work_dir))])
    status_snapshot = main.executor_status()
    assert_true(status_snapshot.online and status_snapshot.status == 'idle', 'executor status endpoint did not read local status')
    pending_path = windows_executor.write_pending_result(executor_work_dir, real_task.id, {'task_id': real_task.id, 'report': {'status': 'rpa_failed'}})
    assert_true(pending_path.exists() and main.executor_status().pending_result_count >= 1, 'pending executor result was not visible')
    restore_script_dir = previous_settings.get('rpa_script_dir')
    main.save_settings([main.AppSetting(key='rpa_script_dir', value=restore_script_dir or '')])

    with main.connect() as db:
        db.execute('DELETE FROM publish_records')
        db.execute(
            'INSERT INTO publish_records (result, skc, title, reason, created_at) VALUES (?, ?, ?, ?, ?)',
            ('Failed', product.skc, product.title, 'smoke retry failure', main.now_text()),
        )
        db.commit()
    retry_task = main.retry_failed_upload_task(main.RetryFailedUploadPayload())
    assert_true(retry_task.total_count == 1, 'retry failed task did not include failed product')
    assert_true(Path(retry_task.export_path).name.startswith('miaoshou_retry_'), 'retry task did not create retry export')

    status_result = main.update_products_status(main.ProductBulkStatusPayload(ids=[product.id], status='待处理'))
    assert_true(status_result['updated_count'] == 1, 'bulk product status update failed')
    pending_products = main.list_products(status='待处理')
    assert_true(any(item.id == product.id for item in pending_products), 'product status filter should include 待处理 products')
    with main.connect() as db:
        db.execute("UPDATE products SET status = ? WHERE id = ?", ('pending_main_image', product.id))
        db.commit()
    pending_alias_products = main.list_products(status='待处理')
    missing_image_status_products = main.list_products(status='待补主图')
    assert_true(any(item.id == product.id for item in pending_alias_products), '待处理 filter should include pending_main_image alias')
    assert_true(any(item.id == product.id for item in missing_image_status_products), '待补主图 filter should include pending_main_image alias')

    temp_task = main.create_upload_task()
    delete_result = main.delete_upload_task(temp_task.id)
    assert_true(delete_result['ok'], 'delete upload task failed')
    cleanup_result = main.cleanup_local_data(main.CleanupPayload(retention_days=7))
    assert_true('deleted_exports' in cleanup_result and 'deleted_logs' in cleanup_result, 'local cleanup result missing counters')

    json_path = Path('data/smoke_external_result.json')
    json_path.write_text(
        json.dumps([
            {
                'title': 'Smoke External JSON Candidate',
                'url': 'https://example.com/smoke-json-candidate',
                'price': 18.8,
                'sales': 999,
                'image_count': 8,
            }
        ], ensure_ascii=False),
        encoding='utf-8',
    )
    rows = main.parse_collection_rows(json_path)
    inserted, skipped = main.insert_collection_rows(rows, 'SmokeJSON')
    assert_true(inserted + skipped >= 1, 'json import path did not process rows')

    main.save_settings(
        [
            main.AppSetting(key='enable_external_collection', value=previous_settings.get('enable_external_collection', 'false')),
            main.AppSetting(key='collection_mode', value=previous_settings.get('collection_mode', '1688')),
        ]
    )

    print('SMOKE_OK')


if __name__ == '__main__':
    main_smoke()
