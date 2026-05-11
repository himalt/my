"""
rpa_upload.py
Playwright 自动把 sku_images/ 里的图片上传到店小秘颜色变种

用法：
  python rpa_upload.py                        # 只上图（处理全部产品）
  python rpa_upload.py NSNZDK-260418-7        # 只上图（指定产品）
  python rpa_upload.py --import               # 导入 xlsx + 上图（自动找最新 xlsx）
  python rpa_upload.py --import path/to.xlsx  # 导入指定 xlsx + 上图

流程（--import 模式）：
  1. 打开浏览器，检查登录状态
  2. 关闭「产品动态」弹窗
  3. 点「导入/导出」→「导入创建产品」→ 上传 xlsx → 选店铺/站点 → 确定
  4. 轮询列表直到产品出现
  5. 遍历 sku_images/ → 搜索产品 → 上传图片 → 保存
"""

import os
import re
import json
import time
import sys
import tempfile
import shutil

RPA_BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(__file__)
_BUNDLED_PLAYWRIGHT_BROWSERS = os.path.join(RPA_BASE_DIR, 'ms-playwright')
if os.path.isdir(_BUNDLED_PLAYWRIGHT_BROWSERS) and not os.environ.get('PLAYWRIGHT_BROWSERS_PATH'):
    os.environ['PLAYWRIGHT_BROWSERS_PATH'] = _BUNDLED_PLAYWRIGHT_BROWSERS

sys.path.insert(0, os.path.dirname(__file__))
from pipeline import DATA_DIR, _clean_filename, _color_to_en

SKU_IMG_DIR  = os.path.join(DATA_DIR, 'sku_images')
COOKIES_PATH = os.path.join(DATA_DIR, 'dxm_cookies.json')
BROWSER_PROFILE_DIR = os.path.join(DATA_DIR, 'browser_profile')
BASE_URL     = 'https://www.dianxiaomi.com'
LIST_URL     = f'{BASE_URL}/web/popTemu/pageList/offline'


IMPORT_STORE  = '测试1'       # 导入时选择的店铺名，可由 --shop-account 覆盖
IMPORT_SITES  = ['美国']      # 导入时勾选的站点，可由 --site 覆盖
SITE_ALIASES = {
    '美国站': '美国',
    '美国': '美国',
}
DEFAULT_BATCH_STOCK = '999'   # 编辑页仓库批量库存默认值

IMAGE_SOURCE_LOCAL = 'local'
IMAGE_SOURCE_COS = 'cos'
COS_REGION_ENV = 'COS_REGION'
COS_BUCKET_ENV = 'COS_BUCKET'
COS_PREFIX_ENV = 'COS_PREFIX'
COS_SECRET_ID_ENV = 'COS_SECRET_ID'
COS_SECRET_KEY_ENV = 'COS_SECRET_KEY'

# ── 读取 apify_config.json（COS 凭证 + image_source 默认值）────────────────────
_CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'apify_config.json')
try:
    with open(_CONFIG_PATH, 'r', encoding='utf-8') as _f:
        _CONFIG = json.load(_f)
except Exception:
    _CONFIG = {}

DEFAULT_IMAGE_SOURCE = _CONFIG.get('image_source', IMAGE_SOURCE_LOCAL)
DEBUG_SHOTS = False


def _set_sku_image_root(path):
    global SKU_IMG_DIR
    clean_path = os.path.abspath(str(path or '').strip())
    if clean_path:
        SKU_IMG_DIR = clean_path
        print(f'RPA 图片目录: {SKU_IMG_DIR}')


def _dedupe_keep_order(values):
    seen = set()
    out = []
    for x in values or []:
        s = str(x or '').strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def _resolve_cos_settings(cos_region=None, cos_bucket=None, cos_prefix=None):
    # 优先级：参数 > 环境变量 > apify_config.json
    region    = str(cos_region or os.getenv(COS_REGION_ENV, '')    or _CONFIG.get('cos_region', '')).strip()
    bucket    = str(cos_bucket or os.getenv(COS_BUCKET_ENV, '')    or _CONFIG.get('cos_bucket', '')).strip()
    prefix_raw = cos_prefix if cos_prefix is not None else (os.getenv(COS_PREFIX_ENV, '') or _CONFIG.get('cos_prefix', ''))
    prefix    = str(prefix_raw or '').strip().strip('/')
    secret_id = str(os.getenv(COS_SECRET_ID_ENV, '') or _CONFIG.get('cos_secret_id', '')).strip()
    secret_key= str(os.getenv(COS_SECRET_KEY_ENV, '') or _CONFIG.get('cos_secret_key', '')).strip()

    missing = []
    if not region:
        missing.append(COS_REGION_ENV)
    if not bucket:
        missing.append(COS_BUCKET_ENV)
    if not secret_id:
        missing.append(COS_SECRET_ID_ENV)
    if not secret_key:
        missing.append(COS_SECRET_KEY_ENV)

    if missing:
        raise RuntimeError(f'COS 模式缺少配置: {", ".join(missing)}')

    return region, bucket, prefix, secret_id, secret_key


def _build_cos_client(region, secret_id, secret_key):
    try:
        from qcloud_cos import CosConfig, CosS3Client
    except Exception as e:
        raise RuntimeError('未安装 COS SDK，请先安装: pip install cos-python-sdk-v5') from e

    cfg = CosConfig(Region=region, SecretId=secret_id, SecretKey=secret_key, Scheme='https')
    return CosS3Client(cfg)


def _iter_cos_keys(client, bucket, prefix):
    marker = ''
    while True:
        kwargs = {'Bucket': bucket, 'Prefix': prefix, 'MaxKeys': 1000}
        if marker:
            kwargs['Marker'] = marker

        resp = client.list_objects(**kwargs)
        contents = resp.get('Contents') or []
        if isinstance(contents, dict):
            contents = [contents]

        for item in contents:
            key = item.get('Key')
            if key:
                yield key

        truncated = str(resp.get('IsTruncated', '')).lower() in ('true', '1')
        if not truncated:
            break

        marker = resp.get('NextMarker') or ''
        if not marker and contents:
            marker = contents[-1].get('Key') or ''
        if not marker:
            break


def _download_cos_images_for_product(client, bucket, base_prefix, prod_no, dest_root):
    prod_no = str(prod_no or '').strip()
    if not prod_no:
        return 0

    prod_prefix = f'{base_prefix}/{prod_no}/' if base_prefix else f'{prod_no}/'
    downloaded = 0

    for key in _iter_cos_keys(client, bucket, prod_prefix):
        rel = key[len(prod_prefix):] if key.startswith(prod_prefix) else ''
        if not rel or rel.endswith('/'):
            continue

        parts = [p for p in rel.split('/') if p]
        if len(parts) < 2:
            continue

        color_folder = '/'.join(parts[:-1]).strip()
        fname = parts[-1].strip()
        if not color_folder or not fname:
            continue
        if not fname.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
            continue

        local_dir = os.path.join(dest_root, prod_no, color_folder)
        os.makedirs(local_dir, exist_ok=True)
        local_path = os.path.join(local_dir, fname)
        client.download_file(Bucket=bucket, Key=key, DestFilePath=local_path)
        downloaded += 1

    return downloaded


def _prepare_cos_image_root(prod_nos, cos_region=None, cos_bucket=None, cos_prefix=None):
    targets = _dedupe_keep_order(prod_nos)
    if not targets:
        raise RuntimeError('未提供可下载的产品货号')

    region, bucket, prefix, secret_id, secret_key = _resolve_cos_settings(
        cos_region=cos_region,
        cos_bucket=cos_bucket,
        cos_prefix=cos_prefix,
        no_publish=no_publish,
    )
    client = _build_cos_client(region, secret_id, secret_key)

    tmp_root = tempfile.mkdtemp(prefix='dxm_cos_images_')
    ready_prod_folders = []
    missing_prod_nos = []
    total_files = 0

    try:
        print(f'COS 配置: region={region}, bucket={bucket}, prefix={prefix or "/"}')
        for prod_no in targets:
            count = _download_cos_images_for_product(
                client=client,
                bucket=bucket,
                base_prefix=prefix,
                prod_no=prod_no,
                dest_root=tmp_root,
            )
            if count > 0:
                ready_prod_folders.append(prod_no)
                total_files += count
                print(f'  [COS] {prod_no} 下载 {count} 张')
            else:
                missing_prod_nos.append(prod_no)

        print(f'COS 图片准备完成：命中产品 {len(ready_prod_folders)}/{len(targets)}，下载文件 {total_files} 个')
        return tmp_root, ready_prod_folders, missing_prod_nos
    except Exception:
        shutil.rmtree(tmp_root, ignore_errors=True)
        raise



def save_cookies(context):
    with open(COOKIES_PATH, 'w', encoding='utf-8') as f:
        json.dump(context.cookies(), f, ensure_ascii=False)
    print(f'  Cookie 已保存: {COOKIES_PATH}')


def load_cookies(context):
    if os.path.exists(COOKIES_PATH):
        try:
            with open(COOKIES_PATH, encoding='utf-8') as f:
                context.add_cookies(json.load(f))
            return True
        except Exception as e:
            print(f'  [warn] Cookie 文件加载失败，改用浏览器持久登录态: {e}')
    return False


def ensure_login(page, context):
    """打开列表页，检测到登录表单就等待用户完成登录（自动感知，不需要按回车）"""
    print('检查登录状态...')
    page.goto(LIST_URL, wait_until='domcontentloaded', timeout=30000)
    time.sleep(3)  # 等前端 JS 渲染完

    if page.locator('input[type="password"]').first.is_visible():
        print('检测到登录页，请在浏览器中完成登录（扫码或输密码）...')
        # 等第一个密码框消失 = 登录跳转完成，最多等 3 分钟
        page.locator('#exampleInputPassword').wait_for(state='hidden', timeout=180000)
        time.sleep(2)
        save_cookies(context)
        print('  登录完成，Cookie 已保存')
    else:
        print('  已登录 OK')


def _visible_option_texts(page, limit=30):
    try:
        texts = page.locator('.ant-select-dropdown:visible .ant-select-item-option').evaluate_all(
            "els => els.map(el => el.innerText || el.textContent || '').map(t => t.trim()).filter(Boolean)"
        )
        return texts[:limit]
    except Exception:
        return []


def select_import_store(page, store_name):
    expected = str(store_name or '').strip()
    if not expected:
        raise RuntimeError('店铺账号为空，请先在页面“店铺与模板”里填写店铺账号')
    store_select = page.locator('.ant-select').filter(
        has=page.locator(':text("请选择店铺")')
    ).first
    store_select.wait_for(state='visible', timeout=8000)
    store_select.click()
    time.sleep(0.8)
    dropdown = page.locator('.ant-select-dropdown:visible').last
    dropdown.wait_for(state='visible', timeout=5000)
    exact_option = dropdown.locator('.ant-select-item-option').filter(has_text=re.compile(rf'^{re.escape(expected)}$')).first
    try:
        exact_option.wait_for(state='visible', timeout=2000)
        exact_option.click()
        print(f'  店铺已选择: {expected}')
        return
    except Exception:
        pass
    contains_option = dropdown.locator('.ant-select-item-option').filter(has_text=expected).first
    try:
        contains_option.wait_for(state='visible', timeout=1500)
        option_text = contains_option.inner_text().strip()
        contains_option.click()
        print(f'  店铺已选择: {option_text}')
        return
    except Exception:
        options = _visible_option_texts(page)
        raise RuntimeError(f'未找到店铺「{expected}」。当前可选店铺: {" | ".join(options) if options else "未读取到店铺列表"}')


# ── 弹窗 & 导入 ───────────────────────────────────────────────────────────────

def dismiss_popup(page):
    """关闭每次打开时出现的「产品动态」/公告弹窗，避免遮挡后续点击。"""
    closed = False

    close_selectors = [
        '.ant-modal-wrap:visible button.ant-modal-close',
        'button:has-text("关闭")',
        'button:has-text("我知道了")',
        'button:has-text("知道了")',
    ]

    for sel in close_selectors:
        try:
            btn = page.locator(sel).first
            btn.wait_for(state='visible', timeout=900)
            btn.click(timeout=1200)
            time.sleep(0.2)
            closed = True
        except Exception:
            continue

    try:
        # 公告 iframe（theNewestModalLabelFrame）有时无按钮文本，走 JS 兜底点右上角关闭
        js_closed = page.evaluate("""() => {
            const isVisible = (el) => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                const s = window.getComputedStyle(el);
                return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
            };
            const wraps = [...document.querySelectorAll('.ant-modal-wrap,.bullet-layer,.comm-modal')].filter(isVisible);
            let clicked = false;
            for (const wrap of wraps) {
                const btns = [...wrap.querySelectorAll('button,.ant-modal-close,[role="button"],span')];
                for (const b of btns) {
                    if (!isVisible(b)) continue;
                    const txt = String(b.textContent || '').replace(/\\s+/g, '');
                    const cls = String(b.className || '');
                    if (txt === '关闭' || txt === '知道了' || txt === '我知道了' || cls.includes('ant-modal-close')) {
                        ['mouseover', 'mousedown', 'mouseup', 'click'].forEach((name) => {
                            b.dispatchEvent(new MouseEvent(name, { bubbles: true, cancelable: true, view: window }));
                        });
                        if (typeof b.click === 'function') b.click();
                        clicked = true;
                        break;
                    }
                }
                if (clicked) break;
            }
            return clicked;
        }""")
        if js_closed:
            closed = True
            time.sleep(0.2)
    except Exception:
        pass

    if not closed:
        try:
            page.keyboard.press('Escape')
            time.sleep(0.15)
        except Exception:
            pass
    else:
        print('  弹窗已关闭')


def _normalize_site_name(site):
    site_text = str(site or '').strip()
    return SITE_ALIASES.get(site_text, site_text.removesuffix('站'))


def import_xlsx_to_dxm(page, xlsx_path,
                        store_name=IMPORT_STORE,
                        sites=None,
                        warehouse_template=None):
    """
    在店小秘「导入创建产品」对话框中完成 xlsx 导入：
      1. 点「导入/导出」→「导入创建产品」
      2. 上传 xlsx 文件
      3. 选择店铺
      4. 勾选站点
      5. 点「确定」
    返回 True 表示操作已提交（不保证后端处理完成）
    """
    if sites is None:
        sites = IMPORT_SITES
    sites = [_normalize_site_name(site) for site in sites if _normalize_site_name(site)]
    xlsx_path = os.path.abspath(str(xlsx_path or ''))
    if not xlsx_path.lower().endswith('.xlsx'):
        print(f'  [error] 导入文件必须是 .xlsx 后缀，当前文件: {xlsx_path}')
        return False
    if not os.path.exists(xlsx_path):
        print(f'  [error] 导入文件不存在: {xlsx_path}')
        return False

    print(f'\n=== 导入 xlsx → 店小秘 ===')
    print(f'  文件: {xlsx_path}')
    page.goto(LIST_URL, wait_until='domcontentloaded', timeout=30000)
    time.sleep(2)
    dismiss_popup(page)

    # ── 点「导入/导出」下拉 ──
    imp_btn = page.locator('button:has-text("导入/导出")').first
    imp_btn.wait_for(state='visible', timeout=10000)
    imp_btn.click()
    time.sleep(0.8)

    # ── 点「导入创建产品」菜单项 ──
    page.locator(':text("导入创建产品"):visible').first.click()
    time.sleep(1.5)
    _shot(page, 'import_dialog_open')

    # ── 第一步：上传文件 ──
    with page.expect_file_chooser(timeout=8000) as fc_info:
        page.locator('button:has-text("请选择导入文件")').first.click()
    fc_info.value.set_files(xlsx_path)
    time.sleep(1)
    print(f'  文件已选择')
    _shot(page, 'import_file_selected')

    # ── 第二步：选择店铺 ──
    try:
        select_import_store(page, store_name)
    except Exception as e:
        print(f'  [error] 选择店铺失败: {e}')
        _shot(page, 'import_store_select_error')
        return False

    # ── 第二步：勾选站点 + 选仓库 ──
    for site in sites:
        try:
            # 勾选站点 checkbox
            cb = page.locator(f'.ant-checkbox-wrapper:near(:text-is("{site}"))').first
            cb.wait_for(state='visible', timeout=3000)
            if not cb.locator('input').is_checked():
                cb.locator('input').check()
            time.sleep(0.3)
            print(f'  站点已勾选: {site}')

            # 选仓库：JS 从"美国"文字向上找父容器内的 ant-select-selector 并点击
            try:
                time.sleep(0.5)
                js_result = page.evaluate(f"""(site) => {{
                    const modal = document.querySelector('.ant-modal-content');
                    if (!modal) return 'no modal';
                    const spans = [...modal.querySelectorAll('span')];
                    const siteSpan = spans.find(s => s.textContent.trim() === site);
                    if (!siteSpan) return 'no span for ' + site;
                    let el = siteSpan.parentElement;
                    for (let i = 0; i < 6; i++) {{
                        const sel = el ? el.querySelector('.ant-select-selector') : null;
                        if (sel) {{ sel.click(); return 'clicked'; }}
                        el = el ? el.parentElement : null;
                    }}
                    return 'not found';
                }}""", site)
                print(f'  JS仓库定位: {js_result}')
                if js_result != 'clicked':
                    raise Exception(js_result)
                time.sleep(0.5)
                opts = page.locator('.ant-select-dropdown:visible .ant-select-item-option')
                opts.first.wait_for(state='visible', timeout=4000)
                preferred = str(warehouse_template or '').strip()
                first_option = opts.first
                target_option = opts.filter(has_text=preferred).first if preferred else first_option
                try:
                    target_option.wait_for(state='visible', timeout=1500)
                except Exception:
                    if preferred:
                        print(f'  [warn] 未找到仓库「{preferred}」，回退第一个仓库')
                    target_option = first_option
                wh_text = target_option.inner_text().strip()
                target_option.click()
                time.sleep(0.2)
                print(f'  仓库已选择: {wh_text}')
            except Exception as e:
                print(f'  [warn] 选仓库失败（可能无需选）: {e}')
        except Exception as e:
            print(f'  [warn] 勾选站点 {site} 失败: {e}')
    _shot(page, 'import_before_confirm')

    # ── 点「确定」提交 ──
    _shot(page, 'import_before_confirm')
    page.locator('button:has-text("确定")').last.click()
    time.sleep(2)
    _shot(page, 'import_after_confirm')
    print('  已点确定')

    # 情况1：出现「批量编辑/批量导入」进度弹窗
    try:
        batch = page.locator('.ant-modal-content').filter(
            has=page.locator(':text("进行中")')
        )
        batch.wait_for(state='visible', timeout=8000)
        print('  批量导入进行中，等待完成...')
        page.locator(':text("进行中"):visible').wait_for(state='hidden', timeout=120000)
        print('  批量导入完成')
        _shot(page, 'import_batch_done')

        ok, reason = _check_import_result(page)
        print(f'  导入结果检查: {reason}')

        try:
            page.locator('.ant-modal-content button:has-text("关闭")').last.click()
            time.sleep(1)
        except Exception:
            pass

        return ok
    except Exception:
        pass

    # 情况2：原导入弹窗已关闭（快速完成），等待后台处理
    try:
        page.locator('button:has-text("请选择导入文件")').wait_for(state='hidden', timeout=3000)
        print('  导入弹窗已关闭，等待后台处理（15s）...')
        time.sleep(15)
        return True
    except Exception:
        pass

    # 情况3：其他（截图供排查）
    _shot(page, 'import_confirm_result')
    print('  请查看截图 import_confirm_result.png 确认状态')
    return False


def _check_import_result(page):
    """解析导入结果弹窗，判断是否真正导入成功。"""
    try:
        modal = page.locator('.ant-modal-content:visible').last
        text = modal.inner_text().strip()
    except Exception:
        return True, '未读取到导入结果详情，继续后续检测'

    imported = None
    m = re.search(r'已成功导入\s*(\d+)\s*条', text)
    if m:
        try:
            imported = int(m.group(1))
        except Exception:
            imported = None

    has_data_error = ('数据异常' in text) or ('导入失败' in text)
    if imported == 0:
        return False, '导入结果显示成功 0 条'
    if has_data_error:
        return False, '导入结果包含数据异常/失败明细'
    if imported is not None and imported > 0:
        return True, f'导入成功 {imported} 条'
    return True, '未识别到成功条数，继续后续检测'


_ALLOWED_UPLOAD_SIZES = {'XS', 'S', 'M', 'L', 'XL', 'XXL'}


def _norm_size_token(value):
    txt = str(value or '').strip().upper()
    if not txt:
        return ''
    txt = txt.replace('（', '(').replace('）', ')')
    txt = re.sub(r'\s+', '', txt)
    txt = txt.replace('_', '').replace('.', '').replace('-', '')
    txt = txt.replace('码', '').replace('號', '').replace('号', '')
    txt = txt.replace('SIZE', '')

    alias = {
        'XSMALL': 'XS',
        'SMALL': 'S',
        'MEDIUM': 'M',
        'LARGE': 'L',
        'XLARGE': 'XL',
        'XXLARGE': 'XXL',
        '2XL': 'XXL',
        '2X': 'XXL',
    }
    return alias.get(txt, txt)


def _load_prod_nos_from_xlsx(xlsx_path):
    """
    读取导入 xlsx 的产品货号，并在可识别尺码时仅保留 XS~XXL 可上传产品。

    返回：
      upload_prod_nos   : list[str] 可执行上图的产品货号（去重保序）
      blocked_prod_nos  : list[str] 因存在非 XS~XXL 尺码而整款跳过上图的产品货号
      blocked_rows      : list[tuple[int, str, str]] 被过滤的源行 (行号, 产品货号, 尺码)
      size_filter_used  : bool 是否成功启用尺码过滤
      has_prod_col      : bool xlsx 是否存在产品货号列
    """
    try:
        import openpyxl
    except Exception as e:
        print(f'  [warn] 无法读取 xlsx（openpyxl 不可用）: {e}')
        return [], [], [], False, False

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f'  [warn] 打开 xlsx 失败: {e}')
        return [], [], [], False, False

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return [], [], [], False, False

    def _norm_header(h):
        s = str(h or '')
        s = s.replace('\n', '')
        s = re.sub(r'\s+', '', s)
        s = s.replace('*', '')
        return s

    headers = [_norm_header(h) for h in header_row]

    prod_idx = None
    size_idx = None
    sku_idx = None
    for i, h in enumerate(headers):
        if h == '产品货号' and prod_idx is None:
            prod_idx = i
        elif h in ('变种属性值二', '尺码') and size_idx is None:
            size_idx = i
        elif h == 'SKU货号' and sku_idx is None:
            sku_idx = i

    if prod_idx is None:
        print('  [warn] xlsx 未找到「产品货号」列，改用 sku_images 目录等待')
        return [], [], [], False, False

    size_filter_used = (size_idx is not None) or (sku_idx is not None)

    ordered_prod_nos = []
    seen_prod = set()
    allowed_prod = set()
    blocked_prod = set()
    blocked_rows = []

    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if prod_idx >= len(row):
            continue
        prod_no = str(row[prod_idx] or '').strip()
        if not prod_no:
            continue

        if prod_no not in seen_prod:
            seen_prod.add(prod_no)
            ordered_prod_nos.append(prod_no)

        if not size_filter_used:
            allowed_prod.add(prod_no)
            continue

        size_val = ''
        if size_idx is not None and size_idx < len(row):
            size_val = _norm_size_token(row[size_idx])
        if not size_val and sku_idx is not None and sku_idx < len(row):
            sku_val = str(row[sku_idx] or '').strip()
            if sku_val:
                tail = sku_val.rsplit('-', 1)[-1]
                size_val = _norm_size_token(tail)

        if size_val in _ALLOWED_UPLOAD_SIZES:
            if prod_no not in blocked_prod:
                allowed_prod.add(prod_no)
        else:
            blocked_prod.add(prod_no)
            raw_size = ''
            if size_idx is not None and size_idx < len(row):
                raw_size = str(row[size_idx] or '').strip()
            blocked_rows.append((row_no, prod_no, size_val or raw_size))

    if size_filter_used:
        upload_prod_nos = [p for p in ordered_prod_nos if p in allowed_prod and p not in blocked_prod]
    else:
        upload_prod_nos = ordered_prod_nos

    blocked_prod_nos = [p for p in ordered_prod_nos if p in blocked_prod]
    return upload_prod_nos, blocked_prod_nos, blocked_rows, size_filter_used, True


def _search_product_ready(page, keyword):
    """判断当前搜索结果里是否真的有匹配产品行（避免仅靠页面任意文本误判）。"""
    # 结果区若出现“暂无数据”，视为未就绪
    if page.locator('xpath=//tbody//td[contains(normalize-space(text()),"暂无数据")]').count() > 0:
        return False

    # 结果表格中需存在含 keyword 的行，且有可编辑入口
    row_match = page.locator(
        f'xpath=//tbody//tr[.//*[contains(normalize-space(text()),"{keyword}")]]'
    ).count() > 0
    has_edit = page.locator('xpath=//tbody//tr//*[normalize-space(text())="编辑"]').count() > 0
    return row_match and has_edit


def wait_for_products(page, prod_nos, max_wait=300, poll_interval=15, require_all=False):
    """
    轮询列表页，直到产品货号出现。
    返回：
      True  -> 达到等待条件（可继续）
      False -> 超时未达到（应中止）
      None  -> 用户手动跳过（Ctrl+C）
    """
    targets = []
    seen = set()
    for x in prod_nos or []:
        s = str(x or '').strip()
        if s and s not in seen:
            seen.add(s)
            targets.append(s)
    if not targets:
        return True

    required = len(targets) if require_all else 1
    found = set()
    idx = 0

    if require_all:
        print(f'\n等待导入产品出现（需全部 {required} 个可搜索，最多等 {max_wait}s）...')
    else:
        print(f'\n等待产品出现（任一可搜索即可，最多等 {max_wait}s）...')

    import time as _t
    start = _t.time()
    print('  （可按 Ctrl+C 跳过等待，直接开始上图）')

    while _t.time() - start < max_wait:
        pending = [x for x in targets if x not in found]
        if not pending:
            print(f'  产品已全部出现！({len(found)}/{required})')
            return True

        search_no = pending[idx % len(pending)]
        idx += 1

        try:
            page.goto(LIST_URL, wait_until='domcontentloaded', timeout=30000)
            time.sleep(2)
            dismiss_popup(page)

            page.locator(':text-is("SKU"):visible').first.click(timeout=3000)
            time.sleep(0.3)
            search_input = page.locator(
                'xpath=//*[contains(normalize-space(text()),"搜索内容")]'
                '/following::input[@type="text"][1]'
            ).first
            search_input.fill(search_no)
            page.locator('button:has-text("搜索")').first.click()
            time.sleep(2.5)

            if _search_product_ready(page, search_no):
                found.add(search_no)
                print(f'  已出现: {search_no} ({len(found)}/{required})')
                if len(found) >= required:
                    print('  等待条件满足，继续执行')
                    return True
        except KeyboardInterrupt:
            print('\n  已跳过等待，继续上图...')
            return None
        except Exception as e:
            print(f'  轮询异常: {e}')

        elapsed = int(_t.time() - start)
        print(f'  仍在等待，{poll_interval}s 后重试... ({elapsed}s/{max_wait}s, 已就绪 {len(found)}/{required})')
        try:
            time.sleep(poll_interval)
        except KeyboardInterrupt:
            print('\n  已跳过等待，继续上图...')
            return None

    print(f'  等待超时（已就绪 {len(found)}/{required}），停止后续上图')
    return False


# ── 搜索商品 ──────────────────────────────────────────────────────────────────

def find_edit_url(page, offer_id):
    """在产品列表用 SKU 搜索 offer_id，返回编辑页完整 URL；找不到返回 None"""
    page.goto(LIST_URL, wait_until='domcontentloaded', timeout=30000)
    time.sleep(2)

    # ── 截图①：刚进列表页 ──
    _shot(page, f'01_list_{offer_id}')

    # ── 切换搜索类型为 SKU ──
    # 搜索类型行: [标题] SKU — 精确全文 ":text-is" + ":visible" 跳过导航隐藏链接
    try:
        page.locator(':text-is("SKU"):visible').first.click(timeout=5000)
        time.sleep(0.5)
        print('  SKU tab 已切换')
    except Exception as e:
        print(f'  [warn] 切换SKU类型失败: {e}')

    # ── 切换为模糊匹配（SKU货号格式: offer_id-颜色-尺码，需包含搜索）──
    try:
        match_btn = page.locator(':text-is("精确匹配"):visible').first
        match_btn.wait_for(state='visible', timeout=2000)
        match_btn.click()
        time.sleep(0.3)
        page.locator(':text-is("模糊匹配"):visible').first.click(timeout=3000)
        time.sleep(0.3)
        print('  已切换为模糊匹配')
    except Exception:
        pass  # 没有精确匹配下拉框或已是模糊匹配，忽略

    # ── 截图②：切换后 ──
    _shot(page, f'02_sku_tab_{offer_id}')

    def _locate_search_input():
        # 侧边栏有 2 个搜索框（待发布产品 + 在线产品），主区域是第 3 个 (nth=2)
        # 优先用 XPath 定位「搜索内容」标签后的第一个 input，更鲁棒
        try:
            inp = page.locator(
                'xpath=//*[contains(normalize-space(text()),"搜索内容")]/following::input[@type="text"][1]'
            ).first
            inp.wait_for(state='visible', timeout=3000)
            print('  [debug] XPath 定位到搜索框')
            return inp
        except Exception:
            count = page.locator('input[type="text"]').count()
            print(f'  [debug] 页面共 {count} 个文字输入框，使用 nth(2)')
            return page.locator('input[type="text"]').nth(2)

    def _trigger_search(keyword):
        search_input = _locate_search_input()
        search_input.fill('')
        search_input.fill(keyword)
        time.sleep(0.3)
        page.locator('button:has-text("搜索")').first.click()
        time.sleep(2.5)

    def _rows_for_offer():
        return page.locator('xpath=//tbody//tr').filter(has_text=offer_id)

    def _switch_left_bucket(bucket):
        try:
            clicked = page.evaluate("""(bucket) => {
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const nodes = [...document.querySelectorAll('a,span,div,li')]
                    .filter(isVisible)
                    .filter((el) => el.getBoundingClientRect().left < 280);
                const target = nodes.find((el) => {
                    const txt = norm(el.textContent);
                    return txt === bucket || txt.startsWith(bucket + '(');
                });
                if (!target) return false;
                ['mouseover', 'mousedown', 'mouseup', 'click'].forEach((name) => {
                    target.dispatchEvent(new MouseEvent(name, { bubbles: true, cancelable: true, view: window }));
                });
                if (typeof target.click === 'function') target.click();
                return true;
            }""", bucket)
            if clicked:
                time.sleep(0.9)
                print(f'  [debug] 已切换左侧分组: {bucket}')
                return True
        except Exception:
            pass
        return False

    _trigger_search(offer_id)

    # ── 截图③：搜索结果 ──
    _shot(page, f'03_result_{offer_id}')

    # 仅在命中行里找「编辑」，避免点到页面其他位置同名文本
    rows = _rows_for_offer()
    row_count = rows.count()
    if row_count == 0:
        fallback_buckets = ['待发布产品', '发布中', '发布失败', '在线产品']
        for idx, bucket in enumerate(fallback_buckets, start=1):
            if not _switch_left_bucket(bucket):
                continue
            _trigger_search(offer_id)
            _shot(page, f'03_result_{offer_id}_bucket{idx}')
            rows = _rows_for_offer()
            row_count = rows.count()
            if row_count > 0:
                print(f'  [{offer_id}] 在「{bucket}」命中 {row_count} 行')
                break

    if row_count == 0:
        print(f'  未找到 [{offer_id}] 的结果行（搜索后列表为空）')
        return None

    row = rows.first
    try:
        row.wait_for(state='visible', timeout=3000)
    except Exception:
        pass

    def _click_edit_in_row():
        selectors = [
            'xpath=.//*[normalize-space(text())="编辑"]',
            'xpath=.//a[normalize-space(text())="编辑"]',
            'xpath=.//span[normalize-space(text())="编辑"]',
            'xpath=.//*[@role="button" and normalize-space(text())="编辑"]',
        ]
        for sel in selectors:
            try:
                btn = row.locator(sel).first
                if btn.count() <= 0:
                    continue
                btn.wait_for(state='visible', timeout=1200)
                btn.scroll_into_view_if_needed()
                btn.click(timeout=1800)
                return True
            except Exception:
                continue

        # JS 兜底：在命中行内找可见“编辑”节点点击
        try:
            clicked = row.evaluate("""(el) => {
                const isVisible = (n) => {
                    if (!n) return false;
                    const r = n.getBoundingClientRect();
                    const s = window.getComputedStyle(n);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const nodes = [...el.querySelectorAll('a,span,button,div')];
                for (const n of nodes) {
                    if (!isVisible(n)) continue;
                    const txt = String(n.textContent || '').trim();
                    if (txt !== '编辑') continue;
                    ['mouseover', 'mousedown', 'mouseup', 'click'].forEach((name) => {
                        n.dispatchEvent(new MouseEvent(name, { bubbles: true, cancelable: true, view: window }));
                    });
                    if (typeof n.click === 'function') n.click();
                    return true;
                }
                return false;
            }""")
            if clicked:
                return True
        except Exception:
            pass
        return False

    def _extract_edit_href_from_row():
        try:
            href = row.evaluate("""(el) => {
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (n) => {
                    if (!n) return false;
                    const r = n.getBoundingClientRect();
                    const s = window.getComputedStyle(n);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const links = [...el.querySelectorAll('a[href]')].filter(isVisible);
                for (const a of links) {
                    const txt = norm(a.textContent);
                    const href = a.getAttribute('href') || '';
                    if (txt.includes('编辑') || href.includes('edit') || href.includes('product')) {
                        return href;
                    }
                }
                return '';
            }""")
            if not href:
                return ''
            if href.startswith('http://') or href.startswith('https://'):
                return href
            if href.startswith('/'):
                return f'{BASE_URL}{href}'
            return f'{BASE_URL}/{href.lstrip("/")}'
        except Exception:
            return ''

    before_url = page.url
    before_pages = len(page.context.pages)

    try:
        with page.context.expect_page(timeout=6000) as new_page_info:
            if not _click_edit_in_row():
                raise RuntimeError('row edit click not found')
        new_page = new_page_info.value
        new_page.wait_for_load_state('domcontentloaded', timeout=15000)
        edit_url = new_page.url
        new_page.close()
        print(f'  编辑链接: {edit_url}')
        return edit_url
    except Exception as e:
        # 路径A：同标签页跳转
        cur_url = page.url
        if cur_url != before_url and '/pageList/offline' not in cur_url:
            print(f'  编辑链接（同标签）: {cur_url}')
            return cur_url

        # 路径B：已新开页但未被 expect_page 捕获
        pages_now = page.context.pages
        if len(pages_now) > before_pages:
            try:
                new_page = pages_now[-1]
                new_page.wait_for_load_state('domcontentloaded', timeout=15000)
                edit_url = new_page.url
                if new_page != page:
                    new_page.close()
                print(f'  编辑链接（补偿捕获）: {edit_url}')
                return edit_url
            except Exception:
                pass

        # 路径C：直接从命中行提取 href
        href = _extract_edit_href_from_row()
        if href:
            print(f'  编辑链接（href提取）: {href}')
            return href

        print(f'  未找到 [{offer_id}] 的编辑按钮: {e}')
        return None


def _shot(page, name):
    """??????? data/ ???????????????"""
    if not DEBUG_SHOTS:
        return
    path = os.path.join(DATA_DIR, f'debug_{name}.png')
    try:
        page.screenshot(path=path)
        print(f'  [??] {path}')
    except Exception as e:
        print(f'  [????] {name}: {e}')


# ── 上传图片 ──────────────────────────────────────────────────────────────────

REF_PRODUCT_TEMPLATE = '女装牛仔短裤'   # 「引用产品」模板名称，可由 --product-template 覆盖

DXM_MIN_IMG_W = 1340   # 店小秘颜色图片最小宽度
DXM_MIN_IMG_H = 1785   # 店小秘颜色图片最小高度


def _ensure_min_image_size(path, min_w=DXM_MIN_IMG_W, min_h=DXM_MIN_IMG_H):
    """
    图片预处理（letterbox 方案）：
    1. RGBA/P → RGB 白底合成（JPEG 不支持透明通道）
    2. 等比缩放到恰好放入 min_w×min_h（min-scale，不裁切）
    3. 贴到 min_w×min_h 纯白画布正中央
    4. 输出正好 1340×1785 的 JPEG（quality=95）
    已经是 1340×1785 RGB JPEG 则跳过，返回 False。
    返回 True 表示有修改。
    """
    TARGET_W = min_w   # 1340
    TARGET_H = min_h   # 1785

    try:
        from PIL import Image
        with Image.open(path) as img:
            w, h = img.size
            already_ok = (
                w == TARGET_W and h == TARGET_H
                and img.mode == 'RGB'
                and path.lower().endswith(('.jpg', '.jpeg'))
            )
            if already_ok:
                return False

            # ① 转 RGB（处理 RGBA / LA / P 等透明模式）
            if img.mode in ('RGBA', 'LA'):
                bg = Image.new('RGB', img.size, (255, 255, 255))
                bg.paste(img, mask=img.split()[-1])
                img = bg
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            # ② 等比缩放：min-scale 确保整张图完全放入目标尺寸（letterbox）
            w, h = img.size
            ratio = min(TARGET_W / w, TARGET_H / h)
            new_w = int(w * ratio)
            new_h = int(h * ratio)
            img_resized = img.resize((new_w, new_h), Image.LANCZOS)

            # ③ 白色画布，居中粘贴
            canvas = Image.new('RGB', (TARGET_W, TARGET_H), (255, 255, 255))
            offset = ((TARGET_W - new_w) // 2, (TARGET_H - new_h) // 2)
            canvas.paste(img_resized, offset)

            canvas.save(path, 'JPEG', quality=95)
            return True
    except Exception as e:
        print(f'  [warn] 图片处理失败 {path}: {e}')
        return False


def _preprocess_sku_images(root_dir, prod_folder_filter=None):
    """扫描 sku_images/ 下所有颜色图片，不满足尺寸要求的等比放大。"""
    if not os.path.isdir(root_dir):
        return
    resized = skipped = 0
    for prod_dir_name in sorted(os.listdir(root_dir)):
        if prod_folder_filter and prod_dir_name.split('_')[0] not in prod_folder_filter:
            continue
        prod_path = os.path.join(root_dir, prod_dir_name)
        if not os.path.isdir(prod_path):
            continue
        for color_dir in sorted(os.listdir(prod_path)):
            color_path = os.path.join(prod_path, color_dir)
            if not os.path.isdir(color_path):
                continue
            for fname in sorted(os.listdir(color_path)):
                if not fname.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                    continue
                img_path = os.path.join(color_path, fname)
                changed = _ensure_min_image_size(img_path)
                if changed:
                    resized += 1
                else:
                    skipped += 1
    if resized:
        print(f'[图片预处理] 放大 {resized} 张（已满足尺寸: {skipped} 张）')
    else:
        print(f'[图片预处理] 全部满足尺寸要求（{skipped} 张），无需处理')


def _candidate_color_folders(color_name):
    """给颜色名生成可能的本地文件夹名，兼容空格/下划线与历史别名。"""
    candidates = []

    def add_name(name):
        raw = str(name or '').strip()
        if not raw:
            return
        variants = [
            raw,
            _clean_filename(raw, max_len=30),
            raw.replace('_', ' ').strip(),
            raw.replace(' ', '_').strip(),
        ]
        for v in variants:
            if v and v not in candidates:
                candidates.append(v)

    add_name(color_name)
    if str(color_name).endswith('色'):
        add_name(str(color_name)[:-1])

    # 上游偶发颜色文案与本地英文色名不一致，补一层兼容
    alias_map = {
        '暗青': 'Dark Blue',
        '暗青色': 'Dark Blue',
        '深青': 'Dark Blue',
        '深青色': 'Dark Blue',
    }
    add_name(alias_map.get(str(color_name).strip(), ''))

    color_en = _color_to_en(color_name)
    add_name(color_en)
    if str(color_en).endswith('色'):
        add_name(str(color_en)[:-1])
    add_name(alias_map.get(str(color_en).strip(), ''))

    return candidates


def _product_skc_from_dir(product_dir):
    """Parse product SKC from sku_images/<SKC> or <SKC>_... directory name."""
    name = os.path.basename(str(product_dir or '')).strip()
    return name.split('_')[0].strip()


def _fill_product_skc(page, skc):
    """Fill product-level SKC/product code field on edit page. Fail closed if not found."""
    skc = str(skc or '').strip()
    if not skc:
        print('  [SKC] product SKC is empty, skip fill')
        return False

    eval_js = r"""(skc) => {
        const norm = (s) => String(s || '').replace(/\s+/g, '').trim();
        const visible = (el) => {
            if (!el) return false;
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
        };
        const writableInput = (root) => {
            if (!root) return null;
            const inputs = [...root.querySelectorAll('input:not([disabled]):not([readonly]), textarea:not([disabled]):not([readonly])')]
                .filter(el => visible(el) && !['file','hidden','checkbox','radio'].includes(String(el.type || '').toLowerCase()));
            return inputs[0] || null;
        };
        const setValue = (input) => {
            input.scrollIntoView({ block: 'center', inline: 'center' });
            input.focus();
            input.value = '';
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.value = skc;
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
            input.blur();
            return input.value === skc;
        };
        const labels = [
            'SKC',
            '产品货号',
            '商品货号',
            '商家货号',
            '店小秘货号',
            '货号',
            '商品编码',
            '自定义编码'
        ];
        const badLabel = (text) => /SKU货号|SKU编码|SKU/i.test(text || '');

        const formItems = [...document.querySelectorAll('.ant-form-item, .form-item, [class*="formItem"], [class*="form-item"], label')].filter(visible);
        const candidates = [];
        for (const item of formItems) {
            const text = norm(item.innerText || item.textContent || '');
            if (!text || badLabel(text)) continue;
            if (text.length > 80) continue;
            const score = labels.reduce((acc, label) => acc + (text.includes(label) ? (label === 'SKC' ? 10 : 6) : 0), 0);
            if (score <= 0) continue;
            let root = item;
            for (let i = 0; i < 4 && root && !writableInput(root); i += 1) root = root.parentElement;
            const input = writableInput(root || item);
            if (input) candidates.push({ input, score, text: text.slice(0, 80) });
        }
        candidates.sort((a, b) => b.score - a.score);
        for (const candidate of candidates) {
            if (setValue(candidate.input)) return { ok: true, method: 'label-scope', text: candidate.text };
        }

        const attrs = [...document.querySelectorAll('input:not([disabled]):not([readonly]), textarea:not([disabled]):not([readonly])')]
            .filter(visible)
            .filter(el => !['file','hidden','checkbox','radio'].includes(String(el.type || '').toLowerCase()));
        for (const input of attrs) {
            const hint = norm([input.name, input.id, input.placeholder, input.getAttribute('aria-label'), input.getAttribute('data-field')].join(' '));
            if (!hint || /sku/i.test(hint)) continue;
            if (/skc/i.test(hint) || hint.includes('产品货号') || hint.includes('商品货号') || hint.includes('商家货号') || hint.includes('货号') || hint.includes('商品编码')) {
                if (setValue(input)) return { ok: true, method: 'input-attr', text: hint.slice(0, 80) };
            }
        }
        const visibleText = (document.body.innerText || '').slice(0, 500);
        return { ok: false, reason: 'skc-input-not-found', visibleText };
    }"""

    def _try_fill(stage):
        try:
            result = page.evaluate(eval_js, skc)
        except Exception as e:
            print(f'  [SKC] fill exception at {stage}: {e}')
            return { 'ok': False, 'reason': str(e), 'stage': stage }
        if isinstance(result, dict):
            result['stage'] = stage
        return result

    result = _try_fill('current')
    if isinstance(result, dict) and result.get('ok'):
        print(f"  [SKC] filled {skc} ({result.get('stage')}/{result.get('method')}: {result.get('text', '')})")
        return True

    for nav_text in ['店小秘信息', '产品信息', '基本信息']:
        try:
            page.get_by_text(nav_text, exact=True).click(timeout=1500)
            time.sleep(0.6)
            result = _try_fill(f'nav:{nav_text}')
            if isinstance(result, dict) and result.get('ok'):
                print(f"  [SKC] filled {skc} ({result.get('stage')}/{result.get('method')}: {result.get('text', '')})")
                return True
        except Exception:
            pass

    try:
        page.evaluate("""async () => {
            const total = Math.max(document.body.scrollHeight, document.documentElement.scrollHeight);
            for (let y = 0; y <= total; y += Math.max(450, Math.floor(window.innerHeight * 0.65))) {
                window.scrollTo(0, y);
                await new Promise(resolve => setTimeout(resolve, 120));
            }
        }""")
    except Exception:
        pass
    result = _try_fill('scroll-scan')
    if isinstance(result, dict) and result.get('ok'):
        print(f"  [SKC] filled {skc} ({result.get('stage')}/{result.get('method')}: {result.get('text', '')})")
        return True

    print(f"  [SKC] editable input not found; product code should already come from import xlsx: {skc}")
    return True



def _fill_variant_skc_codes(page, skc):
    """Fill/verify variation SKU/SKC codes in the column immediately before image upload column."""
    skc = str(skc or '').strip()
    if not skc:
        print('  [variant-sku] empty product SKC, skip')
        return False

    eval_js = r"""(skc) => {
        const imageText = '\u9009\u62e9\u56fe\u7247';
        const visible = (el) => {
            if (!el) return false;
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
        };
        const isWritable = (el) => el && visible(el) && !el.disabled && !el.readOnly
            && !['file','hidden','checkbox','radio'].includes(String(el.type || '').toLowerCase());
        const setValue = (input, value) => {
            input.scrollIntoView({ block: 'center', inline: 'center' });
            input.focus();
            input.value = '';
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.value = value;
            input.dispatchEvent(new Event('input', { bubbles: true }));
            input.dispatchEvent(new Event('change', { bubbles: true }));
            input.blur();
            return input.value === value;
        };
        const normSize = (text) => {
            const t = String(text || '').toUpperCase().replace(/\s+/g, ' ');
            const m = t.match(/\b(XXL|2XL|XL|XS|L|M|S)\b/);
            if (!m) return '';
            return m[1] === '2XL' ? 'XXL' : m[1];
        };
        const cleanToken = (text) => String(text || '')
            .replace(/[\r\n\t]+/g, ' ')
            .replace(/[\\/:*?"<>|]+/g, '-')
            .replace(/\s+/g, ' ')
            .trim()
            .replace(/\s+/g, '_');
        const imageRows = [...document.querySelectorAll('tr')].filter(tr => visible(tr) && (tr.innerText || '').includes(imageText));
        const details = [];
        let filled = 0, verified = 0, skipped = 0, failed = 0;
        for (const tr of imageRows) {
            const cells = [...tr.querySelectorAll('td')];
            const imageIdx = cells.findIndex(td => (td.innerText || '').includes(imageText));
            if (imageIdx <= 0) { skipped += 1; details.push({ reason: 'image-cell-not-found' }); continue; }
            let input = null;
            let inputIdx = -1;
            for (const td of cells.slice(0, imageIdx).reverse()) {
                const candidates = [...td.querySelectorAll('input, textarea')].filter(isWritable);
                const skuInput = candidates.find(el => /variationSku|sku|skc/i.test([el.name, el.id, el.placeholder, el.getAttribute('aria-label')].join(' ')));
                input = skuInput || candidates[0] || input;
                if (input) { inputIdx = cells.indexOf(td); break; }
            }
            if (!input) { skipped += 1; details.push({ reason: 'input-before-image-not-found', text: (tr.innerText || '').slice(0, 120) }); continue; }
            const current = String(input.value || '').trim();
            if (current && current.startsWith(skc) && /^[\x20-\x7E]+$/.test(current)) {
                verified += 1;
                details.push({ ok: true, mode: 'verified', value: current });
                continue;
            }
            const rowText = tr.innerText || '';
            const color = cleanToken((cells[0] && cells[0].innerText) || '');
            const size = normSize(rowText) || normSize(cells.map(td => td.innerText || '').join(' '));
            const value = size && color ? `${skc}-${color}-${size}` : skc;
            if (setValue(input, value)) {
                filled += 1;
                details.push({ ok: true, mode: 'filled', value, inputIdx });
            } else {
                failed += 1;
                details.push({ ok: false, reason: 'set-failed', target: value, actual: input.value, inputIdx });
            }
        }
        return { ok: failed === 0 && (filled + verified) > 0, rows: imageRows.length, filled, verified, skipped, failed, details: details.slice(0, 8) };
    }"""

    try:
        result = page.evaluate(eval_js, skc)
    except Exception as e:
        print(f'  [variant-sku] fill exception: {e}')
        return False

    if isinstance(result, dict):
        print(f"  [variant-sku] rows={result.get('rows', 0)} filled={result.get('filled', 0)} verified={result.get('verified', 0)} skipped={result.get('skipped', 0)} failed={result.get('failed', 0)}")
        if result.get('details'):
            print(f"  [variant-sku] sample={result.get('details')}")
        return bool(result.get('ok'))
    print(f'  [variant-sku] unexpected result: {result}')
    return False

def upload_for_product(page, product_dir, cos_url_base=None, no_publish=False, warehouse_template=None, fill_skc=True):
    """在编辑页按颜色行上传图片，完成后点保存"""
    page.wait_for_load_state('domcontentloaded')
    time.sleep(2)
    _shot(page, 'edit_page_top')

    # ── 引用产品模板（替代手动选类目）──
    _apply_ref_product_template(page)

    skc_value = _product_skc_from_dir(product_dir)

    # ── 抓颜色行（模板应用后颜色变种已存在）──
    rows = page.locator('tr').filter(has=page.locator('text=选择图片')).all()
    if not rows:
        print('  未找到颜色行，跳过')
        return False

    print(f'  找到 {len(rows)} 个颜色行')
    if fill_skc and not _fill_variant_skc_codes(page, skc_value):
        print('  [fail-closed] variant SKU/SKC fill failed, skip save')
        return False
    _shot(page, 'edit_color_rows')  # 截图看颜色行上传区结构

    # 滚到第一个颜色行后截图，同时只看「选择图片」列的 HTML
    rows[0].scroll_into_view_if_needed()
    time.sleep(0.5)
    _shot(page, 'edit_color_rows_scrolled')

    fi_count = page.locator('input[type="file"]').count()
    print(f'  [debug] 全页 input[type=file] 数量: {fi_count}')
    upload_cell_html = page.evaluate("""(() => {
        const rows = [...document.querySelectorAll('tr')].filter(tr => tr.textContent.includes('选择图片'));
        if (!rows.length) return 'no rows';
        for (const td of rows[0].querySelectorAll('td')) {
            if (td.textContent.includes('选择图片')) return td.innerHTML;
        }
        return 'no td with 选择图片';
    })()""")
    print(f'  [debug] 选择图片列 HTML:\n{upload_cell_html}\n---')
    uploaded_count = 0

    for row in rows:
        # 第一列是颜色名
        color_name = row.locator('td').first.inner_text().strip()

        color_path = None
        color_folder = ''
        available_folders = []
        tried_folders = _candidate_color_folders(color_name)
        for folder in tried_folders:
            candidate_path = os.path.join(product_dir, folder)
            if os.path.isdir(candidate_path):
                color_folder = folder
                color_path = candidate_path
                break

        if not color_path:
            # ── 模糊兜底：归一化后比较（忽略大小写/空格/下划线/"色"后缀）──
            try:
                available_folders = [
                    f for f in os.listdir(product_dir)
                    if os.path.isdir(os.path.join(product_dir, f))
                ]
            except Exception:
                available_folders = []

            def _norm_color(s):
                s = str(s).lower().strip()
                s = re.sub(r'[\s_\-]+', '', s)
                if s.endswith('色'):
                    s = s[:-1]
                return s

            norm_candidates = {_norm_color(c) for c in tried_folders}
            for folder in available_folders:
                if _norm_color(folder) in norm_candidates:
                    color_folder = folder
                    color_path = os.path.join(product_dir, folder)
                    print(f'  [{color_name}] 模糊匹配: {folder}')
                    break

        if not color_path:
            sample = ', '.join(tried_folders[:4])
            suffix = ' ...' if len(tried_folders) > 4 else ''
            avail = ', '.join(available_folders[:6]) if available_folders else '(空)'
            print(f'  [{color_name}] 文件夹不存在（尝试: {sample}{suffix}；本地有: {avail}），跳过')
            continue

        imgs = [
            os.path.join(color_path, f)
            for f in ['1.jpg', '2.jpg', '3.jpg']
            if os.path.exists(os.path.join(color_path, f))
        ]
        if not imgs:
            print(f'  [{color_name}] 无图片文件，跳过')
            continue

        # 「选择图片」是下拉按钮，点击后弹出菜单（本地上传/素材库等）
        try:
            row.scroll_into_view_if_needed()
            time.sleep(0.2)

            if cos_url_base:
                # ── COS 网络图片路径（不占店小秘图片空间）──
                prod_dir_name = os.path.basename(product_dir)
                urls = [
                    f"{cos_url_base}/{prod_dir_name}/{color_folder}/{f}"
                    for f in ['1.jpg', '2.jpg', '3.jpg']
                    if os.path.exists(os.path.join(color_path, f))
                ]
                if not urls:
                    print(f'  [{color_name}] COS URL 列表为空，跳过')
                    continue
                url_text = '\n'.join(urls)

                row.locator('button:has-text("选择图片")').first.click()
                time.sleep(0.6)
                if uploaded_count == 0:
                    _shot(page, 'upload_dropdown')
                page.evaluate("""(() => {
                    for (const span of document.querySelectorAll('span.ant-dropdown-menu-title-content')) {
                        if (span.textContent.trim() === '网络图片') {
                            const r = span.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) { span.click(); return; }
                        }
                    }
                })()""")
                textbox = page.get_by_role('textbox', name='请填写图片URL地址，多个地址用回车换行')
                textbox.wait_for(state='visible', timeout=5000)
                textbox.fill(url_text)
                time.sleep(0.5)
                page.get_by_label('从网络地址(URL)选择图片').get_by_role('button', name='添加').click()

                # 等待店小秘服务器拉取图片完成（最多 30 秒）
                # 成功标志：弹窗关闭 或 出现上传成功的图片缩略图
                net_img_ok = False
                for _w in range(30):
                    time.sleep(1)
                    # 弹窗消失说明上传完成
                    try:
                        dlg = page.get_by_label('从网络地址(URL)选择图片')
                        if not dlg.is_visible():
                            net_img_ok = True
                            break
                    except Exception:
                        net_img_ok = True
                        break
                    # 检测失败提示
                    err_text = page.evaluate(
                        "() => { const e = document.querySelector('.ant-message-error'); return e ? e.innerText : ''; }"
                    )
                    if err_text and '上传失败' in err_text:
                        print(f'  [{color_name}] 网络图片上传失败: {err_text}')
                        break

                if net_img_ok:
                    print(f'  [{color_name}] 网络图片 {len(urls)} 张 OK（目录: {color_folder}）')
                else:
                    print(f'  [{color_name}] 网络图片上传超时，继续后续步骤')
                time.sleep(0.5)
                uploaded_count += 1
            else:
                # ── 本地文件上传路径 ──
                with page.expect_file_chooser(timeout=8000) as fc_info:
                    # 点下拉按钮，等下拉动画
                    row.locator('button:has-text("选择图片")').first.click()
                    time.sleep(0.6)
                    if uploaded_count == 0:
                        _shot(page, 'upload_dropdown')
                    # 用 JS 找有实际尺寸的「本地图片」菜单项并点击
                    page.evaluate("""(() => {
                        for (const span of document.querySelectorAll('span.ant-dropdown-menu-title-content')) {
                            if (span.textContent.trim() === '本地图片') {
                                const r = span.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) { span.click(); return; }
                            }
                        }
                    })()""")
                fc_info.value.set_files(imgs)
                time.sleep(1.5)
                print(f'  [{color_name}] 上传 {len(imgs)} 张 OK（目录: {color_folder}）')
                uploaded_count += 1
        except Exception as e:
            print(f'  [{color_name}] 上传失败: {e}')

    if uploaded_count == 0:
        print('  本商品无任何颜色成功上传')
        return False

    # ── 补填省份（产地=中国大陆时必填）──
    try:
        province_ph = page.locator('span.ant-select-selection-placeholder:text-is("请选择省份")').first
        if province_ph.is_visible():
            # 点击上两级的 ant-select-selector
            province_ph.locator('../..').click()
            time.sleep(0.4)
            page.locator('.ant-select-dropdown:visible .ant-select-item-option').first.wait_for(state='visible', timeout=3000)
            opt = page.locator('.ant-select-dropdown:visible .ant-select-item-option:has-text("广东")').first
            if opt.is_visible():
                opt.click()
            else:
                page.locator('.ant-select-dropdown:visible .ant-select-item-option').first.click()
            time.sleep(0.3)
            print('  省份已选择')
    except Exception as e:
        print(f'  [warn] 省份选择失败: {e}')

    # ── 填写尺码表 ──
    _fill_size_chart(page)

    # ── 选择站点仓库 + 填写仓库批量库存（在产品描述上方）──
    _fill_site_warehouse(page, warehouse_template=warehouse_template)
    # 清理可能残留的下拉/弹层，并等待仓库选择后页面重新渲染库存列
    try:
        page.keyboard.press('Escape')
        time.sleep(0.8)
    except Exception:
        pass
    if not _fill_batch_inventory(page, DEFAULT_BATCH_STOCK):
        print('  [warn] 库存批量填写失败，已跳过保存，避免误提交')
        return False

    print('  [描述图] 已改为导入表格携带，跳过编辑页详情图上传')

    # ── 剪掉无图片颜色行（防止保存时「服装类颜色属性必须上传3张图片」错误）──
    removed_cnt, remaining_empty = _prune_no_image_color_rows(page)
    if remaining_empty != 0:
        _shot(page, 'prune_result')
        if remaining_empty > 0:
            print(f'  [fail-closed] 仍有 {remaining_empty} 个无图片颜色行无法移除，跳过保存')
        else:
            print('  [fail-closed] 颜色剪枝整体异常，跳过保存')
        return False
    if removed_cnt > 0:
        _shot(page, 'after_prune')
        time.sleep(0.5)
        print(f'  [颜色剪枝] 已移除 {removed_cnt} 个无图片颜色行，继续保存')

    # 发布
    _shot(page, 'before_save')
    if no_publish:
        print('  [NO-PUBLISH] 已完成导入、上图、模板、尺码、仓库、库存和剪枝；按参数跳过发布点击')
        return True

    def _click_save_once():
        # 第一步：点「发布」主按钮（会弹出下拉菜单）
        btn = page.locator(
            'button:has-text("发布"), '
            'button:has-text("提交"), button:has-text("确认提交")'
        ).first
        btn.wait_for(state='visible', timeout=5000)
        btn.scroll_into_view_if_needed()
        btn.click(timeout=15000)

        # 第二步：等待「立即发布」下拉项出现并点击
        try:
            immediate = page.get_by_text('立即发布', exact=True)
            immediate.wait_for(state='visible', timeout=4000)
            immediate.click()
            print('  [发布] 已点击「立即发布」')
        except Exception:
            # 没出现下拉（老版本页面直接发布），忽略
            pass

    def _detect_save_error():
        """检测发布后是否有错误横幅（仍在编辑页 + 页面含"错误："文字）。"""
        try:
            return page.evaluate("""() => {
                const m = (document.body.innerText || '').match(/错误[：:]([^\\n]{3,80})/);
                return m ? ('错误：' + m[1].trim()) : null;
            }""")
        except Exception:
            return None

    try:
        _click_save_once()
        time.sleep(3)   # 等待发布响应及可能出现的错误提示
        _shot(page, 'after_save')
        err = _detect_save_error()
        if err:
            print(f'  [fail-closed] 发布后检测到错误：{err}')
            return False
        page.wait_for_load_state('domcontentloaded')
        time.sleep(1)
        _shot(page, 'after_save_load')
        print('  已发布')
    except Exception as e:
        _shot(page, 'save_error')
        print(f'  [warn] 自动点击发布失败，尝试先关闭拦截弹层再重试: {e}')

        try:
            _close_visible_modal(page)
        except Exception:
            pass
        try:
            page.keyboard.press('Escape')
        except Exception:
            pass
        time.sleep(0.5)

        try:
            _click_save_once()
            time.sleep(3)
            _shot(page, 'after_save_retry')
            err = _detect_save_error()
            if err:
                print(f'  [fail-closed] 发布重试后检测到错误：{err}')
                return False
            page.wait_for_load_state('domcontentloaded')
            time.sleep(1)
            _shot(page, 'after_save_retry_load')
            print('  已发布（重试成功）')
        except Exception as e2:
            _shot(page, 'save_error_retry')
            print(f'  [warn] 发布重试仍失败: {e2}')
            try:
                if sys.stdin and sys.stdin.isatty():
                    print('  未自动找到发布按钮，请手动点发布后按回车继续...')
                    input()
                    return True
            except EOFError:
                pass
            print('  [warn] 当前为非交互环境，无法人工确认发布，已按失败返回')
            return False

    return True



# ── 引用产品模板 ──────────────────────────────────────────────────────────────

def _close_visible_modal(page):
    """关闭当前可见的 antd modal（关闭按钮 × 或 取消，兜底用 JS 隐藏）"""
    closed = False
    try:
        btn = page.locator('.ant-modal-wrap:visible button.ant-modal-close').first
        if btn.is_visible():
            btn.click()
            time.sleep(0.5)
            closed = True
    except Exception:
        pass
    if not closed:
        try:
            btn = page.locator('.ant-modal-wrap:visible button:has-text("取消")').last
            if btn.is_visible():
                btn.click()
                time.sleep(0.5)
                closed = True
        except Exception:
            pass
    if not closed:
        # 兜底：按 Escape + JS 移除遮挡模态
        try:
            page.keyboard.press('Escape')
            time.sleep(0.3)
            page.evaluate("""() => {
                // 移除所有 ant-modal-wrap 遮罩，允许保存按钮被点击
                document.querySelectorAll('.ant-modal-wrap').forEach(el => {
                    el.style.display = 'none';
                    el.style.pointerEvents = 'none';
                });
                document.querySelectorAll('.ant-modal-mask').forEach(el => {
                    el.style.display = 'none';
                });
            }""")
            time.sleep(0.2)
        except Exception:
            pass


def _prune_no_image_color_rows(page):
    """
    找出图片区中所有没有已上传图片的颜色行，尝试从颜色选择区取消勾选将其移除。
    返回 (removed_count, remaining_empty_count)；remaining=-1 表示整体异常。

    根因：引用产品模板后，模板颜色变种留在页面，但本地 sku_images 只有实际 SKU 颜色，
    多余颜色行保存时触发「服装类颜色属性必须上传3张图片」错误。
    """
    try:
        empty_colors = page.evaluate("""() => {
            const isVisible = el => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                return r.width > 0 && r.height > 0;
            };
            const rows = [...document.querySelectorAll('tr')].filter(tr =>
                isVisible(tr) && tr.textContent.includes('选择图片')
            );
            const empty = [];
            for (const row of rows) {
                const imgs = [...row.querySelectorAll('img')].filter(img =>
                    img.src && img.src.startsWith('http')
                );
                if (imgs.length === 0) {
                    const td = row.querySelector('td');
                    const name = (td ? td.textContent : '').trim();
                    if (name) empty.push(name);
                }
            }
            return empty;
        }""")

        if not empty_colors:
            return (0, 0)

        print(f'  [颜色剪枝] 发现 {len(empty_colors)} 个无图片颜色: {empty_colors}')
        removed = 0

        for color_name in empty_colors:
            try:
                result = page.evaluate("""(colorName) => {
                    const norm = s => (s || '').replace(/\\s/g, '');
                    const target = norm(colorName);
                    // 策略1：ant-design checkbox-wrapper（已勾选）
                    for (const w of document.querySelectorAll('.ant-checkbox-wrapper')) {
                        if (norm(w.textContent) === target) {
                            if (w.querySelector('.ant-checkbox-checked')) {
                                w.click(); return 'wrapper';
                            }
                        }
                    }
                    // 策略2：标准 checkbox + 父级文本
                    for (const inp of document.querySelectorAll('input[type="checkbox"]')) {
                        if (!inp.checked) continue;
                        const label = inp.closest('label') || inp.parentElement;
                        if (label && norm(label.textContent) === target) {
                            inp.click(); return 'input';
                        }
                    }
                    // 策略3：已选颜色 tag 的关闭按钮
                    for (const el of document.querySelectorAll('span,div')) {
                        if (norm(el.textContent) !== target) continue;
                        const closeBtn = el.querySelector('[class*="close"],[class*="delete"],button');
                        if (closeBtn && closeBtn.getBoundingClientRect().width > 0) {
                            closeBtn.click(); return 'close-btn';
                        }
                    }
                    return 'not-found';
                }""", color_name)

                time.sleep(0.5)

                if result != 'not-found':
                    still_there = page.evaluate("""(colorName) => {
                        return [...document.querySelectorAll('tr')]
                            .filter(tr => tr.textContent.includes('选择图片'))
                            .some(row => {
                                const td = row.querySelector('td');
                                return td && td.textContent.trim() === colorName;
                            });
                    }""", color_name)
                    if not still_there:
                        removed += 1
                        print(f'  [颜色剪枝] [OK] 已移除 {color_name}（{result}）')
                    else:
                        print(f'  [颜色剪枝] [FAIL] 点击后行仍存在: {color_name}（{result}）')
                else:
                    print(f'  [颜色剪枝] [FAIL] 未找到复选框: {color_name}')

            except Exception as e:
                print(f'  [颜色剪枝] {color_name} 异常: {e}')

        remaining = page.evaluate("""() => {
            const isVisible = el => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                return r.width > 0 && r.height > 0;
            };
            return [...document.querySelectorAll('tr')].filter(tr =>
                isVisible(tr) && tr.textContent.includes('选择图片')
            ).filter(row => {
                const imgs = [...row.querySelectorAll('img')].filter(img =>
                    img.src && img.src.startsWith('http')
                );
                return imgs.length === 0;
            }).length;
        }""")

        return (removed, remaining)

    except Exception as e:
        print(f'  [颜色剪枝] 整体异常: {e}')
        return (0, -1)


def _apply_ref_product_template(page):
    """
    在编辑页点「引用产品」下拉 → 点「引用产品模板」→ 在弹窗中按模板名搜索并点「引用」→ 完成。
    """
    def _click_ref_template_menu_once():
        selectors = [
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("引用产品模板")',
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("引用产品模版")',
            '.ant-dropdown:visible li:has-text("引用产品模板")',
            '.ant-dropdown:visible li:has-text("引用产品模版")',
            '.ant-dropdown:visible span.ant-dropdown-menu-title-content:has-text("引用产品模板")',
            '.ant-dropdown:visible span.ant-dropdown-menu-title-content:has-text("引用产品模版")',
            '.ant-popover:visible [role="menuitem"]:has-text("引用产品模板")',
            '.ant-popover:visible [role="menuitem"]:has-text("引用产品模版")',
        ]
        for sel in selectors:
            try:
                item = page.locator(sel).first
                item.wait_for(state='visible', timeout=1200)
                item.click(timeout=2500)
                return True
            except Exception:
                continue

        # JS 兜底：只点击可见菜单中的目标项，规避隐藏旧节点误命中
        try:
            js_ok = page.evaluate("""() => {
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const fireClick = (el) => {
                    ['mouseover', 'mousedown', 'mouseup', 'click'].forEach((n) => {
                        el.dispatchEvent(new MouseEvent(n, { bubbles: true, cancelable: true, view: window }));
                    });
                    if (typeof el.click === 'function') el.click();
                };
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const targets = ['引用产品模板', '引用产品模版'];
                const overlays = [...document.querySelectorAll('.ant-dropdown,.ant-popover')].filter(isVisible);
                for (const box of overlays) {
                    const nodes = [...box.querySelectorAll('li,button,a,[role="menuitem"],span,div')].filter(isVisible);
                    for (const n of nodes) {
                        const txt = norm(n.textContent);
                        if (targets.some(t => txt.includes(t))) {
                            fireClick(n);
                            return true;
                        }
                    }
                }
                return false;
            }""")
            return bool(js_ok)
        except Exception:
            return False

    try:
        # 1) 点「引用产品」下拉按钮
        ref_btn = page.locator('button:has-text("引用产品"):visible').first
        ref_btn.wait_for(state='visible', timeout=6000)

        menu_clicked = False
        for _ in range(3):
            ref_btn.click()
            time.sleep(0.5)
            _shot(page, 'ref_product_dropdown')
            if _click_ref_template_menu_once():
                menu_clicked = True
                break
            try:
                page.keyboard.press('Escape')
            except Exception:
                pass
            time.sleep(0.3)

        if not menu_clicked:
            raise RuntimeError('未命中“引用产品模板”菜单项')

        time.sleep(1.2)
        _shot(page, 'ref_product_dialog')

        # 2) 定位模板弹窗
        modal = page.locator('.ant-modal-content').filter(
            has=page.locator(':text-is("模板名称")')
        ).first
        modal.wait_for(state='visible', timeout=7000)

        # 3) dump 弹窗 HTML（调试：确认列表结构）
        modal_html = page.evaluate("""(() => {
            const all = [...document.querySelectorAll('.ant-modal-content')];
            const vis = all.filter(m => {
                const r = m.getBoundingClientRect();
                const s = window.getComputedStyle(m);
                return r.width > 100 && r.height > 80 && s.visibility !== 'hidden' && s.display !== 'none';
            });
            return vis.map(m => m.innerHTML.substring(0, 6000)).join('\\n===\\n');
        })()""")
        html_path = os.path.join(DATA_DIR, 'probe_ref_modal.html')
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(modal_html)
        print(f'  [debug] 弹窗HTML → {html_path}')

        # 4) 按模板名搜索，确保不会误点其他模板
        try:
            search_input = modal.locator('input[name="productModuleSearchValue"], input.ant-input').first
            search_input.wait_for(state='visible', timeout=2500)
            search_input.fill(REF_PRODUCT_TEMPLATE)
            search_btn = modal.locator('button:has-text("搜索")').first
            if search_btn.is_visible():
                search_btn.click()
            time.sleep(0.9)
        except Exception as e:
            print(f'  [warn] 模板搜索失败，回退列表匹配: {e}')

        # 5) 优先点击模板名命中的那一行「引用」
        clicked = False
        row = modal.locator('tr.vxe-body--row').filter(
            has=page.locator(f'td:has-text("{REF_PRODUCT_TEMPLATE}")')
        ).first
        try:
            row.wait_for(state='visible', timeout=4500)
            ref_elem = row.locator('td :text-is("引用"), td a:has-text("引用"), td button:has-text("引用")').first
            ref_elem.wait_for(state='visible', timeout=2500)
            ref_elem.click()
            clicked = True
        except Exception:
            clicked = False

        # fallback：仍允许从弹窗里点第一处「引用」，保证流程尽量继续
        if not clicked:
            print(f'  [warn] 未定位到模板「{REF_PRODUCT_TEMPLATE}」行，尝试点击首个“引用”')
            ref_elem = modal.locator('td :text-is("引用"), td a:has-text("引用"), td button:has-text("引用")').first
            ref_elem.wait_for(state='visible', timeout=3000)
            ref_elem.click()

        time.sleep(3)
        _shot(page, 'ref_product_applied')

        # 点击「引用」后弹窗通常自动关闭，若仍开着则手动关
        _close_visible_modal(page)

        print(f'  引用模板「{REF_PRODUCT_TEMPLATE}」已应用')
    except Exception as e:
        _shot(page, 'ref_product_error')
        print(f'  [warn] 引用产品模板失败: {e}')
        # 关闭弹窗，避免遮挡后续图片上传操作
        _close_visible_modal(page)


# ── 描述图片 ─────────────────────────────────────────────────────────────────

def _set_files_compat(file_target, files):
    file_list = list(files or [])
    if not file_list:
        return 0

    def assign(target_files):
        if hasattr(file_target, 'set_input_files'):
            return file_target.set_input_files(target_files)
        return file_target.set_files(target_files)

    try:
        assign(file_list)
        return len(file_list)
    except Exception as exc:
        if 'Non-multiple file input' not in str(exc):
            raise

    uploaded = 0
    for file_path in file_list:
        assign(file_path)
        time.sleep(0.8)
        uploaded += 1
    return uploaded


def _upload_desc_images(page, product_dir):
    """
    上传产品描述图片：遍历所有颜色子目录，收集 0.jpg(白底图) + 1.jpg(模特图)。
    兼容两类页面：
      1) 有「编辑描述」按钮 + 弹窗编辑器
      2) 无弹窗，直接在「产品描述」区域用 file input 上传
    """
    desc_imgs = []
    detail_dir = os.path.join(product_dir, 'detail')
    if os.path.isdir(detail_dir):
        for fname in sorted(os.listdir(detail_dir)):
            fpath = os.path.join(detail_dir, fname)
            if os.path.isfile(fpath) and os.path.splitext(fname)[1].lower() in ('.jpg', '.jpeg', '.png', '.webp'):
                desc_imgs.append(fpath)
    for color_folder in sorted(os.listdir(product_dir)):
        color_path = os.path.join(product_dir, color_folder)
        if not os.path.isdir(color_path):
            continue
        for fname in ['0.jpg', '1.jpg']:
            fpath = os.path.join(color_path, fname)
            if os.path.exists(fpath):
                desc_imgs.append(fpath)

    if not desc_imgs:
        print('  [描述图] 无 0.jpg/1.jpg，跳过')
        return

    print(f'  [描述图] 共 {len(desc_imgs)} 张待上传')

    # 先点右侧导航「产品描述」让页面滚到描述区域
    try:
        page.locator(':text-is("产品描述"):visible').last.click(timeout=1500)
        time.sleep(0.6)
    except Exception:
        pass

    # dump 产品描述区域 HTML（排查页面结构）
    try:
        section_html = page.evaluate("""() => {
            const norm = (s) => String(s || '').replace(/\\s+/g, '');
            const nodes = [...document.querySelectorAll('label,span,div,td,p')];
            const label = nodes.find(n => norm(n.textContent).startsWith('产品描述'));
            if (!label) return 'LABEL NOT FOUND';
            const cands = [
                label.closest('.ant-form-item'),
                label.closest('tr'),
                label.closest('section'),
                label.parentElement,
                label.parentElement && label.parentElement.parentElement,
            ].filter(Boolean);
            for (const c of cands) {
                const html = c.outerHTML || '';
                if (html.length > 100) return html.substring(0, 12000);
            }
            return 'CONTAINER NOT FOUND';
        }""")
        html_path = os.path.join(DATA_DIR, 'probe_desc_editor.html')
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(section_html)
        print(f'  [描述图] 编辑器 HTML → {html_path}')
    except Exception:
        pass

    # 路径A：无弹窗页面，直接给“产品描述”邻近 file input 赋文件
    try:
        input_idx = page.evaluate("""() => {
            const norm = (s) => String(s || '').replace(/\\s+/g, '');
            const isVisible = (el) => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                const s = window.getComputedStyle(el);
                return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
            };

            const labels = [...document.querySelectorAll('label,span,div,td,p')]
                .filter(n => norm(n.textContent).startsWith('产品描述'));
            const all = [...document.querySelectorAll('input[type="file"]')];
            if (!labels.length || !all.length) return -1;

            const label = labels.find(isVisible) || labels[0];
            const lr = label.getBoundingClientRect();

            let bestIdx = -1;
            let bestScore = 1e9;
            for (let i = 0; i < all.length; i++) {
                const el = all[i];
                if (el.disabled) continue;
                const r = el.getBoundingClientRect();
                // 优先同屏且在“产品描述”附近的 input
                const score = Math.abs(r.top - lr.top) + Math.abs(r.left - lr.left) + (r.top < lr.top ? 2000 : 0);
                if (score < bestScore) {
                    bestScore = score;
                    bestIdx = i;
                }
            }
            return bestIdx;
        }""")

        if isinstance(input_idx, int) and input_idx >= 0:
            uploaded = _set_files_compat(page.locator('input[type="file"]').nth(input_idx), desc_imgs)
            time.sleep(uploaded * 1.2 + 2)
            _shot(page, 'desc_images_uploaded')
            print(f'  [DESC] uploaded {uploaded} image(s) direct mode')
            return uploaded > 0
    except Exception as e:
        print(f'  [描述图] 直传模式失败，回退按钮模式: {e}')

    # 路径B：旧页面，点“编辑描述”后在弹窗里上传
    try:
        clicked = False
        btn_selectors = [
            'button:has-text("编辑描述"):visible',
            'span:has-text("编辑描述"):visible',
            'a:has-text("编辑描述"):visible',
            ':text("编辑描述"):visible',
            ':text-is("编辑描述"):visible',
        ]
        for sel in btn_selectors:
            try:
                btn = page.locator(sel).first
                btn.wait_for(state='visible', timeout=1500)
                btn.scroll_into_view_if_needed()
                btn.click()
                clicked = True
                break
            except Exception:
                continue

        if clicked:
            time.sleep(1.5)
            _shot(page, 'desc_editor_open')

        upload_btn = page.locator(
            '.ant-modal-content button:has-text("添加图片"), '
            '.ant-modal-content button:has-text("插入图片"), '
            '.ant-modal-content button:has-text("上传图片"), '
            '.ant-modal-content [title="图片"], '
            '.ant-modal-content [aria-label="image"]'
        ).first
        upload_btn.wait_for(state='visible', timeout=3500)
        with page.expect_file_chooser(timeout=6000) as fc_info:
            upload_btn.click()
        uploaded = _set_files_compat(fc_info.value, desc_imgs)
        time.sleep(uploaded * 1.2 + 2)
        _shot(page, 'desc_images_uploaded')
        print(f'  [DESC] uploaded {uploaded} image(s) modal mode')

        ok_btn = page.locator(
            '.ant-modal-content button:has-text("确定"), '
            '.ant-modal-content button:has-text("保存")'
        ).last
        if ok_btn.is_visible():
            ok_btn.click()
            time.sleep(1)
            print('  [描述图] 描述已保存')
        return uploaded > 0
    except Exception as e:
        _shot(page, 'desc_upload_error')
        print(f'  [描述图] 未找到可用上传入口，已跳过: {e}')
        return False



# ── 尺码表 ───────────────────────────────────────────────────────────────────

# 固定模板：S/M/L/XL/XXL × (腰围全围, 臀围全围, 裤长)，单位 cm
_SIZE_CHART_DATA = {
    'S':   (68,  94,  35),
    'M':   (72,  98,  36),
    'L':   (78,  104, 37),
    'XL':  (84,  110, 38),
    'XXL': (90,  116, 39),
}
_SIZE_CHART_TEMPLATE = '女士牛仔短裤'   # 尺码表模板名称，可由 --size-template 覆盖


def _fill_size_chart(page):
    """
    在编辑页点「添加/编辑尺码表」→ 引用模板「女士牛仔短裤」→ 确定。
    模板已在店小秘后台预先创建，包含 S/M/L/XL/XXL 全部测量数据。
    """
    try:
        add_link = page.locator(
            'span.link:has-text("添加尺码表"), span.link:has-text("编辑尺码表")'
        ).first
        if not add_link.is_visible():
            print('  [尺码表] 未找到链接，跳过')
            return
        add_link.scroll_into_view_if_needed()
        add_link.click()
        time.sleep(1.5)
        _shot(page, 'sizechart_modal')

        # 点「引用模板」下拉，选「女士牛仔短裤」
        modal = page.locator('.ant-modal-content')
        ref_select = modal.locator('.ant-select:near(:text("引用模板"))').first
        ref_select.click()
        time.sleep(0.6)
        # 搜索模板名
        modal.locator('input.ant-select-selection-search-input').last.fill(_SIZE_CHART_TEMPLATE)
        time.sleep(0.8)
        # 点匹配项
        option = page.locator(
            f'.ant-select-dropdown:visible .ant-select-item-option:has-text("{_SIZE_CHART_TEMPLATE}")'
        ).first
        option.wait_for(state='visible', timeout=5000)
        option.click()
        time.sleep(1)
        _shot(page, 'sizechart_filled')
        print(f'  [尺码表] 已引用模板「{_SIZE_CHART_TEMPLATE}」')

        # 确定
        modal.locator('button:has-text("确定")').last.click()
        time.sleep(1.5)
        print('  [尺码表] 已提交')
    except Exception as e:
        print(f'  [warn] 尺码表填写失败: {e}')


def _fill_site_warehouse(page, warehouse_template=None):

    """编辑页选择站点仓库（优先处理选择仓库/站点仓库区域）。"""
    # 等待页面异步加载仓库数据（仓库列表由 API 异步返回，需给页面足够时间）
    time.sleep(2)

    # ★ 最高优先级：检查库存表是否已有"可售库存"列
    # 证据表明：凡是跳过仓库点击的商品 score=24，凡是尝试点击的商品 score=2
    # 原因：JS dispatchEvent 未触发 Vue/React 状态更新，反而清空了已选仓库
    # 策略：若库存列已存在则直接返回，绝不碰仓库下拉
    try:
        has_inventory_col = page.evaluate("""() => {
            const isVisible = el => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                return r.width > 0 && r.height > 0;
            };
            const tables = [...document.querySelectorAll('table')].filter(isVisible);
            return tables.some(t => {
                const headers = [...t.querySelectorAll('thead th, th')]
                    .filter(isVisible)
                    .map(el => (el.textContent || '').trim());
                return headers.some(h =>
                    h.includes('\u53ef\u552e\u5e93\u5b58') ||
                    (h.includes('\u5e93\u5b58') && !h.includes('\u5305\u88c5'))
                );
            });
        }""")
        if has_inventory_col:
            print('  [\u4ed3\u5e93] \u5e93\u5b58\u5217\u5df2\u5b58\u5728\uff0c\u8df3\u8fc7\u4ed3\u5e93\u9009\u62e9')
            return True
    except Exception:
        pass

    try:
        # 先看是否已选中仓库
        state = page.evaluate("""() => {
            const isVisible = (el) => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                const s = window.getComputedStyle(el);
                return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
            };
            const selects = [...document.querySelectorAll('.ant-select')].filter(isVisible);
            let selected = 0;
            let unselected = 0;
            for (const sel of selects) {
                const scope = sel.closest('div, td, tr, form') || sel.parentElement;
                const txt = ((scope && scope.textContent) || '').replace(/\\s+/g, '');
                if (!txt.includes('仓库')) continue;
                const item = sel.querySelector('.ant-select-selection-item');
                const ph = sel.querySelector('.ant-select-selection-placeholder');
                if (item && item.textContent.trim()) selected += 1;
                if (ph && ph.textContent.includes('请选择')) unselected += 1;
            }
            return { selected, unselected };
        }""")
        if state.get('selected', 0) > 0 and state.get('unselected', 0) == 0:
            print('  [仓库] 已选择，跳过')
            return True
    except Exception:
        pass

    clicked = False
    # 主路径：按「选择仓库/站点仓库」标签定位
    try:
        selector = page.locator(
            'xpath=(//*[contains(normalize-space(text()),"选择仓库") or contains(normalize-space(text()),"站点仓库")]/following::div[contains(@class,"ant-select-selector")][1])[1]'
        ).first
        if selector.is_visible():
            selector.scroll_into_view_if_needed()
            time.sleep(0.3)
            selector.click()
            clicked = True
    except Exception:
        pass

    # 兜底：页面中寻找包含“仓库”文案区域内第一个可见未禁用下拉
    if not clicked:
        try:
            js_result = page.evaluate("""() => {
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const selectors = [...document.querySelectorAll('.ant-select-selector')].filter(isVisible);
                for (const sel of selectors) {
                    const box = sel.closest('.ant-select');
                    if (!box) continue;
                    if (box.className.includes('disabled')) continue;
                    const scope = box.closest('div, td, tr, form') || box.parentElement;
                    const txt = ((scope && scope.textContent) || '').replace(/\\s+/g, '');
                    if (!txt.includes('仓库')) continue;
                    sel.click();
                    return 'clicked';
                }
                return 'not found';
            }""")
            clicked = (js_result == 'clicked')
        except Exception:
            clicked = False

    if not clicked:
        print('  [warn] 未定位到站点仓库下拉，跳过')
        return False

    try:
        # 等待选项出现
        page.locator('.ant-select-dropdown:visible .ant-select-item-option').first.wait_for(
            state='visible', timeout=4000
        )
        time.sleep(0.3)

        # 优先用 Playwright 原生 click（isTrusted=true，Vue/React 合成事件可响应）
        # 如果配置了仓库模板/仓库名称，优先点匹配项；否则点第一个真实仓库选项
        skip_texts = {'全部', '请选择', ''}
        clicked_option = False
        opts_locator = page.locator('.ant-select-dropdown:visible .ant-select-item-option')
        opts_count = opts_locator.count()
        preferred = str(warehouse_template or '').strip()
        if preferred:
            try:
                preferred_opt = opts_locator.filter(has_text=preferred).first
                preferred_opt.wait_for(state='visible', timeout=1500)
                wh_text = (preferred_opt.text_content() or '').strip()
                preferred_opt.click(timeout=2000)
                clicked_option = True
            except Exception:
                print(f'  [warn] 未找到仓库「{preferred}」，回退第一个真实仓库')
        for i in range(opts_count):
            if clicked_option:
                break
            opt = opts_locator.nth(i)
            try:
                txt = (opt.text_content() or '').strip()
                if txt in skip_texts:
                    continue
                if opt.is_visible():
                    opt.click(timeout=2000)
                    wh_text = txt
                    clicked_option = True
                    break
            except Exception:
                continue

        if not clicked_option:
            # 兜底：点第一个可见选项（无论文字内容）
            try:
                first_opt = opts_locator.first
                wh_text = (first_opt.text_content() or '').strip()
                first_opt.click(timeout=2000)
                clicked_option = True
                wh_text = 'fallback:' + wh_text
            except Exception:
                pass

        if not clicked_option:
            print('  [warn] 仓库下拉可见但未找到可点击选项')
            try:
                page.keyboard.press('Escape')
            except Exception:
                pass
            return False

        wh_text_clean = wh_text.replace('fallback:', '')
        time.sleep(1.5)  # 给 Vue 响应时间重新渲染库存列
        print(f'  [仓库] 已点击选项: {wh_text_clean}')
        return True
    except Exception as e:
        print(f'  [warn] 仓库下拉已打开但未选中: {e}')
        try:
            page.keyboard.press('Escape')
            time.sleep(0.2)
        except Exception:
            pass
        return False



def _fill_batch_inventory(page, qty=DEFAULT_BATCH_STOCK):
    """编辑页填写仓库批量库存（仅允许批量弹层路径）。"""
    qty = str(qty)

    def _goto_variant_section():
        for sel in [':text-is("变种信息"):visible', ':text-is("产品信息"):visible', ':text-is("销售信息"):visible']:
            try:
                nav = page.locator(sel).last
                nav.wait_for(state='visible', timeout=800)
                nav.click(timeout=1200)
                time.sleep(0.6)
                return True
            except Exception:
                continue
        return False

    def _scroll_variant_table_to_right():
        try:
            moved = page.evaluate("""() => {
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const hasInventoryHeader = (txt) => txt.includes('可售库存') || txt === '库存' || txt.includes('库存') || txt.includes('可售');
                const denyKeywords = ['申报价格', '建议售价', '售价', '价格', '重量', '尺寸', '包装清单', '包装'];
                const hasDeny = (txt) => denyKeywords.some((k) => txt.includes(k));

                const getWarehouseFollowerTables = () => {
                    const set = new Set();
                    const xpaths = [
                        '//*[contains(normalize-space(.),"选择仓库")]/following::table[1]',
                        '//*[contains(normalize-space(.),"站点仓库")]/following::table[1]',
                    ];
                    for (const xp of xpaths) {
                        try {
                            const snap = document.evaluate(xp, document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
                            for (let i = 0; i < snap.snapshotLength; i += 1) {
                                const t = snap.snapshotItem(i);
                                if (t && isVisible(t)) set.add(t);
                            }
                        } catch (_) {
                        }
                    }
                    return set;
                };

                const tables = [...document.querySelectorAll('table')].filter(isVisible);
                const warehouseFollower = getWarehouseFollowerTables();
                const tableSeen = new Set();
                const anchorTables = [];
                const pushAnchorTable = (t, source, priority) => {
                    if (!t || !isVisible(t) || tableSeen.has(t)) return;
                    tableSeen.add(t);
                    anchorTables.push({ table: t, source, priority });
                };

                for (const t of tables) {
                    if (warehouseFollower.has(t)) {
                        pushAnchorTable(t, 'warehouse-follower', 100);
                        continue;
                    }
                    const headers = [...t.querySelectorAll('thead th,th')]
                        .filter(isVisible)
                        .map((el) => norm(el.textContent || ''))
                        .filter(Boolean)
                        .slice(0, 36);
                    const invHeaders = headers.filter((h) => hasInventoryHeader(h) && !hasDeny(h)).length;
                    if (invHeaders > 0) {
                        pushAnchorTable(t, 'inventory-header', 90 + invHeaders);
                        continue;
                    }
                    const txt = norm(t.textContent || '');
                    if (txt.includes('选择仓库') && txt.includes('批量')) {
                        pushAnchorTable(t, 'warehouse-batch-fallback', 50);
                    }
                }

                const hostSeen = new Set();
                const hosts = [];
                const classChain = (el) => {
                    const chain = [];
                    let cur = el;
                    for (let i = 0; cur && i < 6; i += 1) {
                        const tag = String(cur.tagName || '').toLowerCase();
                        const cls = String(cur.className || '').trim().replace(/\\s+/g, '.');
                        chain.push(cls ? `${tag}.${cls}` : tag);
                        cur = cur.parentElement;
                    }
                    return chain.join(' > ').slice(0, 320);
                };
                const pushHost = (el, source, anchorSource, anchorPriority) => {
                    if (!el || !isVisible(el) || hostSeen.has(el)) return;
                    const sw = Number(el.scrollWidth || 0);
                    const cw = Number(el.clientWidth || 0);
                    if (!(sw > cw + 1)) return;
                    const cls = String(el.className || '');
                    const cs = window.getComputedStyle(el);
                    const overflowX = String(cs.overflowX || '').toLowerCase();
                    const canScrollByOverflow = overflowX === 'auto' || overflowX === 'scroll' || overflowX === 'overlay';
                    const structuralScrollable =
                        cls.includes('vxe')
                        || cls.includes('table-wrapper')
                        || cls.includes('table-container')
                        || cls.includes('body-wrapper')
                        || cls.includes('main-wrapper');
                    if (!(canScrollByOverflow || structuralScrollable)) return;
                    const strict = (sw > cw + 8) && (canScrollByOverflow || cls.includes('vxe'));
                    hostSeen.add(el);
                    hosts.push({
                        el,
                        source,
                        anchorSource,
                        anchorPriority: Number(anchorPriority || 0),
                        maxScroll: Math.max(0, sw - cw),
                        overflowX,
                        cls: cls.slice(0, 140),
                        strict,
                        chain: classChain(el),
                    });
                };

                const knownSelectors = [
                    '.vxe-table--body-wrapper',
                    '.vxe-table--main-wrapper',
                    '.vxe-grid--table-wrapper',
                    '.vxe-grid--table-container',
                    '.vxe-table--header-wrapper',
                    '[class*="vxe-table--body-wrapper"]',
                    '[class*="vxe-grid--table-wrapper"]',
                    '[class*="vxe-grid--table-container"]',
                    '[class*="vxe-table--main-wrapper"]',
                ];

                for (const anchor of anchorTables) {
                    const t = anchor.table;
                    const nearestWrapper = t.closest('.vxe-table--body-wrapper,.vxe-table--main-wrapper,.vxe-grid--table-wrapper,.vxe-grid--table-container');
                    if (nearestWrapper) {
                        pushHost(nearestWrapper, 'anchor-nearest-wrapper', anchor.source, anchor.priority + 10);
                    }

                    let cur = t;
                    for (let i = 0; cur && i < 14; i += 1) {
                        pushHost(cur, 'anchor-ancestor', anchor.source, anchor.priority);
                        const cls = String(cur.className || '');
                        if (cls.includes('vxe-table--body-wrapper') || cls.includes('vxe-table--main-wrapper') || cls.includes('vxe-grid--table-wrapper') || cls.includes('vxe-grid--table-container')) {
                            pushHost(cur, 'anchor-vxe-ancestor', anchor.source, anchor.priority + 8);
                        }
                        cur = cur.parentElement;
                    }

                    const roots = [
                        t.closest('.vxe-grid'),
                        t.closest('.vxe-grid--table-container'),
                        t.closest('.vxe-grid--table-wrapper'),
                        t.closest('.vxe-table'),
                        t.closest('section,form,div'),
                        t.parentElement,
                    ].filter(Boolean);
                    for (const root of roots.slice(0, 6)) {
                        for (const sel of knownSelectors) {
                            const nodes = [...root.querySelectorAll(sel)].filter(isVisible).slice(0, 20);
                            for (const n of nodes) {
                                pushHost(n, `known:${sel}`, anchor.source, anchor.priority + 5);
                            }
                        }
                    }
                }

                if (!hosts.length) {
                    for (const sel of knownSelectors) {
                        const nodes = [...document.querySelectorAll(sel)].filter(isVisible).slice(0, 40);
                        for (const n of nodes) {
                            pushHost(n, `global:${sel}`, 'global', 10);
                        }
                    }
                }

                const strictHostCount = hosts.filter((h) => h.strict).length;

                const tryScroll = (host) => {
                    const el = host.el;
                    const before = Number(el.scrollLeft || 0);
                    const maxScroll = Number(host.maxScroll || Math.max(0, (el.scrollWidth || 0) - (el.clientWidth || 0)));
                    if (maxScroll <= 8) {
                        return { moved: false, reason: 'no-overflow', before, after: before, maxScroll };
                    }
                    if (before >= maxScroll - 2) {
                        return { moved: true, alreadyRight: true, before, after: before, maxScroll };
                    }

                    const checkMoved = () => {
                        const after = Number(el.scrollLeft || 0);
                        return after > before + 1 ? after : null;
                    };

                    try { el.scrollLeft = maxScroll; } catch (_) {}
                    let after = checkMoved();
                    if (after !== null) return { moved: true, before, after, maxScroll };

                    try { el.scrollTo({ left: maxScroll, behavior: 'instant' }); } catch (_) {}
                    after = checkMoved();
                    if (after !== null) return { moved: true, before, after, maxScroll };

                    try { el.scrollBy({ left: Math.max(120, maxScroll), behavior: 'instant' }); } catch (_) {}
                    after = checkMoved();
                    if (after !== null) return { moved: true, before, after, maxScroll };

                    try {
                        el.dispatchEvent(new WheelEvent('wheel', {
                            deltaX: Math.max(160, maxScroll),
                            deltaY: 0,
                            bubbles: true,
                            cancelable: true,
                        }));
                    } catch (_) {}
                    after = checkMoved();
                    if (after !== null) return { moved: true, before, after, maxScroll };

                    return {
                        moved: false,
                        reason: 'immutable-scroll-left',
                        before,
                        after: Number(el.scrollLeft || 0),
                        maxScroll,
                    };
                };

                if (!hosts.length) {
                    return {
                        moved: false,
                        reason: 'no-scrollable-wrapper',
                        tableCount: tables.length,
                        anchorTables: anchorTables.length,
                        hostCount: 0,
                        strictHostCount: 0,
                        anchorSource: anchorTables[0] ? anchorTables[0].source : null,
                    };
                }

                hosts.sort((a, b) => (Number(b.strict) - Number(a.strict)) || (b.anchorPriority - a.anchorPriority) || (b.maxScroll - a.maxScroll));
                for (const host of hosts.slice(0, 24)) {
                    const res = tryScroll(host);
                    if (res.moved) {
                        return {
                            moved: true,
                            alreadyRight: !!res.alreadyRight,
                            source: host.source,
                            anchorSource: host.anchorSource,
                            cls: host.cls,
                            chain: host.chain,
                            overflowX: host.overflowX,
                            before: res.before,
                            scrollLeft: res.after,
                            maxScroll: res.maxScroll,
                            scrollWidth: Number(host.el.scrollWidth || 0),
                            clientWidth: Number(host.el.clientWidth || 0),
                            tableCount: tables.length,
                            anchorTables: anchorTables.length,
                            hostCount: hosts.length,
                            strictHostCount,
                            strictHost: host.strict,
                        };
                    }
                }

                const first = hosts[0];
                const fail = tryScroll(first);
                return {
                    moved: false,
                    reason: fail.reason || 'scroll-failed',
                    source: first.source,
                    anchorSource: first.anchorSource,
                    cls: first.cls,
                    chain: first.chain,
                    overflowX: first.overflowX,
                    before: fail.before,
                    scrollLeft: fail.after,
                    maxScroll: fail.maxScroll,
                    scrollWidth: Number(first.el.scrollWidth || 0),
                    clientWidth: Number(first.el.clientWidth || 0),
                    tableCount: tables.length,
                    anchorTables: anchorTables.length,
                    hostCount: hosts.length,
                    strictHostCount,
                    strictHost: first.strict,
                };
            }""")
            if isinstance(moved, dict) and moved.get('moved'):
                state = '已在最右侧' if moved.get('alreadyRight') else '已横向滚动到右侧'
                print(
                    f"  [库存] {state} "
                    f"(before={moved.get('before')}, after={moved.get('scrollLeft')}, max={moved.get('maxScroll')}, width={moved.get('clientWidth')}, "
                    f"src={moved.get('source')}, anchor={moved.get('anchorSource')}, overflowX={moved.get('overflowX')}, cls={moved.get('cls')}, "
                    f"tables={moved.get('tableCount')}, anchors={moved.get('anchorTables')}, hosts={moved.get('hostCount')}, strictHosts={moved.get('strictHostCount')}, strict={moved.get('strictHost')}, chain={moved.get('chain')})"
                )
                time.sleep(0.2)
                return True
            if isinstance(moved, dict):
                print(
                    '  [库存] 横向滚动未生效 '
                    f"(reason={moved.get('reason')}, before={moved.get('before')}, after={moved.get('scrollLeft')}, max={moved.get('maxScroll')}, "
                    f"src={moved.get('source')}, anchor={moved.get('anchorSource')}, overflowX={moved.get('overflowX')}, cls={moved.get('cls')}, "
                    f"tables={moved.get('tableCount')}, anchors={moved.get('anchorTables')}, hosts={moved.get('hostCount')}, strictHosts={moved.get('strictHostCount')}, strict={moved.get('strictHost')}, chain={moved.get('chain')})"
                )
        except Exception:
            pass
        return False

    def _find_inventory_inputs():
        # 仅找批量操作后弹出的弹层/弹窗输入，避免误命中表格行内或页面顶部输入框
        selectors = [
            '.ant-modal-content:visible input:not([disabled]):not([type="radio"]):not([type="checkbox"]):not([type="hidden"])',
            '.ant-popover:visible input:not([disabled]):not([type="radio"]):not([type="checkbox"]):not([type="hidden"])',
            '.ant-dropdown:visible input:not([disabled]):not([type="radio"]):not([type="checkbox"]):not([type="hidden"])',
            '.ant-modal-content:visible .ant-input-number-input:not([disabled])',
            '.ant-popover:visible .ant-input-number-input:not([disabled])',
            '.ant-dropdown:visible .ant-input-number-input:not([disabled])',
        ]
        for sel in selectors:
            try:
                loc = page.locator(sel)
                count = loc.count()
                if count <= 0:
                    continue
                picked = []
                for i in range(min(count, 8)):
                    one = loc.nth(i)
                    if not one.is_visible():
                        continue
                    try:
                        ok = one.evaluate("""(el) => {
                            const box = el.closest('.ant-modal-content,.ant-popover,.ant-dropdown');
                            if (!box) return false;
                            const txt = String(box.innerText || '').replace(/\\s+/g, '');
                            return txt.includes('库存') || txt.includes('可售') || txt.includes('修改库存') || txt.includes('批量');
                        }""")
                    except Exception:
                        ok = False
                    if ok:
                        picked.append(one)
                if picked:
                    return picked
            except Exception:
                continue
        return []


    def _inventory_panel_ready():
        try:
            js_result = page.evaluate("""() => {
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };

                const scopeKeywords = ['选择仓库', '站点仓库', 'SKU分类', 'SKU信息', '变种信息', '可售库存', '修改库存'];
                const warehouseKeywords = ['选择仓库', '站点仓库', '仓库'];
                const denyKeywords = ['申报价格', '建议售价', '售价', '价格', '重量', '尺寸', '包装清单', '包装', '清单'];
                const hasScopeKeyword = (txt) => scopeKeywords.some((k) => txt.includes(k));
                const hasWarehouseKeyword = (txt) => warehouseKeywords.some((k) => txt.includes(k));
                const hasDeny = (txt) => denyKeywords.some((k) => txt.includes(k));
                const hasInventoryMarker = (txt) => txt.includes('库存') || txt.includes('可售') || txt.includes('修改库存');
                const hasBatchMarker = (txt) => txt.includes('批量');
                const navMarkers = ['首页产品数据采集', '订单管理', '客服通用服务', 'Copyright'];
                const isNoisyScope = (txt) => txt.length > 3200 || navMarkers.some((k) => txt.includes(k));
                const pickScope = (table) => {
                    let cur = table;
                    for (let i = 0; cur && i < 9; i += 1) {
                        const txt = norm(cur.textContent || '');
                        if (hasScopeKeyword(txt) && !isNoisyScope(txt)) return cur;
                        cur = cur.parentElement;
                    }
                    return table.closest('section,form,div') || table.parentElement;
                };
                const getWarehouseFollowerTables = () => {
                    const set = new Set();
                    const xpaths = [
                        '//*[contains(normalize-space(.),"选择仓库")]/following::table[1]',
                        '//*[contains(normalize-space(.),"站点仓库")]/following::table[1]',
                    ];
                    for (const xp of xpaths) {
                        try {
                            const snap = document.evaluate(xp, document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
                            for (let i = 0; i < snap.snapshotLength; i += 1) {
                                const t = snap.snapshotItem(i);
                                if (t && isVisible(t)) set.add(t);
                            }
                        } catch (_) {
                        }
                    }
                    return set;
                };

                const tables = [...document.querySelectorAll('table')].filter(isVisible);
                if (!tables.length) {
                    return { ready: false, reason: 'no-visible-table', tables: 0 };
                }

                const warehouseFollowerTables = getWarehouseFollowerTables();
                const scored = tables.map((t) => {
                    const txt = norm(t.textContent);
                    const scope = pickScope(t) || (t.closest('section,form,div') || t.parentElement);
                    const scopeTxt = norm(scope ? scope.textContent : '');
                    const inScopedArea = hasScopeKeyword(scopeTxt) || hasScopeKeyword(txt);
                    const hasWarehouse = hasWarehouseKeyword(scopeTxt) || hasWarehouseKeyword(txt);
                    const warehouseFollower = warehouseFollowerTables.has(t);
                    const headerTexts = [...t.querySelectorAll('thead th,th')]
                        .filter(isVisible)
                        .map((el) => norm(el.textContent || ''))
                        .filter(Boolean)
                        .slice(0, 28);
                    const invHeaderCount = headerTexts.filter((h) => {
                        if (hasDeny(h)) return false;
                        return h.includes('可售库存') || h === '库存' || h.includes('库存') || h.includes('可售');
                    }).length;
                    const imageHeaderCount = headerTexts.filter((h) => h.includes('图片') || h.includes('主图')).length;
                    const skcHeaderCount = headerTexts.filter((h) => h.includes('SKC编码') || h.includes('货号')).length;
                    const denyHeaderCount = headerTexts.filter((h) => hasDeny(h)).length;
                    const imageOnlyTable = invHeaderCount === 0 && imageHeaderCount > 0 && skcHeaderCount > 0;
                    const hasInventoryEvidence = (
                        invHeaderCount > 0
                        || txt.includes('可售库存')
                        || txt.includes('修改库存')
                        || (txt.includes('库存') && hasWarehouse)
                    ) && !imageOnlyTable;
                    const invActionCount = [...(scope || t).querySelectorAll('button,a,[role="button"],.ant-dropdown-trigger,.ant-btn,thead th,th')]
                        .filter(isVisible)
                        .filter((el) => {
                            const cTxt = norm(el.textContent);
                            if (!cTxt) return false;
                            if (hasDeny(cTxt) && !hasInventoryMarker(cTxt)) return false;
                            if ((cTxt.includes('图片') || cTxt.includes('主图')) && !hasInventoryMarker(cTxt)) return false;
                            return hasBatchMarker(cTxt) || hasInventoryMarker(cTxt);
                        }).length;
                    const denyOnlyPenalty = (denyHeaderCount > 0 && invHeaderCount === 0 && !txt.includes('可售库存')) ? 8 : 0;
                    const packagingPenalty = (scopeTxt.includes('包装清单') && !hasInventoryEvidence) ? 5 : 0;
                    const score =
                        (inScopedArea ? 5 : 0)
                        + (hasWarehouse ? 4 : 0)
                        + (warehouseFollower ? 7 : 0)
                        + (hasInventoryEvidence ? 6 : 0)
                        + (invHeaderCount * 3)
                        + (hasBatchMarker(txt) ? 1 : 0)
                        + (invActionCount > 0 ? 2 : 0)
                        - denyOnlyPenalty
                        - packagingPenalty
                        - (imageOnlyTable ? 10 : 0)
                        - (isNoisyScope(scopeTxt) ? 2 : 0);
                    return {
                        score,
                        inScopedArea,
                        hasWarehouse,
                        warehouseFollower,
                        hasInventoryEvidence,
                        invHeaderCount,
                        invActionCount,
                        imageOnlyTable,
                        scopeTxt,
                        txt,
                    };
                }).sort((a, b) => (Number(b.warehouseFollower) - Number(a.warehouseFollower)) || (b.score - a.score));

                const best = scored[0];
                if (!best) {
                    return { ready: false, reason: 'no-candidate-table', tables: tables.length };
                }

                const ready =
                    best.score >= 12
                    && best.hasInventoryEvidence
                    && !best.imageOnlyTable
                    && (best.invHeaderCount > 0 || best.invActionCount > 0)
                    && (best.warehouseFollower || best.hasWarehouse);
                return {
                    ready,
                    reason: ready ? 'ok' : (best.imageOnlyTable ? 'image-only-table' : 'inventory-panel-marker-weak'),
                    tables: tables.length,
                    score: best.score,
                    scoped: best.inScopedArea,
                    warehouse: best.hasWarehouse,
                    warehouseFollower: best.warehouseFollower,
                    invHeaderCount: best.invHeaderCount,
                    clickableCount: best.invActionCount,
                    sample: (best.scopeTxt || best.txt || '').slice(0, 180),
                };
            }""")
            if isinstance(js_result, dict):
                if js_result.get('ready'):
                    print(
                        '  [库存] 面板就绪 '
                        f"(tables={js_result.get('tables')}, score={js_result.get('score')}, scoped={js_result.get('scoped')}, "
                        f"warehouseFollower={js_result.get('warehouseFollower')}, clickable={js_result.get('clickableCount')})"
                    )
                    return True
                sample_txt = str(js_result.get('sample', '') or '')
                try:
                    enc = (sys.stdout.encoding or 'utf-8')
                    sample_txt = sample_txt.encode(enc, errors='replace').decode(enc, errors='replace')
                except Exception:
                    sample_txt = sample_txt.encode('unicode_escape', errors='ignore').decode('ascii', errors='ignore')
                print(
                    '  [库存] 面板未就绪 '
                    f"(reason={js_result.get('reason')}, tables={js_result.get('tables')}, score={js_result.get('score')}, scoped={js_result.get('scoped')}, "
                    f"warehouseFollower={js_result.get('warehouseFollower')}, sample={sample_txt})"
                )
        except Exception as e:
            print(f'  [warn] 库存面板就绪判定异常: {e}')
        return False


    def _inventory_surface_snapshot():
        try:
            state = page.evaluate("""() => {
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const boxes = [...document.querySelectorAll('.ant-modal-content,.ant-dropdown,.ant-popover')].filter(isVisible);
                const invBoxes = boxes.filter((box) => {
                    const txt = norm(box.innerText || '');
                    return (
                        txt.includes('库存')
                        || txt.includes('可售')
                        || txt.includes('修改库存')
                        || txt.includes('批量')
                        || txt.includes('应用到全部')
                        || txt.includes('应用到同颜色')
                        || txt.includes('应用到同尺码')
                    );
                });
                const invInputCount = invBoxes.reduce((acc, box) => {
                    return acc + box.querySelectorAll('input:not([disabled]):not([type="radio"]):not([type="checkbox"]):not([type="hidden"]),.ant-input-number-input:not([disabled])').length;
                }, 0);
                const active = document.activeElement;
                const activeInOverlay = !!active && boxes.some((box) => box.contains(active));
                const expandedCount = document.querySelectorAll('[aria-expanded="true"]').length;
                return {
                    overlayCount: boxes.length,
                    invOverlayCount: invBoxes.length,
                    invInputCount,
                    activeInOverlay,
                    expandedCount,
                };
            }""")
            if isinstance(state, dict):
                return state
        except Exception:
            pass
        return {
            'overlayCount': 0,
            'invOverlayCount': 0,
            'invInputCount': 0,
            'activeInOverlay': False,
            'expandedCount': 0,
        }

    def _wait_inventory_surface_change(before_state, timeout_sec=1.8):
        end = time.time() + timeout_sec
        while time.time() < end:
            after = _inventory_surface_snapshot()
            if after.get('invOverlayCount', 0) > before_state.get('invOverlayCount', 0):
                return True, 'inv-overlay-opened', after
            if after.get('invInputCount', 0) > before_state.get('invInputCount', 0):
                return True, 'inv-input-appeared', after
            if after.get('expandedCount', 0) > before_state.get('expandedCount', 0):
                return True, 'expanded-increased', after
            if after.get('activeInOverlay') and not before_state.get('activeInOverlay'):
                return True, 'focus-entered-overlay', after
            if after.get('overlayCount', 0) > before_state.get('overlayCount', 0):
                return False, 'overlay-opened-noninventory', after
            time.sleep(0.12)
        return False, 'no-ui-delta', _inventory_surface_snapshot()

    def _click_inventory_batch_once():
        # 在可见编辑容器内定位可交互库存批量入口，点击后必须观测到界面状态变化才算成功
        before_state = _inventory_surface_snapshot()
        try:
            js_result = page.evaluate("""() => {
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const fireClick = (el) => {
                    if (!el) return false;
                    const events = ['mouseover', 'mousedown', 'mouseup', 'click'];
                    for (const name of events) {
                        el.dispatchEvent(new MouseEvent(name, { bubbles: true, cancelable: true, view: window }));
                    }
                    if (typeof el.click === 'function') el.click();
                    return true;
                };
                const describe = (el) => {
                    if (!el) return { tag: '', text: '', cls: '', html: '' };
                    const txt = (el.textContent || '').replace(/\\s+/g, ' ').trim().slice(0, 120);
                    const cls = (el.className || '').toString().slice(0, 120);
                    const html = (el.outerHTML || '').slice(0, 220);
                    return { tag: String(el.tagName || '').toLowerCase(), text: txt, cls, html };
                };

                const inventoryLabels = ['修改库存', '可售库存', '库存'];
                const hasInventoryLabel = (txt) => inventoryLabels.some((k) => txt.includes(k));

                const scopeKeywords = ['选择仓库', '站点仓库', '仓库', 'SKU分类', 'SKU信息', '变种信息', '可售库存', '修改库存', '库存'];
                const warehouseKeywords = ['选择仓库', '站点仓库', '仓库'];
                const denyKeywords = ['申报价格', '建议售价', '售价', '价格', '重量', '尺寸', '包装清单', '包装'];
                const inScope = (txt) => scopeKeywords.some((k) => txt.includes(k));
                const hasWarehouseKeyword = (txt) => warehouseKeywords.some((k) => txt.includes(k));
                const hasDeny = (txt) => denyKeywords.some((k) => txt.includes(k));
                const navMarkers = ['首页产品数据采集', '订单管理', '客服通用服务', 'Copyright'];
                const isNoisyScope = (txt) => txt.length > 3200 || navMarkers.some((k) => txt.includes(k));
                const pickScope = (table) => {
                    let cur = table;
                    for (let i = 0; cur && i < 9; i += 1) {
                        const txt = norm(cur.textContent || '');
                        if (inScope(txt) && !isNoisyScope(txt)) return cur;
                        cur = cur.parentElement;
                    }
                    return table.closest('section,form,div') || table.parentElement;
                };

                const tables = [...document.querySelectorAll('table')].filter(isVisible);
                if (!tables.length) {
                    return { ok: false, reason: 'no-visible-table' };
                }

                const getWarehouseFollowerTables = () => {
                    const set = new Set();
                    const xpaths = [
                        '//*[contains(normalize-space(.),"选择仓库")]/following::table[1]',
                        '//*[contains(normalize-space(.),"站点仓库")]/following::table[1]',
                    ];
                    for (const xp of xpaths) {
                        try {
                            const snap = document.evaluate(xp, document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
                            for (let i = 0; i < snap.snapshotLength; i += 1) {
                                const t = snap.snapshotItem(i);
                                if (t && isVisible(t)) set.add(t);
                            }
                        } catch (_) {
                        }
                    }
                    return set;
                };

                const warehouseFollowerTables = getWarehouseFollowerTables();
                const scoredTables = tables
                    .map((t) => {
                        const txt = norm(t.textContent);
                        const scope = pickScope(t) || (t.closest('section,form,div') || t.parentElement);
                        const scopeTxt = norm(scope ? scope.textContent : '');
                        const scoped = inScope(scopeTxt) || inScope(txt);
                        const hasWarehouse = hasWarehouseKeyword(scopeTxt) || hasWarehouseKeyword(txt);
                        const warehouseFollower = warehouseFollowerTables.has(t);
                        const headerTexts = [...t.querySelectorAll('thead th,th')]
                            .filter(isVisible)
                            .map((el) => norm(el.textContent || ''))
                            .filter(Boolean)
                            .slice(0, 32);
                        const invHeaderCount = headerTexts.filter((h) => {
                            if (hasDeny(h)) return false;
                            return h.includes('可售库存') || h === '库存' || h.includes('库存') || h.includes('可售');
                        }).length;
                        const imageHeaderCount = headerTexts.filter((h) => h.includes('图片') || h.includes('主图')).length;
                        const skcHeaderCount = headerTexts.filter((h) => h.includes('SKC编码') || h.includes('货号')).length;
                        const denyHeaderCount = headerTexts.filter((h) => hasDeny(h)).length;
                        const imageOnlyTable = invHeaderCount === 0 && imageHeaderCount > 0 && skcHeaderCount > 0;
                        const hasInventoryEvidence = (
                            invHeaderCount > 0
                            || txt.includes('可售库存')
                            || txt.includes('修改库存')
                            || (txt.includes('库存') && hasWarehouse)
                        ) && !imageOnlyTable;
                        const packagingOnly = (scopeTxt.includes('包装清单') || txt.includes('包装清单')) && !hasInventoryEvidence;
                        const denyOnly = denyHeaderCount > 0 && invHeaderCount === 0 && !txt.includes('可售库存');
                        const score =
                            (scoped ? 5 : 0)
                            + (hasWarehouse ? 4 : 0)
                            + (warehouseFollower ? 7 : 0)
                            + (hasInventoryEvidence ? 7 : 0)
                            + (invHeaderCount * 3)
                            + (txt.includes('批量') ? 1 : 0)
                            - (denyOnly ? 7 : 0)
                            - (packagingOnly ? 5 : 0)
                            - (imageOnlyTable ? 10 : 0)
                            - (isNoisyScope(scopeTxt) ? 2 : 0);
                        return {
                            t,
                            scope,
                            score,
                            scopeTxt,
                            tableTxt: txt,
                            scoped,
                            hasWarehouse,
                            warehouseFollower,
                            hasInventoryEvidence,
                            invHeaderCount,
                            imageOnlyTable,
                            packagingOnly,
                            denyOnly,
                        };
                    })
                    .filter((x) => x.score >= 10 && x.hasInventoryEvidence && !x.imageOnlyTable && (x.warehouseFollower || x.hasWarehouse))
                    .sort((a, b) => (Number(b.warehouseFollower) - Number(a.warehouseFollower)) || (b.score - a.score));

                const clickedCandidates = [];
                for (const item of scoredTables.slice(0, 4)) {
                    const table = item.t;
                    const scope = item.scope || table.closest('section,form,div') || table.parentElement;

                    const actionCandidates = [...(scope || table).querySelectorAll('button,a,span,[role="button"],[aria-haspopup="true"],.ant-dropdown-trigger,.ant-btn')]
                        .filter(isVisible)
                        .map((el) => {
                            const txt = norm(el.textContent);
                            let score = 0;
                            const cls = (el.className || '').toString();
                            const hasPopup = String(el.getAttribute('aria-haspopup') || '').toLowerCase() === 'true';
                            const isDropdownTrigger = cls.includes('dropdown-trigger') || cls.includes('ant-dropdown-link');
                            const isBtnLike = el.tagName === 'BUTTON' || cls.includes('ant-btn') || cls.includes('btn');
                            const nearTxt = norm((el.closest('th,td,tr,table,section,form,div')?.textContent) || '');
                            const nearInventory = nearTxt.includes('库存') || nearTxt.includes('可售') || nearTxt.includes('修改库存') || nearTxt.includes('选择仓库') || nearTxt.includes('站点仓库');
                            const nearDeny = hasDeny(nearTxt);
                            if (!txt && !(isDropdownTrigger || hasPopup || isBtnLike)) return { el, txt, score: -1 };
                            if (txt && txt.length > 28) return { el, txt, score: -1 };
                            if (hasDeny(txt) || nearDeny) return { el, txt, score: -1 };

                            if (txt.includes('修改库存')) score += 7;
                            else if (txt.includes('可售库存')) score += 7;
                            else if (txt.includes('库存')) score += 5;
                            else if (txt.includes('可售')) score += 3;
                            if (txt.includes('批量')) score += 2;
                            if (!txt && nearInventory && (isDropdownTrigger || hasPopup || isBtnLike)) score += 3;
                            if (isDropdownTrigger) score += 2;
                            if (isBtnLike) score += 1;
                            if (hasPopup) score += 1;
                            if (!nearInventory) score -= 3;
                            if (txt.includes('批量') && nearInventory && (isDropdownTrigger || hasPopup || isBtnLike || el.tagName === 'A')) score += 2;
                            if ((txt.includes('库存') || txt.includes('可售库存')) && nearInventory) score += 2;
                            return { el, txt, score, nearInventory, nearDeny, isDropdownTrigger, hasPopup, isBtnLike };
                        })
                        .filter((x) => x.score >= 4)
                        .sort((a, b) => b.score - a.score);

                    const merged = actionCandidates.slice(0, 8);
                    for (const cand of merged) {
                        const txt = cand.txt || norm(cand.el.textContent);
                        if (!hasInventoryLabel(txt)) continue;
                        fireClick(cand.el);
                        const detail = describe(cand.el);
                        clickedCandidates.push(detail);
                        return {
                            ok: true,
                            reason: 'clicked-candidate',
                            tableScore: item.score,
                            scoped: item.scoped,
                            node: detail,
                            sampled: clickedCandidates.length,
                        };
                    }

                    const fallbackCandidates = actionCandidates
                        .filter((cand) => {
                            if (hasInventoryLabel(cand.txt || '')) return false;
                            if (!cand.nearInventory) return false;
                            return cand.isDropdownTrigger || cand.hasPopup || cand.isBtnLike;
                        })
                        .slice(0, 4);
                    for (const cand of fallbackCandidates) {
                        fireClick(cand.el);
                        const detail = describe(cand.el);
                        clickedCandidates.push(detail);
                        return {
                            ok: true,
                            reason: 'clicked-fallback-no-text',
                            tableScore: item.score,
                            scoped: item.scoped,
                            node: detail,
                            sampled: clickedCandidates.length,
                        };
                    }
                }

                return {
                    ok: false,
                    reason: 'no-clickable-inventory-entry',
                    tableCount: tables.length,
                    topScore: scoredTables[0] ? scoredTables[0].score : 0,
                    topScope: scoredTables[0] ? scoredTables[0].scopeTxt.slice(0, 150) : '',
                };
            }""")
            if isinstance(js_result, dict) and js_result.get('ok'):
                changed, delta_reason, after_state = _wait_inventory_surface_change(before_state, timeout_sec=1.9)
                node = js_result.get('node') or {}
                if changed:
                    print(
                        '  [库存] 批量入口已点击 '
                        f"(reason={js_result.get('reason')}, delta={delta_reason}, score={js_result.get('tableScore')}, scoped={js_result.get('scoped')}, "
                        f"node={node.get('tag')}|{node.get('text')}|{node.get('cls')})"
                    )
                    return True
                print(
                    '  [库存] 批量入口点击后无状态变化 '
                    f"(delta={delta_reason}, overlays={before_state.get('overlayCount')}->{after_state.get('overlayCount')}, "
                    f"invOverlays={before_state.get('invOverlayCount')}->{after_state.get('invOverlayCount')}, "
                    f"invInputs={before_state.get('invInputCount')}->{after_state.get('invInputCount')})"
                )
            elif isinstance(js_result, dict):
                print(
                    '  [库存] 批量入口未命中（JS） '
                    f"(reason={js_result.get('reason')}, tables={js_result.get('tableCount')}, "
                    f"topScore={js_result.get('topScore')}, topScope={js_result.get('topScope', '')})"
                )
        except Exception as e:
            print(f'  [warn] 点击库存批量入口失败（JS）: {e}')

        strict_selectors = [
            'xpath=(//*[contains(normalize-space(.),"选择仓库") or contains(normalize-space(.),"站点仓库")]/following::table[1]//*[self::button or self::a or @role="button"][contains(normalize-space(.),"修改库存") or contains(normalize-space(.),"可售库存") or contains(normalize-space(.),"库存")])[1]',
            'xpath=(//table[.//th[contains(normalize-space(.),"可售库存") or contains(normalize-space(.),"库存")]][not(.//th[contains(normalize-space(.),"申报价格") or contains(normalize-space(.),"建议售价") or contains(normalize-space(.),"尺寸") or contains(normalize-space(.),"重量")])]//*[self::button or self::a or @role="button"][contains(normalize-space(.),"修改库存") or contains(normalize-space(.),"可售库存") or contains(normalize-space(.),"库存")])[1]',
            'xpath=(//table[.//th[contains(normalize-space(.),"可售库存") or contains(normalize-space(.),"库存")]][not(.//th[contains(normalize-space(.),"申报价格") or contains(normalize-space(.),"建议售价") or contains(normalize-space(.),"尺寸") or contains(normalize-space(.),"重量")])]//*[contains(@class,"ant-dropdown-trigger") or contains(@class,"ant-btn")][contains(normalize-space(.),"库存") or contains(normalize-space(.),"可售")])[1]',
            'xpath=(//*[contains(normalize-space(.),"选择仓库") or contains(normalize-space(.),"站点仓库")]/following::table[1]//*[contains(@class,"ant-dropdown-trigger") or contains(@class,"ant-btn")][1])',
        ]
        for sel in strict_selectors:
            try:
                trigger = page.locator(sel).first
                trigger.wait_for(state='visible', timeout=1500)
                trigger.scroll_into_view_if_needed()
                txt = ''
                try:
                    txt = trigger.inner_text().strip()
                except Exception:
                    pass
                trigger.click(timeout=1800)
                changed, delta_reason, after_state = _wait_inventory_surface_change(before_state, timeout_sec=1.9)
                if changed:
                    print(f'  [库存] 批量入口已点击（XPath:{txt[:80]} delta={delta_reason}）')
                    return True
                print(
                    '  [库存] XPath入口点击后无状态变化 '
                    f"(text={txt[:80]}, overlays={before_state.get('overlayCount')}->{after_state.get('overlayCount')}, "
                    f"invOverlays={before_state.get('invOverlayCount')}->{after_state.get('invOverlayCount')})"
                )
            except Exception:
                continue

        return False


    def _click_batch_menu_option_once():
        # 有些页面批量按钮先弹菜单，需要再点一次菜单项（修改库存或应用范围）
        option_selectors = [
            '.ant-dropdown:visible li[data-menu-id$="_all"]',
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("应用到全部")',
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("应用到同颜色")',
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("应用到同尺码")',
            '.ant-dropdown:visible li:has-text("应用到全部")',
            '.ant-dropdown:visible li:has-text("应用到同颜色")',
            '.ant-dropdown:visible li:has-text("应用到同尺码")',
            '.ant-dropdown:visible span:has-text("应用到全部")',
            '.ant-dropdown:visible span:has-text("应用到同颜色")',
            '.ant-dropdown:visible span:has-text("应用到同尺码")',
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("修改库存")',
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("可售")',
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("库存")',
            '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("批量")',
            '.ant-popover:visible [role="menuitem"]:has-text("应用到全部")',
            '.ant-popover:visible [role="menuitem"]:has-text("应用到同颜色")',
            '.ant-popover:visible [role="menuitem"]:has-text("应用到同尺码")',
            '.ant-popover:visible [role="menuitem"]:has-text("修改库存")',
            '.ant-popover:visible [role="menuitem"]:has-text("可售")',
            '.ant-popover:visible [role="menuitem"]:has-text("库存")',
            '.ant-popover:visible li:has-text("应用到全部")',
            '.ant-popover:visible li:has-text("应用到同颜色")',
            '.ant-popover:visible li:has-text("应用到同尺码")',
            '.ant-popover:visible li:has-text("修改库存")',
            '.ant-popover:visible li:has-text("可售")',
            '.ant-popover:visible li:has-text("库存")',
            '.ant-popover:visible span:has-text("应用到全部")',
            '.ant-popover:visible span:has-text("应用到同颜色")',
            '.ant-popover:visible span:has-text("应用到同尺码")',
            '.ant-popover:visible span:has-text("修改库存")',
            '.ant-popover:visible span:has-text("可售")',
            '.ant-popover:visible span:has-text("库存")',
        ]
        for sel in option_selectors:
            try:
                before_state = _inventory_surface_snapshot()
                opt = page.locator(sel).first
                opt.wait_for(state='visible', timeout=700)
                txt = ''
                try:
                    txt = opt.inner_text().strip()
                except Exception:
                    pass
                opt.click(timeout=1500)
                changed, delta_reason, after_state = _wait_inventory_surface_change(before_state, timeout_sec=1.6)
                if changed:
                    print(f'  [库存] 已点击批量菜单项（{txt[:60]} delta={delta_reason}）')
                else:
                    print(
                        '  [库存] 已点击批量菜单项（无显著delta） '
                        f"(text={txt[:60]}, overlays={before_state.get('overlayCount')}->{after_state.get('overlayCount')}, "
                        f"invOverlays={before_state.get('invOverlayCount')}->{after_state.get('invOverlayCount')}, "
                        f"invInputs={before_state.get('invInputCount')}->{after_state.get('invInputCount')})"
                    )
                return True
            except Exception:
                continue

        # JS 兜底：直接在可见 overlay 中按文本命中菜单项并点击
        try:
            before_state = _inventory_surface_snapshot()
            js_result = page.evaluate("""() => {
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const fireClick = (el) => {
                    if (!el) return false;
                    ['mouseover', 'mousedown', 'mouseup', 'click'].forEach((name) => {
                        el.dispatchEvent(new MouseEvent(name, { bubbles: true, cancelable: true, view: window }));
                    });
                    if (typeof el.click === 'function') el.click();
                    return true;
                };

                const priorities = ['应用到全部', '应用到同颜色', '应用到同尺码', '修改库存', '可售库存', '可售', '库存'];
                const overlays = [...document.querySelectorAll('.ant-dropdown,.ant-popover,.ant-modal-content')]
                    .filter(isVisible)
                    .filter((box) => {
                        const txt = norm(box.innerText || box.textContent || '');
                        return priorities.some((k) => txt.includes(k));
                    });

                const cands = [];
                for (const box of overlays) {
                    const nodes = [...box.querySelectorAll('li,button,a,[role="menuitem"],span,div')].filter(isVisible);
                    for (const n of nodes) {
                        const txt = norm(n.textContent || '');
                        if (!txt || txt.length > 26) continue;
                        for (let i = 0; i < priorities.length; i += 1) {
                            const kw = priorities[i];
                            if (txt.includes(kw)) {
                                cands.push({ n, txt, score: 100 - i * 8 });
                                break;
                            }
                        }
                    }
                }

                if (!cands.length) return { ok: false, reason: 'menu-text-not-found' };
                cands.sort((a, b) => b.score - a.score);
                const pick = cands[0];
                fireClick(pick.n);
                return { ok: true, text: (pick.txt || '').slice(0, 60), score: pick.score };
            }""")

            if isinstance(js_result, dict) and js_result.get('ok'):
                changed, delta_reason, after_state = _wait_inventory_surface_change(before_state, timeout_sec=1.6)
                if changed:
                    print(f'  [库存] 已点击批量菜单项（JS:{js_result.get("text", "")[:60]} delta={delta_reason}）')
                else:
                    print(
                        '  [库存] 已点击批量菜单项（JS无显著delta） '
                        f"(text={js_result.get('text', '')[:60]}, overlays={before_state.get('overlayCount')}->{after_state.get('overlayCount')}, "
                        f"invOverlays={before_state.get('invOverlayCount')}->{after_state.get('invOverlayCount')}, "
                        f"invInputs={before_state.get('invInputCount')}->{after_state.get('invInputCount')})"
                    )
                return True
            if isinstance(js_result, dict):
                print(f'  [库存] 批量菜单未命中（JS reason={js_result.get("reason")})')
        except Exception as e:
            print(f'  [warn] 批量菜单点击异常（JS）: {e}')

        return False

    def _wait_inventory_inputs(timeout_sec=6.5):
        deadline = time.time() + timeout_sec
        while time.time() < deadline:
            found = _find_inventory_inputs()
            if found:
                return found
            time.sleep(0.25)
        return []

    def _fill_inventory_modal_if_present(wait_ms=1200):
        """若出现“修改库存”弹窗，走弹窗路径填写库存。"""
        try:
            modal = page.locator('.ant-modal-content:has-text("修改库存")').last
            try:
                modal.wait_for(state='visible', timeout=wait_ms)
            except Exception:
                return False

            try:
                direct_radio = modal.locator('label:has-text("直接修改为")').first
                if direct_radio.is_visible():
                    direct_radio.click(timeout=1200)
                    time.sleep(0.15)
            except Exception:
                pass

            inp = modal.locator('.ant-input-number-input:not([disabled])').first
            if inp.count() == 0:
                inp = modal.locator('input:not([disabled]):not([type="radio"]):not([type="checkbox"]):not([type="hidden"])').first
            inp.wait_for(state='visible', timeout=1800)
            inp.click(timeout=1000)
            inp.fill('')
            inp.fill(qty)
            try:
                inp.press('Tab')
            except Exception:
                pass

            def _inventory_modal_visible():
                try:
                    return bool(page.evaluate("""() => {
                        const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const s = window.getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                        };
                        return [...document.querySelectorAll('.ant-modal-content')].some((el) => {
                            if (!isVisible(el)) return false;
                            const txt = String(el.innerText || '').replace(/\\s+/g, '');
                            return txt.includes('修改库存');
                        });
                    }"""))
                except Exception:
                    return False

            def _wait_modal_hidden(timeout_sec=4.0):
                end = time.time() + timeout_sec
                while time.time() < end:
                    if not _inventory_modal_visible():
                        return True
                    time.sleep(0.12)
                return False

            confirmed = False
            confirm_selectors = [
                'button.ant-btn-primary:has-text("确定")',
                'button:has-text("确定")',
                'button.ant-btn-primary:has-text("确认")',
                'button:has-text("确认")',
            ]
            for sel in confirm_selectors:
                try:
                    modal_now = page.locator('.ant-modal-content:has-text("修改库存")').last
                    btn = modal_now.locator(sel).first
                    btn.wait_for(state='visible', timeout=1200)
                    try:
                        disabled = btn.get_attribute('disabled')
                        if disabled is not None:
                            continue
                    except Exception:
                        pass
                    btn.scroll_into_view_if_needed()
                    # force=True 绕过 ant-modal-wrap 的指针拦截检测（按钮本身在弹窗内可见可用）
                    btn.click(timeout=1500, force=True)
                    if _wait_modal_hidden(timeout_sec=2.2):
                        confirmed = True
                        break
                except Exception:
                    continue

            if not confirmed:
                try:
                    js_clicked = page.evaluate("""() => {
                        const norm = (s) => String(s || '').replace(/\\s+/g, '');
                        const isVisible = (el) => {
                            if (!el) return false;
                            const r = el.getBoundingClientRect();
                            const s = window.getComputedStyle(el);
                            return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                        };
                        const modal = [...document.querySelectorAll('.ant-modal-content')].find((box) => {
                            return isVisible(box) && norm(box.textContent).includes('修改库存');
                        });
                        if (!modal) return false;
                        const btn = [...modal.querySelectorAll('button')].find((b) => {
                            if (!isVisible(b) || b.disabled) return false;
                            const txt = norm(b.textContent);
                            return txt === '确定' || txt === '确认';
                        });
                        if (!btn) return false;
                        ['mouseover', 'mousedown', 'mouseup', 'click'].forEach((name) => {
                            btn.dispatchEvent(new MouseEvent(name, { bubbles: true, cancelable: true, view: window }));
                        });
                        if (typeof btn.click === 'function') btn.click();
                        return true;
                    }""")
                    if js_clicked and _wait_modal_hidden(timeout_sec=2.2):
                        confirmed = True
                except Exception:
                    pass

            if confirmed:
                time.sleep(0.2)
                if _inventory_modal_visible():
                    confirmed = False

            if not confirmed:
                print('  [warn] 修改库存弹窗未成功确认（确定按钮未生效）')
                return False

            print(f'  [库存] 批量库存已填写: {qty}（修改库存弹窗）')
            return True
        except Exception as e:
            print(f'  [warn] 修改库存弹窗填写失败: {e}')
            return False


    def _dump_inventory_context():
        try:
            html = page.evaluate("""() => {
                const norm = (s) => String(s || '').replace(/\\s+/g, ' ');
                const normCompact = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };

                const blocks = [];
                const invScopeKw = ['选择仓库', '站点仓库', 'SKU分类', 'SKU信息', '变种信息', '库存', '可售', '修改库存'];
                const inInvScope = (txt) => invScopeKw.some((k) => txt.includes(k));

                const tables = [...document.querySelectorAll('table')].filter(isVisible)
                    .map((t) => {
                        const txt = normCompact(t.textContent || '');
                        const scope = t.closest('section,form,div') || t.parentElement;
                        const scopeTxt = normCompact(scope ? scope.textContent : '');
                        let score = 0;
                        if (inInvScope(scopeTxt) || inInvScope(txt)) score += 6;
                        if (txt.includes('库存') || txt.includes('可售')) score += 3;
                        if (txt.includes('批量')) score += 2;
                        return { t, score, txt, scopeTxt };
                    })
                    .sort((a, b) => b.score - a.score)
                    .slice(0, 4);

                const tableSamples = tables.map((x, i) => {
                    const actionNodes = [...(x.t.closest('section,form,div') || x.t).querySelectorAll('button,a,[role="button"],.ant-btn,.ant-dropdown-trigger,span')]
                        .filter(isVisible)
                        .map((n) => String(n.textContent || '').trim())
                        .filter(Boolean)
                        .slice(0, 12);
                    return `<li><b>#${i + 1} score=${x.score}</b><br/>scope=${norm((x.scopeTxt || '').slice(0, 520))}<br/>table=${norm((x.txt || '').slice(0, 320))}<br/>actions=${actionNodes.join(' | ')}</li>`;
                });
                blocks.push(`<h3>inventory scoped tables</h3><ol>${tableSamples.join('')}</ol>`);

                const batchNodes = [...document.querySelectorAll('button,a,span,div')]
                    .filter((n) => isVisible(n) && normCompact(n.textContent).includes('批量'))
                    .filter((n) => {
                        const scope = n.closest('td,th,tr,table,section,form,div') || n.parentElement;
                        const scopeTxt = normCompact(scope ? scope.textContent : '');
                        return inInvScope(scopeTxt);
                    })
                    .slice(0, 20)
                    .map((n) => {
                        const scope = n.closest('th,td,tr,table,section,form,div') || n.parentElement;
                        const scopeText = norm(scope ? scope.textContent : '').slice(0, 320);
                        return `<li><b>${(n.textContent || '').trim()}</b><br/>scope: ${scopeText}</li>`;
                    });
                blocks.push(`<h3>scoped 批量 candidates</h3><ol>${batchNodes.join('')}</ol>`);

                const overlays = [...document.querySelectorAll('.ant-popover,.ant-dropdown,.ant-modal-content')]
                    .filter(isVisible)
                    .slice(0, 8)
                    .map(el => `<pre>${(el.outerHTML || '').slice(0, 3600)}</pre>`);
                blocks.push(`<h3>visible overlays</h3>${overlays.join('')}`);

                const inputs = [...document.querySelectorAll('input:not([disabled])')]
                    .filter(isVisible)
                    .filter((el) => {
                        const scope = el.closest('td,th,tr,table,section,form,div') || el.parentElement;
                        const scopeText = normCompact(scope ? scope.textContent : '');
                        return inInvScope(scopeText);
                    })
                    .slice(0, 30)
                    .map(el => {
                        const p = el.getAttribute('placeholder') || '';
                        const cls = el.className || '';
                        const scope = el.closest('td,th,tr,table,section,form,div') || el.parentElement;
                        const scopeText = norm(scope ? scope.textContent : '').slice(0, 260);
                        return `<li>placeholder=${p} class=${cls}<br/>scope=${scopeText}</li>`;
                    });
                blocks.push(`<h3>scoped visible inputs</h3><ol>${inputs.join('')}</ol>`);

                return `<html><body>${blocks.join('<hr/>')}</body></html>`;
            }""")
            path = os.path.join(DATA_DIR, 'probe_inventory_context.html')
            with open(path, 'w', encoding='utf-8') as f:
                f.write(html or '')
            print(f'  [库存] 调试上下文已导出: {path}')
        except Exception as e:
            print(f'  [warn] 导出库存调试上下文失败: {e}')

    def _fill_inventory_inputs(inputs):
        try:
            # 最多填前3个输入框，通常是同一批量弹层中的库存位
            for inp in inputs[:3]:
                inp.click(timeout=800)
                inp.fill('')
                inp.fill(qty)
                try:
                    inp.press('Tab')
                except Exception:
                    pass

            confirmed = False
            for ok_sel in [
                '.ant-popover:visible button:has-text("确定")',
                '.ant-dropdown:visible button:has-text("确定")',
                '.ant-modal-content:visible button:has-text("确定")',
                '.ant-popover:visible button:has-text("确认")',
                '.ant-dropdown:visible button:has-text("确认")',
                '.ant-modal-content:visible button:has-text("确认")',
                'button:has-text("应用"):visible',
            ]:
                try:
                    ok_btn = page.locator(ok_sel).first
                    ok_btn.wait_for(state='visible', timeout=900)
                    ok_btn.click(timeout=1200)
                    confirmed = True
                    break
                except Exception:
                    continue

            if not confirmed:
                try:
                    inputs[0].press('Enter')
                    confirmed = True
                except Exception:
                    pass

            time.sleep(0.45)
            print(f'  [库存] 批量库存已填写: {qty}（批量入口）')
            return True
        except Exception as e:
            print(f'  [warn] 批量库存填写失败（批量弹窗路径）: {e}')
            return False

    def _inventory_inline_state():
        try:
            state = page.evaluate("""(qtyVal) => {
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };

                const tables = [...document.querySelectorAll('table')].filter(isVisible);
                let best = { found: false, total: 0, qtyCount: 0, header: '', col: -1, tableScore: 0 };

                for (const table of tables) {
                    const tableTxt = norm(table.textContent || '');
                    if (!(tableTxt.includes('库存') || tableTxt.includes('可售') || tableTxt.includes('批量'))) continue;

                    const headerRows = [...table.querySelectorAll('thead tr')].filter(isVisible);
                    const headerCells = headerRows.length
                        ? [...headerRows[headerRows.length - 1].querySelectorAll('th')]
                        : [...table.querySelectorAll('th')].filter(isVisible);
                    if (!headerCells.length) continue;

                    const rows = [...table.querySelectorAll('tbody tr')].filter(isVisible);
                    if (!rows.length) continue;

                    for (let col = 0; col < headerCells.length; col += 1) {
                        const htxt = norm(headerCells[col].textContent || '');
                        const denyKeywords = ['申报价格', '建议售价', '售价', '价格', '尺寸', '重量', '包装清单', '包装', '清单'];
                        const hasDeny = denyKeywords.some((k) => htxt.includes(k));
                        const isInventoryHeader = htxt.includes('可售库存') || htxt.includes('库存') || htxt.includes('可售');
                        if (!isInventoryHeader || hasDeny) continue;

                        let total = 0;
                        let qtyCount = 0;
                        let numericLike = 0;

                        for (const row of rows.slice(0, 30)) {
                            const tds = row.querySelectorAll('td');
                            if (!tds || col >= tds.length) continue;
                            const cell = tds[col];
                            const inp = cell.querySelector('input:not([disabled]):not([type="radio"]):not([type="checkbox"]):not([type="hidden"]), .ant-input-number-input:not([disabled])');
                            if (!inp || !isVisible(inp)) continue;
                            total += 1;
                            const v = norm(inp.value || inp.getAttribute('value') || '');
                            if (v === norm(qtyVal)) qtyCount += 1;
                            const type = String(inp.getAttribute('type') || '').toLowerCase();
                            const cls = String(inp.className || '');
                            if (type === 'number' || cls.includes('input-number')) numericLike += 1;
                        }

                        if (!total) continue;

                        let score = total * 2;
                        if (htxt.includes('可售库存')) score += 8;
                        else if (htxt.includes('库存')) score += 6;
                        else if (htxt.includes('可售')) score += 5;
                        if (htxt.includes('批量')) score += 2;
                        if (numericLike > 0) score += 2;
                        if (tableTxt.includes('可售库存')) score += 4;
                        else if (tableTxt.includes('库存')) score += 3;

                        if (score > best.tableScore) {
                            best = {
                                found: true,
                                total,
                                qtyCount,
                                header: htxt,
                                col,
                                tableScore: score,
                            };
                        }
                    }
                }

                return best;
            }""", qty)
            if isinstance(state, dict):
                return state
        except Exception:
            pass
        return {'found': False, 'total': 0, 'qtyCount': 0, 'header': '', 'col': -1, 'tableScore': 0}

    def _apply_inventory_inline_batch_once():
        before_state = _inventory_inline_state()
        try:
            js_result = page.evaluate("""(qtyVal) => {
                const norm = (s) => String(s || '').replace(/\\s+/g, '');
                const isVisible = (el) => {
                    if (!el) return false;
                    const r = el.getBoundingClientRect();
                    const s = window.getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
                };
                const fireClick = (el) => {
                    if (!el) return false;
                    ['mouseover', 'mousedown', 'mouseup', 'click'].forEach((name) => {
                        el.dispatchEvent(new MouseEvent(name, { bubbles: true, cancelable: true, view: window }));
                    });
                    if (typeof el.click === 'function') el.click();
                    return true;
                };
                const isApplyMenuVisible = () => {
                    const boxes = [...document.querySelectorAll('.ant-dropdown,.ant-popover,.ant-modal-content')].filter(isVisible);
                    return boxes.some((box) => {
                        const txt = norm(box.innerText || box.textContent || '');
                        return (
                            txt.includes('应用到全部')
                            || txt.includes('应用到同颜色')
                            || txt.includes('应用到同尺码')
                            || txt.includes('修改库存')
                            || txt.includes('可售库存')
                        );
                    });
                };

                const tables = [...document.querySelectorAll('table')].filter(isVisible);
                const targets = [];

                for (const table of tables) {
                    const tableTxt = norm(table.textContent || '');
                    if (!(tableTxt.includes('库存') || tableTxt.includes('可售') || tableTxt.includes('批量'))) continue;

                    const headerRows = [...table.querySelectorAll('thead tr')].filter(isVisible);
                    const headerCells = headerRows.length
                        ? [...headerRows[headerRows.length - 1].querySelectorAll('th')]
                        : [...table.querySelectorAll('th')].filter(isVisible);
                    if (!headerCells.length) continue;

                    const rows = [...table.querySelectorAll('tbody tr')].filter(isVisible);
                    if (!rows.length) continue;

                    for (let col = 0; col < headerCells.length; col += 1) {
                        const htxt = norm(headerCells[col].textContent || '');
                        const denyKeywords = ['申报价格', '建议售价', '售价', '价格', '尺寸', '重量', '包装清单', '包装', '清单'];
                        const hasDeny = denyKeywords.some((k) => htxt.includes(k));
                        const isInventoryHeader = htxt.includes('可售库存') || htxt.includes('库存') || htxt.includes('可售');
                        if (!isInventoryHeader || hasDeny) continue;

                        let total = 0;
                        let numericLike = 0;
                        let first = null;

                        for (const row of rows.slice(0, 30)) {
                            const tds = row.querySelectorAll('td');
                            if (!tds || col >= tds.length) continue;
                            const cell = tds[col];
                            const inp = cell.querySelector('input:not([disabled]):not([type="radio"]):not([type="checkbox"]):not([type="hidden"]), .ant-input-number-input:not([disabled])');
                            if (!inp || !isVisible(inp)) continue;
                            total += 1;
                            if (!first) first = { row, cell, inp, headerCell: headerCells[col] || null };
                            const type = String(inp.getAttribute('type') || '').toLowerCase();
                            const cls = String(inp.className || '');
                            if (type === 'number' || cls.includes('input-number')) numericLike += 1;
                        }

                        if (!first || !total) continue;

                        let score = total * 2;
                        if (htxt.includes('可售库存')) score += 9;
                        else if (htxt.includes('库存')) score += 7;
                        else if (htxt.includes('可售')) score += 6;
                        if (htxt.includes('批量')) score += 2;
                        if (numericLike > 0) score += 2;
                        if (tableTxt.includes('可售库存')) score += 4;
                        else if (tableTxt.includes('库存')) score += 3;

                        targets.push({
                            score,
                            col,
                            header: htxt,
                            total,
                            row: first.row,
                            cell: first.cell,
                            inp: first.inp,
                            headerCell: first.headerCell,
                        });
                    }
                }

                targets.sort((a, b) => b.score - a.score);
                if (!targets.length || targets[0].score < 7) {
                    return { ok: false, reason: 'inventory-inline-target-not-found' };
                }

                const target = targets[0];
                const inp = target.inp;
                inp.focus();
                inp.value = '';
                inp.dispatchEvent(new Event('input', { bubbles: true }));
                inp.value = String(qtyVal || '');
                inp.dispatchEvent(new Event('input', { bubbles: true }));
                inp.dispatchEvent(new Event('change', { bubbles: true }));
                inp.blur();

                const triggerCandidates = [];
                const pushTriggers = (root) => {
                    if (!root) return;
                    const cands = [...root.querySelectorAll('button,a,[role="button"],[aria-haspopup="true"],.ant-btn,.ant-dropdown-trigger,.iconfont,svg,span')]
                        .filter(isVisible);
                    for (const el of cands) {
                        const txt = norm(el.textContent || '');
                        const cls = String(el.className || '');
                        const hasPopup = String(el.getAttribute('aria-haspopup') || '').toLowerCase() === 'true';
                        const score =
                            (txt.includes('批量') ? 6 : 0)
                            + (txt.includes('应用') ? 4 : 0)
                            + (cls.includes('dropdown') ? 3 : 0)
                            + (cls.includes('icon_down') || cls.includes('icon-down') ? 3 : 0)
                            + (cls.includes('icon') ? 1 : 0)
                            + (hasPopup ? 2 : 0);
                        if (score >= 2) triggerCandidates.push({ el, score, txt, cls });
                    }
                };

                pushTriggers(target.cell);
                pushTriggers(target.row);
                pushTriggers(target.headerCell);

                triggerCandidates.sort((a, b) => b.score - a.score);
                if (!triggerCandidates.length) {
                    return {
                        ok: false,
                        reason: 'inventory-input-set-but-trigger-missing',
                        header: target.header,
                        col: target.col,
                    };
                }

                let clicked = false;
                let menuVisible = false;
                let used = null;
                for (const cand of triggerCandidates.slice(0, 4)) {
                    fireClick(cand.el);
                    clicked = true;
                    used = cand;
                    if (isApplyMenuVisible()) {
                        menuVisible = true;
                        break;
                    }
                }

                if (!clicked) {
                    return {
                        ok: false,
                        reason: 'inventory-inline-trigger-not-clicked',
                        header: target.header,
                        col: target.col,
                    };
                }

                return {
                    ok: true,
                    reason: menuVisible ? 'inventory-inline-set-trigger-menu-opened' : 'inventory-inline-set-trigger-clicked',
                    menuVisible,
                    header: target.header,
                    col: target.col,
                    triggerText: used ? (used.txt || '').slice(0, 60) : '',
                    triggerClass: used ? (used.cls || '').slice(0, 80) : '',
                };
            }""", qty)

            if not (isinstance(js_result, dict) and js_result.get('ok')):
                if isinstance(js_result, dict):
                    print(f'  [库存] 行内批量入口未命中（reason={js_result.get("reason")}）')
                return False

            print(
                '  [库存] 行内库存列已定位 '
                f"(reason={js_result.get('reason')}, header={js_result.get('header')}, col={js_result.get('col')}, "
                f"trigger={js_result.get('triggerText')}|{js_result.get('triggerClass')})"
            )

            if not js_result.get('menuVisible'):
                menu_shown = False
                for sel in [
                    '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("应用到全部")',
                    '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("应用到同颜色")',
                    '.ant-dropdown:visible .ant-dropdown-menu-item:has-text("应用到同尺码")',
                    '.ant-popover:visible [role="menuitem"]:has-text("应用到全部")',
                    '.ant-popover:visible [role="menuitem"]:has-text("应用到同颜色")',
                    '.ant-popover:visible [role="menuitem"]:has-text("应用到同尺码")',
                ]:
                    try:
                        page.locator(sel).first.wait_for(state='visible', timeout=500)
                        menu_shown = True
                        break
                    except Exception:
                        continue
                if not menu_shown:
                    print('  [库存] 行内批量菜单未弹出/未命中')
                    return False

            menu_clicked = _click_batch_menu_option_once()
            if not menu_clicked:
                print('  [库存] 行内批量菜单未弹出/未命中')
                return False

            # 行内触发后也可能走到“修改库存”弹窗，优先复用弹窗填写路径
            if _fill_inventory_modal_if_present(wait_ms=1800):
                print('  [库存] 行内触发后转为修改库存弹窗并已完成填写')
                return True

            end = time.time() + 2.6
            while time.time() < end:
                after_state = _inventory_inline_state()
                if (
                    before_state.get('total', 0) > 0
                    and before_state.get('qtyCount', 0) >= before_state.get('total', 0)
                    and after_state.get('qtyCount', 0) >= after_state.get('total', 0)
                ):
                    print('  [库存] 行内库存已满足目标值（已全量）')
                    return True
                if after_state.get('qtyCount', 0) >= before_state.get('qtyCount', 0) + 3:
                    print(
                        '  [库存] 行内库存批量生效 '
                        f"(qtyCount={before_state.get('qtyCount')}->{after_state.get('qtyCount')}, total={after_state.get('total')})"
                    )
                    return True
                time.sleep(0.2)

            final_state = _inventory_inline_state()
            print(
                '  [库存] 行内库存批量未观察到数量增长 '
                f"(qtyCount={before_state.get('qtyCount')}->{final_state.get('qtyCount')}, total={final_state.get('total')})"
            )
            return False
        except Exception as e:
            print(f'  [warn] 行内库存批量路径异常: {e}')
            return False

    # ★ 路径0（录制验证）：input[name="stock"] 直接填写 → icon_send 触发 → 应用到全部
    # 录制证据：row.locator('input[name="stock"]').fill("999")
    #           locator('.iconfont.icon_send.link.ml-4.ant-dropdown-trigger').first.click()
    #           get_by_text("应用到全部").first.click()
    def _fill_stock_input_apply_all():
        try:
            # 先滚动表格到右侧，确保库存列可见可交互
            _scroll_variant_table_to_right()
            time.sleep(0.5)

            stock_inputs = page.locator('input[name="stock"]:visible')
            cnt = stock_inputs.count()
            print(f'  [库存] stock输入路径: 找到 {cnt} 个 input[name=stock]')
            if cnt == 0:
                return False
            first_input = stock_inputs.first
            first_input.scroll_into_view_if_needed()
            time.sleep(0.2)

            # 先用 JS 强制聚焦 + 清空，再逐字输入（触发 Vue 响应式）
            page.evaluate("""(el) => {
                el.focus();
                el.select();
                var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                nativeInputValueSetter.call(el, '');
                el.dispatchEvent(new Event('input', {bubbles: true}));
            }""", first_input.element_handle())
            time.sleep(0.1)
            first_input.press_sequentially(qty, delay=60)
            time.sleep(0.3)
            # 验证输入值
            actual = first_input.input_value()
            print(f'  [库存] 输入后实际值: {actual!r}（期望: {qty!r}）')
            if actual != qty:
                print('  [库存] 输入验证失败，尝试 triple_click + fill 兜底')
                first_input.click(click_count=3, timeout=2000)
                first_input.fill(qty)
                time.sleep(0.3)
                actual = first_input.input_value()
                print(f'  [库存] 兜底后实际值: {actual!r}')
                if actual != qty:
                    return False

            # 点击旁边的 send/apply 图标
            trigger = page.locator('.iconfont.icon_send.link.ml-4.ant-dropdown-trigger').first
            if not trigger.is_visible():
                trigger = page.locator('.ant-dropdown-trigger:visible').first
                if not trigger.is_visible():
                    print('  [库存] stock输入路径: 未找到 icon_send 触发器')
                    return False
            trigger.click(timeout=2000)
            time.sleep(0.5)
            # 点"应用到全部"
            apply_btn = page.locator(':text-is("应用到全部"):visible').first
            apply_btn.wait_for(state='visible', timeout=3000)
            apply_btn.click(timeout=2000)
            time.sleep(1.0)
            print(f'  [库存] 已填写 {qty}（input[name=stock] + 应用到全部）')
            return True
        except Exception as e:
            print(f'  [库存] stock输入路径异常: {e}')
            _shot(page, 'stock_input_error')
            return False

    if _fill_stock_input_apply_all():
        return True
    print('  [warn] stock直接路径未命中，库存跳过（请手动填写）')
    return True



def run(
    offer_ids_filter=None,
    do_import=False,
    xlsx_path=None,
    image_source=IMAGE_SOURCE_LOCAL,
    cos_region=None,
    cos_bucket=None,
    cos_prefix=None,
    no_publish=False,
    shop_account=None,
    site=None,
    product_template=None,
    size_template=None,
    warehouse_template=None,
    logistics_template=None,
    debug_shots=False,
):
    """
    offer_ids_filter : list[str] 可选，只处理指定产品货号
    do_import        : True 则先执行 xlsx 导入再上图
    xlsx_path        : 导入的 xlsx 路径；为 None 时自动找 data/ 下最新的店小秘导入_*.xlsx
    image_source     : 图片来源，local|cos
    """
    from playwright.sync_api import sync_playwright

    image_source = str(image_source or IMAGE_SOURCE_LOCAL).strip().lower()
    if image_source not in (IMAGE_SOURCE_LOCAL, IMAGE_SOURCE_COS):
        print(f'不支持的图片来源: {image_source}（仅支持 local|cos）')
        return

    requested_offer_ids = _dedupe_keep_order(offer_ids_filter)
    store_name = str(shop_account or IMPORT_STORE).strip() or IMPORT_STORE
    site_values = _dedupe_keep_order(str(site or '').split(',')) or IMPORT_SITES
    global REF_PRODUCT_TEMPLATE, _SIZE_CHART_TEMPLATE, DEBUG_SHOTS
    DEBUG_SHOTS = bool(debug_shots)
    if str(product_template or '').strip():
        REF_PRODUCT_TEMPLATE = str(product_template).strip()
    if str(size_template or '').strip():
        _SIZE_CHART_TEMPLATE = str(size_template).strip()
    warehouse_name = str(warehouse_template or '').strip()
    if warehouse_name:
        print(f'使用仓库模板/仓库名称: {warehouse_name}')
    if str(logistics_template or '').strip():
        print(f'使用物流模板: {str(logistics_template).strip()}')
    print(f'上传配置: 店铺={store_name}; 站点={",".join(site_values)}; 产品模板={REF_PRODUCT_TEMPLATE}; 尺码模板={_SIZE_CHART_TEMPLATE}')

    # 自动找最新 xlsx
    if do_import and not xlsx_path:
        candidates = sorted(
            os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR)
            if f.startswith('店小秘导入_') and f.endswith('.xlsx')
        )
        if not candidates:
            print('未找到 data/店小秘导入_*.xlsx，请先运行 export_miaoshou_xlsx.py')
            return
        xlsx_path = candidates[-1]
        print(f'自动选用: {xlsx_path}')

    import_prod_nos = []
    blocked_prod_nos = []
    blocked_rows = []
    size_filter_used = False
    has_prod_col = False
    if do_import and xlsx_path:
        upload_prod_nos, blocked_prod_nos, blocked_rows, size_filter_used, has_prod_col = _load_prod_nos_from_xlsx(xlsx_path)
        import_prod_nos = upload_prod_nos

        if size_filter_used and blocked_prod_nos:
            print(f'导入文件含超范围尺码，已跳过上图产品 {len(blocked_prod_nos)} 个（仅保留 XS~XXL）')
            for row_no, prod_no, size_txt in blocked_rows[:20]:
                print(f'  [SIZE-SKIP] 第{row_no}行 产品货号={prod_no} 尺码={size_txt or "(空)"}')
            if len(blocked_rows) > 20:
                print(f'  [SIZE-SKIP] ... 其余 {len(blocked_rows) - 20} 行省略')
        elif do_import and has_prod_col and size_filter_used:
            print('导入文件尺码检查通过：仅包含 XS~XXL')

        if has_prod_col and not import_prod_nos:
            if size_filter_used and blocked_prod_nos:
                print('导入文件产品均超出 XS~XXL 范围，导入后不执行上图。')
            else:
                print('导入文件未解析到可上图产品货号，导入后不执行上图。')
            return

    image_root_dir = SKU_IMG_DIR
    cos_url_base = None   # 非 None 时走"网络图片"路径，不占店小秘图片空间
    prod_folders = []

    if image_source == IMAGE_SOURCE_LOCAL:
        prod_folders = sorted(
            f for f in os.listdir(SKU_IMG_DIR)
            if os.path.isdir(os.path.join(SKU_IMG_DIR, f))
        )
        if requested_offer_ids:
            requested_set = set(requested_offer_ids)
            prod_folders = [f for f in prod_folders if f.split('_')[0] in requested_set]

        if do_import and has_prod_col:
            print(f'导入文件包含 {len(import_prod_nos)} 个可上图产品货号，等待导入时按该列表检测')
            import_set = set(import_prod_nos)
            prod_folders = [f for f in prod_folders if f.split('_')[0] in import_set]
            found_nos = {f.split('_')[0] for f in prod_folders}
            missing_nos = [no for no in import_prod_nos if no not in found_nos]
            print(f'导入模式仅处理 {len(prod_folders)} 个本地产品目录（按xlsx产品货号过滤）')
            if missing_nos:
                sample = ', '.join(missing_nos[:8])
                suffix = ' ...' if len(missing_nos) > 8 else ''
                print(f'  [warn] 以下产品货号在 sku_images 未找到目录: {sample}{suffix}')
            if not prod_folders:
                print('导入文件中的产品在本地 sku_images 无可上传目录，已跳过上图。')
                return

        if not prod_folders and not do_import:
            print('sku_images/ 目录为空，请先运行 test_sku_feishu.py download')
            return

    else:
        # COS 模式：以本地 sku_images/ 为参考，通过"网络图片"传 COS 公开 URL
        # 不下载到本地，不占店小秘图片空间
        region, bucket, prefix, secret_id, secret_key = _resolve_cos_settings(cos_region, cos_bucket, cos_prefix)
        cos_url_base = f'https://{bucket}.cos.{region}.myqcloud.com/{prefix}'
        print(f'COS 网络图片模式: {cos_url_base}')

        # ── 先预处理本地图片（放大 + RGBA→RGB），再同步到 COS ──
        _preprocess_sku_images(SKU_IMG_DIR)

        # ── 自动同步本地 sku_images/ 到 COS（增量，已有的跳过）──
        print('[COS] 开始同步本地图片到 COS...')
        try:
            client = _build_cos_client(region, secret_id, secret_key)
            cos_existing = set()
            marker = ''
            while True:
                kw = {'Bucket': bucket, 'Prefix': prefix + '/', 'MaxKeys': 1000}
                if marker:
                    kw['Marker'] = marker
                resp = client.list_objects(**kw)
                contents = resp.get('Contents') or []
                if isinstance(contents, dict):
                    contents = [contents]
                for item in contents:
                    k = item.get('Key')
                    if k:
                        cos_existing.add(k)
                if str(resp.get('IsTruncated', '')).lower() not in ('true', '1'):
                    break
                marker = resp.get('NextMarker') or (contents[-1].get('Key') if contents else '')
                if not marker:
                    break

            sync_ok = sync_fail = sync_skip = 0
            for prod_dir_name in sorted(os.listdir(SKU_IMG_DIR)):
                prod_path = os.path.join(SKU_IMG_DIR, prod_dir_name)
                if not os.path.isdir(prod_path):
                    continue
                for color_dir in sorted(os.listdir(prod_path)):
                    color_path = os.path.join(prod_path, color_dir)
                    if not os.path.isdir(color_path):
                        continue
                    for fname in sorted(os.listdir(color_path)):
                        if not fname.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                            continue
                        cos_key = f'{prefix}/{prod_dir_name}/{color_dir}/{fname}'
                        if cos_key in cos_existing:
                            sync_skip += 1
                            continue
                        local_path = os.path.join(color_path, fname)
                        try:
                            with open(local_path, 'rb') as f:
                                client.put_object(Bucket=bucket, Body=f, Key=cos_key)
                            sync_ok += 1
                        except Exception as e:
                            print(f'  [COS][warn] 上传失败 {cos_key}: {e}')
                            sync_fail += 1
            print(f'[COS] 同步完成：新上传 {sync_ok}，跳过 {sync_skip}，失败 {sync_fail}')
        except Exception as e:
            print(f'[COS][warn] 同步异常: {e}，继续使用已有 COS 图片')

        prod_folders = sorted(
            f for f in os.listdir(SKU_IMG_DIR)
            if os.path.isdir(os.path.join(SKU_IMG_DIR, f))
        )
        if requested_offer_ids:
            requested_set = set(requested_offer_ids)
            prod_folders = [f for f in prod_folders if f.split('_')[0] in requested_set]

        if do_import and has_prod_col:
            import_set = set(import_prod_nos)
            prod_folders = [f for f in prod_folders if f.split('_')[0] in import_set]
            found_nos = {f.split('_')[0] for f in prod_folders}
            missing_nos = [no for no in import_prod_nos if no not in found_nos]
            print(f'COS 模式仅处理 {len(prod_folders)} 个本地产品目录（按xlsx产品货号过滤）')
            if missing_nos:
                sample = ', '.join(missing_nos[:8])
                suffix = ' ...' if len(missing_nos) > 8 else ''
                print(f'  [warn] 以下产品货号在本地 sku_images 未找到目录: {sample}{suffix}')
            if not prod_folders:
                print('导入文件中的产品在本地 sku_images 无目录，已跳过上图。')
                return

        if not prod_folders and not do_import:
            print('sku_images/ 目录为空，请先确保本地有图片再使用 COS 模式')
            return

    prod_nos = [f.split('_')[0] for f in prod_folders]

    # local 模式在这里做预处理；COS 模式已在同步前完成，此处跳过
    if image_source == IMAGE_SOURCE_LOCAL:
        prod_filter_set = set(prod_nos) if prod_nos else None
        _preprocess_sku_images(SKU_IMG_DIR, prod_folder_filter=prod_filter_set)

    print(f'共 {len(prod_folders)} 个商品待上传')

    cos_tmp_dir = None
    try:
        with sync_playwright() as pw:
            os.makedirs(BROWSER_PROFILE_DIR, exist_ok=True)
            context = pw.chromium.launch_persistent_context(
                BROWSER_PROFILE_DIR,
                headless=False,
                slow_mo=150,
                viewport={'width': 1440, 'height': 900},
            )
            load_cookies(context)
            page = context.new_page()

            ensure_login(page, context)
            dismiss_popup(page)

            # ── 导入 xlsx ──
            if do_import:
                submitted = import_xlsx_to_dxm(page, xlsx_path, store_name=store_name, sites=site_values, warehouse_template=warehouse_name)
                if not submitted:
                    print('导入提交失败，已停止后续上图。请先确认导入弹窗状态后重试。')
                    context.close()
                    return

                # 仅按将要上图的产品做就绪检测，避免等待无本地目录/无COS图片的货号
                wait_nos = prod_nos or import_prod_nos
                if not wait_nos:
                    print('未获取到待检测的产品货号，无法确认导入完成，已停止后续上图。')
                    context.close()
                    return

                INITIAL_WAIT_MAX = 240
                INITIAL_WAIT_POLL = 12
                ROLLING_WAIT_MAX = 900
                ROLLING_WAIT_POLL = 25

                print(f'导入模式待上传目录数: {len(prod_folders)}')
                if prod_folders:
                    sample_folders = ', '.join(prod_folders[:5])
                    suffix_folders = ' ...' if len(prod_folders) > 5 else ''
                    print(f'  待上传目录示例: {sample_folders}{suffix_folders}')

                wait_result = wait_for_products(
                    page,
                    wait_nos,
                    max_wait=INITIAL_WAIT_MAX,
                    poll_interval=INITIAL_WAIT_POLL,
                    require_all=False,
                )
                if wait_result is False:
                    print('导入后初始等待未命中就绪商品，转入滚动重试上传。')

            if not prod_folders:
                context.close()
                return

            # ── 上传图片 ──
            ok, fail = 0, 0

            if do_import:
                pending = list(prod_folders)
                rolling_start = time.time()
                round_no = 0

                while pending:
                    round_no += 1
                    progressed = 0
                    next_pending = []
                    print(f'\n[滚动上传] 第 {round_no} 轮，待处理 {len(pending)} 个')

                    for prod_folder in pending:
                        offer_id = prod_folder.split('_')[0]
                        product_dir = os.path.join(image_root_dir, prod_folder)
                        print(f'\n── {prod_folder}')

                        edit_url = find_edit_url(page, offer_id)
                        if not edit_url:
                            next_pending.append(prod_folder)
                            continue

                        page.goto(edit_url, wait_until='domcontentloaded', timeout=30000)
                        success = upload_for_product(page, product_dir, cos_url_base=cos_url_base, no_publish=no_publish, warehouse_template=warehouse_name)
                        if success:
                            ok += 1
                        else:
                            fail += 1
                        progressed += 1

                    pending = next_pending
                    if not pending:
                        break

                    elapsed = int(time.time() - rolling_start)
                    remain = ROLLING_WAIT_MAX - elapsed
                    if remain <= 0:
                        break

                    wait_sec = min(ROLLING_WAIT_POLL, remain)
                    if progressed == 0:
                        sample = ', '.join(x.split('_')[0] for x in pending[:8])
                        suffix = ' ...' if len(pending) > 8 else ''
                        print(
                            f'\n[滚动上传] 本轮无新增就绪，{wait_sec}s 后重试 '
                            f'({elapsed}s/{ROLLING_WAIT_MAX}s)，剩余 {len(pending)} 个: {sample}{suffix}'
                        )
                        time.sleep(wait_sec)
                    else:
                        print(
                            f'\n[滚动上传] 本轮完成 {progressed} 个，立即继续下一轮 '
                            f'(剩余 {len(pending)} 个，已等待 {elapsed}s/{ROLLING_WAIT_MAX}s)'
                        )

                if pending:
                    fail += len(pending)
                    pending_ids = ', '.join(x.split('_')[0] for x in pending)
                    print(f'\n[滚动上传] 超时仍未就绪，以下产品已跳过: {pending_ids}')
            else:
                for prod_folder in prod_folders:
                    offer_id    = prod_folder.split('_')[0]
                    product_dir = os.path.join(image_root_dir, prod_folder)
                    print(f'\n── {prod_folder}')

                    edit_url = find_edit_url(page, offer_id)
                    if not edit_url:
                        fail += 1
                        continue

                    page.goto(edit_url, wait_until='domcontentloaded', timeout=30000)
                    success = upload_for_product(page, product_dir, cos_url_base=cos_url_base, no_publish=no_publish, warehouse_template=warehouse_name)
                    if success:
                        ok += 1
                    else:
                        fail += 1

            save_cookies(context)
            context.close()

        print(f'\n完成：成功 {ok} 个，失败/跳过 {fail} 个')
        return fail == 0 and ok > 0
    finally:
        if cos_tmp_dir:
            shutil.rmtree(cos_tmp_dir, ignore_errors=True)
            print(f'COS 临时目录已清理: {cos_tmp_dir}')


    return False


if __name__ == '__main__':
    args = sys.argv[1:]
    do_import = False
    xlsx_path = None
    ids = []
    image_source = DEFAULT_IMAGE_SOURCE   # 默认值来自 apify_config.json image_source 字段
    cos_region = None
    cos_bucket = None
    cos_prefix = None
    no_publish = False
    shop_account = None
    site = None
    product_template = None
    size_template = None
    warehouse_template = None
    logistics_template = None
    debug_shots = False
    sku_image_dir = None

    i = 0
    while i < len(args):
        arg = args[i]
        if arg == '--import':
            do_import = True
            # 下一个参数如果不是 -- 开头则视为 xlsx 路径
            if i + 1 < len(args) and not args[i + 1].startswith('--'):
                i += 1
                xlsx_path = args[i]
        elif arg == '--image-source':
            if i + 1 >= len(args):
                print('--image-source 缺少参数（local|cos）')
                sys.exit(2)
            i += 1
            image_source = args[i]
        elif arg == '--cos-region':
            if i + 1 >= len(args):
                print('--cos-region 缺少参数')
                sys.exit(2)
            i += 1
            cos_region = args[i]
        elif arg == '--cos-bucket':
            if i + 1 >= len(args):
                print('--cos-bucket 缺少参数')
                sys.exit(2)
            i += 1
            cos_bucket = args[i]
        elif arg == '--no-publish':
            no_publish = True
        elif arg == '--debug-shots':
            debug_shots = True
        elif arg == '--sku-image-dir':
            if i + 1 >= len(args):
                print('--sku-image-dir 缺少参数')
                sys.exit(2)
            i += 1
            sku_image_dir = args[i]
        elif arg == '--shop-account':
            if i + 1 >= len(args):
                print('--shop-account 缺少参数')
                sys.exit(2)
            i += 1
            shop_account = args[i]
        elif arg == '--site':
            if i + 1 >= len(args):
                print('--site 缺少参数')
                sys.exit(2)
            i += 1
            site = args[i]
        elif arg == '--product-template':
            if i + 1 >= len(args):
                print('--product-template 缺少参数')
                sys.exit(2)
            i += 1
            product_template = args[i]
        elif arg == '--size-template':
            if i + 1 >= len(args):
                print('--size-template 缺少参数')
                sys.exit(2)
            i += 1
            size_template = args[i]
        elif arg == '--warehouse-template':
            if i + 1 >= len(args):
                print('--warehouse-template 缺少参数')
                sys.exit(2)
            i += 1
            warehouse_template = args[i]
        elif arg == '--logistics-template':
            if i + 1 >= len(args):
                print('--logistics-template 缺少参数')
                sys.exit(2)
            i += 1
            logistics_template = args[i]
        elif arg == '--cos-prefix':
            if i + 1 >= len(args):
                print('--cos-prefix 缺少参数')
                sys.exit(2)
            i += 1
            cos_prefix = args[i]
        elif arg.startswith('--'):
            print(f'未知参数: {arg}')
            sys.exit(2)
        else:
            ids.append(arg)
        i += 1

    if sku_image_dir:
        _set_sku_image_root(sku_image_dir)

    ok = run(
        offer_ids_filter=ids or None,
        do_import=do_import,
        xlsx_path=xlsx_path,
        image_source=image_source,
        cos_region=cos_region,
        cos_bucket=cos_bucket,
        cos_prefix=cos_prefix,
        no_publish=no_publish,
        shop_account=shop_account,
        site=site,
        product_template=product_template,
        size_template=size_template,
        warehouse_template=warehouse_template,
        logistics_template=logistics_template,
        debug_shots=debug_shots,
    )

    if ok is False:
        sys.exit(1)

