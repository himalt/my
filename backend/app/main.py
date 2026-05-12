from __future__ import annotations

import os
import base64
import json
import csv
import mimetypes
import hashlib
import math
import re
import shutil
import subprocess
import sqlite3
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from html import escape, unescape
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

from fastapi import Body, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
import xlrd

def resolve_resource_dir() -> Path:
    frozen_base = getattr(sys, "_MEIPASS", None)
    if frozen_base:
        return Path(frozen_base)
    return Path(__file__).resolve().parents[2]


def resolve_runtime_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parents[2]


RESOURCE_DIR = resolve_resource_dir()
BASE_DIR = resolve_runtime_dir()
DATA_DIR = Path(os.environ.get("UPLOAD_ASSISTANT_DATA_DIR", BASE_DIR / "data"))
IMAGE_DIR = DATA_DIR / "images"
DB_PATH = Path(os.environ.get("UPLOAD_ASSISTANT_DB", DATA_DIR / "app.db"))
FRONTEND_DIR = RESOURCE_DIR / "frontend"
STATIC_DIR = FRONTEND_DIR / "static"
RELEASE_DIR = BASE_DIR / "release"
LOCAL_DATA_DIRECTORIES = {
    "data": DATA_DIR,
    "database_parent": DB_PATH.parent,
    "images": IMAGE_DIR,
    "logs": DATA_DIR / "logs",
    "exports": DATA_DIR / "export",
    "collection_requests": DATA_DIR / "collection_requests",
    "collection_results": DATA_DIR / "collection_results",
    "imports": DATA_DIR / "imports",
    "templates": DATA_DIR / "templates",
    "uploads": DATA_DIR / "uploads",
    "rpa_images": DATA_DIR / "rpa_images",
}
WINDOWS_APP_COS_URL = "https://temutp423-1320875862.cos.ap-beijing.myqcloud.com/temu/release/upload-assistant-windows.zip"
SCRIPT_SOURCE_DIR = Path(r"C:\Users\zyf\.accio\accounts\1753260534\agents\MID-27260534U1775454-06F6F0-2183-EB2CBD\project")
NANOBANANA_SUBMIT_URL = "https://api.wuyinkeji.com/api/async/image_nanoBanana2"
NANOBANANA_DETAIL_URL = "https://api.wuyinkeji.com/api/async/detail"
BUILTIN_API_CONFIGS_PATH = Path(os.environ.get("UPLOAD_ASSISTANT_BUILTIN_API_CONFIGS", RESOURCE_DIR / "config" / "builtin_api_configs.json"))
BUILTIN_SETTINGS_PATH = Path(os.environ.get("UPLOAD_ASSISTANT_BUILTIN_SETTINGS", RESOURCE_DIR / "config" / "builtin_settings.json"))
DEFAULT_IMAGE_PROMPT = "纯白底，仅商品本身，无模特，无人体部位。平铺拍摄，影棚照明，干净产品图，高分辨率。保留商品结构细节，去掉腰带及可拆卸配饰。背景色纯白#FFFFFF。"
VISION_CACHE_VERSION = "doubao-color-v2"
COLLECTION_TASK_QUEUE_LOCK = threading.Lock()
COLLECTION_TASK_QUEUE_ACTIVE = False
COLLECTION_IMPORT_TASK_QUEUE_LOCK = threading.Lock()
COLLECTION_IMPORT_TASK_QUEUE_ACTIVE = False
COLLECTION_IMAGE_TASK_QUEUE_LOCK = threading.Lock()
COLLECTION_IMAGE_TASK_QUEUE_ACTIVE = False
PROCESSING_TITLE_TASK_QUEUE_LOCK = threading.Lock()
PROCESSING_TITLE_TASK_QUEUE_ACTIVE = False
PROCESSING_FIELD_TASK_QUEUE_LOCK = threading.Lock()
PROCESSING_FIELD_TASK_QUEUE_ACTIVE = False
COLLECTION_TARGET_MAX = 200
COLLECTION_BATCH_TARGET_MAX = 1000
ONEBOUND_PAGE_SIZE_MAX = 50

PLATFORM_COLOR_GROUPS = [("白色系", ["白色", "米白色", "乳白色", "象牙白"]), ("红色系", ["红色", "桔红色", "玫红色", "粉红色", "桃色", "蔷薇色", "深粉红", "胭脂色", "藕色", "西瓜红", "酒红色", "猩红色", "亮粉色", "洋红色", "珊瑚色", "橙红色", "砖红色", "深红色"]), ("黑色系", ["黑色"]), ("花色系", ["花色"]), ("黄色系", ["卡其色", "姜黄色", "明黄色", "杏色", "柠檬黄", "荧光黄", "金色", "香槟色", "黄色", "浅黄色", "小麦色", "橙色", "土黄色", "向日葵色", "肤色"]), ("灰色系", ["深灰色", "浅灰色", "奶奶灰", "灰色", "铅色", "石板灰", "银灰色", "深空灰", "石墨色"]), ("蓝色系", ["天蓝色", "孔雀蓝", "宝蓝色", "浅蓝色", "蓝色", "湖蓝色", "亮钢蓝", "道奇蓝", "深天蓝", "军服蓝", "蔚蓝色", "淡青色", "青色", "暗青色", "藏青色", "粉蓝色", "海蓝色", "中蓝色", "深蓝色"]), ("绿色系", ["军绿色", "墨绿色", "浅绿色", "绿色", "翠绿色", "孔雀绿", "荧光绿", "绿黄色", "草绿色", "黄绿色", "碧绿", "绿宝石", "海洋绿", "橄榄色", "橄榄绿", "抹茶色"]), ("透明系", ["透明"]), ("紫色系", ["浅紫色", "深紫色", "紫红色", "紫罗兰", "暗紫罗兰", "紫色", "深洋红色", "紫兰色", "熏衣草淡紫"]), ("棕色系", ["咖啡色", "巧克力色", "栗色", "深卡其布色", "浅棕色", "深棕色", "褐色", "棕褐色", "驼色", "橄榄棕色", "红褐色", "茶色", "亚麻色", "黄土赭色", "琥珀色", "小豆色", "焦糖色", "肉桂色", "玫瑰棕色", "玫瑰金", "古铜色"]), ("其他", ["混合色"])]
PLATFORM_STANDARD_COLORS = [color for _, colors in PLATFORM_COLOR_GROUPS for color in colors]
COLOR_ALIAS_ENTRIES = [("白色", "白色"), ("白", "白色"), ("米白色", "米白色"), ("米白", "米白色"), ("乳白色", "乳白色"), ("乳白", "乳白色"), ("象牙白", "象牙白"), ("红色", "红色"), ("红", "红色"), ("桔红色", "桔红色"), ("桔红", "桔红色"), ("玫红色", "玫红色"), ("玫红", "玫红色"), ("粉红色", "粉红色"), ("粉红", "粉红色"), ("桃色", "桃色"), ("桃", "桃色"), ("蔷薇色", "蔷薇色"), ("蔷薇", "蔷薇色"), ("深粉红", "深粉红"), ("胭脂色", "胭脂色"), ("胭脂", "胭脂色"), ("藕色", "藕色"), ("藕", "藕色"), ("西瓜红", "西瓜红"), ("酒红色", "酒红色"), ("酒红", "酒红色"), ("猩红色", "猩红色"), ("猩红", "猩红色"), ("亮粉色", "亮粉色"), ("亮粉", "亮粉色"), ("洋红色", "洋红色"), ("洋红", "洋红色"), ("珊瑚色", "珊瑚色"), ("珊瑚", "珊瑚色"), ("橙红色", "橙红色"), ("橙红", "橙红色"), ("砖红色", "砖红色"), ("砖红", "砖红色"), ("深红色", "深红色"), ("深红", "深红色"), ("黑色", "黑色"), ("黑", "黑色"), ("花色", "花色"), ("花", "花色"), ("卡其色", "卡其色"), ("卡其", "卡其色"), ("姜黄色", "姜黄色"), ("姜黄", "姜黄色"), ("明黄色", "明黄色"), ("明黄", "明黄色"), ("杏色", "杏色"), ("杏", "杏色"), ("柠檬黄", "柠檬黄"), ("荧光黄", "荧光黄"), ("金色", "金色"), ("金", "金色"), ("香槟色", "香槟色"), ("香槟", "香槟色"), ("黄色", "黄色"), ("黄", "黄色"), ("浅黄色", "浅黄色"), ("浅黄", "浅黄色"), ("小麦色", "小麦色"), ("小麦", "小麦色"), ("橙色", "橙色"), ("橙", "橙色"), ("土黄色", "土黄色"), ("土黄", "土黄色"), ("向日葵色", "向日葵色"), ("向日葵", "向日葵色"), ("肤色", "肤色"), ("肤", "肤色"), ("深灰色", "深灰色"), ("深灰", "深灰色"), ("浅灰色", "浅灰色"), ("浅灰", "浅灰色"), ("奶奶灰", "奶奶灰"), ("灰色", "灰色"), ("灰", "灰色"), ("铅色", "铅色"), ("铅", "铅色"), ("石板灰", "石板灰"), ("银灰色", "银灰色"), ("银灰", "银灰色"), ("深空灰", "深空灰"), ("石墨色", "石墨色"), ("石墨", "石墨色"), ("天蓝色", "天蓝色"), ("天蓝", "天蓝色"), ("孔雀蓝", "孔雀蓝"), ("宝蓝色", "宝蓝色"), ("宝蓝", "宝蓝色"), ("浅蓝色", "浅蓝色"), ("浅蓝", "浅蓝色"), ("蓝色", "蓝色"), ("蓝", "蓝色"), ("湖蓝色", "湖蓝色"), ("湖蓝", "湖蓝色"), ("亮钢蓝", "亮钢蓝"), ("道奇蓝", "道奇蓝"), ("深天蓝", "深天蓝"), ("军服蓝", "军服蓝"), ("蔚蓝色", "蔚蓝色"), ("蔚蓝", "蔚蓝色"), ("淡青色", "淡青色"), ("淡青", "淡青色"), ("青色", "青色"), ("青", "青色"), ("暗青色", "暗青色"), ("暗青", "暗青色"), ("藏青色", "藏青色"), ("藏青", "藏青色"), ("粉蓝色", "粉蓝色"), ("粉蓝", "粉蓝色"), ("海蓝色", "海蓝色"), ("海蓝", "海蓝色"), ("中蓝色", "中蓝色"), ("中蓝", "中蓝色"), ("深蓝色", "深蓝色"), ("深蓝", "深蓝色"), ("军绿色", "军绿色"), ("军绿", "军绿色"), ("墨绿色", "墨绿色"), ("墨绿", "墨绿色"), ("浅绿色", "浅绿色"), ("浅绿", "浅绿色"), ("绿色", "绿色"), ("绿", "绿色"), ("翠绿色", "翠绿色"), ("翠绿", "翠绿色"), ("孔雀绿", "孔雀绿"), ("荧光绿", "荧光绿"), ("绿黄色", "绿黄色"), ("绿黄", "绿黄色"), ("草绿色", "草绿色"), ("草绿", "草绿色"), ("黄绿色", "黄绿色"), ("黄绿", "黄绿色"), ("碧绿", "碧绿"), ("绿宝石", "绿宝石"), ("海洋绿", "海洋绿"), ("橄榄色", "橄榄色"), ("橄榄", "橄榄色"), ("橄榄绿", "橄榄绿"), ("抹茶色", "抹茶色"), ("抹茶", "抹茶色"), ("透明", "透明"), ("浅紫色", "浅紫色"), ("浅紫", "浅紫色"), ("深紫色", "深紫色"), ("深紫", "深紫色"), ("紫红色", "紫红色"), ("紫红", "紫红色"), ("紫罗兰", "紫罗兰"), ("暗紫罗兰", "暗紫罗兰"), ("紫色", "紫色"), ("紫", "紫色"), ("深洋红色", "深洋红色"), ("深洋红", "深洋红色"), ("紫兰色", "紫兰色"), ("紫兰", "紫兰色"), ("熏衣草淡紫", "熏衣草淡紫"), ("咖啡色", "咖啡色"), ("咖啡", "咖啡色"), ("巧克力色", "巧克力色"), ("巧克力", "巧克力色"), ("栗色", "栗色"), ("栗", "栗色"), ("深卡其布色", "深卡其布色"), ("深卡其布", "深卡其布色"), ("浅棕色", "浅棕色"), ("浅棕", "浅棕色"), ("深棕色", "深棕色"), ("深棕", "深棕色"), ("褐色", "褐色"), ("褐", "褐色"), ("棕褐色", "棕褐色"), ("棕褐", "棕褐色"), ("驼色", "驼色"), ("驼", "驼色"), ("橄榄棕色", "橄榄棕色"), ("橄榄棕", "橄榄棕色"), ("红褐色", "红褐色"), ("红褐", "红褐色"), ("茶色", "茶色"), ("茶", "茶色"), ("亚麻色", "亚麻色"), ("亚麻", "亚麻色"), ("黄土赭色", "黄土赭色"), ("黄土赭", "黄土赭色"), ("琥珀色", "琥珀色"), ("琥珀", "琥珀色"), ("小豆色", "小豆色"), ("小豆", "小豆色"), ("焦糖色", "焦糖色"), ("焦糖", "焦糖色"), ("肉桂色", "肉桂色"), ("肉桂", "肉桂色"), ("玫瑰棕色", "玫瑰棕色"), ("玫瑰棕", "玫瑰棕色"), ("玫瑰金", "玫瑰金"), ("古铜色", "古铜色"), ("古铜", "古铜色"), ("混合色", "混合色"), ("混合", "混合色"), ("white", "白色"), ("purewhite", "白色"), ("offwhite", "米白色"), ("milkywhite", "乳白色"), ("cream", "乳白色"), ("ivory", "象牙白"), ("red", "红色"), ("orangered", "桔红色"), ("rose red", "玫红色"), ("rosered", "玫红色"), ("pink", "粉红色"), ("hotpink", "深粉红"), ("magenta", "洋红色"), ("coral", "珊瑚色"), ("burgundy", "酒红色"), ("wine red", "酒红色"), ("winered", "酒红色"), ("scarlet", "猩红色"), ("black", "黑色"), ("blk", "黑色"), ("multi", "花色"), ("multicolor", "花色"), ("mixed", "混合色"), ("mixedcolor", "混合色"), ("khaki", "卡其色"), ("mustard", "姜黄色"), ("lemonyellow", "柠檬黄"), ("neonyellow", "荧光黄"), ("gold", "金色"), ("champagne", "香槟色"), ("yellow", "黄色"), ("lightyellow", "浅黄色"), ("wheat", "小麦色"), ("orange", "橙色"), ("earthyellow", "土黄色"), ("skin", "肤色"), ("nude", "肤色"), ("darkgray", "深灰色"), ("darkgrey", "深灰色"), ("lightgray", "浅灰色"), ("lightgrey", "浅灰色"), ("gray", "灰色"), ("grey", "灰色"), ("silvergray", "银灰色"), ("spacegray", "深空灰"), ("graphite", "石墨色"), ("skyblue", "天蓝色"), ("peacockblue", "孔雀蓝"), ("royalblue", "宝蓝色"), ("lightblue", "浅蓝色"), ("blue", "蓝色"), ("lakeblue", "湖蓝色"), ("dodgerblue", "道奇蓝"), ("deepskyblue", "深天蓝"), ("navy", "藏青色"), ("navyblue", "藏青色"), ("azure", "蔚蓝色"), ("cyan", "青色"), ("darkcyan", "暗青色"), ("powderblue", "粉蓝色"), ("oceanblue", "海蓝色"), ("mediumblue", "中蓝色"), ("darkblue", "深蓝色"), ("denimblue", "蓝色"), ("armygreen", "军绿色"), ("darkgreen", "墨绿色"), ("lightgreen", "浅绿色"), ("green", "绿色"), ("emerald", "绿宝石"), ("neongreen", "荧光绿"), ("grassgreen", "草绿色"), ("yellowgreen", "黄绿色"), ("seagreen", "海洋绿"), ("olive", "橄榄色"), ("olivegreen", "橄榄绿"), ("matcha", "抹茶色"), ("clear", "透明"), ("transparent", "透明"), ("lightpurple", "浅紫色"), ("darkpurple", "深紫色"), ("purple", "紫色"), ("violet", "紫罗兰"), ("darkviolet", "暗紫罗兰"), ("lavender", "熏衣草淡紫"), ("brown", "褐色"), ("coffee", "咖啡色"), ("chocolate", "巧克力色"), ("maroon", "栗色"), ("lightbrown", "浅棕色"), ("darkbrown", "深棕色"), ("camel", "驼色"), ("linen", "亚麻色"), ("amber", "琥珀色"), ("caramel", "焦糖色"), ("cinnamon", "肉桂色"), ("rosegold", "玫瑰金"), ("bronze", "古铜色")]

MIAOSHOU_COLUMNS = [
    "*产品标题",
    "*英文标题",
    "产品描述",
    "产品货号",
    "*变种属性名称一",
    "*变种属性值一",
    "变种属性名称二",
    "变种属性值二",
    "预览图",
    "*申报价格\n(店铺币种)",
    "SKU货号",
    "*长（cm）",
    "*宽（cm）",
    "*高（cm）",
    "*重量（g）",
    "识别码类型",
    "识别码",
    "站外产品链接",
    "*轮播图",
    "*产品素材图",
    "外包装形状",
    "外包装类型",
    "外包装图片",
    "建议售价（USD）",
    "库存",
    "发货时效（天）",
]
RPA_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
STANDARD_UPLOAD_SIZES = ["XS", "S", "M", "L", "XL", "XXL"]
STANDARD_UPLOAD_SIZE_RANK = {size: index for index, size in enumerate(STANDARD_UPLOAD_SIZES)}
RPA_ALLOWED_UPLOAD_SIZES = set(STANDARD_UPLOAD_SIZES)
SIZE_ALIAS_MAP = {
    "XS": "XS",
    "S": "S",
    "M": "M",
    "L": "L",
    "XL": "XL",
    "XXL": "XXL",
    "2XL": "XXL",
    "XXXL": "XXL",
    "3XL": "XXL",
    "4XL": "XXL",
}

def ensure_runtime_directories() -> dict[str, str]:
    created: dict[str, str] = {}
    for key, path in LOCAL_DATA_DIRECTORIES.items():
        path.mkdir(parents=True, exist_ok=True)
        created[key] = str(path)
    return created


ensure_runtime_directories()

app = FastAPI(title="上货助手", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
app.mount("/images", StaticFiles(directory=IMAGE_DIR), name="images")


class Product(BaseModel):
    id: int
    title: str
    skc: str
    sku_summary: str
    main_image: Optional[str] = None
    purchase_price: float = 0
    platform_quote_price: float = 0
    weight_g: int = 0
    first_mile: float = 0
    warehouse_fee: float = 15
    last_mile: float = 0
    platform_cost: float = 0
    total_cost: float = 0
    estimated_profit: float = 0
    gross_margin: float = 0
    status: str
    created_at: str
    updated_at: str


class ProductPayload(BaseModel):
    title: str
    skc: str
    sku_summary: str = ""
    purchase_price: float = 0
    platform_quote_price: float = 0
    weight_g: int = 0
    first_mile: float = 0
    warehouse_fee: float = 15
    last_mile: float = 0
    platform_cost: float = 0
    status: str = "利润正常"


class QuoteImportPayload(BaseModel):
    text: str


class QuoteImportResult(BaseModel):
    updated_count: int = 0
    unmatched_count: int = 0
    invalid_count: int = 0
    duplicate_count: int = 0
    rows_count: int = 0
    unmatched: list[dict[str, object]] = []
    invalid: list[dict[str, object]] = []
    duplicates: list[str] = []


class FreightImportResult(BaseModel):
    first_mile_count: int = 0
    last_mile_count: int = 0
    default_zone: str = "zone5"
    saved: bool = False


class FreightFirstMileRule(BaseModel):
    channel: str = "空运"
    max_weight_g: int
    price_per_kg: float
    fixed_fee: float = 0


class FreightLastMileRule(BaseModel):
    channel: str = "尾程"
    max_weight_g: int
    zones: dict[str, float]


class FreightRulesPayload(BaseModel):
    default_zone: str = "zone5"
    warehouse_fee: float = 15
    first_mile: list[FreightFirstMileRule] = []
    last_mile: list[FreightLastMileRule] = []



class CollectionItem(BaseModel):
    id: int
    title: str
    source: str
    source_url: str = ""
    price: float = 0
    sales: int = 0
    image_count: int = 0
    image_url: str = ""
    selected: bool = False
    status: str
    created_at: str


class CollectionItemPayload(BaseModel):
    title: str
    source: str = "1688"
    source_url: str = ""
    price: float = 0
    sales: int = 0
    image_count: int = 0
    image_url: str = ""


class CollectionItemPage(BaseModel):
    items: list[CollectionItem]
    total: int
    page: int
    page_size: int
    total_pages: int


class CollectionTask(BaseModel):
    id: int
    keyword: str
    source: str = "1688"
    mode: str = "1688"
    collector: str = "1688_public_search"
    target_count: int = 10
    min_price: float = 0
    max_price: float = 0
    blacklist: str = ""
    status: str
    note: str = ""
    request_path: str = ""
    result_count: int = 0
    parent_task_id: int = 0
    batch_index: int = 1
    batch_total: int = 1
    created_at: str
    updated_at: str


class CollectionTaskPayload(BaseModel):
    keyword: str
    source: str = "1688"
    mode: str = "1688"
    collector: str = "1688_public_search"
    target_count: int = 10
    min_price: float = 0
    max_price: float = 0
    blacklist: str = ""


class CollectionTaskEvent(BaseModel):
    id: int
    task_id: int
    stage: str
    status: str
    message: str = ""
    page_no: int = 0
    parsed_count: int = 0
    imported_count: int = 0
    skipped_count: int = 0
    created_at: str


class CollectionImportResult(BaseModel):
    imported_count: int
    skipped_count: int
    source: str
    filename: str


class CollectorCandidate(BaseModel):
    title: str
    source: str
    source_url: str = ""
    price: float = 0
    sales: int = 0
    image_count: int = 0
    image_url: str = ""


class CollectorRunResult(BaseModel):
    mode: str
    collector: str
    note: str = ""
    candidates: list[CollectorCandidate]


class ImportCollectionPayload(BaseModel):
    ids: list[int]


class CollectionImportTask(BaseModel):
    id: int
    status: str
    total_count: int = 0
    processed_count: int = 0
    imported_count: int = 0
    skipped_count: int = 0
    failed_count: int = 0
    imported_product_ids: list[int] = []
    note: str = ""
    created_at: str
    updated_at: str


class CollectionImageTaskStatus(BaseModel):
    queued_count: int = 0
    running_count: int = 0
    completed_count: int = 0
    failed_count: int = 0
    active: bool = False


class CollectionBulkPayload(BaseModel):
    ids: list[int]


class ProductBulkStatusPayload(BaseModel):
    ids: list[int]
    status: str


class GenerateTitlePayload(BaseModel):
    product_id: int


class BulkGenerateTitlePayload(BaseModel):
    ids: list[int]


class BulkGenerateTitleItemResult(BaseModel):
    product_id: int
    ok: bool
    chinese_title: str = ""
    english_title: str = ""
    error: str = ""


class BulkGenerateTitleResult(BaseModel):
    success_count: int
    failed_count: int
    results: list[BulkGenerateTitleItemResult]


class ProcessingTitleTask(BaseModel):
    id: int
    status: str
    ids_json: str = "[]"
    failed_ids_json: str = "[]"
    total_count: int = 0
    processed_count: int = 0
    success_count: int = 0
    failed_count: int = 0
    cache_hit_count: int = 0
    note: str = ""
    created_at: str
    updated_at: str


class ProcessingFieldTask(BaseModel):
    id: int
    status: str
    failed_ids_json: str = "[]"
    total_count: int = 0
    processed_count: int = 0
    success_count: int = 0
    failed_count: int = 0
    note: str = ""
    created_at: str
    updated_at: str


class ProcessingTaskHistoryItem(BaseModel):
    id: int
    task_type: str
    task_label: str
    status: str
    total_count: int = 0
    processed_count: int = 0
    success_count: int = 0
    failed_count: int = 0
    cache_hit_count: int = 0
    retryable_count: int = 0
    note: str = ""
    created_at: str = ""
    updated_at: str = ""


class ProcessingTaskHistoryPage(BaseModel):
    items: list[ProcessingTaskHistoryItem]
    total: int = 0
    page: int = 1
    page_size: int = 20
    total_pages: int = 1


class ProcessingExceptionItem(BaseModel):
    product_id: int
    skc: str = ""
    title: str = ""
    status: str = "open"
    issues: list[str] = []
    warnings: list[str] = []
    updated_at: str = ""


class SpecAliasItem(BaseModel):
    id: int = 0
    kind: str
    alias: str
    target: str
    updated_at: str = ""


class SpecAliasPayload(BaseModel):
    kind: str
    alias: str
    target: str


class SpecExceptionSummaryItem(BaseModel):
    kind: str
    value: str
    count: int = 0
    product_ids: list[int] = []


class ProcessImagePayload(BaseModel):
    product_id: int
    source_image: str = ""


class ImageOption(BaseModel):
    label: str
    url: str
    kind: str = ""
    preview_url: str = ""
    color: str = ""
    size: str = ""
    sku_id: str = ""


class ColorImageAssignment(BaseModel):
    color: str
    image_url: str
    preview_url: str = ""
    source: str = ""
    sort_order: int = 0
    is_auto: bool = True
    confidence: float = 0


class DetailImageAssignment(BaseModel):
    image_url: str
    preview_url: str = ""
    source: str = ""
    sort_order: int = 0


class AutoAssignColorImagesPayload(BaseModel):
    product_id: int
    count_per_color: int = 3
    use_vision: bool = False
    user_confirmed_vision: bool = False


class ManualColorImageAssignmentPayload(BaseModel):
    product_id: int
    color: str
    slot_index: int
    image_url: str = ""


class ManualDetailImageAssignmentPayload(BaseModel):
    product_id: int
    slot_index: int
    image_url: str = ""
    remove_image_url: str = ""
    remaining_image_urls: list[str] = []


class DedupeProcessingImagesPayload(BaseModel):
    ids: list[int]


class DedupeProcessingImagesResult(BaseModel):
    product_count: int
    scanned_count: int
    removed_count: int
    detail_removed_count: int
    color_removed_count: int


class AutoAssignColorImagesResult(BaseModel):
    product_id: int
    count_per_color: int
    assigned_count: int
    colors: dict[str, int]


class ImageTask(BaseModel):
    id: int
    product_id: int
    provider: str
    task_id: str
    status: str
    source_image: str = ""
    result_url: str = ""
    local_path: str = ""
    prompt: str = ""
    note: str = ""
    created_at: str
    updated_at: str


class ProcessingItem(BaseModel):
    product_id: int
    title: str
    skc: str
    english_title: str
    color: str = ""
    size: str = ""
    sku_code: str
    declared_price: float = 0
    weight_g: int = 350
    length_cm: int = 15
    width_cm: int = 10
    height_cm: int = 2
    source_url: str = ""
    main_image: str = ""
    image_options: list[ImageOption] = []
    color_image_assignments: list[ColorImageAssignment] = []
    detail_image_assignments: list[DetailImageAssignment] = []
    detail_status: str = ""
    detail_note: str = ""
    image_status: str
    stock: int = 999
    ship_days: int = 9
    status: str


class ProcessingItemPage(BaseModel):
    items: list[ProcessingItem]
    total: int = 0
    page: int = 1
    page_size: int = 50
    total_pages: int = 1


class ProcessingItemPayload(BaseModel):
    english_title: str = ""
    color: str = ""
    size: str = ""
    sku_code: str = ""
    declared_price: float = 0
    declared_price_mode: str = "set"
    weight_g: int = 350
    length_cm: int = 15
    width_cm: int = 10
    height_cm: int = 2
    source_url: str = ""
    stock: int = 999
    ship_days: int = 9


class ProcessingBulkPayload(BaseModel):
    ids: list[int]
    fields: ProcessingItemPayload
    start_skc: str = ""


class UploadTask(BaseModel):
    id: int
    name: str
    status: str
    total_count: int = 0
    success_count: int = 0
    failed_count: int = 0
    export_path: str = ""
    run_log: str = ""
    failure_stage: str = ""
    failure_reason: str = ""
    evidence_path: str = ""
    retry_count: int = 0
    executor_id: str = ""
    claimed_at: str = ""
    heartbeat_at: str = ""
    created_at: str
    updated_at: str


class ExecutorClaimPayload(BaseModel):
    executor_id: str = ""
    version: str = ""


class ExecutorHeartbeatPayload(BaseModel):
    executor_id: str = ""


class ExecutorStatus(BaseModel):
    online: bool = False
    executor_id: str = ""
    status: str = "offline"
    message: str = ""
    task_id: int = 0
    updated_at: str = ""
    pending_result_count: int = 0
    status_path: str = ""


class ExecutorReportPayload(BaseModel):
    executor_id: str = ""
    status: str
    success_count: int = 0
    failed_count: int = 0
    run_log: str = ""
    stdout: str = ""
    failure_stage: str = ""
    failure_reason: str = ""
    evidence_path: str = ""


class RetryFailedUploadPayload(BaseModel):
    ids: list[int] = []


class CleanupPayload(BaseModel):
    retention_days: int = 7
    clean_success_images: bool = True
    clean_exports: bool = True
    clean_logs: bool = True


class PublishRecord(BaseModel):
    id: int
    result: str
    skc: str
    title: str
    reason: str = ""
    created_at: str


class ApiConfig(BaseModel):
    key: str
    name: str
    enabled: bool = True
    base_url: str = ""
    model: str = ""
    api_key: str = ""
    usage: str = ""
    updated_at: str = ""


class AppSetting(BaseModel):
    key: str
    value: str
    updated_at: str = ""


class PromptTemplate(BaseModel):
    id: int
    name: str
    category: str = ""
    prompt_type: str = ""
    usage: str = ""
    content: str = ""
    status: str = "启用中"
    updated_at: str


class PromptTemplatePayload(BaseModel):
    name: str
    category: str = ""
    prompt_type: str = ""
    usage: str = ""
    content: str = ""
    status: str = "启用中"



def configured_script_dir() -> Path:
    default_script_dir = BASE_DIR / "rpa" if (BASE_DIR / "rpa").exists() else SCRIPT_SOURCE_DIR
    configured = get_setting_value("script_dir", str(default_script_dir)).strip()
    if configured == str(SCRIPT_SOURCE_DIR) and (BASE_DIR / "rpa").exists():
        return BASE_DIR / "rpa"
    return Path(configured) if configured else SCRIPT_SOURCE_DIR


def legacy_rpa_config() -> dict[str, object]:
    config_path = configured_script_dir() / "apify_config.json"
    if not config_path.exists():
        return {}
    try:
        data = json.loads(config_path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def upload_image_source() -> str:
    configured = get_setting_value("upload_image_source", "").strip().lower()
    if configured in {"local", "cos"}:
        return configured
    legacy = str(legacy_rpa_config().get("image_source") or "local").strip().lower()
    return legacy if legacy in {"local", "cos"} else "local"


def cos_setting_value(key: str) -> str:
    configured = get_setting_value(key, "").strip()
    if configured:
        return configured
    legacy_key = key.removeprefix("cos_") if key.startswith("cos_") else key
    return str(legacy_rpa_config().get(key) or legacy_rpa_config().get(f"cos_{legacy_key}") or "").strip()


def cos_preflight() -> dict[str, object]:
    config = legacy_rpa_config()
    return {
        "image_source": upload_image_source(),
        "config_path": str(configured_script_dir() / "apify_config.json"),
        "config_exists": bool(config),
        "region": cos_setting_value("cos_region"),
        "bucket": cos_setting_value("cos_bucket"),
        "prefix": cos_setting_value("cos_prefix"),
        "secret_id_configured": bool(cos_setting_value("cos_secret_id")),
        "secret_key_configured": bool(cos_setting_value("cos_secret_key")),
    }


def cos_public_image_base() -> str:
    if upload_image_source() != "cos":
        return ""
    region = cos_setting_value("cos_region")
    bucket = cos_setting_value("cos_bucket")
    if not region or not bucket:
        return ""
    return f"https://{bucket}.cos.{region}.myqcloud.com"


def cos_public_image_url(*segments: str) -> str:
    base = cos_public_image_base()
    if not base:
        return ""
    prefix = cos_setting_value("cos_prefix").strip().strip("/")
    path_parts = [part for part in [*prefix.split("/"), *segments] if str(part or "").strip()]
    encoded_path = "/".join(urllib.parse.quote(str(part).strip(), safe="") for part in path_parts)
    return f"{base}/{encoded_path}" if encoded_path else base


def cos_required_settings() -> dict[str, str]:
    values = {
        "region": cos_setting_value("cos_region"),
        "bucket": cos_setting_value("cos_bucket"),
        "prefix": cos_setting_value("cos_prefix").strip().strip("/"),
        "secret_id": cos_setting_value("cos_secret_id"),
        "secret_key": cos_setting_value("cos_secret_key"),
    }
    missing = [label for key, label in [("region", "COS Region"), ("bucket", "COS Bucket"), ("secret_id", "COS SecretId"), ("secret_key", "COS SecretKey")] if not values[key]]
    if missing:
        raise HTTPException(status_code=400, detail=f"COS 模式导出前必须配置：{'、'.join(missing)}")
    return values


def build_cos_client(region: str, secret_id: str, secret_key: str):
    try:
        from qcloud_cos import CosConfig, CosS3Client
    except Exception as exc:
        raise HTTPException(status_code=500, detail="当前环境未安装 COS SDK，请先安装 cos-python-sdk-v5") from exc
    config = CosConfig(Region=region, SecretId=secret_id, SecretKey=secret_key, Scheme="https")
    return CosS3Client(config)


def cos_object_key(prefix: str, *segments: str) -> str:
    parts = [part for part in [*prefix.strip("/").split("/"), *segments] if str(part or "").strip()]
    return "/".join(str(part).strip().strip("/") for part in parts)


def sync_rpa_images_to_cos(items: list[ProcessingItem]) -> dict[str, int]:
    settings = cos_required_settings()
    root = configured_rpa_sku_image_dir()
    safe_skc_values = {safe_path_segment(item.skc, f"product_{item.product_id}") for item in items}
    if not root.exists():
        raise HTTPException(status_code=400, detail="COS 模式导出前未生成本地 RPA 图片目录")
    client = build_cos_client(settings["region"], settings["secret_id"], settings["secret_key"])
    uploaded_count = 0
    failed: list[str] = []
    for product_dir in sorted(root.iterdir()):
        if not product_dir.is_dir() or product_dir.name not in safe_skc_values:
            continue
        for image_path in sorted(product_dir.rglob("*")):
            if not image_path.is_file() or image_path.suffix.lower() not in RPA_IMAGE_EXTENSIONS:
                continue
            relative_parts = image_path.relative_to(product_dir).parts
            key = cos_object_key(settings["prefix"], product_dir.name, *relative_parts)
            try:
                with image_path.open("rb") as file:
                    client.put_object(Bucket=settings["bucket"], Body=file, Key=key)
                uploaded_count += 1
            except Exception as exc:
                failed.append(f"{key}: {exc}")
    if failed:
        raise HTTPException(status_code=502, detail=f"COS 图片上传失败：{failed[0]}")
    return {"uploaded_count": uploaded_count, "failed_count": 0}


def upload_rpa_command(export_path: str) -> list[str]:
    command = ["python", "rpa_upload.py", "--import", export_path]
    if get_setting_value("upload_save_screenshots", "false").strip().lower() == "true":
        command.append("--debug-shots")
    for setting_key, cli_arg in [
        ("temu_shop_account", "--shop-account"),
        ("temu_site", "--site"),
        ("temu_product_template", "--product-template"),
        ("temu_size_template", "--size-template"),
        ("temu_warehouse_template", "--warehouse-template"),
        ("temu_logistics_template", "--logistics-template"),
    ]:
        value = get_setting_value(setting_key, "").strip()
        if value:
            command.extend([cli_arg, value])
    if get_setting_value("upload_auto_submit", "false").strip().lower() != "true":
        command.append("--no-publish")
    if upload_image_source() == "cos":
        command.extend(["--image-source", "cos"])
        for setting_key, cli_arg in [
            ("cos_region", "--cos-region"),
            ("cos_bucket", "--cos-bucket"),
            ("cos_prefix", "--cos-prefix"),
        ]:
            value = cos_setting_value(setting_key)
            if value:
                command.extend([cli_arg, value])
    return command


def display_upload_rpa_command(export_path: str) -> str:
    command = upload_rpa_command(export_path)
    redacted: list[str] = []
    skip_next = False
    for part in command:
        if skip_next:
            redacted.append("***")
            skip_next = False
            continue
        redacted.append(part)
    return " ".join(redacted)


def configured_rpa_sku_image_dir() -> Path:
    return configured_script_dir() / "data" / "sku_images"


DEFAULT_IMAGE_RECOGNITION_PROMPT = (
    "\u4f60\u662f\u7535\u5546\u5546\u54c1\u56fe\u7247\u989c\u8272\u8bc6\u522b\u52a9\u624b\uff0c"
    "\u53ea\u5224\u65ad\u56fe\u7247\u4e2d\u5546\u54c1\u4e3b\u4f53\u989c\u8272\uff0c"
    "\u4e0d\u8981\u53d7\u80cc\u666f\u3001\u6a21\u7279\u80a4\u8272\u3001\u6587\u5b57\u3001\u6c34\u5370\u5f71\u54cd\u3002\n\n"
    "\u5019\u9009\u989c\u8272 JSON\uff1a\n{color_json}\n\n"
    "\u4efb\u52a1\uff1a\n"
    "1. \u5224\u65ad\u56fe\u7247\u91cc\u7684\u5546\u54c1\u4e3b\u4f53\u6700\u63a5\u8fd1\u54ea\u4e2a\u5019\u9009\u989c\u8272\u3002\n"
    "2. matched_color \u5fc5\u987b\u4ece\u5019\u9009\u989c\u8272 JSON \u4e2d\u539f\u6837\u9009\u62e9\u4e00\u4e2a\u503c\u3002\n"
    "3. \u5982\u679c\u56fe\u7247\u4e0d\u662f\u5546\u54c1\u56fe\uff0cis_product_image=false\uff0cmatched_color=\"\"\u3002\n"
    "4. \u5982\u679c\u65e0\u6cd5\u5224\u65ad\u6216\u6ca1\u6709\u63a5\u8fd1\u989c\u8272\uff0cmatched_color=\"\"\u3002\n"
    "5. confidence \u8fd4\u56de 0 \u5230 1 \u7684\u6570\u5b57\u3002\n\n"
    "\u53ea\u8fd4\u56de\u4e00\u4e2a JSON \u5bf9\u8c61\uff0c\u4e0d\u8981\u8fd4\u56de\u89e3\u91ca\u6587\u5b57\uff0c"
    "\u4e0d\u8981\u4f7f\u7528 Markdown\u3002\n"
    "JSON \u5b57\u6bb5\u56fa\u5b9a\u4e3a\uff1amatched_color, confidence, is_product_image, reason\u3002"
)

DEFAULT_PROMPT_TEMPLATES = [
    ("Temu \u82f1\u6587\u6807\u9898\u751f\u6210", "\u5973\u88c5\u77ed\u88e4", "\u6807\u9898\u63d0\u793a\u8bcd", "\u6807\u9898\u751f\u6210", "\u6839\u636e\u5546\u54c1\u6807\u9898\u3001\u7c7b\u76ee\u3001\u989c\u8272\u548c\u5356\u70b9\u751f\u6210\u81ea\u7136\u51c6\u786e\u7684 Temu \u82f1\u6587\u6807\u9898\u3002", "\u542f\u7528\u4e2d"),
    ("\u725b\u4ed4\u77ed\u88e4\u4e3b\u56fe\u7cbe\u4fee", "\u5973\u88c5\u77ed\u88e4", "\u56fe\u751f\u56fe\u63d0\u793a\u8bcd", "\u4e3b\u56fe\u4f18\u5316", "\u4fdd\u7559\u5546\u54c1\u4e3b\u4f53\uff0c\u4f18\u5316\u5149\u7ebf\u3001\u8d28\u611f\u548c\u80cc\u666f\uff0c\u4e0d\u6539\u53d8\u5546\u54c1\u6b3e\u5f0f\u3002", "\u542f\u7528\u4e2d"),
    ("\u5546\u54c1\u989c\u8272\u56fe\u7247\u8bc6\u522b", "\u901a\u7528\u5546\u54c1", "\u56fe\u7247\u8bc6\u522b\u63d0\u793a\u8bcd", "\u989c\u8272\u5206\u56fe", DEFAULT_IMAGE_RECOGNITION_PROMPT, "\u542f\u7528\u4e2d"),
]
DEFAULT_PROMPT_TEMPLATE_KEYS = {(name, category, prompt_type, usage) for name, category, prompt_type, usage, _content, _status in DEFAULT_PROMPT_TEMPLATES}
DEFAULT_PROMPT_TEMPLATE_KEYS.add(("服饰中英文标题生成", "服饰通用", "标题提示词", "标题生成"))


DEFAULT_API_CONFIGS = {
    "deepseek": {
        "name": "DeepSeek 文案生成",
        "enabled": True,
        "base_url": "https://api.deepseek.com",
        "model": "deepseek-chat",
        "api_key": "",
        "usage": "标题生成 / 文案优化",
    },
    "image": {
        "name": "Nano Banana 图生图",
        "enabled": True,
        "base_url": NANOBANANA_SUBMIT_URL,
        "model": "nano-banana",
        "api_key": "",
        "usage": "图片处理 / 图生图",
    },
}


def load_builtin_api_configs() -> dict[str, dict[str, object]]:
    configs = {key: dict(value) for key, value in DEFAULT_API_CONFIGS.items()}
    if not BUILTIN_API_CONFIGS_PATH.exists():
        return configs
    try:
        raw = json.loads(BUILTIN_API_CONFIGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return configs
    if not isinstance(raw, dict):
        return configs
    for key, value in raw.items():
        if isinstance(value, dict):
            current = configs.get(str(key), {})
            current.update(value)
            configs[str(key)] = current
    return configs


def sync_builtin_api_configs(db: sqlite3.Connection, force_keys: set[str] | None = None) -> None:
    timestamp = now_text()
    force_keys = force_keys or set()
    for key, config in load_builtin_api_configs().items():
        api_key = str(config.get("api_key") or "").strip()
        if not api_key and key not in DEFAULT_API_CONFIGS:
            continue
        row = db.execute("SELECT * FROM api_configs WHERE key = ?", (key,)).fetchone()
        should_update_secret = bool(api_key) and (key in force_keys or not row or not str(row["api_key"] or "").strip())
        if row:
            db.execute(
                """
                UPDATE api_configs
                SET name = ?, enabled = ?, base_url = ?, model = ?, api_key = CASE WHEN ? THEN ? ELSE api_key END,
                    usage = ?, updated_at = ?
                WHERE key = ?
                """,
                (
                    str(config.get("name") or row["name"] or key),
                    int(bool(config.get("enabled", True))),
                    str(config.get("base_url") or row["base_url"] or ""),
                    str(config.get("model") or row["model"] or ""),
                    int(should_update_secret),
                    api_key,
                    str(config.get("usage") or row["usage"] or ""),
                    timestamp,
                    key,
                ),
            )
        else:
            db.execute(
                "INSERT INTO api_configs (key, name, enabled, base_url, model, api_key, usage, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    key,
                    str(config.get("name") or key),
                    int(bool(config.get("enabled", True))),
                    str(config.get("base_url") or ""),
                    str(config.get("model") or ""),
                    api_key,
                    str(config.get("usage") or ""),
                    timestamp,
                ),
            )


def load_builtin_settings() -> dict[str, str]:
    if not BUILTIN_SETTINGS_PATH.exists():
        return {}
    try:
        raw = json.loads(BUILTIN_SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(raw, dict):
        return {}
    settings: dict[str, str] = {}
    for key, value in raw.items():
        if value is None:
            continue
        settings[str(key)] = str(value)
    return settings


def sync_builtin_settings(db: sqlite3.Connection, force_keys: set[str] | None = None) -> None:
    timestamp = now_text()
    force_keys = force_keys or set()
    for key, value in load_builtin_settings().items():
        clean_value = str(value or "").strip()
        if not clean_value:
            continue
        row = db.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
        if row and key not in force_keys and str(row["value"] or "").strip():
            continue
        db.execute(
            """
            INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
            """,
            (key, clean_value, timestamp),
        )


def looks_corrupted_text(value: str) -> bool:
    text = str(value or "")
    if not text:
        return False
    question_count = text.count("?")
    mojibake_markers = ["Ã", "Â", "Ð", "Ñ", "æ", "ç", "è", "é", "å", "¤", "½", "�"]
    return question_count >= max(3, len(text) // 3) or any(marker in text for marker in mojibake_markers)


def connect() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def chunked_ids(ids: list[int], size: int = 800) -> list[list[int]]:
    return [ids[index:index + size] for index in range(0, len(ids), size)]


def product_from_row(row: sqlite3.Row) -> Product:
    data = dict(row)
    data.setdefault("platform_quote_price", 0)
    data.setdefault("weight_g", 0)
    data.setdefault("warehouse_fee", 15)
    data.setdefault("last_mile", data.get("platform_cost", 0))
    main_image = str(data.get("main_image") or "").strip()
    if main_image.startswith("/images/"):
        local_path = IMAGE_DIR / main_image.removeprefix("/images/")
        if not local_path.exists():
            with connect() as image_db:
                fallback = image_db.execute(
                    """
                    SELECT image_url FROM product_color_image_assignments
                    WHERE product_id = ? AND image_url != ''
                    ORDER BY sort_order, id LIMIT 1
                    """,
                    (data["id"],),
                ).fetchone()
                if not fallback:
                    fallback = image_db.execute(
                        """
                        SELECT image_url FROM product_sku_images
                        WHERE product_id = ? AND image_url != ''
                        ORDER BY sort_order, id LIMIT 1
                        """,
                        (data["id"],),
                    ).fetchone()
            if fallback:
                data["main_image"] = fallback["image_url"]
    if float(data.get("platform_quote_price") or 0) <= 0:
        data["estimated_profit"] = 0
        data["gross_margin"] = 0
    return Product(**data)


def collection_item_from_row(row: sqlite3.Row) -> CollectionItem:
    data = dict(row)
    data["selected"] = bool(data["selected"])
    return CollectionItem(**data)


def collection_task_from_row(row: sqlite3.Row) -> CollectionTask:
    return CollectionTask(**dict(row))


def collection_task_event_from_row(row: sqlite3.Row) -> CollectionTaskEvent:
    return CollectionTaskEvent(**dict(row))



def collection_task_batches(total_count: int) -> list[int]:
    total_count = max(1, min(int(total_count or 1), COLLECTION_BATCH_TARGET_MAX))
    batches: list[int] = []
    remaining = total_count
    while remaining > 0:
        batch_count = min(COLLECTION_TARGET_MAX, remaining)
        batches.append(batch_count)
        remaining -= batch_count
    return batches
def collection_import_task_from_row(row: sqlite3.Row) -> CollectionImportTask:
    data = dict(row)
    note = str(data.get("note") or "")
    match = re.search(r"imported_product_ids=([^;]+)", note)
    if match:
        data["imported_product_ids"] = [int(value) for value in match.group(1).split(",") if value.strip().isdigit()]
    return CollectionImportTask(**data)


def processing_title_task_from_row(row: sqlite3.Row) -> ProcessingTitleTask:
    return ProcessingTitleTask(**dict(row))


def processing_field_task_from_row(row: sqlite3.Row) -> ProcessingFieldTask:
    return ProcessingFieldTask(**dict(row))


class CollectionTaskCancelled(Exception):
    pass


def collection_task_is_cancel_requested(task_id: int) -> bool:
    with connect() as db:
        row = db.execute("SELECT status FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
    return bool(row and row["status"] == "cancel_requested")


def raise_if_collection_task_cancelled(task_id: int | None) -> None:
    if task_id and collection_task_is_cancel_requested(task_id):
        raise CollectionTaskCancelled()




def schedule_collection_task_execution(task_id: int) -> None:
    global COLLECTION_TASK_QUEUE_ACTIVE
    with connect() as db:
        row = db.execute("SELECT status FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
        if row and row["status"] not in {"queued", "running"}:
            db.execute("UPDATE collection_tasks SET status = ?, updated_at = ? WHERE id = ?", ("queued", now_text(), task_id))
            add_collection_task_event(db, task_id, "queued", "info", "\u4efb\u52a1\u5df2\u52a0\u5165\u91c7\u96c6\u961f\u5217")
    with COLLECTION_TASK_QUEUE_LOCK:
        if COLLECTION_TASK_QUEUE_ACTIVE:
            return
        COLLECTION_TASK_QUEUE_ACTIVE = True
    thread = threading.Thread(target=run_collection_task_queue_worker, daemon=True)
    thread.start()



def resume_collection_task_queue() -> None:
    with connect() as db:
        stale_running = db.execute(
            "UPDATE collection_tasks SET status = 'queued', updated_at = ? WHERE status = 'running' AND mode = '1688'",
            (now_text(),),
        ).rowcount
        row = db.execute("SELECT id FROM collection_tasks WHERE status = 'queued' AND mode = '1688' ORDER BY id ASC LIMIT 1").fetchone()
        if stale_running:
            add_collection_task_event(db, int(row["id"]) if row else 0, "resume", "info", f"reset stale running tasks: {stale_running}") if row else None
    if row:
        schedule_collection_task_execution(int(row["id"]))

def run_collection_task_queue_worker() -> None:
    global COLLECTION_TASK_QUEUE_ACTIVE
    try:
        while True:
            with connect() as db:
                row = db.execute(
                    "SELECT * FROM collection_tasks WHERE status = 'queued' AND mode = '1688' ORDER BY id ASC LIMIT 1"
                ).fetchone()
            if not row:
                break
            run_collection_task_background(int(row["id"]))
    finally:
        with COLLECTION_TASK_QUEUE_LOCK:
            COLLECTION_TASK_QUEUE_ACTIVE = False
        with connect() as db:
            row = db.execute(
                "SELECT id FROM collection_tasks WHERE status = 'queued' AND mode = '1688' ORDER BY id ASC LIMIT 1"
            ).fetchone()
        if row:
            schedule_collection_task_execution(int(row["id"]))


def run_collection_task_background(task_id: int) -> None:
    with connect() as db:
        row = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
        if not row or row["mode"] != "1688":
            return
        if row["status"] not in {"queued", "running", "cancel_requested"}:
            return
        if row["status"] == "cancel_requested":
            note = "采集任务已取消，未开始执行"
            db.execute("UPDATE collection_tasks SET status = ?, note = ?, updated_at = ? WHERE id = ?", ("cancelled", note, now_text(), task_id))
            add_collection_task_event(db, task_id, "cancelled", "info", note)
            return
        db.execute("UPDATE collection_tasks SET status = ?, updated_at = ? WHERE id = ?", ("running", now_text(), task_id))
        add_collection_task_event(db, task_id, "background_start", "running", "\u91c7\u96c6\u4efb\u52a1\u5f00\u59cb\u6267\u884c")
        payload = CollectionTaskPayload(
            keyword=row["keyword"],
            source=row["source"],
            mode=row["mode"],
            collector=row["collector"],
            target_count=row["target_count"],
            min_price=row["min_price"],
            max_price=row["max_price"],
            blacklist=row["blacklist"],
        )
        skipped = 0
        try:
            provider = "onebound" if get_setting_value("onebound_key") and get_setting_value("onebound_secret") else "1688_cookie"
            add_collection_task_event(db, task_id, "execute", "running", f"\u5f00\u59cb\u91c7\u96c6\uff1a\u6765\u6e90={provider}\uff0c\u76ee\u6807={payload.target_count} \u6761")
            rows = run_onebound_1688_search(payload, task_id) if provider == "onebound" else run_1688_public_search(payload, task_id)
            raise_if_collection_task_cancelled(task_id)
            imported, skipped = insert_collection_rows_with_db(db, normalize_collection_rows(rows), row["source"])
            add_collection_task_event(db, task_id, "write", "success" if imported else "info", f"\u89e3\u6790 {len(rows)} \u6761\uff0c\u65b0\u589e {imported} \u6761\uff0c\u8df3\u8fc7 {skipped} \u6761", parsed_count=len(rows), imported_count=imported, skipped_count=skipped)
            status = "completed" if imported else "empty"
            note = f"{provider} \u91c7\u96c6\u5b8c\u6210\uff1a\u89e3\u6790 {len(rows)} \u6761\uff0c\u65b0\u589e {imported} \u6761\uff0c\u8df3\u8fc7 {skipped} \u6761\u3002"
        except HTTPException as exc:
            status = "failed"
            imported = 0
            note = str(exc.detail)
            add_collection_task_event(db, task_id, "error", "error", note)
        except CollectionTaskCancelled:
            status = "cancelled"
            imported = 0
            note = "采集任务已取消，已停止后续写入"
            add_collection_task_event(db, task_id, "cancelled", "info", note)
        db.execute("UPDATE collection_tasks SET status = ?, collector = ?, note = ?, result_count = ?, updated_at = ? WHERE id = ?", (status, "1688_public_search", note, imported, now_text(), task_id))
        add_collection_task_event(db, task_id, "finished", "success" if imported else "info", note, imported_count=imported, skipped_count=skipped)


def schedule_collection_import_task_execution(task_id: int) -> None:
    global COLLECTION_IMPORT_TASK_QUEUE_ACTIVE
    with connect() as db:
        row = db.execute("SELECT status FROM collection_import_tasks WHERE id = ?", (task_id,)).fetchone()
        if row and row["status"] not in {"queued", "running"}:
            db.execute("UPDATE collection_import_tasks SET status = ?, updated_at = ? WHERE id = ?", ("queued", now_text(), task_id))
    with COLLECTION_IMPORT_TASK_QUEUE_LOCK:
        if COLLECTION_IMPORT_TASK_QUEUE_ACTIVE:
            return
        COLLECTION_IMPORT_TASK_QUEUE_ACTIVE = True
    thread = threading.Thread(target=run_collection_import_task_queue_worker, daemon=True)
    thread.start()


def resume_collection_import_task_queue() -> None:
    with connect() as db:
        db.execute(
            "UPDATE collection_import_tasks SET status = 'queued', note = '服务重启后继续排队入库', updated_at = ? WHERE status = 'running'",
            (now_text(),),
        )
        row = db.execute("SELECT id FROM collection_import_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
    if row:
        schedule_collection_import_task_execution(int(row["id"]))


def run_collection_import_task_queue_worker() -> None:
    global COLLECTION_IMPORT_TASK_QUEUE_ACTIVE
    try:
        while True:
            with connect() as db:
                row = db.execute("SELECT id FROM collection_import_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
            if not row:
                break
            run_collection_import_task_background(int(row["id"]))
    finally:
        with COLLECTION_IMPORT_TASK_QUEUE_LOCK:
            COLLECTION_IMPORT_TASK_QUEUE_ACTIVE = False
        with connect() as db:
            row = db.execute("SELECT id FROM collection_import_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
        if row:
            schedule_collection_import_task_execution(int(row["id"]))


def run_collection_import_task_background(task_id: int) -> None:
    with connect() as db:
        task = db.execute("SELECT * FROM collection_import_tasks WHERE id = ?", (task_id,)).fetchone()
        if not task or task["status"] != "queued":
            return
        ids = [int(item_id) for item_id in json.loads(task["ids_json"] or "[]") if int(item_id) > 0]
        db.execute(
            "UPDATE collection_import_tasks SET status = ?, total_count = ?, note = ?, updated_at = ? WHERE id = ?",
            ("running", len(ids), "正在加入商品库", now_text(), task_id),
        )
    imported_count = 0
    skipped_count = 0
    failed_count = 0
    processed_count = 0
    imported_product_ids: list[int] = []
    for item_id in ids:
        try:
            with connect() as db:
                row = db.execute("SELECT * FROM collection_items WHERE id = ?", (item_id,)).fetchone()
                if not row or row["status"] == "imported":
                    skipped_count += 1
                else:
                    imported_product = import_collection_row_as_product(db, row)
                    imported_product_ids.append(int(imported_product.id))
                    imported_count += 1
                processed_count += 1
                db.execute(
                    """
                    UPDATE collection_import_tasks
                    SET processed_count = ?, imported_count = ?, skipped_count = ?, failed_count = ?, note = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (processed_count, imported_count, skipped_count, failed_count, f"已处理 {processed_count}/{len(ids)} 条", now_text(), task_id),
                )
        except Exception as exc:
            failed_count += 1
            processed_count += 1
            with connect() as db:
                db.execute(
                    """
                    UPDATE collection_import_tasks
                    SET processed_count = ?, imported_count = ?, skipped_count = ?, failed_count = ?, note = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (processed_count, imported_count, skipped_count, failed_count, f"部分入库失败：{str(exc)[:120]}", now_text(), task_id),
                )
    status = "completed" if imported_count else ("failed" if failed_count else "empty")
    note = f"入库完成：新增 {imported_count} 条，跳过 {skipped_count} 条，失败 {failed_count} 条"
    if imported_product_ids:
        note = f"{note}; imported_product_ids={','.join(str(product_id) for product_id in imported_product_ids)}"
    with connect() as db:
        db.execute(
            """
            UPDATE collection_import_tasks
            SET status = ?, processed_count = ?, imported_count = ?, skipped_count = ?, failed_count = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, processed_count, imported_count, skipped_count, failed_count, note, now_text(), task_id),
        )


def enqueue_collection_image_download(db: sqlite3.Connection, collection_item_id: int, image_url: str) -> None:
    clean_url = str(image_url or "").strip()
    if not clean_url or clean_url.startswith("/images/") or not clean_url.startswith("http"):
        return
    timestamp = now_text()
    db.execute(
        """
        INSERT OR IGNORE INTO collection_image_tasks (collection_item_id, source_url, status, note, created_at, updated_at)
        VALUES (?, ?, 'queued', ?, ?, ?)
        """,
        (collection_item_id, clean_url, "等待后台下载采集图", timestamp, timestamp),
    )


def schedule_collection_image_task_execution() -> None:
    global COLLECTION_IMAGE_TASK_QUEUE_ACTIVE
    with COLLECTION_IMAGE_TASK_QUEUE_LOCK:
        if COLLECTION_IMAGE_TASK_QUEUE_ACTIVE:
            return
        COLLECTION_IMAGE_TASK_QUEUE_ACTIVE = True
    thread = threading.Thread(target=run_collection_image_task_queue_worker, daemon=True)
    thread.start()


def resume_collection_image_task_queue() -> None:
    with connect() as db:
        db.execute(
            "UPDATE collection_image_tasks SET status = 'queued', note = '服务重启后重新排队下载', updated_at = ? WHERE status = 'running'",
            (now_text(),),
        )
        row = db.execute("SELECT id FROM collection_image_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
    if row:
        schedule_collection_image_task_execution()


def run_collection_image_task_queue_worker() -> None:
    global COLLECTION_IMAGE_TASK_QUEUE_ACTIVE
    try:
        while True:
            with connect() as db:
                task = db.execute("SELECT * FROM collection_image_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
                if not task:
                    break
                db.execute(
                    "UPDATE collection_image_tasks SET status = 'running', attempts = attempts + 1, note = ?, updated_at = ? WHERE id = ?",
                    ("正在下载采集图", now_text(), task["id"]),
                )
            try:
                local_image = save_collection_image_from_url(int(task["collection_item_id"]), task["source_url"])
                with connect() as db:
                    if local_image:
                        db.execute("UPDATE collection_items SET image_url = ? WHERE id = ?", (local_image, task["collection_item_id"]))
                        db.execute(
                            "UPDATE collection_image_tasks SET status = 'completed', local_path = ?, note = ?, updated_at = ? WHERE id = ?",
                            (local_image, "采集图已下载到本地", now_text(), task["id"]),
                        )
                    else:
                        db.execute(
                            "UPDATE collection_image_tasks SET status = 'failed', note = ?, updated_at = ? WHERE id = ?",
                            ("采集图下载失败", now_text(), task["id"]),
                        )
            except Exception as exc:
                with connect() as db:
                    db.execute(
                        "UPDATE collection_image_tasks SET status = 'failed', note = ?, updated_at = ? WHERE id = ?",
                        (f"采集图下载异常：{str(exc)[:120]}", now_text(), task["id"]),
                    )
    finally:
        with COLLECTION_IMAGE_TASK_QUEUE_LOCK:
            COLLECTION_IMAGE_TASK_QUEUE_ACTIVE = False
        with connect() as db:
            row = db.execute("SELECT id FROM collection_image_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
        if row:
            schedule_collection_image_task_execution()


def schedule_processing_title_task_execution() -> None:
    global PROCESSING_TITLE_TASK_QUEUE_ACTIVE
    with PROCESSING_TITLE_TASK_QUEUE_LOCK:
        if PROCESSING_TITLE_TASK_QUEUE_ACTIVE:
            return
        PROCESSING_TITLE_TASK_QUEUE_ACTIVE = True
    thread = threading.Thread(target=run_processing_title_task_queue_worker, daemon=True)
    thread.start()


def resume_processing_title_task_queue() -> None:
    with connect() as db:
        db.execute(
            "UPDATE processing_title_tasks SET status = 'queued', note = '服务重启后重新排队处理标题', updated_at = ? WHERE status = 'running'",
            (now_text(),),
        )
        row = db.execute("SELECT id FROM processing_title_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
    if row:
        schedule_processing_title_task_execution()


def run_processing_title_task_queue_worker() -> None:
    global PROCESSING_TITLE_TASK_QUEUE_ACTIVE
    try:
        while True:
            with connect() as db:
                task = db.execute("SELECT * FROM processing_title_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
                if not task:
                    break
                db.execute("UPDATE processing_title_tasks SET status = 'running', note = ?, updated_at = ? WHERE id = ?", ("正在批量处理标题", now_text(), task["id"]))
            run_processing_title_task_background(int(task["id"]))
    finally:
        with PROCESSING_TITLE_TASK_QUEUE_LOCK:
            PROCESSING_TITLE_TASK_QUEUE_ACTIVE = False
        with connect() as db:
            row = db.execute("SELECT id FROM processing_title_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
        if row:
            schedule_processing_title_task_execution()


def run_processing_title_task_background(task_id: int) -> None:
    with connect() as db:
        task = db.execute("SELECT * FROM processing_title_tasks WHERE id = ?", (task_id,)).fetchone()
    if not task:
        return
    ids = [int(product_id) for product_id in json.loads(task["ids_json"] or "[]") if int(product_id) > 0]
    processed_count = 0
    success_count = 0
    failed_count = 0
    cache_hit_count = 0
    failed_ids: list[int] = []
    request_delay_seconds = float_setting("ai_title_request_delay_seconds", 0.35, 0.0, 10.0)
    max_retries = int_setting("ai_title_max_retries", 1, 0, 5)
    retry_delay_seconds = float_setting("ai_title_retry_delay_seconds", 2.0, 0.0, 60.0)
    for index, product_id in enumerate(ids, start=1):
        try:
            item = get_processing_item(product_id)
            with connect() as db:
                db.execute(
                    """
                    UPDATE processing_title_tasks
                    SET note = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (f"正在处理 {index}/{len(ids)}：商品 #{product_id} 标题", now_text(), task_id),
                )
            had_cache = title_cache_exists(item)
            chinese_title = ""
            english_title = ""
            last_error: Exception | None = None
            for attempt in range(max_retries + 1):
                try:
                    chinese_title, english_title = generate_title_pair_for_processing_item(item)
                    validate_title_pair(chinese_title, english_title)
                    break
                except Exception as exc:
                    chinese_title = ""
                    english_title = ""
                    last_error = exc
                    if attempt >= max_retries:
                        raise
                    if retry_delay_seconds > 0:
                        time.sleep(retry_delay_seconds)
            if not chinese_title or not english_title:
                raise last_error or ValueError("标题生成结果为空")
            update_product_title(product_id, chinese_title)
            save_processing_override(
                product_id,
                ProcessingItemPayload(
                    english_title=english_title,
                    color=item.color,
                    size=item.size,
                    sku_code=item.sku_code,
                    declared_price=item.declared_price,
                    weight_g=item.weight_g,
                    length_cm=item.length_cm,
                    width_cm=item.width_cm,
                    height_cm=item.height_cm,
                    source_url=item.source_url,
                    stock=item.stock,
                    ship_days=item.ship_days,
                ),
            )
            success_count += 1
            if had_cache:
                cache_hit_count += 1
            elif request_delay_seconds > 0:
                time.sleep(request_delay_seconds)
        except Exception as exc:
            failed_count += 1
            failed_ids.append(product_id)
            note = f"商品 #{product_id} 标题处理失败：{str(exc)[:120]}"
            with connect() as db:
                db.execute(
                    """
                    INSERT OR REPLACE INTO app_settings (key, value, updated_at)
                    VALUES (?, ?, ?)
                    """,
                    (f"processing_title_task_{task_id}_last_error", note, now_text()),
                )
        processed_count += 1
        with connect() as db:
            db.execute(
                """
                UPDATE processing_title_tasks
                SET processed_count = ?, success_count = ?, failed_count = ?, cache_hit_count = ?, failed_ids_json = ?, note = ?, updated_at = ?
                WHERE id = ?
                """,
                (processed_count, success_count, failed_count, cache_hit_count, json.dumps(failed_ids), f"已处理 {processed_count}/{len(ids)} 个商品标题", now_text(), task_id),
            )
    status = "completed" if success_count else "failed"
    note = f"标题处理完成：成功 {success_count} 个，失败 {failed_count} 个，命中缓存 {cache_hit_count} 个"
    with connect() as db:
        db.execute(
            """
            UPDATE processing_title_tasks
            SET status = ?, processed_count = ?, success_count = ?, failed_count = ?, cache_hit_count = ?, failed_ids_json = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, processed_count, success_count, failed_count, cache_hit_count, json.dumps(failed_ids), note, now_text(), task_id),
        )


def schedule_processing_field_task_execution() -> None:
    global PROCESSING_FIELD_TASK_QUEUE_ACTIVE
    with PROCESSING_FIELD_TASK_QUEUE_LOCK:
        if PROCESSING_FIELD_TASK_QUEUE_ACTIVE:
            return
        PROCESSING_FIELD_TASK_QUEUE_ACTIVE = True
    thread = threading.Thread(target=run_processing_field_task_queue_worker, daemon=True)
    thread.start()


def resume_processing_field_task_queue() -> None:
    with connect() as db:
        db.execute(
            "UPDATE processing_field_tasks SET status = 'queued', note = '服务重启后重新排队批量字段处理', updated_at = ? WHERE status = 'running'",
            (now_text(),),
        )
        row = db.execute("SELECT id FROM processing_field_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
    if row:
        schedule_processing_field_task_execution()


def run_processing_field_task_queue_worker() -> None:
    global PROCESSING_FIELD_TASK_QUEUE_ACTIVE
    try:
        while True:
            with connect() as db:
                task = db.execute("SELECT * FROM processing_field_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
                if not task:
                    break
                db.execute("UPDATE processing_field_tasks SET status = 'running', note = ?, updated_at = ? WHERE id = ?", ("正在批量保存处理字段", now_text(), task["id"]))
            run_processing_field_task_background(int(task["id"]))
    finally:
        with PROCESSING_FIELD_TASK_QUEUE_LOCK:
            PROCESSING_FIELD_TASK_QUEUE_ACTIVE = False
        with connect() as db:
            row = db.execute("SELECT id FROM processing_field_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
        if row:
            schedule_processing_field_task_execution()


def run_processing_field_task_background(task_id: int) -> None:
    with connect() as db:
        task = db.execute("SELECT * FROM processing_field_tasks WHERE id = ?", (task_id,)).fetchone()
    if not task:
        return
    ids = [int(product_id) for product_id in json.loads(task["ids_json"] or "[]") if int(product_id) > 0]
    fields = ProcessingItemPayload(**json.loads(task["fields_json"] or "{}"))
    start_skc = str(task["start_skc"] or "").strip()
    processed_count = 0
    success_count = 0
    failed_count = 0
    failed_ids: list[int] = []
    if start_skc:
        timestamp = now_text()
        with connect() as db:
            for index, product_id in enumerate(ids):
                desired_skc = increment_skc_value(start_skc, index)
                skc = ensure_unique_skc(db, desired_skc, product_id)
                db.execute("UPDATE products SET skc = ?, updated_at = ? WHERE id = ?", (skc, timestamp, product_id))
    for product_id in ids:
        try:
            current = get_processing_item(product_id)
            with connect() as db:
                product_row = db.execute("SELECT purchase_price FROM products WHERE id = ?", (product_id,)).fetchone()
            base_price = float(product_row["purchase_price"] or 0) if product_row else float(current.declared_price or 0)
            declared_price_mode = str(getattr(fields, "declared_price_mode", "set") or "set").strip().lower()
            if declared_price_mode in {"add", "+"}:
                declared_price = round(base_price + float(fields.declared_price or 0), 2)
            elif declared_price_mode in {"multiply", "*", "x", "×"}:
                declared_price = round(base_price * float(fields.declared_price or 0), 2)
            else:
                declared_price = float(fields.declared_price or 0)
            merged_fields = ProcessingItemPayload(
                english_title=current.english_title,
                color=current.color,
                size=current.size,
                sku_code=current.sku_code,
                declared_price=declared_price,
                weight_g=fields.weight_g,
                length_cm=fields.length_cm,
                width_cm=fields.width_cm,
                height_cm=fields.height_cm,
                source_url=current.source_url,
                stock=fields.stock,
                ship_days=fields.ship_days,
            )
            save_processing_override(product_id, merged_fields)
            success_count += 1
        except Exception:
            failed_count += 1
            failed_ids.append(product_id)
        processed_count += 1
        with connect() as db:
            db.execute(
                """
                UPDATE processing_field_tasks
                SET processed_count = ?, success_count = ?, failed_count = ?, failed_ids_json = ?, note = ?, updated_at = ?
                WHERE id = ?
                """,
                (processed_count, success_count, failed_count, json.dumps(failed_ids), f"已保存 {processed_count}/{len(ids)} 个商品字段", now_text(), task_id),
            )
    status = "completed" if success_count else "failed"
    note = f"批量字段处理完成：成功 {success_count} 个，失败 {failed_count} 个"
    with connect() as db:
        db.execute(
            """
            UPDATE processing_field_tasks
            SET status = ?, processed_count = ?, success_count = ?, failed_count = ?, failed_ids_json = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, processed_count, success_count, failed_count, json.dumps(failed_ids), note, now_text(), task_id),
        )

def add_collection_task_event(
    db: sqlite3.Connection,
    task_id: int,
    stage: str,
    status: str = "info",
    message: str = "",
    page_no: int = 0,
    parsed_count: int = 0,
    imported_count: int = 0,
    skipped_count: int = 0,
) -> None:
    db.execute(
        """
        INSERT INTO collection_task_events (
            task_id, stage, status, message, page_no, parsed_count, imported_count, skipped_count, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (task_id, stage, status, message, page_no, parsed_count, imported_count, skipped_count, now_text()),
    )


def collection_failure_note(provider: str, parsed_count: int, imported_count: int, skipped_count: int, base_note: str = "") -> str:
    prefix = f"{provider} 真实采集已执行，解析 {parsed_count} 条，写入 {imported_count} 条，跳过重复/无标题 {skipped_count} 条。"
    if imported_count > 0:
        return prefix
    if parsed_count <= 0:
        reason = "失败原因：接口没有返回商品，可能是关键词无结果、筛选条件过窄、API/Cookie 被限制，或平台页面结构变化。"
    elif skipped_count >= parsed_count:
        reason = "失败原因：解析到的商品全部被跳过，主要原因是重复商品或缺少标题，没有新增候选商品。"
    else:
        reason = "失败原因：本次没有写入新增候选商品，请检查关键词、价格区间、黑名单和采集来源配置。"
    return f"{prefix} {reason}"


def collection_empty_note(base_note: str, created_count: int, skipped_count: int, candidate_count: int) -> str:
    prefix = f"{base_note} 写入 {created_count} 条，跳过重复 {skipped_count} 条。".strip()
    if created_count > 0:
        return prefix
    if candidate_count <= 0:
        reason = "失败原因：采集器没有返回候选商品，请检查关键词、采集来源配置或外部接口状态。"
    elif skipped_count >= candidate_count:
        reason = "失败原因：候选商品全部是重复记录，没有新增采集结果。"
    else:
        reason = "失败原因：本次没有新增采集结果，请检查过滤条件和采集数据格式。"
    return f"{prefix} {reason}"


def collection_exception_note(error_detail: object) -> str:
    message = str(error_detail or "采集执行失败").strip()
    return message if message.startswith("失败原因：") else f"失败原因：{message}"


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def first_value(row: dict[str, object], aliases: list[str], default: object = "") -> object:
    normalized_row = {normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        key = normalize_header(alias)
        if key in normalized_row and normalized_row[key] not in (None, ""):
            return normalized_row[key]
    return default


def to_float(value: object, default: float = 0) -> float:
    try:
        return float(str(value).replace("¥", "").replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def to_int(value: object, default: int = 0) -> int:
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return default


def parse_collection_rows(file_path: Path) -> list[dict[str, object]]:
    if file_path.suffix.lower() == ".json":
        data = json.loads(file_path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data = data.get("items") or data.get("data") or data.get("results") or [data]
        if not isinstance(data, list):
            return []
        return [{normalize_header(key): value for key, value in item.items()} for item in data if isinstance(item, dict)]
    if file_path.suffix.lower() == ".csv":
        text = file_path.read_text(encoding="utf-8-sig", errors="replace")
        reader = csv.DictReader(text.splitlines())
        return [{normalize_header(key): value for key, value in row.items()} for row in reader]
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_header(value) for value in rows[0]]
    parsed: list[dict[str, object]] = []
    for values in rows[1:]:
        parsed.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
    return parsed


QUOTE_SKC_ALIASES = ["skc", "SKC货号", "SKC 货号", "货号", "产品货号", "商品编码", "商家编码"]
QUOTE_SKU_ALIASES = ["SKU货号", "SKU 货号", "sku货号", "seller sku", "sku"]
QUOTE_PRICE_ALIASES = ["调整后申报价格", "调整后申报价", "申报价格", "原申报价格", "平台核价", "核价", "平台价", "核定价", "结算价", "供货价", "价格", "price"]
QUOTE_WEIGHT_ALIASES = ["重量", "重量g", "重量（g）", "weight", "weightg"]
FREIGHT_RULES_SETTING_KEY = "freight_rules_json"
DEFAULT_FREIGHT_ZONE_KEY = "freight_default_zone"
DEFAULT_WAREHOUSE_FEE_KEY = "freight_default_warehouse_fee"


def save_setting_value(db: sqlite3.Connection, key: str, value: str) -> None:
    timestamp = now_text()
    db.execute(
        """
        INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
        """,
        (key, value, timestamp),
    )


def parse_quote_rows_from_text(text: str) -> list[dict[str, object]]:
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    if not lines:
        return []
    delimiter = "\t" if "\t" in lines[0] else ("," if "," in lines[0] else None)
    if delimiter:
        headers = [normalize_header(value) for value in lines[0].split(delimiter)]
        has_known_header = any(header == normalize_header(alias) for header in headers for alias in QUOTE_SKC_ALIASES + QUOTE_PRICE_ALIASES)
        if has_known_header:
            reader = csv.DictReader(lines, delimiter=delimiter)
            return [{normalize_header(key): value for key, value in row.items()} for row in reader]
    parsed: list[dict[str, object]] = []
    for line in lines:
        parts = re.split(r"\s+|,|\t", line)
        parts = [part for part in parts if part]
        if len(parts) >= 2:
            parsed.append({"skc": parts[0], "平台核价": parts[1], "重量": parts[2] if len(parts) >= 3 else ""})
    return parsed


def parse_quote_rows_from_file(file_path: Path) -> list[dict[str, object]]:
    if file_path.suffix.lower() in {".txt", ".csv"}:
        return parse_quote_rows_from_text(file_path.read_text(encoding="utf-8-sig", errors="replace"))
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    parsed: list[dict[str, object]] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [normalize_header(value) for value in rows[0]]
        if not any(header for header in headers):
            continue
        for values in rows[1:]:
            parsed.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
    return parsed


def worksheet_rows_from_file(file_path: Path) -> dict[str, list[list[object]]]:
    suffix = file_path.suffix.lower()
    if suffix == ".xls":
        workbook = xlrd.open_workbook(str(file_path))
        return {
            sheet.name: [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
            for sheet in workbook.sheets()
        }
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    return {sheet.title: [list(row) for row in sheet.iter_rows(values_only=True)] for sheet in workbook.worksheets}


def parse_freight_rules(file_path: Path) -> dict[str, object]:
    sheets = worksheet_rows_from_file(file_path)
    first_mile_rules: list[dict[str, object]] = []
    last_mile_rules: list[dict[str, object]] = []
    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        headers = [normalize_header(value) for value in rows[0]]
        has_last_mile_header = any(header.startswith("zone") for header in headers)
        has_first_mile_header = not has_last_mile_header and ("头程渠道" in headers or "价格/kg" in headers or "价格kg" in headers)
        sheet_key = normalize_header(sheet_name)

        if has_first_mile_header:
            for values in rows[1:]:
                row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
                channel = str(first_value(row, ["头程渠道", "渠道"], "")).strip() or "空运"
                max_weight_g = to_int(first_value(row, ["重量段（g）", "重量段g", "重量g"], 0), 0)
                price_per_kg = to_float(first_value(row, ["价格/kg", "价格kg", "单价"], 0), 0)
                fixed_fee = to_float(first_value(row, ["附加费", "固定费"], 0), 0)
                if max_weight_g > 0 and price_per_kg > 0:
                    first_mile_rules.append({"channel": channel, "max_weight_g": max_weight_g, "price_per_kg": price_per_kg, "fixed_fee": fixed_fee})
            continue

        if has_last_mile_header:
            for values in rows[1:]:
                row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
                channel = str(first_value(row, ["尾程渠道", "渠道"], "")).strip() or "Temu"
                max_weight_g = to_int(first_value(row, ["重量（g）", "重量g", "重量"], 0), 0)
                zones = {header: to_float(row.get(header), 0) for header in headers if header.startswith("zone")}
                if max_weight_g > 0 and any(value > 0 for value in zones.values()):
                    last_mile_rules.append({"channel": channel, "max_weight_g": max_weight_g, "zones": zones})
            continue

        # No-header fallback: first mile defaults to columns channel/max_weight_g/price_per_kg/fixed_fee;
        # last mile defaults to channel/imperial/max_weight_g/zone1..zone9.
        is_first_mile_sheet = "1" in sheet_key or "first" in sheet_key or "头程" in sheet_key
        is_last_mile_sheet = "2" in sheet_key or "last" in sheet_key or "尾程" in sheet_key
        if not is_first_mile_sheet and not is_last_mile_sheet:
            numeric_width = max((sum(1 for value in row if to_float(value, 0) > 0) for row in rows[:5]), default=0)
            is_last_mile_sheet = numeric_width >= 4
            is_first_mile_sheet = not is_last_mile_sheet

        if is_first_mile_sheet:
            for values in rows:
                channel = str(values[0] if len(values) > 0 and values[0] else "空运").strip()
                max_weight_g = to_int(values[1] if len(values) > 1 else 0, 0)
                price_per_kg = to_float(values[2] if len(values) > 2 else 0, 0)
                fixed_fee = to_float(values[3] if len(values) > 3 else 0, 0)
                if max_weight_g > 0 and price_per_kg > 0:
                    first_mile_rules.append({"channel": channel or "空运", "max_weight_g": max_weight_g, "price_per_kg": price_per_kg, "fixed_fee": fixed_fee})
            continue

        if is_last_mile_sheet:
            for values in rows:
                channel = str(values[0] if len(values) > 0 and values[0] else "Temu").strip()
                max_weight_g = to_int(values[2] if len(values) > 2 else 0, 0)
                zones = {f"zone{index - 2}": to_float(values[index], 0) for index in range(3, min(len(values), 12))}
                if max_weight_g > 0 and any(value > 0 for value in zones.values()):
                    last_mile_rules.append({"channel": channel or "Temu", "max_weight_g": max_weight_g, "zones": zones})
    first_mile_rules.sort(key=lambda item: int(item["max_weight_g"]))
    last_mile_rules.sort(key=lambda item: int(item["max_weight_g"]))
    return {"first_mile": first_mile_rules, "last_mile": last_mile_rules}

def freight_rules() -> dict[str, object]:
    raw = get_setting_value(FREIGHT_RULES_SETTING_KEY, "")
    if raw:
        try:
            data = json.loads(raw)
            if isinstance(data, dict):
                return data
        except json.JSONDecodeError:
            pass
    return {"first_mile": [{"channel": "空运", "max_weight_g": 10000000, "price_per_kg": 69, "fixed_fee": 4}], "last_mile": []}


def default_warehouse_fee() -> float:
    return to_float(get_setting_value(DEFAULT_WAREHOUSE_FEE_KEY, "15"), 15)


def normalize_freight_zone(zone: str | None) -> str:
    zone_key = normalize_header(zone or "zone5") or "zone5"
    if not re.fullmatch(r"zone[1-9]", zone_key):
        raise HTTPException(status_code=400, detail="默认 Zone 必须是 zone1~zone9")
    return zone_key


def freight_payload_to_rules(payload: FreightRulesPayload) -> dict[str, object]:
    first_mile: list[dict[str, object]] = []
    for rule in payload.first_mile:
        if rule.max_weight_g <= 0 or rule.price_per_kg < 0 or rule.fixed_fee < 0:
            raise HTTPException(status_code=400, detail="头程规则的重量和费用不能为负数")
        first_mile.append({
            "channel": rule.channel.strip() or "空运",
            "max_weight_g": int(rule.max_weight_g),
            "price_per_kg": float(rule.price_per_kg),
            "fixed_fee": float(rule.fixed_fee),
        })
    last_mile: list[dict[str, object]] = []
    for rule in payload.last_mile:
        if rule.max_weight_g <= 0:
            raise HTTPException(status_code=400, detail="尾程规则的重量必须大于 0")
        zones: dict[str, float] = {}
        for key, value in (rule.zones or {}).items():
            zone_key = normalize_freight_zone(str(key))
            fee = float(value or 0)
            if fee < 0:
                raise HTTPException(status_code=400, detail="尾程费用不能为负数")
            zones[zone_key] = fee
        last_mile.append({
            "channel": rule.channel.strip() or "尾程",
            "max_weight_g": int(rule.max_weight_g),
            "zones": zones,
        })
    first_mile.sort(key=lambda item: int(item.get("max_weight_g") or 0))
    last_mile.sort(key=lambda item: int(item.get("max_weight_g") or 0))
    if not first_mile:
        raise HTTPException(status_code=400, detail="至少需要 1 条头程规则")
    return {"first_mile": first_mile, "last_mile": last_mile}


def calculate_first_mile(weight_g: int) -> float:
    if weight_g <= 0:
        return 0
    rules = freight_rules().get("first_mile", [])
    for rule in rules if isinstance(rules, list) else []:
        if weight_g <= int(rule.get("max_weight_g") or 0):
            return float(math.ceil(float(rule.get("price_per_kg") or 0) * weight_g / 1000 + float(rule.get("fixed_fee") or 0)))
    return 0


def calculate_last_mile(weight_g: int, zone: str | None = None) -> float:
    if weight_g <= 0:
        return 0
    zone_key = normalize_header(zone or get_setting_value(DEFAULT_FREIGHT_ZONE_KEY, "zone5") or "zone5")
    rules = freight_rules().get("last_mile", [])
    for rule in rules if isinstance(rules, list) else []:
        if weight_g <= int(rule.get("max_weight_g") or 0):
            zones = rule.get("zones") if isinstance(rule.get("zones"), dict) else {}
            return float(zones.get(zone_key) or 0)
    return 0


def quote_row_value(row: dict[str, object], aliases: list[str]) -> object:
    return first_value(row, aliases, "")


def skc_from_sku_code(sku_code: str, known_skc_values: list[str]) -> str:
    clean_sku = str(sku_code or "").strip()
    if not clean_sku:
        return ""
    upper_sku = clean_sku.upper()
    for skc in sorted(known_skc_values, key=len, reverse=True):
        upper_skc = skc.upper()
        if upper_sku == upper_skc or upper_sku.startswith(f"{upper_skc}-"):
            return skc
    match = re.match(r"^([A-Za-z]{1,8}\d{6,}(?:-\d+)?)", clean_sku)
    return match.group(1) if match else clean_sku


def apply_quote_rows(rows: list[dict[str, object]]) -> QuoteImportResult:
    timestamp = now_text()
    updated_count = 0
    unmatched: list[dict[str, object]] = []
    invalid: list[dict[str, object]] = []
    duplicates: list[str] = []
    seen: set[str] = set()
    with connect() as db:
        known_skc_values = [str(row["skc"] or "") for row in db.execute("SELECT skc FROM products WHERE skc != ''").fetchall()]
        for index, row in enumerate(rows, start=1):
            skc = str(quote_row_value(row, QUOTE_SKC_ALIASES) or "").strip()
            sku_code = str(quote_row_value(row, QUOTE_SKU_ALIASES) or "").strip()
            if not skc:
                skc = skc_from_sku_code(sku_code, known_skc_values)
            quote_price = to_float(quote_row_value(row, QUOTE_PRICE_ALIASES), 0)
            weight_g = to_int(quote_row_value(row, QUOTE_WEIGHT_ALIASES), 0)
            if not skc or quote_price <= 0:
                invalid.append({"row": index, "skc": skc, "reason": "缺少 SKC 或平台核价"})
                continue
            normalized_skc = skc.upper()
            if normalized_skc in seen:
                duplicates.append(skc)
                continue
            seen.add(normalized_skc)
            product = db.execute("SELECT * FROM products WHERE upper(skc) = ?", (normalized_skc,)).fetchone()
            if not product:
                unmatched.append({"row": index, "skc": skc, "platform_quote_price": quote_price})
                continue
            payload = ProductPayload(
                title=product["title"],
                skc=product["skc"],
                sku_summary=product["sku_summary"],
                purchase_price=float(product["purchase_price"] or 0),
                platform_quote_price=quote_price,
                weight_g=weight_g or int(product["weight_g"] or 0),
                first_mile=0 if weight_g else float(product["first_mile"] or 0),
                warehouse_fee=default_warehouse_fee(),
                last_mile=0 if weight_g else float(product["last_mile"] or product["platform_cost"] or 0),
                platform_cost=float(product["platform_cost"] or 0),
                status=product["status"],
            )
            payload = product_payload_with_freight(payload)
            total_cost, estimated_profit, gross_margin = recalc(payload)
            status = quote_status(estimated_profit, gross_margin, product["status"])
            db.execute(
                """
                UPDATE products SET platform_quote_price = ?, weight_g = ?, first_mile = ?, warehouse_fee = ?,
                last_mile = ?, platform_cost = ?, total_cost = ?, estimated_profit = ?, gross_margin = ?,
                status = ?, updated_at = ? WHERE id = ?
                """,
                (
                    payload.platform_quote_price, payload.weight_g, payload.first_mile, payload.warehouse_fee,
                    payload.last_mile, payload.last_mile, total_cost, estimated_profit, gross_margin, status,
                    timestamp, product["id"],
                ),
            )
            updated_count += 1
    return QuoteImportResult(
        updated_count=updated_count,
        unmatched_count=len(unmatched),
        invalid_count=len(invalid),
        duplicate_count=len(duplicates),
        rows_count=len(rows),
        unmatched=unmatched[:50],
        invalid=invalid[:50],
        duplicates=duplicates[:50],
    )


def sync_product_weight_from_processing(db: sqlite3.Connection, product_id: int, weight_g: int) -> None:
    clean_weight = int(weight_g or 0)
    if clean_weight <= 0:
        return
    product = db.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    if not product:
        return
    payload = ProductPayload(
        title=product["title"],
        skc=product["skc"],
        sku_summary=product["sku_summary"],
        purchase_price=float(product["purchase_price"] or 0),
        platform_quote_price=float(product["platform_quote_price"] or 0),
        weight_g=clean_weight,
        first_mile=0,
        warehouse_fee=default_warehouse_fee(),
        last_mile=0,
        platform_cost=0,
        status=product["status"],
    )
    payload = product_payload_with_freight(payload)
    total_cost, estimated_profit, gross_margin = recalc(payload)
    status = quote_status(estimated_profit, gross_margin, product["status"])
    db.execute(
        """
        UPDATE products SET weight_g = ?, first_mile = ?, warehouse_fee = ?, last_mile = ?, platform_cost = ?,
        total_cost = ?, estimated_profit = ?, gross_margin = ?, status = ?, updated_at = ? WHERE id = ?
        """,
        (
            payload.weight_g, payload.first_mile, payload.warehouse_fee, payload.last_mile, payload.last_mile,
            total_cost, estimated_profit, gross_margin, status, now_text(), product_id,
        ),
    )


def recalculate_all_product_freight(db: sqlite3.Connection) -> int:
    timestamp = now_text()
    rows = db.execute("SELECT * FROM products").fetchall()
    updated = 0
    for product in rows:
        payload = ProductPayload(
            title=product["title"],
            skc=product["skc"],
            sku_summary=product["sku_summary"],
            purchase_price=float(product["purchase_price"] or 0),
            platform_quote_price=float(product["platform_quote_price"] or 0),
            weight_g=int(product["weight_g"] or 0),
            first_mile=0,
            warehouse_fee=default_warehouse_fee(),
            last_mile=0,
            platform_cost=0,
            status=product["status"],
        )
        payload = product_payload_with_freight(payload)
        total_cost, estimated_profit, gross_margin = recalc(payload)
        status = quote_status(estimated_profit, gross_margin, product["status"])
        db.execute(
            """
            UPDATE products SET first_mile = ?, warehouse_fee = ?, last_mile = ?, platform_cost = ?,
            total_cost = ?, estimated_profit = ?, gross_margin = ?, status = ?, updated_at = ? WHERE id = ?
            """,
            (payload.first_mile, payload.warehouse_fee, payload.last_mile, payload.last_mile, total_cost, estimated_profit, gross_margin, status, timestamp, product["id"]),
        )
        updated += 1
    return updated


def collection_preflight() -> dict[str, object]:
    cookie_path = Path(get_setting_value("1688_cookie_path", "data/1688_cookies.json"))
    if not cookie_path.is_absolute():
        cookie_path = BASE_DIR / cookie_path
    inline_cookie = get_setting_value("1688_cookie")
    onebound_key = get_setting_value("onebound_key")
    onebound_secret = get_setting_value("onebound_secret")
    return {
        "default_mode": get_setting_value("collection_mode", "1688"),
        "1688": {
            "cookie_path": str(cookie_path),
            "cookie_exists": cookie_path.exists(),
            "inline_cookie_configured": bool(inline_cookie),
            "onebound_configured": bool(onebound_key and onebound_secret),
            "configured": bool(onebound_key and onebound_secret) or bool(inline_cookie) or cookie_path.exists(),
            "status": "ready" if (onebound_key and onebound_secret) else ("cookie_ready" if (inline_cookie or cookie_path.exists()) else "missing_credentials"),
        },
    }


def run_local_collector(payload: CollectionTaskPayload, task_id: int, mode: str, collector: str, note: str) -> CollectorRunResult:
    raise HTTPException(status_code=410, detail="本地模拟采集已移除，请使用 1688 / 万邦 API 真实采集或文件导入")


def write_collection_request(task_id: int, payload: CollectionTaskPayload, mode: str, collector: str) -> str:
    request_dir = DATA_DIR / "collection_requests"
    request_dir.mkdir(parents=True, exist_ok=True)
    target = request_dir / f"collection_task_{task_id}.json"
    request_payload = {
        "task_id": task_id,
        "mode": mode,
        "collector": collector,
        "keyword": payload.keyword,
        "source": payload.source,
        "target_count": payload.target_count,
        "min_price": payload.min_price,
        "max_price": payload.max_price,
        "blacklist": payload.blacklist,
        "created_at": now_text(),
        "execution_policy": "manual_review_required",
    }
    if mode == "1688":
        request_payload["1688"] = {
            "provider": "onebound" if get_setting_value("onebound_key") and get_setting_value("onebound_secret") else "cookie",
            "onebound_configured": bool(get_setting_value("onebound_key") and get_setting_value("onebound_secret")),
            "cookie_path": get_setting_value("1688_cookie_path", "data/1688_cookies.json"),
        }
    target.write_text(json.dumps(request_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(target)


def ensure_collection_request(task: sqlite3.Row) -> str:
    request_path = task["request_path"]
    if request_path and Path(request_path).exists():
        return request_path
    payload = CollectionTaskPayload(
        keyword=task["keyword"],
        source=task["source"],
        mode=task["mode"],
        collector=task["collector"],
        target_count=task["target_count"],
        min_price=task["min_price"],
        max_price=task["max_price"],
        blacklist=task["blacklist"],
    )
    return write_collection_request(task["id"], payload, task["mode"], task["collector"])


def resolve_collector(payload: CollectionTaskPayload) -> tuple[str, str, str]:
    requested_mode = payload.mode or get_setting_value("collection_mode", "1688")
    collector = payload.collector or requested_mode or "local"
    note = ""
    preflight = collection_preflight()
    if requested_mode == "1688":
        return "1688", "1688_public_search", note
    if requested_mode == "simulate":
        raise HTTPException(status_code=410, detail="本地模拟采集已移除，请选择 1688 / 万邦 API 采集")
    return requested_mode, collector, note


def run_collector(payload: CollectionTaskPayload, task_id: int) -> CollectorRunResult:
    mode, collector, note = resolve_collector(payload)
    if mode == "1688":
        return CollectorRunResult(
            mode=mode,
            collector=collector,
            note="真实采集请求已生成，需人工确认后再执行。",
            candidates=[],
        )
    raise HTTPException(status_code=400, detail=f"不支持的采集模式：{mode}")


def call_openai_compatible_chat(base_url: str, api_key: str, model: str, messages: list[dict[str, str]], timeout: int = 20) -> str:
    endpoint = base_url.rstrip("/")
    if not endpoint.endswith("/chat/completions"):
        endpoint = f"{endpoint}/v1/chat/completions"
    body = json.dumps({"model": model, "messages": messages, "temperature": 0.4}).encode("utf-8")
    request = urllib.request.Request(
        endpoint,
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:500]
        raise HTTPException(status_code=502, detail=f"AI 接口返回错误：{exc.code} {detail}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"AI 接口连接失败：{exc.reason}") from exc
    except TimeoutError as exc:
        raise HTTPException(status_code=504, detail="AI 接口请求超时") from exc
    try:
        return payload["choices"][0]["message"]["content"].strip()
    except (KeyError, IndexError, TypeError) as exc:
        raise HTTPException(status_code=502, detail="AI 接口返回格式异常") from exc


def extract_onebound_items(payload: object) -> list[dict[str, object]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return []
    for key in ["items", "item", "data", "result", "results"]:
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
        if isinstance(value, dict):
            nested = extract_onebound_items(value)
            if nested:
                return nested
    return []


def normalize_onebound_items(items: list[dict[str, object]], source: str) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for item in items:
        normalized = {normalize_header(key): value for key, value in item.items()}
        title = first_value(normalized, ["title", "name", "subject", "商品标题"], "")
        num_iid = str(first_value(normalized, ["num_iid", "offer_id", "offerId", "id", "item_id"], "")).strip()
        url = str(first_value(normalized, ["detail_url", "url", "item_url", "link", "商品链接"], "")).strip()
        if not url and num_iid:
            url = f"https://detail.1688.com/offer/{num_iid}.html"
        image_url = str(first_value(normalized, ["pic_url", "image", "img", "main_image", "picurl", "image_url"], "")).strip()
        rows.append(
            {
                "title": title,
                "source": source,
                "url": url,
                "image_url": image_url,
                "price": first_value(normalized, ["price", "promotion_price", "sale_price", "orginal_price", "价格"], 0),
                "sales": first_value(normalized, ["sales", "sales_volume", "sale_count", "month_sales", "销量"], 0),
                "image_count": 1 if image_url else 0,
            }
        )
    return rows



def collection_existing_dedupe_sets(db: sqlite3.Connection | None = None) -> tuple[set[str], set[str]]:
    close_db = False
    if db is None:
        db = connect()
        close_db = True
    try:
        existing_urls = {
            str(row["source_url"] or "").strip()
            for row in db.execute("SELECT source_url FROM collection_items WHERE source_url != ''").fetchall()
        }
        existing_keys = {
            f"{row['source']}::{row['title']}".lower()
            for row in db.execute("SELECT source, title FROM collection_items").fetchall()
        }
        return existing_urls, existing_keys
    finally:
        if close_db:
            db.close()


def collection_row_dedupe_key(row: dict[str, object], fallback_source: str = "") -> tuple[str, str]:
    source = str(row.get("source") or fallback_source or "").strip()
    title = str(row.get("title") or "").strip()
    source_url = str(row.get("source_url") or row.get("url") or "").strip()
    return source_url, f"{source}::{title}".lower()


def run_onebound_1688_search(payload: CollectionTaskPayload, task_id: int | None = None) -> list[dict[str, object]]:
    key = get_setting_value("onebound_key", "").strip()
    secret = get_setting_value("onebound_secret", "").strip()
    if not key or not secret:
        raise HTTPException(status_code=400, detail="Onebound API Key/Secret is not configured")
    target_count = max(1, min(int(payload.target_count or 1), COLLECTION_TARGET_MAX))
    page_size = min(ONEBOUND_PAGE_SIZE_MAX, target_count)
    existing_urls, existing_keys = collection_existing_dedupe_sets()
    collected: list[dict[str, object]] = []
    seen_urls: set[str] = set()
    seen_keys: set[str] = set()
    max_pages = max(1, min(20, ((target_count + page_size - 1) // page_size) + 8))
    last_error = ""
    for page_no in range(1, max_pages + 1):
        raise_if_collection_task_cancelled(task_id)
        params = {
            "key": key,
            "secret": secret,
            "q": payload.keyword.strip(),
            "page": str(page_no),
            "page_size": str(page_size),
        }
        if payload.min_price:
            params["start_price"] = str(payload.min_price)
        if payload.max_price:
            params["end_price"] = str(payload.max_price)
        endpoint = "https://api-gw.onebound.cn/1688/item_search?" + urllib.parse.urlencode(params)
        request = urllib.request.Request(endpoint, headers={"Accept": "application/json"}, method="GET")
        try:
            with urllib.request.urlopen(request, timeout=45) as response:
                data = json.loads(response.read().decode("utf-8", errors="replace"))
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:500]
            raise HTTPException(status_code=502, detail=f"Onebound API returned error: {exc.code} {detail}") from exc
        except urllib.error.URLError as exc:
            raise HTTPException(status_code=502, detail=f"Onebound API connection failed: {exc.reason}") from exc
        except TimeoutError as exc:
            raise HTTPException(status_code=504, detail="Onebound API request timed out") from exc
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=502, detail="Onebound API returned non-JSON content") from exc
        error = first_value(data, ["error", "error_msg", "msg", "message"], "")
        code = str(first_value(data, ["error_code", "code"], "")).strip()
        items = extract_onebound_items(data)
        if not items:
            if error:
                last_error = f"{code} {error}".strip()
                if not collected:
                    raise HTTPException(status_code=502, detail=f"Onebound API error: {last_error}".strip())
            break
        rows = normalize_onebound_items(items, payload.source or "1688")
        new_count_this_page = 0
        for row in rows:
            raise_if_collection_task_cancelled(task_id)
            source_url, dedupe_key = collection_row_dedupe_key(row, payload.source or "1688")
            if (source_url and source_url in existing_urls) or dedupe_key in existing_keys:
                continue
            if (source_url and source_url in seen_urls) or dedupe_key in seen_keys:
                continue
            if source_url:
                seen_urls.add(source_url)
            seen_keys.add(dedupe_key)
            collected.append(row)
            new_count_this_page += 1
            if len(collected) >= target_count:
                return collected[:target_count]
        if len(items) < page_size:
            break
        if new_count_this_page == 0 and page_no >= max_pages:
            break
        raise_if_collection_task_cancelled(task_id)
        time.sleep(0.2)
    if not collected and last_error:
        raise HTTPException(status_code=502, detail=f"Onebound API error: {last_error}".strip())
    return collected[:target_count]

def extract_offer_id(source_url: str) -> str:
    match = re.search(r"offer/(\d+)\.html", source_url or "")
    if match:
        return match.group(1)
    match = re.search(r"(?:num_iid|item_id|offer_id|id)=(\d+)", source_url or "")
    return match.group(1) if match else ""


def onebound_request(endpoint: str, params: dict[str, str], label: str) -> dict[str, object]:
    key = get_setting_value("onebound_key", "").strip()
    secret = get_setting_value("onebound_secret", "").strip()
    if not key or not secret:
        raise HTTPException(status_code=400, detail="万邦 API Key/Secret 未配置")
    query = {"key": key, "secret": secret, **params}
    url = f"https://api-gw.onebound.cn/1688/{endpoint}?" + urllib.parse.urlencode(query)
    request = urllib.request.Request(url, headers={"Accept": "application/json"}, method="GET")
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            payload = json.loads(response.read().decode("utf-8", errors="replace"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:500]
        raise HTTPException(status_code=502, detail=f"万邦 {label} 返回错误：{exc.code} {detail}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"万邦 {label} 连接失败：{exc.reason}") from exc
    except TimeoutError as exc:
        raise HTTPException(status_code=504, detail=f"万邦 {label} 请求超时") from exc
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=502, detail=f"万邦 {label} 返回非 JSON 内容") from exc
    error = first_value(payload, ["error", "error_msg", "msg", "message"], "")
    code = str(first_value(payload, ["error_code", "code"], "")).strip()
    if error and not extract_onebound_items(payload):
        raise HTTPException(status_code=502, detail=f"万邦 {label} 错误：{code} {error}".strip())
    return payload


def run_onebound_1688_detail(source_url: str) -> dict[str, object]:
    offer_id = extract_offer_id(source_url)
    if not offer_id:
        raise HTTPException(status_code=400, detail="无法从商品链接提取 1688 offer id")
    return onebound_request("item_get", {"num_iid": offer_id}, "详情")


def first_dict(value: object) -> dict[str, object]:
    if isinstance(value, dict):
        if any(key in value for key in ["title", "item_imgs", "pic_url", "skus", "props_name", "desc_img"]):
            return value
        for key in ["item", "data", "result"]:
            if isinstance(value.get(key), dict):
                return first_dict(value[key])
        return value
    return {}


def flatten_strings(value: object) -> list[str]:
    values: list[str] = []
    if isinstance(value, str):
        if value.strip():
            values.append(value.strip())
    elif isinstance(value, list):
        for item in value:
            values.extend(flatten_strings(item))
    elif isinstance(value, dict):
        for item in value.values():
            values.extend(flatten_strings(item))
    return values


def extract_image_urls(value: object) -> list[str]:
    urls: list[str] = []
    seen: set[str] = set()
    for text in flatten_strings(value):
        if text.startswith("http") and re.search(r"\.(jpg|jpeg|png|webp|gif)(\?|$)", text, re.I) and text not in seen:
            seen.add(text)
            urls.append(text)
        for url in re.findall(r"https?://[^\s'\"<>]+", text):
            clean_url = url.rstrip(",;)]}")
            if re.search(r"\.(jpg|jpeg|png|webp|gif)(\?|$)", clean_url, re.I) and clean_url not in seen:
                seen.add(clean_url)
                urls.append(clean_url)
    return urls


def parse_1688_properties_name(value: str) -> dict[str, str]:
    props: dict[str, str] = {}
    for part in str(value or "").split(";"):
        pieces = part.split(":", 3)
        if len(pieces) == 4:
            props[f"{pieces[0]}:{pieces[1]}"] = pieces[3].strip()
    return props


def extract_named_prop(value: str, name: str) -> str:
    for part in str(value or "").split(";"):
        pieces = part.split(":", 3)
        if len(pieces) == 4 and pieces[2] == name:
            return pieces[3].strip()
    return ""


def normalize_prop_images(value: object) -> dict[str, str]:
    if isinstance(value, dict):
        value = value.get("prop_img") or value.get("props_img") or value.get("list") or value.get("data") or value
    if isinstance(value, dict):
        value = list(value.values())
    if not isinstance(value, list):
        return {}
    images: dict[str, str] = {}
    for item in value:
        if not isinstance(item, dict):
            continue
        key = str(item.get("properties") or item.get("property") or item.get("props") or "").strip()
        urls = extract_image_urls(item)
        if key and urls:
            images[key] = urls[0]
    return images


def normalize_detail_payload(payload: dict[str, object]) -> dict[str, object]:
    item = first_dict(payload)
    title = str(first_value(item, ["title", "name", "subject"], "")).strip()
    price = to_float(first_value(item, ["price", "promotion_price", "sale_price", "orginal_price"], 0), 0)
    main_images = extract_image_urls(first_value(item, ["item_imgs", "item_img", "pic_url", "pic_urls", "images", "imgs"], []))
    desc_images = extract_image_urls(first_value(item, ["desc_img", "desc_imgs", "desc_images", "detail_imgs", "detail_images"], []))
    all_images = extract_image_urls(item)
    prop_name_map = parse_1688_properties_name(str(item.get("props_name") or item.get("property_alias") or ""))
    prop_image_map = normalize_prop_images(item.get("prop_imgs") or item.get("props_img") or item.get("prop_images"))
    sku_rows: list[dict[str, object]] = []
    sku_source = item.get("skus") or item.get("sku") or item.get("sku_infos") or item.get("props_img") or []
    if isinstance(sku_source, dict):
        sku_source = sku_source.get("sku") or sku_source.get("skus") or sku_source.get("list") or sku_source.get("data") or list(sku_source.values())
    if not isinstance(sku_source, list):
        sku_source = []
    for index, sku in enumerate(sku_source):
        if not isinstance(sku, dict):
            continue
        normalized = {normalize_header(key): value for key, value in sku.items()}
        properties_name = str(first_value(normalized, ["properties_name", "propertiesname", "props_name"], "")).strip()
        properties = str(first_value(normalized, ["properties", "props"], "")).strip()
        color = extract_named_prop(properties_name, "颜色") or str(first_value(normalized, ["color", "颜色", "颜色分类", "prop_value", "propvalue"], "")).strip()
        size = extract_named_prop(properties_name, "尺码") or str(first_value(normalized, ["size", "尺码", "规格", "spec"], "")).strip()
        sku_id = str(first_value(normalized, ["sku_id", "skuid", "id", "spec_id"], "")).strip()
        images = extract_image_urls(sku)
        for prop_key in properties.split(";"):
            if prop_key in prop_image_map and prop_image_map[prop_key] not in images:
                images.append(prop_image_map[prop_key])
            if not color and prop_key in prop_name_map and prop_key.startswith("0:"):
                color = prop_name_map[prop_key]
            if not size and prop_key in prop_name_map and prop_key.startswith("1:"):
                size = prop_name_map[prop_key]
        sku_rows.append({"sku_id": sku_id, "color": color, "size": size, "images": images, "sort_order": index})
    return {"title": title, "price": price, "main_images": main_images, "desc_images": desc_images, "all_images": all_images, "skus": sku_rows}


def apply_detail_to_product(db: sqlite3.Connection, product_id: int, source_url: str, payload: dict[str, object]) -> dict[str, object]:
    detail = normalize_detail_payload(payload)
    timestamp = now_text()
    db.execute(
        """
        INSERT INTO product_detail_snapshots (product_id, provider, source_url, raw_json, status, note, updated_at)
        VALUES (?, 'onebound', ?, ?, 'completed', ?, ?)
        ON CONFLICT(product_id) DO UPDATE SET provider=excluded.provider, source_url=excluded.source_url,
        raw_json=excluded.raw_json, status=excluded.status, note=excluded.note, updated_at=excluded.updated_at
        """,
        (product_id, source_url, json.dumps(payload, ensure_ascii=False), "万邦详情采集完成", timestamp),
    )
    db.execute("DELETE FROM product_sku_images WHERE product_id = ?", (product_id,))
    image_order = 0
    inserted_keys: set[tuple[str, str, str, str]] = set()

    def insert_detail_image(url: str, source: str, color: str = "", size: str = "", sku_id: str = "") -> None:
        nonlocal image_order
        clean_url = str(url or "").strip()
        if not clean_url:
            return
        key = (source, clean_url, str(color or ""), str(size or ""))
        if key in inserted_keys:
            return
        inserted_keys.add(key)
        db.execute(
            """
            INSERT INTO product_sku_images (product_id, sku_id, color, size, image_url, sort_order, source, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (product_id, str(sku_id or ""), str(color or ""), str(size or ""), clean_url, image_order, source, timestamp),
        )
        image_order += 1

    for sku in detail["skus"]:
        for url in sku["images"]:
            insert_detail_image(url, "detail_sku_image", sku["color"], sku["size"], sku["sku_id"])
    for url in detail["main_images"]:
        insert_detail_image(url, "detail_main_image")
    desc_urls = detail.get("desc_images") or []
    fallback_detail_urls = [url for url in detail["all_images"] if url not in set(detail["main_images"])]
    for url in desc_urls or fallback_detail_urls[:40]:
        insert_detail_image(url, "detail_desc_image")
    colors = sorted({str(sku["color"]).strip() for sku in detail["skus"] if str(sku["color"]).strip()})
    sizes = sorted({str(sku["size"]).strip() for sku in detail["skus"] if str(sku["size"]).strip()})
    product = db.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    if product:
        title = detail["title"] or product["title"]
        purchase_price = float(detail["price"] or product["purchase_price"] or 0)
        payload_for_cost = ProductPayload(
            title=title,
            skc=product["skc"],
            sku_summary=f"{len(colors) or 1}色 × {len(sizes) or 1}码" if colors or sizes else product["sku_summary"],
            purchase_price=purchase_price,
            first_mile=float(product["first_mile"]),
            platform_cost=float(product["platform_cost"]),
            status=product["status"],
        )
        total_cost, estimated_profit, gross_margin = recalc(payload_for_cost)
        db.execute(
            """
            UPDATE products SET title = ?, sku_summary = ?, purchase_price = ?, total_cost = ?,
            estimated_profit = ?, gross_margin = ?, updated_at = ? WHERE id = ?
            """,
            (payload_for_cost.title, payload_for_cost.sku_summary, payload_for_cost.purchase_price, total_cost, estimated_profit, gross_margin, timestamp, product_id),
        )
    override = db.execute("SELECT * FROM processing_overrides WHERE product_id = ?", (product_id,)).fetchone()
    color_text = " / ".join(colors)
    size_text = " / ".join(sizes)
    default_sku_code = f"{product['skc']}-{(colors[0] if colors else 'COLOR')}-{(sizes[0] if sizes else 'SIZE')}" if product else ""
    if override:
        color_text = color_text or override["color"]
        size_text = size_text or override["size"]
        sku_code = override["sku_code"]
        if (not sku_code or sku_code.endswith("COLOR-SIZE")) and default_sku_code:
            sku_code = default_sku_code
        db.execute(
            "UPDATE processing_overrides SET color = ?, size = ?, sku_code = ?, source_url = ?, updated_at = ? WHERE product_id = ?",
            (color_text, size_text, sku_code, source_url, timestamp, product_id),
        )
    elif product and (color_text or size_text):
        default_declared_price = round(float(product["total_cost"] or 0) + 239, 2)
        db.execute(
            """
            INSERT INTO processing_overrides (
                product_id, english_title, color, size, sku_code, declared_price,
                weight_g, length_cm, width_cm, height_cm, source_url, stock, ship_days, updated_at
            ) VALUES (?, '', ?, ?, ?, ?, 350, 15, 10, 2, ?, 999, 9, ?)
            """,
            (product_id, color_text, size_text, default_sku_code, default_declared_price, source_url, timestamp),
        )
    return detail


def try_enrich_product_from_onebound_detail(db: sqlite3.Connection, product_id: int, source_url: str) -> str:
    if not source_url.startswith("http"):
        return "无详情链接，跳过万邦详情采集"
    if not get_setting_value("onebound_key") or not get_setting_value("onebound_secret"):
        db.execute(
            """
            INSERT INTO product_detail_snapshots (product_id, provider, source_url, raw_json, status, note, updated_at)
            VALUES (?, 'onebound', ?, '', 'skipped', '万邦 API Key/Secret 未配置，跳过详情采集', ?)
            ON CONFLICT(product_id) DO UPDATE SET source_url=excluded.source_url, status=excluded.status, note=excluded.note, updated_at=excluded.updated_at
            """,
            (product_id, source_url, now_text()),
        )
        return "万邦 API Key/Secret 未配置，跳过详情采集"
    try:
        detail_payload = run_onebound_1688_detail(source_url)
        detail = apply_detail_to_product(db, product_id, source_url, detail_payload)
        return f"万邦详情采集完成：{len(detail['skus'])} 个 SKU，{len(detail['all_images'])} 张图片"
    except HTTPException as exc:
        db.execute(
            """
            INSERT INTO product_detail_snapshots (product_id, provider, source_url, raw_json, status, note, updated_at)
            VALUES (?, 'onebound', ?, '', 'failed', ?, ?)
            ON CONFLICT(product_id) DO UPDATE SET source_url=excluded.source_url, status=excluded.status, note=excluded.note, updated_at=excluded.updated_at
            """,
            (product_id, source_url, str(exc.detail), now_text()),
        )
        return f"万邦详情采集失败：{exc.detail}"


def fetch_text(url: str, timeout: int = 25) -> str:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        },
        method="GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read()
            charset = response.headers.get_content_charset() or "utf-8"
            return raw.decode(charset, errors="replace")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:300]
        raise HTTPException(status_code=502, detail=f"真实采集页面返回错误：{exc.code} {detail}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"真实采集连接失败：{exc.reason}") from exc
    except TimeoutError as exc:
        raise HTTPException(status_code=504, detail="真实采集请求超时") from exc


def load_cookie_header(cookie_path: Path) -> str:
    inline_cookie = get_setting_value("1688_cookie", "").strip()
    if inline_cookie:
        return inline_cookie
    if not cookie_path.exists():
        raise HTTPException(status_code=400, detail=f"1688 Cookie 文件不存在：{cookie_path}")
    text = cookie_path.read_text(encoding="utf-8", errors="replace").strip()
    if not text:
        raise HTTPException(status_code=400, detail=f"1688 Cookie 文件为空：{cookie_path}")
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return text
    if isinstance(data, dict):
        if isinstance(data.get("cookies"), list):
            data = data["cookies"]
        else:
            return "; ".join(f"{key}={value}" for key, value in data.items() if value not in (None, ""))
    if isinstance(data, list):
        parts = []
        for item in data:
            if isinstance(item, dict) and item.get("name") and item.get("value") is not None:
                parts.append(f"{item['name']}={item['value']}")
        return "; ".join(parts)
    raise HTTPException(status_code=400, detail="1688 Cookie 文件格式不支持，请使用浏览器导出的 cookies JSON 或 Cookie 字符串")


def fetch_text_with_cookie(url: str, cookie_header: str, timeout: int = 25) -> str:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Cookie": cookie_header,
        },
        method="GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read()
            charset = response.headers.get_content_charset() or "utf-8"
            return raw.decode(charset, errors="replace")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:300]
        raise HTTPException(status_code=502, detail=f"1688 返回错误：{exc.code} {detail}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"1688 连接失败：{exc.reason}") from exc
    except TimeoutError as exc:
        raise HTTPException(status_code=504, detail="1688 请求超时") from exc


def clean_text(value: object) -> str:
    text = unescape(str(value or ""))
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def find_near_price(text: str, start: int) -> float:
    window = text[max(0, start - 300): start + 800]
    matches = re.findall(r'(?:(?:price|salePrice|discountPrice|价格|￥|¥)["\'\s:=]*)(\d+(?:\.\d+)?)', window, flags=re.I)
    for match in matches:
        price = to_float(match)
        if 0 < price < 100000:
            return price
    return 0


def parse_1688_public_search(html: str, source: str, target_count: int, min_price: float, max_price: float, blacklist: str) -> list[dict[str, object]]:
    blacklist_words = [word.strip().lower() for word in blacklist.replace("，", ",").split(",") if word.strip()]
    rows: list[dict[str, object]] = []
    seen: set[str] = set()
    link_pattern = re.compile(r'href=["\'](?P<url>(?:https?:)?//(?:detail|offer|sale)\.1688\.com/[^"\']+)["\'][^>]*>(?P<title>.*?)</a>', re.I | re.S)
    for match in link_pattern.finditer(html):
        url = match.group("url")
        if url.startswith("//"):
            url = f"https:{url}"
        url = url.split("?")[0]
        title = clean_text(match.group("title"))
        if not title or len(title) < 4 or url in seen:
            continue
        if blacklist_words and any(word in title.lower() for word in blacklist_words):
            continue
        price = find_near_price(html, match.start())
        if min_price and price and price < min_price:
            continue
        if max_price and price and price > max_price:
            continue
        rows.append({"title": title, "source": source, "url": url, "price": price, "sales": 0, "image_count": 1})
        seen.add(url)
        if len(rows) >= target_count:
            return rows
    json_title_pattern = re.compile(r'["\'](?:subject|title|shortTitle|offerTitle)["\']\s*:\s*["\'](?P<title>[^"\']{4,160})["\']', re.I)
    urls = re.findall(r'(?:https?:)?//(?:detail|offer|sale)\.1688\.com/[^"\'\\\s]+', html, flags=re.I)
    url_index = 0
    for match in json_title_pattern.finditer(html):
        title = clean_text(match.group("title"))
        while url_index < len(urls) and urls[url_index].split("?")[0] in seen:
            url_index += 1
        url = urls[url_index].split("?")[0] if url_index < len(urls) else ""
        if url.startswith("//"):
            url = f"https:{url}"
        if not title or title in seen:
            continue
        if blacklist_words and any(word in title.lower() for word in blacklist_words):
            continue
        price = find_near_price(html, match.start())
        if min_price and price and price < min_price:
            continue
        if max_price and price and price > max_price:
            continue
        rows.append({"title": title, "source": source, "url": url, "price": price, "sales": 0, "image_count": 1})
        seen.add(url or title)
        if len(rows) >= target_count:
            break
    return rows


def run_1688_public_search(payload: CollectionTaskPayload, task_id: int | None = None) -> list[dict[str, object]]:
    keyword = payload.keyword.strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="Please enter collection keyword")
    cookie_path = Path(get_setting_value("1688_cookie_path", "data/1688_cookies.json"))
    if not cookie_path.is_absolute():
        cookie_path = BASE_DIR / cookie_path
    cookie_header = load_cookie_header(cookie_path)
    target_count = max(1, min(int(payload.target_count or 1), COLLECTION_TARGET_MAX))
    min_price = max(float(payload.min_price or 0), 0)
    max_price = max(float(payload.max_price or 0), 0)
    existing_urls, existing_keys = collection_existing_dedupe_sets()
    collected: list[dict[str, object]] = []
    seen_urls: set[str] = set()
    seen_keys: set[str] = set()
    last_error = ""
    for page_no in range(1, 11):
        raise_if_collection_task_cancelled(task_id)
        query = urllib.parse.urlencode({"keywords": keyword, "beginPage": page_no}, encoding="utf-8")
        search_urls = [
            f"https://s.1688.com/selloffer/offer_search.htm?{query}",
            f"https://search.1688.com/selloffer/offer_search.htm?{query}",
        ]
        page_had_response = False
        for url in search_urls:
            raise_if_collection_task_cancelled(task_id)
            try:
                html = fetch_text_with_cookie(url, cookie_header)
                page_had_response = True
                rows = parse_1688_public_search(html, payload.source or "1688", target_count, min_price, max_price, payload.blacklist)
                for row in rows:
                    raise_if_collection_task_cancelled(task_id)
                    source_url, dedupe_key = collection_row_dedupe_key(row, payload.source or "1688")
                    if (source_url and source_url in existing_urls) or dedupe_key in existing_keys:
                        continue
                    if (source_url and source_url in seen_urls) or dedupe_key in seen_keys:
                        continue
                    if source_url:
                        seen_urls.add(source_url)
                    seen_keys.add(dedupe_key)
                    collected.append(row)
                    if len(collected) >= target_count:
                        return collected[:target_count]
                if rows:
                    break
                if "login.1688.com" in html or "_____tmd_____" in html or "x5sec" in html:
                    last_error = "Cookie expired or blocked by 1688 risk control; please refresh login cookie"
                    break
                last_error = "Page returned but no product was parsed; page structure may have changed"
            except HTTPException as exc:
                last_error = str(exc.detail)
        if len(collected) >= target_count:
            break
        if not page_had_response:
            break
        raise_if_collection_task_cancelled(task_id)
        time.sleep(0.2)
    if collected:
        return collected[:target_count]
    raise HTTPException(status_code=502, detail=f"1688 Cookie collection failed: {last_error or 'no products returned'}")

def normalize_collection_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    return [{normalize_header(key): value for key, value in row.items()} for row in rows if isinstance(row, dict)]


def insert_collection_rows_with_db(db: sqlite3.Connection, rows: list[dict[str, object]], source: str) -> tuple[int, int]:
    imported = 0
    skipped = 0
    timestamp = now_text()
    existing_urls, existing_keys = collection_existing_dedupe_sets(db)
    for row in rows:
        title = str(first_value(row, ["title", "商品标题", "标题", "name", "productname", "producttitle"], "")).strip()
        if not title:
            skipped += 1
            continue
        row_source = str(first_value(row, ["source", "来源", "platform"], source)).strip() or source
        source_url = str(first_value(row, ["url", "source_url", "sourceurl", "链接", "商品链接", "link", "producturl"], "")).strip()
        dedupe_key = f"{row_source}::{title}".lower()
        if (source_url and source_url in existing_urls) or dedupe_key in existing_keys:
            skipped += 1
            continue
        price = to_float(first_value(row, ["price", "价格", "采购价", "saleprice", "currentprice"], 0))
        sales = to_int(first_value(row, ["sales", "销量", "月销量", "sold", "orders"], 0))
        image_count = to_int(first_value(row, ["image_count", "imagecount", "图片数", "图片数量", "images"], 0))
        image_url = str(first_value(row, ["image_url", "imageurl", "pic_url", "picurl", "图片", "主图", "main_image", "mainimage", "img"], "")).strip()
        if not image_url:
            skipped += 1
            continue
        cursor = db.execute(
            """
            INSERT INTO collection_items (title, source, source_url, image_url, price, sales, image_count, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', ?)
            """,
            (title, row_source, source_url, image_url, price, sales, image_count, timestamp),
        )
        enqueue_collection_image_download(db, cursor.lastrowid, image_url)
        if source_url:
            existing_urls.add(source_url)
        existing_keys.add(dedupe_key)
        imported += 1
    if imported:
        schedule_collection_image_task_execution()
    return imported, skipped


def insert_collection_rows(rows: list[dict[str, object]], source: str) -> tuple[int, int]:
    with connect() as db:
        return insert_collection_rows_with_db(db, rows, source)


def insert_collection_candidates(candidates: list[CollectorCandidate]) -> tuple[int, int]:
    imported = 0
    skipped = 0
    timestamp = now_text()
    with connect() as db:
        existing_urls, existing_keys = collection_existing_dedupe_sets(db)
        for candidate in candidates:
            source_url, dedupe_key = collection_row_dedupe_key(candidate.model_dump(), candidate.source)
            if not candidate.image_url:
                skipped += 1
                continue
            if (source_url and source_url in existing_urls) or dedupe_key in existing_keys:
                skipped += 1
                continue
            cursor = db.execute(
                """
                INSERT INTO collection_items (title, source, source_url, image_url, price, sales, image_count, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', ?)
                """,
                (candidate.title, candidate.source, candidate.source_url, candidate.image_url, candidate.price, candidate.sales, candidate.image_count, timestamp),
            )
            enqueue_collection_image_download(db, cursor.lastrowid, candidate.image_url)
            if source_url:
                existing_urls.add(source_url)
            existing_keys.add(dedupe_key)
            imported += 1
    if imported:
        schedule_collection_image_task_execution()
    return imported, skipped


def insert_collection_candidates_with_db(db: sqlite3.Connection, candidates: list[CollectorCandidate]) -> tuple[int, int]:
    imported = 0
    skipped = 0
    timestamp = now_text()
    existing_urls, existing_keys = collection_existing_dedupe_sets(db)
    for candidate in candidates:
        source_url, dedupe_key = collection_row_dedupe_key(candidate.model_dump(), candidate.source)
        if not candidate.image_url:
            skipped += 1
            continue
        if (source_url and source_url in existing_urls) or dedupe_key in existing_keys:
            skipped += 1
            continue
        cursor = db.execute(
            """
            INSERT INTO collection_items (title, source, source_url, image_url, price, sales, image_count, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', ?)
            """,
            (candidate.title, candidate.source, candidate.source_url, candidate.image_url, candidate.price, candidate.sales, candidate.image_count, timestamp),
        )
        enqueue_collection_image_download(db, cursor.lastrowid, candidate.image_url)
        if source_url:
            existing_urls.add(source_url)
        existing_keys.add(dedupe_key)
        imported += 1
    if imported:
        schedule_collection_image_task_execution()
    return imported, skipped


def recalc(payload: ProductPayload) -> tuple[float, float, float]:
    last_mile = payload.last_mile if payload.last_mile else payload.platform_cost
    total_cost = round(payload.purchase_price + payload.first_mile + payload.warehouse_fee + last_mile, 2)
    estimated_profit = round(payload.platform_quote_price - total_cost, 2) if payload.platform_quote_price > 0 else 0
    gross_margin = round((estimated_profit / payload.platform_quote_price * 100) if payload.platform_quote_price else 0, 1)
    return total_cost, estimated_profit, gross_margin


def quote_status(profit: float, margin: float, fallback: str = "利润正常") -> str:
    if profit < 0:
        return "亏损"
    if margin == 0:
        return fallback
    if margin < 10:
        return "低利润"
    if margin < 20:
        return "利润偏低"
    return "利润正常"


def product_payload_with_freight(payload: ProductPayload) -> ProductPayload:
    if payload.weight_g <= 0:
        return payload
    data = payload.model_dump()
    if float(data.get("first_mile") or 0) <= 0:
        data["first_mile"] = calculate_first_mile(payload.weight_g)
    if float(data.get("last_mile") or 0) <= 0:
        data["last_mile"] = calculate_last_mile(payload.weight_g)
        data["platform_cost"] = data["last_mile"]
    if float(data.get("warehouse_fee") or 0) <= 0:
        data["warehouse_fee"] = default_warehouse_fee()
    return ProductPayload(**data)


def init_db() -> None:
    with connect() as db:
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                skc TEXT NOT NULL UNIQUE,
                sku_summary TEXT NOT NULL DEFAULT '',
                main_image TEXT,
                purchase_price REAL NOT NULL DEFAULT 0,
                platform_quote_price REAL NOT NULL DEFAULT 0,
                weight_g INTEGER NOT NULL DEFAULT 0,
                first_mile REAL NOT NULL DEFAULT 0,
                warehouse_fee REAL NOT NULL DEFAULT 15,
                last_mile REAL NOT NULL DEFAULT 0,
                platform_cost REAL NOT NULL DEFAULT 0,
                total_cost REAL NOT NULL DEFAULT 0,
                estimated_profit REAL NOT NULL DEFAULT 0,
                gross_margin REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT '利润正常',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS collection_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                source TEXT NOT NULL DEFAULT '1688',
                source_url TEXT NOT NULL DEFAULT '',
                image_url TEXT NOT NULL DEFAULT '',
                price REAL NOT NULL DEFAULT 0,
                sales INTEGER NOT NULL DEFAULT 0,
                image_count INTEGER NOT NULL DEFAULT 0,
                selected INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'pending',
                created_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS upload_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                total_count INTEGER NOT NULL DEFAULT 0,
                success_count INTEGER NOT NULL DEFAULT 0,
                failed_count INTEGER NOT NULL DEFAULT 0,
                export_path TEXT NOT NULL DEFAULT '',
                run_log TEXT NOT NULL DEFAULT '',
                failure_stage TEXT NOT NULL DEFAULT '',
                failure_reason TEXT NOT NULL DEFAULT '',
                evidence_path TEXT NOT NULL DEFAULT '',
                retry_count INTEGER NOT NULL DEFAULT 0,
                executor_id TEXT NOT NULL DEFAULT '',
                claimed_at TEXT NOT NULL DEFAULT '',
                heartbeat_at TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS collection_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT NOT NULL,
                source TEXT NOT NULL DEFAULT '1688',
                mode TEXT NOT NULL DEFAULT '1688',
                collector TEXT NOT NULL DEFAULT '1688_public_search',
                target_count INTEGER NOT NULL DEFAULT 10,
                min_price REAL NOT NULL DEFAULT 0,
                max_price REAL NOT NULL DEFAULT 0,
                blacklist TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'pending',
                note TEXT NOT NULL DEFAULT '',
                request_path TEXT NOT NULL DEFAULT '',
                result_count INTEGER NOT NULL DEFAULT 0,
                parent_task_id INTEGER NOT NULL DEFAULT 0,
                batch_index INTEGER NOT NULL DEFAULT 1,
                batch_total INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS collection_task_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER NOT NULL,
                stage TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'info',
                message TEXT NOT NULL DEFAULT '',
                page_no INTEGER NOT NULL DEFAULT 0,
                parsed_count INTEGER NOT NULL DEFAULT 0,
                imported_count INTEGER NOT NULL DEFAULT 0,
                skipped_count INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                FOREIGN KEY(task_id) REFERENCES collection_tasks(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS collection_import_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                status TEXT NOT NULL DEFAULT 'queued',
                ids_json TEXT NOT NULL DEFAULT '[]',
                total_count INTEGER NOT NULL DEFAULT 0,
                processed_count INTEGER NOT NULL DEFAULT 0,
                imported_count INTEGER NOT NULL DEFAULT 0,
                skipped_count INTEGER NOT NULL DEFAULT 0,
                failed_count INTEGER NOT NULL DEFAULT 0,
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS collection_image_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                collection_item_id INTEGER NOT NULL,
                source_url TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'queued',
                local_path TEXT NOT NULL DEFAULT '',
                attempts INTEGER NOT NULL DEFAULT 0,
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(collection_item_id) REFERENCES collection_items(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS ai_title_cache (
                cache_key TEXT PRIMARY KEY,
                source_title TEXT NOT NULL DEFAULT '',
                chinese_title TEXT NOT NULL DEFAULT '',
                english_title TEXT NOT NULL DEFAULT '',
                provider TEXT NOT NULL DEFAULT 'deepseek',
                model TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS processing_title_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                status TEXT NOT NULL DEFAULT 'queued',
                ids_json TEXT NOT NULL DEFAULT '[]',
                failed_ids_json TEXT NOT NULL DEFAULT '[]',
                total_count INTEGER NOT NULL DEFAULT 0,
                processed_count INTEGER NOT NULL DEFAULT 0,
                success_count INTEGER NOT NULL DEFAULT 0,
                failed_count INTEGER NOT NULL DEFAULT 0,
                cache_hit_count INTEGER NOT NULL DEFAULT 0,
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS processing_exceptions (
                product_id INTEGER PRIMARY KEY,
                skc TEXT NOT NULL DEFAULT '',
                title TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'open',
                issues_json TEXT NOT NULL DEFAULT '[]',
                warnings_json TEXT NOT NULL DEFAULT '[]',
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS processing_field_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                status TEXT NOT NULL DEFAULT 'queued',
                ids_json TEXT NOT NULL DEFAULT '[]',
                failed_ids_json TEXT NOT NULL DEFAULT '[]',
                fields_json TEXT NOT NULL DEFAULT '{}',
                start_skc TEXT NOT NULL DEFAULT '',
                total_count INTEGER NOT NULL DEFAULT 0,
                processed_count INTEGER NOT NULL DEFAULT 0,
                success_count INTEGER NOT NULL DEFAULT 0,
                failed_count INTEGER NOT NULL DEFAULT 0,
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS processing_overrides (
                product_id INTEGER PRIMARY KEY,
                english_title TEXT NOT NULL DEFAULT '',
                color TEXT NOT NULL DEFAULT '',
                size TEXT NOT NULL DEFAULT '',
                sku_code TEXT NOT NULL DEFAULT '',
                declared_price REAL NOT NULL DEFAULT 0,
                weight_g INTEGER NOT NULL DEFAULT 350,
                length_cm INTEGER NOT NULL DEFAULT 15,
                width_cm INTEGER NOT NULL DEFAULT 10,
                height_cm INTEGER NOT NULL DEFAULT 2,
                source_url TEXT NOT NULL DEFAULT '',
                stock INTEGER NOT NULL DEFAULT 999,
                ship_days INTEGER NOT NULL DEFAULT 9,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS publish_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                result TEXT NOT NULL,
                skc TEXT NOT NULL,
                title TEXT NOT NULL,
                reason TEXT NOT NULL DEFAULT '',
                screenshot TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS api_configs (
                key TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 1,
                base_url TEXT NOT NULL DEFAULT '',
                model TEXT NOT NULL DEFAULT '',
                api_key TEXT NOT NULL DEFAULT '',
                usage TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS image_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                provider TEXT NOT NULL DEFAULT 'nanobanana',
                task_id TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'submitted',
                source_image TEXT NOT NULL DEFAULT '',
                result_url TEXT NOT NULL DEFAULT '',
                local_path TEXT NOT NULL DEFAULT '',
                prompt TEXT NOT NULL DEFAULT '',
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS product_detail_snapshots (
                product_id INTEGER PRIMARY KEY,
                provider TEXT NOT NULL DEFAULT '',
                source_url TEXT NOT NULL DEFAULT '',
                raw_json TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT '',
                note TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS product_sku_images (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                sku_id TEXT NOT NULL DEFAULT '',
                color TEXT NOT NULL DEFAULT '',
                size TEXT NOT NULL DEFAULT '',
                image_url TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                source TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS product_color_image_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                color TEXT NOT NULL DEFAULT '',
                image_url TEXT NOT NULL DEFAULT '',
                source TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                is_auto INTEGER NOT NULL DEFAULT 1,
                confidence REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS product_detail_image_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                image_url TEXT NOT NULL DEFAULT '',
                source TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS product_image_vision_cache (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                image_identity TEXT NOT NULL,
                image_url TEXT NOT NULL DEFAULT '',
                matched_color TEXT NOT NULL DEFAULT '',
                confidence REAL NOT NULL DEFAULT 0,
                is_product_image INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT '',
                provider TEXT NOT NULL DEFAULT '',
                raw_json TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(product_id, image_identity),
                FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS spec_aliases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kind TEXT NOT NULL,
                alias TEXT NOT NULL,
                target TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(kind, alias)
            )
            """
        )
        db.execute(
            """
            CREATE TABLE IF NOT EXISTS prompt_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                category TEXT NOT NULL DEFAULT '',
                prompt_type TEXT NOT NULL DEFAULT '',
                usage TEXT NOT NULL DEFAULT '',
                content TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT '启用中',
                updated_at TEXT NOT NULL
            )
            """
        )
        title_task_columns = [row[1] for row in db.execute("PRAGMA table_info(processing_title_tasks)").fetchall()]
        if "failed_ids_json" not in title_task_columns:
            db.execute("ALTER TABLE processing_title_tasks ADD COLUMN failed_ids_json TEXT NOT NULL DEFAULT '[]'")
        field_task_columns = [row[1] for row in db.execute("PRAGMA table_info(processing_field_tasks)").fetchall()]
        if "failed_ids_json" not in field_task_columns:
            db.execute("ALTER TABLE processing_field_tasks ADD COLUMN failed_ids_json TEXT NOT NULL DEFAULT '[]'")
        upload_columns = [row[1] for row in db.execute("PRAGMA table_info(upload_tasks)").fetchall()]
        if "run_log" not in upload_columns:
            db.execute("ALTER TABLE upload_tasks ADD COLUMN run_log TEXT NOT NULL DEFAULT ''")
        if "failure_stage" not in upload_columns:
            db.execute("ALTER TABLE upload_tasks ADD COLUMN failure_stage TEXT NOT NULL DEFAULT ''")
        if "failure_reason" not in upload_columns:
            db.execute("ALTER TABLE upload_tasks ADD COLUMN failure_reason TEXT NOT NULL DEFAULT ''")
        if "evidence_path" not in upload_columns:
            db.execute("ALTER TABLE upload_tasks ADD COLUMN evidence_path TEXT NOT NULL DEFAULT ''")
        if "retry_count" not in upload_columns:
            db.execute("ALTER TABLE upload_tasks ADD COLUMN retry_count INTEGER NOT NULL DEFAULT 0")
        if "executor_id" not in upload_columns:
            db.execute("ALTER TABLE upload_tasks ADD COLUMN executor_id TEXT NOT NULL DEFAULT ''")
        if "claimed_at" not in upload_columns:
            db.execute("ALTER TABLE upload_tasks ADD COLUMN claimed_at TEXT NOT NULL DEFAULT ''")
        if "heartbeat_at" not in upload_columns:
            db.execute("ALTER TABLE upload_tasks ADD COLUMN heartbeat_at TEXT NOT NULL DEFAULT ''")
        collection_columns = [row[1] for row in db.execute("PRAGMA table_info(collection_items)").fetchall()]
        if "image_url" not in collection_columns:
            db.execute("ALTER TABLE collection_items ADD COLUMN image_url TEXT NOT NULL DEFAULT ''")
        db.execute("CREATE INDEX IF NOT EXISTS idx_collection_items_status_id ON collection_items(status, id DESC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_collection_items_source_id ON collection_items(source, id DESC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_collection_items_price_id ON collection_items(price, id DESC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_collection_items_source_url ON collection_items(source_url)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_collection_items_created_id ON collection_items(created_at, id DESC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_collection_task_events_task_id ON collection_task_events(task_id, id DESC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_collection_import_tasks_status_id ON collection_import_tasks(status, id ASC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_collection_image_tasks_status_id ON collection_image_tasks(status, id ASC)")
        db.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_collection_image_tasks_item_url ON collection_image_tasks(collection_item_id, source_url)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_processing_title_tasks_status_id ON processing_title_tasks(status, id ASC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_processing_exceptions_status ON processing_exceptions(status, updated_at DESC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_processing_field_tasks_status_id ON processing_field_tasks(status, id ASC)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_product_sku_images_product_order ON product_sku_images(product_id, sort_order, id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_product_sku_images_product_url ON product_sku_images(product_id, image_url)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_product_color_assignments_product_order ON product_color_image_assignments(product_id, color, sort_order, id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_product_color_assignments_product_url ON product_color_image_assignments(product_id, image_url)")
        product_columns = [row[1] for row in db.execute("PRAGMA table_info(products)").fetchall()]
        for column_name, ddl in [
            ("platform_quote_price", "ALTER TABLE products ADD COLUMN platform_quote_price REAL NOT NULL DEFAULT 0"),
            ("weight_g", "ALTER TABLE products ADD COLUMN weight_g INTEGER NOT NULL DEFAULT 0"),
            ("warehouse_fee", "ALTER TABLE products ADD COLUMN warehouse_fee REAL NOT NULL DEFAULT 15"),
            ("last_mile", "ALTER TABLE products ADD COLUMN last_mile REAL NOT NULL DEFAULT 0"),
            ("processing_hidden", "ALTER TABLE products ADD COLUMN processing_hidden INTEGER NOT NULL DEFAULT 0"),
        ]:
            if column_name not in product_columns:
                db.execute(ddl)
        collection_task_columns = [row[1] for row in db.execute("PRAGMA table_info(collection_tasks)").fetchall()]
        for column_name, ddl in [
            ("mode", "ALTER TABLE collection_tasks ADD COLUMN mode TEXT NOT NULL DEFAULT '1688'"),
            ("collector", "ALTER TABLE collection_tasks ADD COLUMN collector TEXT NOT NULL DEFAULT '1688_public_search'"),
            ("note", "ALTER TABLE collection_tasks ADD COLUMN note TEXT NOT NULL DEFAULT ''"),
            ("request_path", "ALTER TABLE collection_tasks ADD COLUMN request_path TEXT NOT NULL DEFAULT ''"),
            ("parent_task_id", "ALTER TABLE collection_tasks ADD COLUMN parent_task_id INTEGER NOT NULL DEFAULT 0"),
            ("batch_index", "ALTER TABLE collection_tasks ADD COLUMN batch_index INTEGER NOT NULL DEFAULT 1"),
            ("batch_total", "ALTER TABLE collection_tasks ADD COLUMN batch_total INTEGER NOT NULL DEFAULT 1"),
        ]:
            if column_name not in collection_task_columns:
                db.execute(ddl)
        api_count = db.execute("SELECT COUNT(*) AS total FROM api_configs").fetchone()["total"]
        if api_count == 0:
            sync_builtin_api_configs(db, force_keys={"deepseek", "image", "vision"})
        else:
            sync_builtin_api_configs(db)
            deepseek_api = db.execute("SELECT * FROM api_configs WHERE key = 'deepseek'").fetchone()
            if deepseek_api and (looks_corrupted_text(deepseek_api["name"]) or looks_corrupted_text(deepseek_api["usage"])):
                db.execute(
                    "UPDATE api_configs SET name = ?, usage = ?, updated_at = ? WHERE key = 'deepseek'",
                    ("DeepSeek 文案生成", "标题生成 / 文案优化", now_text()),
                )
            image_api = db.execute("SELECT * FROM api_configs WHERE key = 'image'").fetchone()
            if image_api and not image_api["base_url"] and image_api["model"] == "image-to-image":
                db.execute(
                    "UPDATE api_configs SET name = ?, base_url = ?, model = ?, usage = ?, updated_at = ? WHERE key = 'image'",
                    ("Nano Banana 图生图", NANOBANANA_SUBMIT_URL, "nano-banana", "图片处理 / 图生图", now_text()),
                )
        setting_count = db.execute("SELECT COUNT(*) AS total FROM app_settings").fetchone()["total"]
        if setting_count == 0:
            timestamp = now_text()
            db.executemany(
                "INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)",
                [
                    ("database", "data/app.db", timestamp),
                    ("image_dir", "data/images", timestamp),
                    ("run_mode", "headless", timestamp),
                    ("max_retries", "2", timestamp),
                    ("cookie_path", "data/dxm_cookies.json", timestamp),
                    ("script_dir", str(SCRIPT_SOURCE_DIR), timestamp),
                    ("enable_real_rpa", "false", timestamp),
                    ("collection_mode", "1688", timestamp),
                    ("enable_external_collection", "false", timestamp),
                    ("onebound_key", "", timestamp),
                    ("onebound_secret", "", timestamp),
                    ("1688_cookie", "", timestamp),
                    ("1688_cookie_path", "data/1688_cookies.json", timestamp),
                    ("miaoshou_default_suggested_price_usd", "60", timestamp),
                    ("upload_image_source", "", timestamp),
                    ("cos_region", "", timestamp),
                    ("cos_bucket", "", timestamp),
                    ("cos_prefix", "", timestamp),
                    ("executor_mode", "local_python", timestamp),
                    ("executor_server_url", "http://127.0.0.1:8000", timestamp),
                    ("executor_bind_code", "", timestamp),
                    ("executor_download_url", "", timestamp),
                    ("executor_version", "0.1.0", timestamp),
                    ("executor_poll_seconds", "5", timestamp),
                    ("executor_heartbeat_timeout", "60", timestamp),
                    ("executor_task_scope", "manual", timestamp),
                    ("upload_fill_skc", "true", timestamp),
                    ("upload_skc_missing_policy", "pause", timestamp),
                    ("upload_auto_submit", "false", timestamp),
                    ("upload_error_policy", "skip", timestamp),
                    ("upload_save_screenshots", "false", timestamp),
                    ("upload_save_html", "false", timestamp),
                    ("upload_trace", "off", timestamp),
                    ("upload_step_delay_ms", "500", timestamp),
                    ("temu_shop_account", "", timestamp),
                    ("temu_site", "美国站", timestamp),
                    ("temu_product_template", "", timestamp),
                    ("temu_size_category", "", timestamp),
                    ("temu_size_template", "", timestamp),
                    ("temu_warehouse_template", "", timestamp),
                    ("temu_logistics_template", "", timestamp),
                    ("temu_ship_days", "9", timestamp),
                    ("temu_declare_markup", "239.0", timestamp),
                    ("temu_default_weight_g", "350", timestamp),
                    ("temu_default_stock", "999", timestamp),
                    ("temu_package_length_cm", "10", timestamp),
                    ("temu_package_width_cm", "5", timestamp),
                    ("temu_package_height_cm", "1", timestamp),
                    ("temu_1688_excel_path", "", timestamp),
                    ("temu_batch_limit", "", timestamp),
                    ("temu_start_skc", "", timestamp),
                    ("temu_append_sku_suffix", "true", timestamp),
                    ("temu_add_model_info", "false", timestamp),
                    ("temu_model_index", "2", timestamp),
                    ("default_processing_declared_price_mode", "add", timestamp),
                    ("default_processing_declared_price", "0", timestamp),
                    ("default_processing_weight_g", "350", timestamp),
                    ("default_processing_length_cm", "15", timestamp),
                    ("default_processing_width_cm", "10", timestamp),
                    ("default_processing_height_cm", "2", timestamp),
                    ("default_processing_stock", "999", timestamp),
                    ("default_processing_ship_days", "9", timestamp),
                    ("default_processing_start_skc", "", timestamp),
                ],
            )
            sync_builtin_settings(db, force_keys={"onebound_key", "onebound_secret", "enable_external_collection"})
        else:
            timestamp = now_text()
            defaults = {
                "collection_mode": "1688",
                "enable_external_collection": "false",
                "onebound_key": "",
                "onebound_secret": "",
                "1688_cookie": "",
                "1688_cookie_path": "data/1688_cookies.json",
                "script_dir": str(SCRIPT_SOURCE_DIR),
                "miaoshou_default_suggested_price_usd": "60",
                "upload_image_source": "",
                "cos_region": "",
                "cos_bucket": "",
                "cos_prefix": "",
                "executor_mode": "local_python",
                "executor_server_url": "http://127.0.0.1:8000",
                "executor_bind_code": "",
                "executor_download_url": "",
                "executor_version": "0.1.0",
                "executor_poll_seconds": "5",
                "executor_heartbeat_timeout": "60",
                "executor_task_scope": "manual",
                "upload_fill_skc": "true",
                "upload_skc_missing_policy": "pause",
                "upload_auto_submit": "false",
                "upload_error_policy": "skip",
                "upload_save_screenshots": "false",
                "upload_save_html": "false",
                "upload_trace": "off",
                "upload_step_delay_ms": "500",
                "temu_shop_account": "",
                "temu_site": "美国站",
                "temu_product_template": "",
                "temu_size_category": "",
                "temu_size_template": "",
                "temu_warehouse_template": "",
                "temu_logistics_template": "",
                "temu_ship_days": "9",
                "temu_declare_markup": "239.0",
                "temu_default_weight_g": "350",
                "temu_default_stock": "999",
                "temu_package_length_cm": "10",
                "temu_package_width_cm": "5",
                "temu_package_height_cm": "1",
                "temu_1688_excel_path": "",
                "temu_batch_limit": "",
                "temu_start_skc": "",
                "temu_append_sku_suffix": "true",
                "temu_add_model_info": "false",
                "temu_model_index": "2",
                "default_processing_declared_price_mode": "add",
                "default_processing_declared_price": "0",
                "default_processing_weight_g": "350",
                "default_processing_length_cm": "15",
                "default_processing_width_cm": "10",
                "default_processing_height_cm": "2",
                "default_processing_stock": "999",
                "default_processing_ship_days": "9",
                "default_processing_start_skc": "",
            }
            for key, value in defaults.items():
                exists = db.execute("SELECT 1 FROM app_settings WHERE key = ?", (key,)).fetchone()
                if not exists:
                    db.execute("INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)", (key, value, timestamp))
            sync_builtin_settings(db)
        prompt_count = db.execute("SELECT COUNT(*) AS total FROM prompt_templates").fetchone()["total"]
        corrupt_prompt_count = db.execute(
            """
            SELECT COUNT(*) AS total FROM prompt_templates
            WHERE name LIKE '%??%' OR category LIKE '%??%' OR prompt_type LIKE '%??%' OR usage LIKE '%??%' OR content LIKE '%??%' OR status LIKE '%??%'
            """
        ).fetchone()["total"]
        if prompt_count and corrupt_prompt_count:
            db.execute("DELETE FROM prompt_templates")
            prompt_count = 0
        if prompt_count == 0:
            timestamp = now_text()
            db.executemany(
                "INSERT INTO prompt_templates (name, category, prompt_type, usage, content, status, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                [(*template, timestamp) for template in DEFAULT_PROMPT_TEMPLATES],
            )


@app.on_event("startup")
def startup() -> None:
    init_db()
    resume_collection_task_queue()
    resume_collection_import_task_queue()
    resume_collection_image_task_queue()
    resume_processing_title_task_queue()
    resume_processing_field_task_queue()


@app.get("/")
def index() -> FileResponse:
    return FileResponse(FRONTEND_DIR / "index.html")


@app.get("/api/health")
def health() -> dict[str, object]:
    return {"ok": True, "app": "upload-assistant", "data_dir": str(DATA_DIR), "time": now_text()}


def spec_alias_from_row(row: sqlite3.Row) -> SpecAliasItem:
    return SpecAliasItem(
        id=int(row["id"]),
        kind=row["kind"],
        alias=row["alias"],
        target=row["target"],
        updated_at=row["updated_at"] or "",
    )


@app.get("/api/spec-mappings/colors")
def spec_mapping_colors() -> dict[str, object]:
    custom_aliases = load_spec_alias_map("color")
    return {
        "groups": [{"name": name, "colors": colors} for name, colors in PLATFORM_COLOR_GROUPS],
        "standard_colors": PLATFORM_STANDARD_COLORS,
        "alias_count": len(COLOR_ALIAS_MAP) + len(custom_aliases),
        "custom_aliases": custom_aliases,
    }


@app.get("/api/spec-mappings/sizes")
def spec_mapping_sizes() -> dict[str, object]:
    custom_aliases = load_spec_alias_map("size")
    return {
        "name": "女士牛仔短裤",
        "sizes": STANDARD_UPLOAD_SIZES,
        "alias_map": size_alias_map(),
        "custom_aliases": custom_aliases,
    }


def extract_spec_exception_values(lines: list[str]) -> list[tuple[str, str]]:
    values: list[tuple[str, str]] = []
    for line in lines:
        text = str(line or "")
        if text.startswith("颜色无法匹配标准色："):
            raw_values = text.split("：", 1)[1]
            for value in re.split(r"[、,，/]+", raw_values):
                clean_value = value.strip()
                if clean_value:
                    values.append(("color", clean_value))
        elif text.startswith("不支持尺码："):
            raw_values = text.split("：", 1)[1]
            for value in re.split(r"[、,，/]+", raw_values):
                clean_value = value.strip().upper()
                if clean_value:
                    values.append(("size", clean_value))
    return values


@app.get("/api/spec-mappings/exceptions", response_model=list[SpecExceptionSummaryItem])
def list_spec_exception_summaries() -> list[SpecExceptionSummaryItem]:
    summary: dict[tuple[str, str], set[int]] = {}
    with connect() as db:
        rows = db.execute("SELECT product_id, issues_json, warnings_json FROM processing_exceptions WHERE status = 'open'").fetchall()
    for row in rows:
        product_id = int(row["product_id"])
        lines = json.loads(row["issues_json"] or "[]") + json.loads(row["warnings_json"] or "[]")
        for kind, value in extract_spec_exception_values(lines):
            summary.setdefault((kind, value), set()).add(product_id)
    items = [
        SpecExceptionSummaryItem(kind=kind, value=value, count=len(product_ids), product_ids=sorted(product_ids))
        for (kind, value), product_ids in summary.items()
    ]
    return sorted(items, key=lambda item: (-item.count, item.kind, item.value))


@app.get("/api/spec-mappings/aliases", response_model=list[SpecAliasItem])
def list_spec_aliases(kind: str = "") -> list[SpecAliasItem]:
    clauses: list[str] = []
    params: list[object] = []
    if kind:
        clauses.append("kind = ?")
        params.append(kind)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as db:
        rows = db.execute(f"SELECT * FROM spec_aliases {where_sql} ORDER BY updated_at DESC, id DESC", params).fetchall()
    return [spec_alias_from_row(row) for row in rows]


@app.post("/api/spec-mappings/aliases", response_model=SpecAliasItem)
def save_spec_alias(payload: SpecAliasPayload) -> SpecAliasItem:
    kind = payload.kind.strip().lower()
    alias = payload.alias.strip()
    target = payload.target.strip()
    if kind not in {"color", "size"}:
        raise HTTPException(status_code=400, detail="规格类型必须是 color 或 size")
    if not alias:
        raise HTTPException(status_code=400, detail="供应商写法不能为空")
    if kind == "color" and target not in PLATFORM_STANDARD_COLORS:
        raise HTTPException(status_code=400, detail="目标颜色不在标准色表中")
    if kind == "size" and target.upper() not in RPA_ALLOWED_UPLOAD_SIZES:
        raise HTTPException(status_code=400, detail="目标尺码不在标准尺码中")
    if kind == "size":
        alias = alias.upper()
        target = target.upper()
    timestamp = now_text()
    with connect() as db:
        db.execute(
            """
            INSERT INTO spec_aliases (kind, alias, target, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(kind, alias) DO UPDATE SET target = excluded.target, updated_at = excluded.updated_at
            """,
            (kind, alias, target, timestamp),
        )
        row = db.execute("SELECT * FROM spec_aliases WHERE kind = ? AND alias = ?", (kind, alias)).fetchone()
    return spec_alias_from_row(row)


@app.delete("/api/spec-mappings/aliases/{alias_id}")
def delete_spec_alias(alias_id: int) -> dict[str, object]:
    with connect() as db:
        row = db.execute("SELECT * FROM spec_aliases WHERE id = ?", (alias_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="规格别名不存在")
        db.execute("DELETE FROM spec_aliases WHERE id = ?", (alias_id,))
    return {"deleted": True, "id": alias_id}


@app.get("/api/image-proxy")
def image_proxy(url: str) -> Response:
    parsed = urllib.parse.urlparse(url)
    allowed_hosts = {"cbu01.alicdn.com", "cbu02.alicdn.com", "cbu03.alicdn.com", "img.alicdn.com"}
    if parsed.scheme not in {"http", "https"} or parsed.netloc not in allowed_hosts:
        raise HTTPException(status_code=400, detail="不支持的图片域名")
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            "Referer": "https://detail.1688.com/",
        },
        method="GET",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            content = response.read()
            content_type = response.headers.get("Content-Type", "image/jpeg")
    except urllib.error.HTTPError as exc:
        raise HTTPException(status_code=exc.code, detail="图片代理请求失败") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"图片代理连接失败：{exc.reason}") from exc
    return Response(content=content, media_type=content_type, headers={"Cache-Control": "public, max-age=86400"})


def fetch_external_image(url: str) -> tuple[bytes, str]:
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        raise ValueError("unsupported image url")
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            "Referer": "https://detail.1688.com/",
        },
        method="GET",
    )
    with urllib.request.urlopen(request, timeout=25) as response:
        content_type = response.headers.get("Content-Type", "image/jpeg")
        return response.read(), content_type


def save_collection_image_from_url(collection_item_id: int, image_url: str) -> str:
    clean_url = str(image_url or "").strip()
    if not clean_url:
        return ""
    if clean_url.startswith("/images/"):
        return clean_url
    try:
        content, content_type = fetch_external_image(clean_url)
    except (urllib.error.URLError, TimeoutError, ValueError):
        return ""
    extension = image_extension_from_content_type(content_type, clean_url)
    target_dir = IMAGE_DIR / "collection" / str(collection_item_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"main{extension}"
    for old_file in target_dir.glob("main.*"):
        old_file.unlink(missing_ok=True)
    target.write_bytes(content)
    return f"/images/collection/{collection_item_id}/{target.name}"


def copy_local_image_to_product(product_id: int, image_url: str, name: str = "main") -> str:
    source = local_image_path_from_url(image_url)
    if not source:
        return ""
    product_dir = IMAGE_DIR / str(product_id)
    product_dir.mkdir(parents=True, exist_ok=True)
    suffix = source.suffix.lower() if source.suffix.lower() in RPA_IMAGE_EXTENSIONS else ".jpg"
    target = product_dir / f"{name}{suffix}"
    for old_file in product_dir.glob(f"{name}.*"):
        old_file.unlink(missing_ok=True)
    shutil.copyfile(source, target)
    return f"/images/{product_id}/{target.name}"


def save_product_main_image_from_url(product_id: int, image_url: str) -> str:
    if not image_url:
        return ""
    if str(image_url).strip().startswith("/images/"):
        return copy_local_image_to_product(product_id, image_url)
    try:
        content, content_type = fetch_external_image(image_url)
    except (urllib.error.URLError, TimeoutError, ValueError):
        return ""
    extension = ".jpg"
    if "png" in content_type:
        extension = ".png"
    elif "webp" in content_type:
        extension = ".webp"
    product_dir = IMAGE_DIR / str(product_id)
    product_dir.mkdir(parents=True, exist_ok=True)
    target = product_dir / f"main{extension}"
    for old_file in product_dir.glob("main.*"):
        old_file.unlink(missing_ok=True)
    target.write_bytes(content)
    return f"/images/{product_id}/{target.name}"


def image_extension_from_content_type(content_type: str, fallback_url: str = "") -> str:
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    if "gif" in content_type:
        return ".gif"
    suffix = Path(urllib.parse.urlparse(fallback_url).path).suffix.lower()
    return suffix if suffix in {".jpg", ".jpeg", ".png", ".webp", ".gif"} else ".jpg"


def save_processed_image_from_url(product_id: int, image_url: str) -> str:
    content, content_type = fetch_external_image(image_url)
    extension = image_extension_from_content_type(content_type, image_url)
    product_dir = IMAGE_DIR / str(product_id)
    product_dir.mkdir(parents=True, exist_ok=True)
    target = product_dir / f"processed{extension}"
    for old_file in product_dir.glob("processed.*"):
        old_file.unlink(missing_ok=True)
    target.write_bytes(content)
    return f"/images/{product_id}/{target.name}"


def save_processed_image_from_local(product_id: int, image_url: str) -> str:
    return copy_local_image_to_product(product_id, image_url, name="processed")


def image_task_from_row(row: sqlite3.Row) -> ImageTask:
    return ImageTask(**dict(row))


def image_preview_url(image_url: str) -> str:
    clean_url = str(image_url or "").strip()
    if clean_url.startswith("http"):
        return f"/api/image-proxy?url={urllib.parse.quote(clean_url, safe='')}"
    return clean_url


def add_image_option(options: list[ImageOption], seen: set[str], label: str, url: str, kind: str, color: str = "", size: str = "", sku_id: str = "") -> None:
    clean_url = str(url or "").strip()
    dedupe_key = "::".join([kind, clean_url, str(color or ""), str(size or "")])
    if not clean_url or dedupe_key in seen:
        return
    if clean_url.startswith("http") and not re.search(r"\.(jpg|jpeg|png|webp|gif)(\?|$)", clean_url, re.I):
        return
    seen.add(dedupe_key)
    options.append(ImageOption(label=label, url=clean_url, kind=kind, preview_url=image_preview_url(clean_url), color=color, size=size, sku_id=sku_id))


def build_product_image_options(product_id: int, main_image: str = "", source_url: str = "") -> list[ImageOption]:
    options: list[ImageOption] = []
    seen: set[str] = set()
    add_image_option(options, seen, "商品主图", main_image, "main_image")
    product_dir = IMAGE_DIR / str(product_id)
    if product_dir.exists():
        for image_path in sorted(product_dir.iterdir(), key=lambda item: item.stat().st_mtime, reverse=True):
            if image_path.is_file() and image_path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
                add_image_option(options, seen, f"本地图片：{image_path.name}", f"/images/{product_id}/{image_path.name}", "local_file")
    if source_url.startswith("http"):
        add_image_option(options, seen, "处理字段站外链接", source_url, "source_url")
    with connect() as db:
        sku_images = db.execute(
            """
            SELECT sku_id, color, size, image_url, source FROM product_sku_images
            WHERE product_id = ? AND image_url != ''
            ORDER BY sort_order, id
            """,
            (product_id,),
        ).fetchall()
        for row in sku_images:
            parts = [part for part in [row["color"], row["size"]] if part]
            label = "详情图" + (f"：{' / '.join(parts)}" if parts else "")
            add_image_option(options, seen, label, row["image_url"], row["source"] or "detail_image", row["color"], row["size"], row["sku_id"])
        product = db.execute("SELECT title FROM products WHERE id = ?", (product_id,)).fetchone()
        if product:
            rows = db.execute(
                """
                SELECT source, image_url FROM collection_items
                WHERE image_url != '' AND (source_url = ? OR title = ?)
                ORDER BY id DESC LIMIT 12
                """,
                (source_url, product["title"]),
            ).fetchall()
            for row in rows:
                add_image_option(options, seen, f"采集图：{row['source']}", row["image_url"], "collection_image")
    return options


def color_image_assignment_from_row(row: sqlite3.Row) -> ColorImageAssignment:
    return ColorImageAssignment(
        color=row["color"],
        image_url=row["image_url"],
        preview_url=image_preview_url(row["image_url"]),
        source=row["source"],
        sort_order=int(row["sort_order"]),
        is_auto=bool(row["is_auto"]),
        confidence=float(row["confidence"] or 0),
    )


def list_color_image_assignments(product_id: int) -> list[ColorImageAssignment]:
    with connect() as db:
        rows = db.execute(
            """
            SELECT * FROM product_color_image_assignments
            WHERE product_id = ?
            ORDER BY color, sort_order, id
            """,
            (product_id,),
        ).fetchall()
    return [color_image_assignment_from_row(row) for row in rows]


def detail_image_assignment_from_row(row: sqlite3.Row) -> DetailImageAssignment:
    image_url = row["image_url"] or ""
    return DetailImageAssignment(
        image_url=image_url,
        preview_url=image_preview_url(image_url) if image_url else "",
        source=row["source"],
        sort_order=int(row["sort_order"]),
    )


def list_detail_image_assignments(product_id: int) -> list[DetailImageAssignment]:
    with connect() as db:
        rows = db.execute(
            """
            SELECT * FROM product_detail_image_assignments
            WHERE product_id = ?
            ORDER BY sort_order, id
            """,
            (product_id,),
        ).fetchall()
    return [detail_image_assignment_from_row(row) for row in rows]


def image_identity_server(url: str) -> str:
    clean_url = str(url or "").strip()
    if not clean_url:
        return ""
    parsed = urllib.parse.urlsplit(clean_url)
    if parsed.scheme and parsed.netloc:
        return urllib.parse.urlunsplit((parsed.scheme, parsed.netloc, parsed.path, "", ""))
    return clean_url.split("?", 1)[0]


def normalize_color_alias(value: object) -> str:
    return re.sub(r"[\s\-_\[\](){}:\u003a\uff1a,\uff0c.\u3002/\\|;\uff1b\u3001]+", "", str(value or "").strip().lower())


def load_spec_alias_map(kind: str) -> dict[str, str]:
    try:
        with connect() as db:
            rows = db.execute("SELECT alias, target FROM spec_aliases WHERE kind = ?", (kind,)).fetchall()
        return {normalize_color_alias(row["alias"]): row["target"] for row in rows if row["alias"] and row["target"]}
    except sqlite3.Error:
        return {}


COLOR_ALIAS_MAP = {normalize_color_alias(alias): color for alias, color in COLOR_ALIAS_ENTRIES}


def color_alias_map() -> dict[str, str]:
    return {**COLOR_ALIAS_MAP, **load_spec_alias_map("color")}


def size_alias_map() -> dict[str, str]:
    custom = {key.upper(): value.upper() for key, value in load_spec_alias_map("size").items()}
    return {**SIZE_ALIAS_MAP, **custom}


def resolve_known_color(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = normalize_color_alias(text)
    aliases = color_alias_map()
    if normalized in aliases:
        return aliases[normalized]
    for part in re.split(r"[\r\n,，、/]+", text):
        color = aliases.get(normalize_color_alias(part))
        if color:
            return color
    for alias, color in aliases.items():
        if alias and alias in normalized:
            return color
    return ""


def clean_color_label(value: str) -> str:
    text = str(value or "").strip()
    match = re.search(r"(?:\u989c\u8272)[:\uff1a]([^;/\\|]+)", text)
    if match:
        text = match.group(1)
    text = text.replace("\uff08", "(").replace("\uff09", ")").strip()
    return resolve_known_color(text)


def safe_path_segment(value: str, fallback: str = "item") -> str:
    text = str(value or "").strip() or fallback
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", text)
    text = re.sub(r"\s+", "_", text).strip(" ._")
    return text[:80] or fallback


COLOR_ENGLISH_MAP: dict[str, str] = {}
for alias, color in COLOR_ALIAS_ENTRIES:
    if re.fullmatch(r"[A-Za-z][A-Za-z\s_-]*", str(alias or "")) and color not in COLOR_ENGLISH_MAP:
        COLOR_ENGLISH_MAP[color] = safe_path_segment(str(alias), color).lower()


def color_english_label(value: str) -> str:
    color = clean_color_label(value) or str(value or "").strip()
    if not color:
        return "COLOR"
    return COLOR_ENGLISH_MAP.get(color, safe_path_segment(color, "COLOR")).upper()


def contains_cjk(value: str) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", str(value or "")))


def english_title_fallback(source_title: str) -> str:
    text = str(source_title or "").lower()
    parts: list[str] = []
    if any(keyword in text for keyword in ["牛仔", "denim", "jean"]):
        parts.append("Denim")
    if any(keyword in text for keyword in ["短裤", "short"]):
        parts.append("Shorts")
    elif any(keyword in text for keyword in ["裤", "pant"]):
        parts.append("Pants")
    else:
        parts.append("Fashion Item")
    if any(keyword in text for keyword in ["女", "女士", "women", "ladies"]):
        parts.insert(0, "Women's")
    if any(keyword in text for keyword in ["高腰", "high waist", "high-waist"]):
        parts.append("High Waist")
    if any(keyword in text for keyword in ["宽松", "loose"]):
        parts.append("Loose Fit")
    if any(keyword in text for keyword in ["夏", "summer"]):
        parts.append("Summer")
    title = " ".join(dict.fromkeys(part for part in parts if part)).strip()
    if len(title) < 18:
        title = f"{title} Casual Everyday Bottoms"
    return title[:120]


def valid_english_title(value: str) -> bool:
    text = str(value or "").strip()
    return bool(text) and not contains_cjk(text)


def local_image_path_from_url(image_url: str) -> Path | None:
    clean_url = str(image_url or "").strip()
    if not clean_url.startswith("/images/"):
        return None
    relative_path = clean_url.removeprefix("/images/").split("?", 1)[0]
    candidate = (IMAGE_DIR / relative_path).resolve()
    try:
        candidate.relative_to(IMAGE_DIR.resolve())
    except ValueError:
        return None
    return candidate if candidate.is_file() else None


def materialize_image_for_rpa(image_url: str, target: Path) -> bool:
    local_path = local_image_path_from_url(image_url)
    target.parent.mkdir(parents=True, exist_ok=True)
    if local_path:
        shutil.copyfile(local_path, target)
        return True
    if str(image_url or "").startswith("http"):
        try:
            content, _content_type = fetch_external_image(image_url)
        except (urllib.error.URLError, TimeoutError, ValueError):
            return False
        target.write_bytes(content)
        return True
    return False


def extract_json_object(text: str) -> dict[str, object]:
    try:
        parsed = json.loads(text)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        pass
    match = re.search(r"\{.*\}", text, re.S)
    if not match:
        return {}
    try:
        parsed = json.loads(match.group(0))
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def get_vision_api_config() -> sqlite3.Row | None:
    with connect() as db:
        return db.execute("SELECT * FROM api_configs WHERE key = 'vision'").fetchone()


def normalize_chat_completions_endpoint(base_url: str) -> str:
    endpoint = (base_url or "https://ark.cn-beijing.volces.com/api/v3").strip().rstrip("/")
    if endpoint.endswith("/chat/completions"):
        return endpoint
    if endpoint.endswith("/responses"):
        endpoint = endpoint[: -len("/responses")]
    return f"{endpoint}/chat/completions"


def normalize_color_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-_【】\[\]（）(){}:：,，/\\|]+", "", text)
    return text


def resolve_matched_color(candidate: str, colors: list[str]) -> str:
    candidate_text = str(candidate or "").strip()
    if not candidate_text:
        return ""
    if candidate_text in colors:
        return candidate_text
    normalized_candidate = normalize_color_text(candidate_text)
    if not normalized_candidate:
        return ""
    for color in colors:
        if normalize_color_text(color) == normalized_candidate:
            return color
    for color in colors:
        normalized_color = normalize_color_text(color)
        if normalized_candidate and normalized_candidate in normalized_color:
            return color
    for color in colors:
        normalized_color = normalize_color_text(color)
        if normalized_color and normalized_color in normalized_candidate:
            return color
    return ""


def active_prompt_template(prompt_type: str, usage: str = "") -> str:
    enabled_status = chr(0x542f) + chr(0x7528) + chr(0x4e2d)
    clauses = ["prompt_type = ?", "status = ?"]
    params: list[object] = [prompt_type, enabled_status]
    if usage:
        clauses.append("usage = ?")
        params.append(usage)
    with connect() as db:
        row = db.execute(
            f"SELECT content FROM prompt_templates WHERE {' AND '.join(clauses)} ORDER BY id DESC LIMIT 1",
            params,
        ).fetchone()
    return row["content"] if row else ""


def active_title_prompt_template() -> str:
    title_prompt_type = chr(0x6807) + chr(0x9898) + chr(0x63d0) + chr(0x793a) + chr(0x8bcd)
    title_usage = chr(0x6807) + chr(0x9898) + chr(0x751f) + chr(0x6210)
    enabled_status = chr(0x542f) + chr(0x7528) + chr(0x4e2d)
    with connect() as db:
        row = db.execute(
            """
            SELECT content FROM prompt_templates
            WHERE status = ? AND prompt_type = ? AND usage = ?
            ORDER BY id DESC LIMIT 1
            """,
            (enabled_status, title_prompt_type, title_usage),
        ).fetchone()
        if not row:
            row = db.execute(
                """
                SELECT content FROM prompt_templates
                WHERE status = ? AND content LIKE '%{orig_title}%' AND content LIKE '%{attrs_text}%'
                ORDER BY id DESC LIMIT 1
                """,
                (enabled_status,),
            ).fetchone()
    return row["content"] if row else ""


def render_title_prompt(template: str, item: ProcessingItem) -> str:
    attrs = [
        f"SKC: {item.skc}",
        f"Color: {item.color}",
        f"Size: {item.size}",
        f"SKU Code: {item.sku_code}",
        f"Source URL: {item.source_url}",
    ]
    attrs_text = "\n".join(part for part in attrs if not part.endswith(": "))
    prompt = template or ""
    replacements = {
        "orig_title": item.title,
        "attrs_text": attrs_text,
        "skc": item.skc,
        "color": item.color,
        "size": item.size,
        "sku_code": item.sku_code,
        "source_url": item.source_url,
    }
    for key, value in replacements.items():
        prompt = prompt.replace("{" + key + "}", str(value or ""))
    if "{orig_title}" not in template and "{attrs_text}" not in template:
        prompt = f"{prompt}\n\nSource title: {item.title}\nAttributes:\n{attrs_text}".strip()
    return prompt


def clean_generated_title(value: str, limit: int = 120) -> str:
    text = str(value or "").strip().strip("` ").strip()
    if text.lower().startswith("json"):
        text = text[4:].strip()
    text = re.sub(r"^[-*\d.\u3001\s]+", "", text)
    text = text.strip('\\"\u201c\u201d\u2018\u2019')
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def extract_title_from_model_response(response_text: str, fallback_source_title: str) -> str:
    text = clean_generated_title(response_text)
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            title = str(parsed.get("title_en") or parsed.get("english_title") or parsed.get("title") or "").strip()
            if title:
                return clean_generated_title(title)
    except Exception:
        pass
    return text or english_title_fallback(fallback_source_title)


def extract_batch_title_pairs(response_text: str) -> list[dict[str, str]]:
    text = str(response_text or "").strip().strip("` ")
    if text.lower().startswith("json"):
        text = text[4:].strip()
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return [item for item in parsed if isinstance(item, dict)]
        if isinstance(parsed, dict):
            items = parsed.get("items") or parsed.get("titles") or parsed.get("results")
            if isinstance(items, list):
                return [item for item in items if isinstance(item, dict)]
    except Exception:
        pass
    return []

def image_url_for_vision(image_url: str) -> str:
    clean_url = str(image_url or "").strip()
    if clean_url.startswith(("http://", "https://", "data:image/")):
        return clean_url
    if clean_url.startswith("/images/"):
        relative_path = clean_url.removeprefix("/images/").replace("/", os.sep)
        image_path = (IMAGE_DIR / relative_path).resolve()
        if not str(image_path).startswith(str(IMAGE_DIR.resolve())) or not image_path.is_file():
            raise HTTPException(status_code=400, detail=f"本地图片不存在，无法视觉识别：{clean_url}")
        mime_type = mimetypes.guess_type(str(image_path))[0] or "image/jpeg"
        return f"data:{mime_type};base64,{base64.b64encode(image_path.read_bytes()).decode('ascii')}"
    raise HTTPException(status_code=400, detail=f"豆包视觉识别只支持 http/https 或本地 /images 图片：{clean_url}")


def call_doubao_vision_color(image_url: str, colors: list[str]) -> dict[str, object]:
    config = get_vision_api_config()
    if not config or not config["enabled"] or not config["api_key"]:
        raise HTTPException(status_code=400, detail="豆包视觉识别 API 未配置或未启用")
    endpoint = normalize_chat_completions_endpoint(config["base_url"] or "https://ark.cn-beijing.volces.com/api/v3")
    model = config["model"] or "doubao-1.5-vision-pro-32k"
    color_json = json.dumps(colors, ensure_ascii=False)
    prompt_template = active_prompt_template("图片识别提示词", "颜色分图")
    if not prompt_template:
        prompt_template = next((item[4] for item in DEFAULT_PROMPT_TEMPLATES if item[2] == "图片识别提示词" and item[3] == "颜色分图"), "")
    prompt = prompt_template.replace("{color_json}", color_json).strip()

    body = json.dumps(
        {
            "model": model,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": image_url_for_vision(image_url)}},
                    ],
                }
            ],
            "temperature": 0,
            "max_tokens": 300,
        },
        ensure_ascii=False,
    ).encode("utf-8")
    request = urllib.request.Request(
        endpoint,
        data=body,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {config['api_key']}"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:500]
        raise HTTPException(status_code=502, detail=f"豆包视觉识别失败：{exc.code} {detail}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"豆包视觉识别连接失败：{exc.reason}") from exc
    content = (((payload.get("choices") or [{}])[0].get("message") or {}).get("content") or "").strip()
    result = extract_json_object(content)
    matched_color = resolve_matched_color(str(result.get("matched_color") or "").strip(), colors)
    return {
        "matched_color": matched_color,
        "confidence": to_float(result.get("confidence"), 0),
        "is_product_image": bool(result.get("is_product_image", True)),
        "reason": str(result.get("reason") or ""),
        "raw": payload,
    }


def get_or_create_vision_match(product_id: int, option: ImageOption, colors: list[str], allow_api: bool) -> dict[str, object] | None:
    identity = image_identity_server(option.url)
    if not identity:
        return None
    with connect() as db:
        cached = db.execute(
            "SELECT * FROM product_image_vision_cache WHERE product_id = ? AND image_identity = ?",
            (product_id, identity),
        ).fetchone()
    if cached:
        try:
            cached_raw = json.loads(cached["raw_json"] or "{}")
        except json.JSONDecodeError:
            cached_raw = {}
        if cached_raw.get("cache_version") == VISION_CACHE_VERSION:
            return {"matched_color": cached["matched_color"], "confidence": float(cached["confidence"] or 0), "is_product_image": bool(cached["is_product_image"]), "status": cached["status"]}
    if not allow_api:
        return None
    result = call_doubao_vision_color(option.url, colors)
    result["cache_version"] = VISION_CACHE_VERSION
    timestamp = now_text()
    with connect() as db:
        db.execute(
            """
            INSERT INTO product_image_vision_cache
            (product_id, image_identity, image_url, matched_color, confidence, is_product_image, status, provider, raw_json, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'doubao', ?, ?, ?)
            ON CONFLICT(product_id, image_identity) DO UPDATE SET
              image_url=excluded.image_url, matched_color=excluded.matched_color, confidence=excluded.confidence,
              is_product_image=excluded.is_product_image, status=excluded.status, provider=excluded.provider,
              raw_json=excluded.raw_json, updated_at=excluded.updated_at
            """,
            (
                product_id,
                identity,
                option.url,
                result["matched_color"],
                result["confidence"],
                1 if result["is_product_image"] else 0,
                "matched" if result["matched_color"] else "unmatched",
                json.dumps(result, ensure_ascii=False),
                timestamp,
                timestamp,
            ),
        )
        db.commit()
    return result


def unique_options_for_assignment(options: list[ImageOption]) -> list[ImageOption]:
    seen: set[str] = set()
    unique_options: list[ImageOption] = []
    for option in options:
        key = image_identity_server(option.url)
        if not key or key in seen:
            continue
        seen.add(key)
        unique_options.append(option)
    return unique_options


def auto_assign_color_images(payload: AutoAssignColorImagesPayload) -> AutoAssignColorImagesResult:
    if payload.use_vision and not payload.user_confirmed_vision:
        raise HTTPException(status_code=409, detail="图片识别会调用视觉 API 并可能产生费用，请先确认后再执行")
    count_per_color = max(1, min(int(payload.count_per_color or 3), 10))
    with connect() as db:
        product = db.execute("SELECT * FROM products WHERE id = ?", (payload.product_id,)).fetchone()
        override = db.execute("SELECT source_url FROM processing_overrides WHERE product_id = ?", (payload.product_id,)).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    source_url = override["source_url"] if override and override["source_url"] else ""
    image_options = build_product_image_options(payload.product_id, product["main_image"] or "", source_url)
    sku_options = unique_options_for_assignment([option for option in image_options if option.kind == "detail_sku_image" and option.color])
    supplement_options = unique_options_for_assignment([
        option for option in image_options
        if option.kind in {"detail_desc_image", "detail_main_image", "collection_image", "main_image"}
    ])
    colors = sorted({clean_color_label(option.color) for option in sku_options if clean_color_label(option.color)})
    if not colors:
        return AutoAssignColorImagesResult(product_id=payload.product_id, count_per_color=count_per_color, assigned_count=0, colors={})
    vision_by_color: dict[str, list[tuple[ImageOption, float]]] = {color: [] for color in colors}
    if payload.use_vision and colors:
        for option in supplement_options[:40]:
            match = get_or_create_vision_match(payload.product_id, option, colors, payload.user_confirmed_vision)
            if not match or not match.get("is_product_image"):
                continue
            matched_color = str(match.get("matched_color") or "")
            confidence = float(match.get("confidence") or 0)
            if matched_color in vision_by_color and confidence >= 0.75:
                vision_by_color[matched_color].append((option, confidence))
    timestamp = now_text()
    assigned_by_color: dict[str, list[tuple[ImageOption, float]]] = {}
    used_global: set[str] = set()
    for color in colors:
        direct = [option for option in sku_options if clean_color_label(option.color) == color]
        picked: list[tuple[ImageOption, float]] = []
        used_local: set[str] = set()
        for option in direct:
            key = image_identity_server(option.url)
            if key and key not in used_local:
                picked.append((option, 0.98))
                used_local.add(key)
            if len(picked) >= count_per_color:
                break
        candidate_supplements = vision_by_color.get(color) if payload.use_vision else []
        if not candidate_supplements:
            candidate_supplements = [(option, 0.55) for option in supplement_options]
        for option, confidence in candidate_supplements:
            key = image_identity_server(option.url)
            if not key or key in used_local or key in used_global:
                continue
            picked.append((option, confidence))
            used_local.add(key)
            if len(picked) >= count_per_color:
                break
        for option, _confidence in picked:
            key = image_identity_server(option.url)
            if key:
                used_global.add(key)
        assigned_by_color[color] = picked[:count_per_color]
    with connect() as db:
        db.execute("DELETE FROM product_color_image_assignments WHERE product_id = ?", (payload.product_id,))
        assigned_count = 0
        for color, picked in assigned_by_color.items():
            for sort_order, (option, confidence) in enumerate(picked):
                db.execute(
                    """
                    INSERT INTO product_color_image_assignments
                    (product_id, color, image_url, source, sort_order, is_auto, confidence, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
                    """,
                    (payload.product_id, color, option.url, option.kind, sort_order, confidence, timestamp, timestamp),
                )
                assigned_count += 1
        db.commit()
    return AutoAssignColorImagesResult(
        product_id=payload.product_id,
        count_per_color=count_per_color,
        assigned_count=assigned_count,
        colors={color: len(picked) for color, picked in assigned_by_color.items()},
    )


def save_manual_color_image_assignment(payload: ManualColorImageAssignmentPayload) -> list[ColorImageAssignment]:
    color = normalize_assignment_color(payload.color)
    image_url = str(payload.image_url or "").strip()
    slot_index = max(0, min(99, int(payload.slot_index or 0)))
    with connect() as db:
        product = db.execute("SELECT id FROM products WHERE id = ?", (payload.product_id,)).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="商品不存在")
        timestamp = now_text()
        db.execute(
            "DELETE FROM product_color_image_assignments WHERE product_id = ? AND color = ? AND sort_order = ?",
            (payload.product_id, color, slot_index),
        )
        if image_url:
            db.execute(
                """
                INSERT INTO product_color_image_assignments
                (product_id, color, image_url, source, sort_order, is_auto, confidence, created_at, updated_at)
                VALUES (?, ?, ?, 'manual_drag', ?, 0, 1, ?, ?)
                """,
                (payload.product_id, color, image_url, slot_index, timestamp, timestamp),
            )
        if not image_url:
            compact_color_image_assignments(db, payload.product_id, color)
        db.commit()
    return list_color_image_assignments(payload.product_id)


def compact_detail_image_assignments(db: sqlite3.Connection, product_id: int) -> None:
    rows = db.execute(
        """
        SELECT id FROM product_detail_image_assignments
        WHERE product_id = ? AND image_url != ''
        ORDER BY sort_order, id
        """,
        (product_id,),
    ).fetchall()
    timestamp = now_text()
    for index, row in enumerate(rows):
        db.execute(
            "UPDATE product_detail_image_assignments SET sort_order = ?, updated_at = ? WHERE id = ?",
            (index, timestamp, row["id"]),
        )


def save_detail_image_slot(product_id: int, slot_index: int, image_url: str, remove_image_url: str = "") -> list[DetailImageAssignment]:
    clean_url = str(image_url or "").strip()
    safe_slot_index = max(0, min(99, int(slot_index or 0)))
    with connect() as db:
        product = db.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="商品不存在")
        timestamp = now_text()
        result = db.execute(
            "DELETE FROM product_detail_image_assignments WHERE product_id = ? AND sort_order = ?",
            (product_id, safe_slot_index),
        )
        deleted_count = result.rowcount
        remove_url = str(remove_image_url or "").strip()
        if not clean_url and deleted_count == 0 and remove_url:
            row = db.execute(
                """
                SELECT id FROM product_detail_image_assignments
                WHERE product_id = ? AND image_url = ?
                ORDER BY sort_order, id
                LIMIT 1
                """,
                (product_id, remove_url),
            ).fetchone()
            if row:
                delete_result = db.execute("DELETE FROM product_detail_image_assignments WHERE id = ?", (row["id"],))
                deleted_count = delete_result.rowcount
        if not clean_url and deleted_count == 0:
            row = db.execute(
                """
                SELECT id FROM product_detail_image_assignments
                WHERE product_id = ? AND image_url != ''
                ORDER BY sort_order, id
                LIMIT 1 OFFSET ?
                """,
                (product_id, safe_slot_index),
            ).fetchone()
            if row:
                db.execute("DELETE FROM product_detail_image_assignments WHERE id = ?", (row["id"],))
        if clean_url:
            db.execute(
                """
                INSERT INTO product_detail_image_assignments
                (product_id, image_url, source, sort_order, created_at, updated_at)
                VALUES (?, ?, 'manual_drag', ?, ?, ?)
                """,
                (product_id, clean_url, safe_slot_index, timestamp, timestamp),
            )
        else:
            compact_detail_image_assignments(db, product_id)
        db.commit()
    return list_detail_image_assignments(product_id)


def replace_detail_image_assignments(product_id: int, image_urls: list[str]) -> list[DetailImageAssignment]:
    clean_urls = [str(url or "").strip() for url in image_urls if str(url or "").strip()]
    with connect() as db:
        product = db.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="商品不存在")
        timestamp = now_text()
        db.execute("DELETE FROM product_detail_image_assignments WHERE product_id = ?", (product_id,))
        for index, url in enumerate(clean_urls[:100]):
            db.execute(
                """
                INSERT INTO product_detail_image_assignments
                (product_id, image_url, source, sort_order, created_at, updated_at)
                VALUES (?, ?, 'manual_drag', ?, ?, ?)
                """,
                (product_id, url, index, timestamp, timestamp),
            )
        db.commit()
    return list_detail_image_assignments(product_id)


def save_manual_detail_image_assignment(payload: ManualDetailImageAssignmentPayload) -> list[DetailImageAssignment]:
    if payload.remaining_image_urls:
        return replace_detail_image_assignments(payload.product_id, payload.remaining_image_urls)
    return save_detail_image_slot(payload.product_id, payload.slot_index, payload.image_url, payload.remove_image_url)


def normalize_image_dedupe_key(image_url: str) -> str:
    clean_url = str(image_url or "").strip()
    if not clean_url:
        return ""
    parsed = urllib.parse.urlsplit(clean_url)
    if parsed.scheme and parsed.netloc:
        return urllib.parse.urlunsplit((parsed.scheme.lower(), parsed.netloc.lower(), parsed.path, "", ""))
    return clean_url.replace("\\", "/").strip().lower()


def dedupe_processing_images(product_ids: list[int]) -> DedupeProcessingImagesResult:
    safe_ids = [int(product_id) for product_id in product_ids if int(product_id) > 0]
    if not safe_ids:
        raise HTTPException(status_code=400, detail="请选择要去重的商品")
    scanned_count = 0
    detail_removed_count = 0
    color_removed_count = 0
    with connect() as db:
        for product_id in safe_ids:
            seen: set[str] = set()
            detail_rows = db.execute(
                "SELECT id, image_url FROM product_detail_image_assignments WHERE product_id = ? AND image_url != '' ORDER BY sort_order, id",
                (product_id,),
            ).fetchall()
            color_rows = db.execute(
                "SELECT id, image_url FROM product_color_image_assignments WHERE product_id = ? AND image_url != '' ORDER BY color, sort_order, id",
                (product_id,),
            ).fetchall()
            for row in list(detail_rows) + list(color_rows):
                dedupe_key = normalize_image_dedupe_key(row["image_url"])
                if dedupe_key:
                    scanned_count += 1
                    seen.add(dedupe_key)
            duplicate_detail_ids: list[int] = []
            duplicate_color_ids: list[int] = []
            seen.clear()
            for row in detail_rows:
                dedupe_key = normalize_image_dedupe_key(row["image_url"])
                if not dedupe_key:
                    continue
                if dedupe_key in seen:
                    duplicate_detail_ids.append(int(row["id"]))
                else:
                    seen.add(dedupe_key)
            for row in color_rows:
                dedupe_key = normalize_image_dedupe_key(row["image_url"])
                if not dedupe_key:
                    continue
                if dedupe_key in seen:
                    duplicate_color_ids.append(int(row["id"]))
                else:
                    seen.add(dedupe_key)
            if duplicate_detail_ids:
                placeholders = ",".join("?" for _ in duplicate_detail_ids)
                db.execute(f"DELETE FROM product_detail_image_assignments WHERE id IN ({placeholders})", duplicate_detail_ids)
                detail_removed_count += len(duplicate_detail_ids)
            if duplicate_color_ids:
                placeholders = ",".join("?" for _ in duplicate_color_ids)
                db.execute(f"DELETE FROM product_color_image_assignments WHERE id IN ({placeholders})", duplicate_color_ids)
                color_removed_count += len(duplicate_color_ids)
    removed_count = detail_removed_count + color_removed_count
    return DedupeProcessingImagesResult(
        product_count=len(safe_ids),
        scanned_count=scanned_count,
        removed_count=removed_count,
        detail_removed_count=detail_removed_count,
        color_removed_count=color_removed_count,
    )


def get_image_api_config() -> sqlite3.Row | None:
    with connect() as db:
        return db.execute("SELECT * FROM api_configs WHERE key = 'image'").fetchone()


def nanobanana_submit(api_key: str, image_url: str, prompt: str, size: str = "1344x1792", aspect_ratio: str = "3:4") -> str:
    body = json.dumps({"key": api_key, "prompt": prompt, "size": size, "aspectRatio": aspect_ratio, "urls": [image_url]}).encode("utf-8")
    request = urllib.request.Request(NANOBANANA_SUBMIT_URL, data=body, headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:500]
        raise HTTPException(status_code=502, detail=f"图生图提交失败：{exc.code} {detail}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"图生图连接失败：{exc.reason}") from exc
    except TimeoutError as exc:
        raise HTTPException(status_code=504, detail="图生图提交超时") from exc
    task_id = (payload.get("data") or {}).get("id")
    if not task_id:
        raise HTTPException(status_code=502, detail=f"图生图提交失败：服务商未返回任务 ID，返回：{payload.get('msg') or payload.get('message') or payload.get('code')}")
    return str(task_id)


def nanobanana_query(api_key: str, task_id: str) -> dict[str, object]:
    query = urllib.parse.urlencode({"key": api_key, "id": task_id})
    request = urllib.request.Request(f"{NANOBANANA_DETAIL_URL}?{query}", headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(request, timeout=45) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:500]
        raise HTTPException(status_code=502, detail=f"图生图查询失败：{exc.code} {detail}") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail=f"图生图查询连接失败：{exc.reason}") from exc
    except TimeoutError as exc:
        raise HTTPException(status_code=504, detail="图生图查询超时") from exc
    data = payload.get("data") or {}
    status_code = data.get("status")
    result_urls = data.get("result") or []
    if isinstance(result_urls, str):
        result_urls = [result_urls]
    if status_code == 2:
        return {"status": "completed", "result_urls": [url for url in result_urls if url], "note": "图生图已完成"}
    if status_code == 3:
        return {"status": "failed", "result_urls": [], "note": data.get("message") or "图生图任务失败"}
    return {"status": "processing", "result_urls": [], "note": f"图生图处理中，服务商状态：{status_code}"}


def default_english_title(title: str) -> str:
    return f"{title} for Temu listing"


def normalize_assignment_color(value: str) -> str:
    return clean_color_label(value) or "混合色"


def compact_color_image_assignments(db: sqlite3.Connection, product_id: int, color: str) -> None:
    rows = db.execute(
        """
        SELECT id FROM product_color_image_assignments
        WHERE product_id = ? AND color = ? AND image_url != ''
        ORDER BY sort_order, id
        """,
        (product_id, color),
    ).fetchall()
    timestamp = now_text()
    for index, row in enumerate(rows):
        db.execute(
            "UPDATE product_color_image_assignments SET sort_order = ?, updated_at = ? WHERE id = ?",
            (index, timestamp, row["id"]),
        )


def seed_color_first_image_for_import(db: sqlite3.Connection, product_id: int, image_url: str) -> None:
    clean_url = str(image_url or "").strip()
    if not clean_url:
        return
    timestamp = now_text()
    db.execute(
        """
        INSERT INTO product_color_image_assignments
        (product_id, color, image_url, source, sort_order, is_auto, confidence, created_at, updated_at)
        VALUES (?, ?, ?, ?, 0, 1, 0.6, ?, ?)
        ON CONFLICT DO NOTHING
        """,
        (product_id, "混合色", clean_url, "collection_first_image", timestamp, timestamp),
    )


def create_processing_override_for_import(db: sqlite3.Connection, product_id: int, product: ProductPayload, source_url: str, total_cost: float) -> None:
    declared_price = round(float(total_cost) + 239, 2)
    db.execute(
        """
        INSERT INTO processing_overrides (
            product_id, english_title, color, size, sku_code, declared_price,
            weight_g, length_cm, width_cm, height_cm, source_url, stock, ship_days, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(product_id) DO UPDATE SET
            english_title = excluded.english_title,
            sku_code = excluded.sku_code,
            declared_price = excluded.declared_price,
            source_url = excluded.source_url,
            updated_at = excluded.updated_at
        """,
        (
            product_id,
            default_english_title(product.title),
            "pending",
            "pending",
            f"{product.skc}-COLOR-SIZE",
            declared_price,
            350,
            15,
            10,
            2,
            source_url,
            999,
            9,
            now_text(),
        ),
    )


def import_collection_row_as_product(db: sqlite3.Connection, row: sqlite3.Row) -> Product:
    skc = f"DS{datetime.now().strftime('%y%m%d')}{row['id']:03d}"
    offset = 0
    base_skc = skc
    while db.execute("SELECT 1 FROM products WHERE skc = ?", (skc,)).fetchone():
        offset += 1
        skc = increment_skc_value(base_skc, offset)
    has_collection_image = bool(row["image_url"])
    item = ProductPayload(
        title=row["title"],
        skc=skc,
        sku_summary="pending",
        purchase_price=float(row["price"]),
        platform_quote_price=0,
        weight_g=int(float(get_setting_value("default_processing_weight_g", "350") or 350)),
        first_mile=28.0,
        warehouse_fee=default_warehouse_fee(),
        last_mile=6.5,
        platform_cost=6.5,
        status="待处理" if has_collection_image else "pending_main_image",
    )
    total_cost, estimated_profit, gross_margin = recalc(item)
    timestamp = now_text()
    cursor = db.execute(
        """
        INSERT INTO products (
            title, skc, sku_summary, purchase_price, first_mile, platform_cost,
            platform_quote_price, weight_g, warehouse_fee, last_mile,
            total_cost, estimated_profit, gross_margin, status, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            item.title,
            item.skc,
            item.sku_summary,
            item.purchase_price,
            item.first_mile,
            item.platform_cost,
            item.platform_quote_price,
            item.weight_g,
            item.warehouse_fee,
            item.last_mile,
            total_cost,
            estimated_profit,
            gross_margin,
            item.status,
            timestamp,
            timestamp,
        ),
    )
    product_id = cursor.lastrowid
    main_image = save_product_main_image_from_url(product_id, row["image_url"])
    if main_image:
        db.execute("UPDATE products SET main_image = ?, status = ?, updated_at = ? WHERE id = ?", (main_image, "待处理", now_text(), product_id))
    elif has_collection_image:
        db.execute("UPDATE products SET status = ?, updated_at = ? WHERE id = ?", ("pending_main_image", now_text(), product_id))
    image_source_url = row["image_url"] or row["source_url"]
    create_processing_override_for_import(db, product_id, item, image_source_url, total_cost)
    seed_color_first_image_for_import(db, product_id, main_image or row["image_url"])
    if row["source_url"]:
        try_enrich_product_from_onebound_detail(db, product_id, row["source_url"])
    db.execute("UPDATE collection_items SET selected = 1, status = 'imported' WHERE id = ?", (row["id"],))
    product_row = db.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    return product_from_row(product_row)


@app.get("/api/products", response_model=list[Product])
def list_products(q: str = "", status: str = "", image: str = "") -> list[Product]:
    keyword = f"%{q.strip()}%"
    clauses: list[str] = []
    params: list[object] = []
    if q.strip():
        clauses.append("(title LIKE ? OR skc LIKE ? OR sku_summary LIKE ?)")
        params.extend([keyword, keyword, keyword])
    if status.strip():
        status_value = status.strip()
        status_groups = {
            "待处理": ["待处理", "pending", "needs_review", "pending_main_image", "待补主图"],
            "待补主图": ["待补主图", "pending_main_image", "needs_image"],
            "待上货": ["待上货", "ready_upload", "ready_to_export", "export_ready", "利润正常"],
            "已上货": ["已上货", "uploaded", "rpa_success"],
            "上货失败": ["上货失败", "upload_failed", "rpa_failed"],
            "利润异常": ["亏损", "低利润", "利润偏低"],
        }
        matched_statuses = status_groups.get(status_value, [status_value])
        placeholders = ",".join("?" for _ in matched_statuses)
        clauses.append(f"status IN ({placeholders})")
        params.extend(matched_statuses)
    if image == "has":
        clauses.append("main_image IS NOT NULL AND main_image != ''")
    elif image == "missing":
        clauses.append("(main_image IS NULL OR main_image = '')")
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as db:
        rows = db.execute(f"SELECT * FROM products {where_sql} ORDER BY id DESC", params).fetchall()
    return [product_from_row(row) for row in rows]


@app.get("/api/dashboard")
def dashboard() -> dict[str, int]:
    with connect() as db:
        product_count = db.execute("SELECT COUNT(*) AS total FROM products").fetchone()["total"]
        missing_images = db.execute("SELECT COUNT(*) AS total FROM products WHERE main_image IS NULL OR main_image = ''").fetchone()["total"]
        ready_products = db.execute("SELECT COUNT(DISTINCT skc) AS total FROM publish_records WHERE result = 'Success' AND skc != ''").fetchone()["total"]
        task_count = db.execute("SELECT COUNT(*) AS total FROM upload_tasks").fetchone()["total"]
        latest_ready_tasks = db.execute("SELECT COUNT(*) AS total FROM upload_tasks WHERE status IN ('export_ready', 'rpa_success')").fetchone()["total"]
    return {
        "product_count": product_count,
        "ready_products": ready_products,
        "missing_images": missing_images,
        "task_count": task_count,
        "ready_tasks": latest_ready_tasks,
    }


@app.get("/api/system/status")
def system_status() -> dict[str, object]:
    runtime_dirs = ensure_runtime_directories()
    upload = upload_preflight()
    collection = collection_preflight()
    script_dir = configured_script_dir()
    cos = cos_preflight()
    directory_checks = [
        {"key": key, "path": str(path), "exists": path.exists(), "writable": os.access(path, os.W_OK)}
        for key, path in LOCAL_DATA_DIRECTORIES.items()
    ]
    checklist = [
        {"key": "database", "label": "本地数据库", "ok": DB_PATH.exists(), "action": "无需处理" if DB_PATH.exists() else "启动应用会自动创建数据库"},
        {"key": "image_dir", "label": "图片目录", "ok": IMAGE_DIR.exists(), "action": "无需处理" if IMAGE_DIR.exists() else "创建 data/images 目录"},
        {"key": "runtime_dirs", "label": "本地运行目录", "ok": all(item["exists"] and item["writable"] for item in directory_checks), "action": "本地数据/日志/导出/采集目录可写"},
        {"key": "real_rpa", "label": "真实 RPA", "ok": True, "action": "点击正式上货直接执行"},
        {"key": "safe_collection", "label": "外部采集安全", "ok": get_setting_value("enable_external_collection", "false") != "true", "action": "默认关闭，安全" if get_setting_value("enable_external_collection", "false") != "true" else "确认要执行外部采集后再保持开启"},
        {"key": "upload_preflight", "label": "上货预检", "ok": bool(upload["ready"]), "action": "脚本和图片目录就绪" if upload["ready"] else "检查店小秘脚本目录和图片目录"},
        {"key": "executor", "label": "本地执行器", "ok": bool(get_setting_value("executor_mode", "local_python")), "action": f"模式：{get_setting_value('executor_mode', 'local_python')}，版本：{get_setting_value('executor_version', '0.1.0')}"},
        {"key": "fill_skc", "label": "自动填写 SKC", "ok": get_setting_value("upload_fill_skc", "true") == "true", "action": "上货流程会填写商品 SKC" if get_setting_value("upload_fill_skc", "true") == "true" else "当前关闭，可能导致 SKC 空缺"},
        {"key": "collector_1688", "label": "1688 采集", "ok": bool(collection["1688"]["configured"]), "action": "万邦 API/Cookie 已配置" if collection["1688"]["configured"] else "如需 1688 真实采集，优先填写万邦 Key/Secret"},
    ]
    return {
        "database": str(DB_PATH),
        "database_exists": DB_PATH.exists(),
        "image_dir": str(IMAGE_DIR),
        "image_dir_exists": IMAGE_DIR.exists(),
        "data_dir": str(DATA_DIR),
        "runtime_dirs": runtime_dirs,
        "directory_checks": directory_checks,
        "enable_real_rpa": "true",
        "enable_external_collection": get_setting_value("enable_external_collection", "false"),
        "collection_mode": get_setting_value("collection_mode", "1688"),
        "executor": {
            "mode": get_setting_value("executor_mode", "local_python"),
            "server_url": get_setting_value("executor_server_url", "http://127.0.0.1:8000"),
            "download_url": get_setting_value("executor_download_url", ""),
            "version": get_setting_value("executor_version", "0.1.0"),
            "poll_seconds": get_setting_value("executor_poll_seconds", "5"),
            "task_scope": get_setting_value("executor_task_scope", "manual"),
        },
        "upload_flow": {
            "fill_skc": get_setting_value("upload_fill_skc", "true"),
            "skc_missing_policy": get_setting_value("upload_skc_missing_policy", "pause"),
            "auto_submit": get_setting_value("upload_auto_submit", "false"),
            "error_policy": get_setting_value("upload_error_policy", "skip"),
            "save_screenshots": get_setting_value("upload_save_screenshots", "false"),
            "save_html": get_setting_value("upload_save_html", "false"),
            "trace": get_setting_value("upload_trace", "off"),
            "step_delay_ms": get_setting_value("upload_step_delay_ms", "500"),
        },
        "script_dir": str(script_dir),
        "script_dir_exists": script_dir.exists(),
        "cos_preflight": cos,
        "upload_preflight": upload,
        "collection_preflight": collection,
        "checklist": checklist,
    }


@app.get("/api/downloads/windows-app")
def download_windows_app() -> FileResponse:
    target = (RELEASE_DIR / "upload-assistant-windows.zip").resolve()
    if target.parent != RELEASE_DIR.resolve() or not target.exists():
        raise HTTPException(status_code=404, detail="上货软件安装包还没有上传到服务器")
    return FileResponse(target, filename="upload-assistant-windows.zip", media_type="application/zip")


@app.get("/api/downloads")
def list_downloads() -> dict[str, object]:
    target = RELEASE_DIR / "upload-assistant-windows.zip"
    return {
        "windows_app": {
            "available": target.exists(),
            "filename": target.name,
            "size": target.stat().st_size if target.exists() else 0,
            "download_url": WINDOWS_APP_COS_URL,
            "fallback_url": "/api/downloads/windows-app" if target.exists() else "",
        }
    }


@app.post("/api/products", response_model=Product)
def create_product(payload: ProductPayload) -> Product:
    payload = product_payload_with_freight(payload)
    total_cost, estimated_profit, gross_margin = recalc(payload)
    timestamp = now_text()
    try:
        with connect() as db:
            cursor = db.execute(
                """
                INSERT INTO products (
                    title, skc, sku_summary, purchase_price, first_mile, platform_cost,
                    platform_quote_price, weight_g, warehouse_fee, last_mile,
                    total_cost, estimated_profit, gross_margin, status, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    payload.title,
                    payload.skc,
                    payload.sku_summary,
                    payload.purchase_price,
                    payload.first_mile,
                    payload.platform_cost,
                    payload.platform_quote_price,
                    payload.weight_g,
                    payload.warehouse_fee,
                    payload.last_mile,
                    total_cost,
                    estimated_profit,
                    gross_margin,
                    payload.status,
                    timestamp,
                    timestamp,
                ),
            )
            row = db.execute("SELECT * FROM products WHERE id = ?", (cursor.lastrowid,)).fetchone()
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="SKC 已存在") from exc
    return product_from_row(row)


@app.put("/api/products/{product_id}", response_model=Product)
def update_product(product_id: int, payload: ProductPayload) -> Product:
    payload = product_payload_with_freight(payload)
    total_cost, estimated_profit, gross_margin = recalc(payload)
    try:
        with connect() as db:
            result = db.execute(
                """
                UPDATE products
                SET title = ?, skc = ?, sku_summary = ?, purchase_price = ?, first_mile = ?,
                    platform_cost = ?, platform_quote_price = ?, weight_g = ?, warehouse_fee = ?, last_mile = ?,
                    total_cost = ?, estimated_profit = ?, gross_margin = ?,
                    status = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    payload.title,
                    payload.skc,
                    payload.sku_summary,
                    payload.purchase_price,
                    payload.first_mile,
                    payload.platform_cost,
                    payload.platform_quote_price,
                    payload.weight_g,
                    payload.warehouse_fee,
                    payload.last_mile,
                    total_cost,
                    estimated_profit,
                    gross_margin,
                    payload.status,
                    now_text(),
                    product_id,
                ),
            )
            if result.rowcount == 0:
                raise HTTPException(status_code=404, detail="商品不存在")
            row = db.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="SKC 已存在") from exc
    return product_from_row(row)


@app.post("/api/products/quotes/import-text", response_model=QuoteImportResult)
def import_product_quotes_text(payload: QuoteImportPayload) -> QuoteImportResult:
    return apply_quote_rows(parse_quote_rows_from_text(payload.text))


@app.post("/api/products/quotes/import-file", response_model=QuoteImportResult)
def import_product_quotes_file(file: UploadFile = File(...)) -> QuoteImportResult:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".xlsx", ".csv", ".txt"}:
        raise HTTPException(status_code=400, detail="请上传 xlsx、csv 或 txt 核价表")
    import_dir = DATA_DIR / "imports"
    import_dir.mkdir(parents=True, exist_ok=True)
    target = import_dir / f"quote_{datetime.now().strftime('%Y%m%d_%H%M%S')}{suffix}"
    target.write_bytes(file.file.read())
    rows = parse_quote_rows_from_file(target)
    return apply_quote_rows(rows)


@app.post("/api/products/freight/import-file", response_model=FreightImportResult)
def import_freight_file(file: UploadFile = File(...), default_zone: str = Form("zone5")) -> FreightImportResult:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail="请上传 xlsx 或 xls 运费表")
    import_dir = DATA_DIR / "imports"
    import_dir.mkdir(parents=True, exist_ok=True)
    target = import_dir / f"freight_{datetime.now().strftime('%Y%m%d_%H%M%S')}{suffix}"
    target.write_bytes(file.file.read())
    rules = parse_freight_rules(target)
    result = FreightImportResult(
        first_mile_count=len(rules.get("first_mile", [])),
        last_mile_count=len(rules.get("last_mile", [])),
        default_zone=normalize_header(default_zone or "zone5") or "zone5",
        saved=True,
    )
    if result.first_mile_count <= 0 or result.last_mile_count <= 0:
        raise HTTPException(status_code=400, detail="运费表缺少头程或尾程规则，请检查 Sheet1/Sheet2 表头")
    with connect() as db:
        save_setting_value(db, FREIGHT_RULES_SETTING_KEY, json.dumps(rules, ensure_ascii=False))
        save_setting_value(db, DEFAULT_FREIGHT_ZONE_KEY, result.default_zone)
        save_setting_value(db, DEFAULT_WAREHOUSE_FEE_KEY, str(default_warehouse_fee()))
        recalculate_all_product_freight(db)
    return result


@app.get("/api/products/freight/rules")
def get_freight_rules() -> dict[str, object]:
    rules = freight_rules()
    first_mile = rules.get("first_mile", []) if isinstance(rules.get("first_mile"), list) else []
    last_mile = rules.get("last_mile", []) if isinstance(rules.get("last_mile"), list) else []
    return {
        "default_zone": get_setting_value(DEFAULT_FREIGHT_ZONE_KEY, "zone5"),
        "warehouse_fee": default_warehouse_fee(),
        "first_mile_count": len(first_mile),
        "last_mile_count": len(last_mile),
        "first_mile": first_mile[:20],
        "last_mile": last_mile[:50],
    }


@app.put("/api/products/freight/rules")
def save_freight_rules(payload: FreightRulesPayload) -> dict[str, object]:
    zone = normalize_freight_zone(payload.default_zone)
    warehouse_fee = float(payload.warehouse_fee or 0)
    if warehouse_fee < 0:
        raise HTTPException(status_code=400, detail="仓库操作费不能为负数")
    rules = freight_payload_to_rules(payload)
    with connect() as db:
        save_setting_value(db, FREIGHT_RULES_SETTING_KEY, json.dumps(rules, ensure_ascii=False))
        save_setting_value(db, DEFAULT_FREIGHT_ZONE_KEY, zone)
        save_setting_value(db, DEFAULT_WAREHOUSE_FEE_KEY, str(warehouse_fee))
        updated_count = recalculate_all_product_freight(db)
    first_mile = rules.get("first_mile", []) if isinstance(rules.get("first_mile"), list) else []
    last_mile = rules.get("last_mile", []) if isinstance(rules.get("last_mile"), list) else []
    return {
        "saved": True,
        "updated_count": updated_count,
        "default_zone": zone,
        "warehouse_fee": warehouse_fee,
        "first_mile_count": len(first_mile),
        "last_mile_count": len(last_mile),
        "first_mile": first_mile[:20],
        "last_mile": last_mile[:50],
    }


@app.post("/api/products/{product_id}/image", response_model=Product)
def upload_product_image(product_id: int, image: UploadFile = File(...)) -> Product:
    if not image.content_type or not image.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="请上传图片文件")
    suffix = Path(image.filename or "main.jpg").suffix.lower() or ".jpg"
    product_dir = IMAGE_DIR / str(product_id)
    product_dir.mkdir(parents=True, exist_ok=True)
    target = product_dir / f"main{suffix}"
    for old_file in product_dir.glob("main.*"):
        old_file.unlink(missing_ok=True)
    with target.open("wb") as file:
        shutil.copyfileobj(image.file, file)
    image_path = f"/images/{product_id}/{target.name}"
    with connect() as db:
        result = db.execute(
            "UPDATE products SET main_image = ?, status = ?, updated_at = ? WHERE id = ?",
            (image_path, "利润正常", now_text(), product_id),
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="商品不存在")
        row = db.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()
    return product_from_row(row)


def write_placeholder_image(product_id: int, skc: str, title: str) -> str:
    product_dir = IMAGE_DIR / str(product_id)
    product_dir.mkdir(parents=True, exist_ok=True)
    target = product_dir / "main.svg"
    safe_skc = escape(skc[:32])
    safe_title = escape(title[:42])
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="900" height="1200" viewBox="0 0 900 1200">
  <defs>
    <linearGradient id="bg" x1="0" x2="1" y1="0" y2="1">
      <stop offset="0" stop-color="#fff7ed"/>
      <stop offset="1" stop-color="#ffedd5"/>
    </linearGradient>
  </defs>
  <rect width="900" height="1200" fill="url(#bg)"/>
  <rect x="90" y="150" width="720" height="900" rx="48" fill="#ffffff" stroke="#fed7aa" stroke-width="8"/>
  <text x="450" y="460" text-anchor="middle" font-family="Arial, sans-serif" font-size="54" font-weight="700" fill="#c2410c">上货助手</text>
  <text x="450" y="550" text-anchor="middle" font-family="Arial, sans-serif" font-size="40" fill="#7c2d12">内部占位主图</text>
  <text x="450" y="660" text-anchor="middle" font-family="Arial, sans-serif" font-size="34" fill="#111827">{safe_skc}</text>
  <text x="450" y="735" text-anchor="middle" font-family="Arial, sans-serif" font-size="28" fill="#6b7280">{safe_title}</text>
  <text x="450" y="910" text-anchor="middle" font-family="Arial, sans-serif" font-size="24" fill="#9ca3af">仅用于本地 MVP 闭环验证，请替换真实商品图后再真实上货</text>
</svg>
"""
    for old_file in product_dir.glob("main.*"):
        old_file.unlink(missing_ok=True)
    target.write_text(svg, encoding="utf-8")
    return f"/images/{product_id}/{target.name}"


@app.post("/api/products/images/placeholders")
def generate_missing_image_placeholders() -> dict[str, object]:
    raise HTTPException(status_code=410, detail="本地占位主图功能已移除，请上传真实商品图")


@app.post("/api/products/images/process")
def process_product_image(payload: ProcessImagePayload) -> ImageTask:
    with connect() as db:
        product = db.execute("SELECT * FROM products WHERE id = ?", (payload.product_id,)).fetchone()
    if not product:
        raise HTTPException(status_code=404, detail="商品不存在")
    processing_item = get_processing_item(payload.product_id)
    image_options = build_product_image_options(payload.product_id, product["main_image"] or "", processing_item.source_url.strip())
    selected_image = payload.source_image.strip() or (image_options[0].url if image_options else "")
    allowed_images = {option.url for option in image_options}
    if not selected_image:
        raise HTTPException(status_code=400, detail="商品没有可用于图生图的图片")
    if selected_image not in allowed_images:
        raise HTTPException(status_code=400, detail="请选择当前商品关联的图片进行图生图")
    if selected_image.startswith("/images/"):
        source_path = local_image_path_from_url(selected_image)
        if not source_path:
            raise HTTPException(status_code=404, detail="所选本地图片文件不存在")
        local_path = save_processed_image_from_local(payload.product_id, selected_image)
        timestamp = now_text()
        with connect() as db:
            cursor = db.execute(
                """
                INSERT INTO image_tasks (product_id, provider, task_id, status, source_image, result_url, local_path, prompt, note, created_at, updated_at)
                VALUES (?, 'local', ?, 'completed', ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    payload.product_id,
                    f"local-{payload.product_id}-{int(time.time())}",
                    selected_image,
                    selected_image,
                    local_path,
                    "本地图片处理",
                    "已复制本地图片作为处理图，未调用外部图生图 API",
                    timestamp,
                    timestamp,
                ),
            )
            db.execute("UPDATE products SET main_image = ?, status = ?, updated_at = ? WHERE id = ?", (local_path, "待处理", timestamp, payload.product_id))
            row = db.execute("SELECT * FROM image_tasks WHERE id = ?", (cursor.lastrowid,)).fetchone()
        return image_task_from_row(row)
    if not selected_image.startswith("http"):
        raise HTTPException(status_code=400, detail="所选图片不是公网 URL，不能提交图生图")
    image_api = get_image_api_config()
    if not image_api or not image_api["enabled"] or not image_api["api_key"]:
        raise HTTPException(status_code=400, detail="图生图 API 未配置：请在设置中心填写 Nano Banana API Key")
    prompt = DEFAULT_IMAGE_PROMPT
    task_id = nanobanana_submit(image_api["api_key"], selected_image, prompt)
    timestamp = now_text()
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO image_tasks (product_id, provider, task_id, status, source_image, prompt, note, created_at, updated_at)
            VALUES (?, 'nanobanana', ?, 'submitted', ?, ?, ?, ?, ?)
            """,
            (payload.product_id, task_id, selected_image, prompt, "已提交 Nano Banana 图生图任务", timestamp, timestamp),
        )
        row = db.execute("SELECT * FROM image_tasks WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return image_task_from_row(row)


@app.get("/api/products/{product_id}/image-tasks", response_model=list[ImageTask])
def list_product_image_tasks(product_id: int) -> list[ImageTask]:
    with connect() as db:
        rows = db.execute("SELECT * FROM image_tasks WHERE product_id = ? ORDER BY id DESC", (product_id,)).fetchall()
    return [image_task_from_row(row) for row in rows]


@app.post("/api/image-tasks/{task_row_id}/refresh", response_model=ImageTask)
def refresh_image_task(task_row_id: int) -> ImageTask:
    image_api = get_image_api_config()
    if not image_api or not image_api["enabled"] or not image_api["api_key"]:
        raise HTTPException(status_code=400, detail="图生图 API 未配置：请在设置中心填写 Nano Banana API Key")
    with connect() as db:
        task = db.execute("SELECT * FROM image_tasks WHERE id = ?", (task_row_id,)).fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="图生图任务不存在")
    if task["status"] == "completed" and task["local_path"]:
        return image_task_from_row(task)
    result = nanobanana_query(image_api["api_key"], task["task_id"])
    status = str(result["status"])
    result_urls = result["result_urls"]
    result_url = result_urls[0] if result_urls else task["result_url"]
    local_path = task["local_path"]
    note = str(result["note"])
    if status == "completed" and result_url:
        try:
            local_path = save_processed_image_from_url(task["product_id"], result_url)
        except (urllib.error.URLError, TimeoutError, ValueError) as exc:
            status = "download_failed"
            note = f"图生图完成，但下载结果图失败：{exc}"
    with connect() as db:
        db.execute(
            "UPDATE image_tasks SET status = ?, result_url = ?, local_path = ?, note = ?, updated_at = ? WHERE id = ?",
            (status, result_url, local_path, note, now_text(), task_row_id),
        )
        row = db.execute("SELECT * FROM image_tasks WHERE id = ?", (task_row_id,)).fetchone()
    return image_task_from_row(row)


@app.post("/api/products/bulk/status")
def update_products_status(payload: ProductBulkStatusPayload) -> dict[str, int]:
    if not payload.ids:
        return {"updated_count": 0}
    placeholders = ",".join("?" for _ in payload.ids)
    with connect() as db:
        result = db.execute(
            f"UPDATE products SET status = ?, updated_at = ? WHERE id IN ({placeholders})",
            [payload.status, now_text(), *payload.ids],
        )
    return {"updated_count": result.rowcount}


@app.post("/api/products/bulk/delete")
def delete_products_bulk(payload: CollectionBulkPayload) -> dict[str, int]:
    deleted = 0
    for product_id in payload.ids:
        try:
            delete_product(product_id)
            deleted += 1
        except HTTPException:
            continue
    return {"deleted_count": deleted}


@app.post("/api/dev/cleanup-test-data")
def cleanup_test_data() -> dict[str, int]:
    patterns = ["%smoke test%", "%example.local%", "%??%", "%??%"]
    deleted_products = 0
    deleted_collection_items = 0
    deleted_collection_tasks = 0
    with connect() as db:
        product_rows = db.execute(
            """
            SELECT id FROM products
            WHERE lower(title) LIKE ? OR title LIKE ? OR title LIKE ? OR title LIKE ?
            """,
            patterns,
        ).fetchall()
        product_ids = [row["id"] for row in product_rows]
        for product_id in product_ids:
            db.execute("DELETE FROM processing_overrides WHERE product_id = ?", (product_id,))
            result = db.execute("DELETE FROM products WHERE id = ?", (product_id,))
            deleted_products += result.rowcount
        deleted_collection_items = db.execute(
            """
            DELETE FROM collection_items
            WHERE lower(title) LIKE ? OR source_url LIKE ? OR title LIKE ? OR title LIKE ?
            """,
            patterns,
        ).rowcount
        deleted_collection_tasks = db.execute(
            """
            DELETE FROM collection_tasks
            WHERE lower(keyword) LIKE ? OR keyword LIKE ? OR keyword LIKE ? OR keyword LIKE ?
            """,
            patterns,
        ).rowcount
    for product_id in product_ids:
        product_dir = IMAGE_DIR / str(product_id)
        if product_dir.exists():
            shutil.rmtree(product_dir)
    return {
        "deleted_products": deleted_products,
        "deleted_collection_items": deleted_collection_items,
        "deleted_collection_tasks": deleted_collection_tasks,
    }


def remove_file_if_old(path: Path, cutoff_timestamp: float) -> bool:
    if not path.is_file() or path.stat().st_mtime >= cutoff_timestamp:
        return False
    path.unlink(missing_ok=True)
    return True


@app.post("/api/local/cleanup")
def cleanup_local_data(payload: CleanupPayload) -> dict[str, int]:
    retention_days = max(int(payload.retention_days or 7), 1)
    cutoff_timestamp = time.time() - retention_days * 86400
    cutoff_text = datetime.fromtimestamp(cutoff_timestamp).strftime("%Y-%m-%d %H:%M:%S")
    deleted_exports = 0
    deleted_logs = 0
    deleted_images = 0
    deleted_upload_tasks = 0
    deleted_publish_records = 0
    if payload.clean_exports:
        export_dir = DATA_DIR / "export"
        if export_dir.exists():
            for path in export_dir.glob("*.xlsx"):
                if remove_file_if_old(path, cutoff_timestamp):
                    deleted_exports += 1
    if payload.clean_logs:
        logs_dir = DATA_DIR / "logs"
        if logs_dir.exists():
            for pattern in ("*.log", "*.json"):
                for path in logs_dir.glob(pattern):
                    if remove_file_if_old(path, cutoff_timestamp):
                        deleted_logs += 1
    with connect() as db:
        old_success_tasks = db.execute(
            "SELECT id FROM upload_tasks WHERE status = 'rpa_success' AND updated_at < ?",
            (cutoff_text,),
        ).fetchall()
        old_task_ids = [row["id"] for row in old_success_tasks]
        if old_task_ids:
            placeholders = ",".join("?" for _ in old_task_ids)
            deleted_upload_tasks = db.execute(f"DELETE FROM upload_tasks WHERE id IN ({placeholders})", old_task_ids).rowcount
        if payload.clean_success_images:
            rows = db.execute(
                """
                SELECT DISTINCT p.id
                FROM products p
                JOIN publish_records pr ON pr.skc = p.skc
                WHERE pr.result IN ('Success', '成功') AND pr.created_at < ?
                """,
                (cutoff_text,),
            ).fetchall()
            for row in rows:
                product_dir = IMAGE_DIR / str(row["id"])
                if not product_dir.exists():
                    continue
                for path in product_dir.iterdir():
                    if path.is_file() and path.suffix.lower() in RPA_IMAGE_EXTENSIONS:
                        path.unlink(missing_ok=True)
                        deleted_images += 1
        deleted_publish_records = db.execute(
            "DELETE FROM publish_records WHERE result IN ('Success', '成功', 'Queued') AND created_at < ?",
            (cutoff_text,),
        ).rowcount
    return {
        "retention_days": retention_days,
        "deleted_exports": deleted_exports,
        "deleted_logs": deleted_logs,
        "deleted_success_images": deleted_images,
        "deleted_upload_tasks": deleted_upload_tasks,
        "deleted_publish_records": deleted_publish_records,
    }










def get_setting_value(key: str, default: str = "") -> str:
    with connect() as db:
        row = db.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def int_setting(key: str, default: int, min_value: int = 0, max_value: int = 999) -> int:
    try:
        value = int(float(get_setting_value(key, str(default))))
    except (TypeError, ValueError):
        return default
    return max(min_value, min(max_value, value))


def float_setting(key: str, default: float, min_value: float = 0.0, max_value: float = 60.0) -> float:
    try:
        value = float(get_setting_value(key, str(default)))
    except (TypeError, ValueError):
        return default
    return max(min_value, min(max_value, value))


def write_run_log(task_id: int, content: str) -> str:
    log_dir = DATA_DIR / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"upload_task_{task_id}.log"
    log_path.write_text(content, encoding="utf-8")
    return str(log_path)


def processing_item_issues(item: ProcessingItem) -> list[str]:
    issues: list[str] = []
    colors = item_colors(item)
    unknown_colors = item_unknown_color_values(item)
    unsupported_sizes = item_unsupported_upload_sizes(item)
    if item.image_status != "has_main_image":
        issues.append("缺少主图")
    if not item.english_title.strip():
        issues.append("英文标题为空")
    elif contains_cjk(item.english_title):
        issues.append("英文标题包含中文")
    if not colors:
        issues.append("颜色为空")
    elif item.color.strip().lower() == "pending":
        issues.append("颜色待确认")
    if unknown_colors:
        issues.append(f"颜色无法匹配标准色：{'、'.join(unknown_colors[:6])}")
    if not item.size.strip():
        issues.append("尺码为空")
    elif item.size.strip().lower() == "pending":
        issues.append("尺码待确认")
    else:
        if unsupported_sizes:
            issues.append(f"不支持尺码：{'、'.join(unsupported_sizes[:6])}")
        if not item_supported_upload_sizes(item):
            issues.append("无 RPA 支持尺码")
    if not item.sku_code.strip():
        issues.append("SKU Code 为空")
    elif "COLOR-SIZE" in item.sku_code.upper():
        issues.append("SKU Code 使用默认占位")
    if item.declared_price <= 0:
        issues.append("申报价必须大于 0")
    if item.weight_g <= 0:
        issues.append("重量必须大于 0")
    if item.length_cm <= 0 or item.width_cm <= 0 or item.height_cm <= 0:
        issues.append("尺寸必须大于 0")
    if item.stock <= 0:
        issues.append("库存必须大于 0")
    if item.ship_days <= 0:
        issues.append("发货时效必须大于 0")
    for color, image_count in item_color_image_counts(item).items():
        if image_count < 3:
            issues.append(f"{color} 颜色图不足 3 张")
    return issues


def processing_item_warnings(item: ProcessingItem) -> list[str]:
    warnings: list[str] = []
    if not item.source_url.strip():
        warnings.append("站外链接为空")
    return warnings


def upload_item_diagnostics(items: list[ProcessingItem]) -> list[dict[str, object]]:
    diagnostics: list[dict[str, object]] = []
    for item in items:
        issues = processing_item_issues(item)
        diagnostics.append(
            {
                "product_id": item.product_id,
                "skc": item.skc,
                "title": item.title,
                "ready": not issues,
                "issues": issues,
                "warnings": processing_item_warnings(item),
            }
        )
    return diagnostics


def upload_preflight_summary(items: list[ProcessingItem]) -> dict[str, object]:
    diagnostics = upload_item_diagnostics(items)
    blocked_items = [item for item in diagnostics if not item["ready"]]
    issue_counts: dict[str, int] = {}
    warning_counts: dict[str, int] = {}
    for item in blocked_items:
        for issue in item["issues"]:
            issue_counts[issue] = issue_counts.get(issue, 0) + 1
    for item in diagnostics:
        for warning in item["warnings"]:
            warning_counts[warning] = warning_counts.get(warning, 0) + 1
    return {
        "total_count": len(items),
        "ready_count": len(items) - len(blocked_items),
        "blocked_count": len(blocked_items),
        "issue_counts": issue_counts,
        "warning_counts": warning_counts,
        "blocked_items": blocked_items,
        "warning_items": [item for item in diagnostics if item["warnings"]],
    }


def sync_processing_exceptions(items: list[ProcessingItem]) -> dict[str, object]:
    summary = upload_preflight_summary(items)
    diagnostics = upload_item_diagnostics(items)
    timestamp = now_text()
    active_ids: set[int] = set()
    with connect() as db:
        for item in diagnostics:
            product_id = int(item["product_id"])
            issues = list(item["issues"] or [])
            warnings = list(item["warnings"] or [])
            if not issues and not warnings:
                continue
            active_ids.add(product_id)
            db.execute(
                """
                INSERT INTO processing_exceptions (product_id, skc, title, status, issues_json, warnings_json, note, created_at, updated_at)
                VALUES (?, ?, ?, 'open', ?, ?, ?, ?, ?)
                ON CONFLICT(product_id) DO UPDATE SET
                    skc = excluded.skc,
                    title = excluded.title,
                    status = 'open',
                    issues_json = excluded.issues_json,
                    warnings_json = excluded.warnings_json,
                    note = excluded.note,
                    updated_at = excluded.updated_at
                """,
                (
                    product_id,
                    str(item["skc"] or ""),
                    str(item["title"] or ""),
                    json.dumps(issues, ensure_ascii=False),
                    json.dumps(warnings, ensure_ascii=False),
                    "；".join([*issues, *warnings])[:500],
                    timestamp,
                    timestamp,
                ),
            )
        if active_ids:
            placeholders = ",".join("?" for _ in active_ids)
            db.execute(
                f"UPDATE processing_exceptions SET status = 'resolved', updated_at = ? WHERE status = 'open' AND product_id NOT IN ({placeholders})",
                [timestamp, *active_ids],
            )
        else:
            db.execute("UPDATE processing_exceptions SET status = 'resolved', updated_at = ? WHERE status = 'open'", (timestamp,))
    return summary


def processing_exception_from_row(row: sqlite3.Row) -> ProcessingExceptionItem:
    return ProcessingExceptionItem(
        product_id=int(row["product_id"]),
        skc=row["skc"] or "",
        title=row["title"] or "",
        status=row["status"] or "open",
        issues=json.loads(row["issues_json"] or "[]"),
        warnings=json.loads(row["warnings_json"] or "[]"),
        updated_at=row["updated_at"] or "",
    )


def product_ids_for_failed_publish_records(db: sqlite3.Connection, ids: list[int] | None = None) -> list[int]:
    params: list[object] = []
    where = "WHERE pr.result IN ('Failed', '失败')"
    if ids:
        placeholders = ",".join("?" for _ in ids)
        where += f" AND pr.id IN ({placeholders})"
        params.extend(ids)
    rows = db.execute(
        f"""
        SELECT DISTINCT p.id
        FROM publish_records pr
        JOIN products p ON p.skc = pr.skc
        {where}
        ORDER BY p.id DESC
        """,
        params,
    ).fetchall()
    return [int(row["id"]) for row in rows]


def build_processing_items_for_ids(product_ids: list[int]) -> list[ProcessingItem]:
    if not product_ids:
        return []
    wanted = set(int(product_id) for product_id in product_ids)
    items_by_id = {item.product_id: item for item in build_processing_items() if item.product_id in wanted}
    return [items_by_id[product_id] for product_id in product_ids if product_id in items_by_id]


def normalize_upload_size(value: str) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return ""
    text = re.sub(r"[【\[].*?[】\]]", "", text).strip()
    aliases = size_alias_map()
    compact_text = normalize_color_alias(text).upper()
    if text in aliases:
        return aliases[text]
    if compact_text in aliases:
        return aliases[compact_text]
    match = re.search(r"(?<![A-Z0-9])(XS|S|M|L|XL|XXL|2XL|3XL|4XL|XXXL)(?![A-Z0-9])", text)
    if not match:
        match = re.search(r"(XS|S|M|L|XL|XXL|2XL|3XL|4XL|XXXL)", text)
    if not match:
        return text
    return aliases.get(match.group(1), match.group(1))


def item_sizes(item: ProcessingItem) -> list[str]:
    sizes: list[str] = []
    for part in re.split(r"[,，/、\n]+", item.size or ""):
        size = normalize_upload_size(part)
        if size and size not in sizes:
            sizes.append(size)
    return sizes or [normalize_upload_size(str(item.size or ""))]


def sort_upload_sizes(sizes: list[str]) -> list[str]:
    return sorted(sizes, key=lambda size: (STANDARD_UPLOAD_SIZE_RANK.get(size, len(STANDARD_UPLOAD_SIZES)), size))


def item_supported_upload_sizes(item: ProcessingItem) -> list[str]:
    return sort_upload_sizes([size for size in item_sizes(item) if size in RPA_ALLOWED_UPLOAD_SIZES])


def item_unsupported_upload_sizes(item: ProcessingItem) -> list[str]:
    return [size for size in item_sizes(item) if size and size not in RPA_ALLOWED_UPLOAD_SIZES]


def item_raw_color_values(item: ProcessingItem) -> list[str]:
    values: list[str] = []
    for value in re.split(r"[,，/、\n]+", item.color or ""):
        text = str(value or "").strip()
        if text and text.lower() != "pending" and text not in values:
            values.append(text)
    for assignment in item.color_image_assignments:
        text = str(assignment.color or "").strip()
        if text and text.lower() != "pending" and text not in values:
            values.append(text)
    for option in item.image_options:
        text = str(option.color or "").strip()
        if text and text.lower() != "pending" and text not in values:
            values.append(text)
    return values


def item_unknown_color_values(item: ProcessingItem) -> list[str]:
    unknown: list[str] = []
    for value in item_raw_color_values(item):
        if not clean_color_label(value) and value not in unknown:
            unknown.append(value)
    return unknown


def item_colors(item: ProcessingItem) -> list[str]:
    colors: list[str] = []
    for value in re.split(r"[,，/、\n]+", item.color or ""):
        color = clean_color_label(value)
        if color and color not in colors:
            colors.append(color)
    for assignment in item.color_image_assignments:
        color = clean_color_label(assignment.color)
        if color and color not in colors:
            colors.append(color)
    specific_colors = [color for color in colors if color != "混合色"]
    if specific_colors:
        return specific_colors
    if colors:
        return colors
    for option in item.image_options:
        color = clean_color_label(option.color)
        if color and color not in colors:
            colors.append(color)
    return colors


def color_upload_images(item: ProcessingItem, color: str) -> list[str]:
    clean_color = clean_color_label(color)
    if not clean_color:
        return []
    assigned = sorted(
        [assignment for assignment in item.color_image_assignments if clean_color_label(assignment.color) == clean_color],
        key=lambda assignment: assignment.sort_order,
    )
    urls = [assignment.image_url for assignment in assigned if assignment.image_url]
    if item.main_image:
        urls.append(item.main_image)
    for option in item.image_options:
        option_color = clean_color_label(option.color)
        if option.kind in {"detail_desc_image", "detail_image", "source_url"}:
            continue
        if option.url and (option_color == clean_color or (not option_color and option.kind in {"main_image", "local_file", "collection_image"})):
            urls.append(option.url)
    unique_urls: list[str] = []
    seen: set[str] = set()
    for url in urls:
        identity = image_identity_server(url)
        if identity and identity not in seen:
            seen.add(identity)
            unique_urls.append(url)
    return unique_urls[:3]


def item_upload_images(item: ProcessingItem) -> list[str]:
    colors = item_colors(item)
    if not colors:
        return []
    return color_upload_images(item, colors[0])


def detail_upload_images(item: ProcessingItem) -> list[str]:
    assigned = sorted(
        [assignment for assignment in item.detail_image_assignments if assignment.image_url],
        key=lambda assignment: assignment.sort_order,
    )
    unique_urls: list[str] = []
    seen: set[str] = set()
    for assignment in assigned:
        identity = image_identity_server(assignment.image_url)
        if identity and identity not in seen:
            seen.add(identity)
            unique_urls.append(assignment.image_url)
    return unique_urls


def miaoshou_image_url(value: str) -> str:
    text = str(value or "").strip()
    if text.startswith("http"):
        return text
    if text.startswith("/images/"):
        port = os.environ.get("UPLOAD_ASSISTANT_PORT", "8000")
        return f"http://127.0.0.1:{port}{text}"
    return text


def miaoshou_detail_description(item: ProcessingItem, detail_urls: list[str] | None = None) -> str:
    urls = detail_urls if detail_urls is not None else [miaoshou_image_url(url) for url in detail_upload_images(item)]
    urls = [url for url in urls if url]
    if not urls:
        return ""
    return "".join(f'<p><img src="{escape(url, quote=True)}" /></p>' for url in urls)


def cos_export_url_resolver(items: list[ProcessingItem]) -> Callable[[ProcessingItem, str, str | None, int], str]:
    if not cos_public_image_base():
        return lambda _item, _source_url, _color, _index: ""
    item_ids = {item.product_id for item in items}
    skc_by_product_id = {item.product_id: safe_path_segment(item.skc, f"product_{item.product_id}") for item in items}
    detail_urls_by_product_id = {item.product_id: detail_upload_images(item) for item in items}
    color_urls_by_key = {
        (item.product_id, clean_color_label(color)): color_upload_images(item, color)
        for item in items
        for color in item_colors(item)
    }
    cos_urls: dict[tuple[int, str], str] = {}
    root = configured_rpa_sku_image_dir()
    if not root.exists():
        return lambda _item, _source_url, _color, _index: ""
    for product_dir in root.iterdir():
        if not product_dir.is_dir():
            continue
        matching_ids = [product_id for product_id, safe_skc in skc_by_product_id.items() if safe_skc == product_dir.name]
        for product_id in matching_ids:
            if product_id not in item_ids:
                continue
            detail_dir = product_dir / "detail"
            if detail_dir.exists():
                for index, source_url in enumerate(detail_urls_by_product_id.get(product_id, []), start=1):
                    target = detail_dir / f"{index}.jpg"
                    if target.is_file():
                        cos_urls[(product_id, image_identity_server(source_url))] = cos_public_image_url(product_dir.name, "detail", target.name)
            for color_key, source_urls in [
                (key_color, urls)
                for (key_product_id, key_color), urls in color_urls_by_key.items()
                if key_product_id == product_id
            ]:
                color_dir = product_dir / safe_path_segment(color_key, "Color")
                if not color_dir.exists():
                    continue
                for index, source_url in enumerate(source_urls, start=1):
                    target = color_dir / f"{index}.jpg"
                    if target.is_file():
                        cos_urls[(product_id, image_identity_server(source_url))] = cos_public_image_url(product_dir.name, color_dir.name, target.name)

    def resolve(item: ProcessingItem, source_url: str, _color: str | None = None, _index: int = 0) -> str:
        return cos_urls.get((item.product_id, image_identity_server(source_url)), "")

    return resolve


def item_color_image_counts(item: ProcessingItem) -> dict[str, int]:
    return {color: len(color_upload_images(item, color)) for color in item_colors(item)}


def first_public_image_url(item: ProcessingItem) -> str:
    for image_url in item_upload_images(item):
        if str(image_url or "").startswith("http"):
            return image_url
    return ""


def first_public_or_blank(item: ProcessingItem) -> str:
    return first_public_image_url(item)


def prepare_rpa_sku_images(items: list[ProcessingItem]) -> dict[str, object]:
    target_root = configured_rpa_sku_image_dir()
    target_root.mkdir(parents=True, exist_ok=True)
    prepared: list[dict[str, object]] = []
    prepared_details: list[dict[str, object]] = []
    missing: list[dict[str, str]] = []
    for item in items:
        safe_skc = safe_path_segment(item.skc, f"product_{item.product_id}")
        product_root = target_root / safe_skc
        detail_urls = detail_upload_images(item)
        detail_dir = product_root / "detail"
        if detail_dir.exists():
            for old_file in detail_dir.iterdir():
                if old_file.is_file() and old_file.suffix.lower() in RPA_IMAGE_EXTENSIONS:
                    old_file.unlink(missing_ok=True)
        detail_written = 0
        for index, image_url in enumerate(detail_urls, start=1):
            target = detail_dir / f"{index}.jpg"
            if materialize_image_for_rpa(image_url, target):
                detail_written += 1
        if detail_written:
            prepared_details.append({"skc": item.skc, "count": detail_written, "dir": str(detail_dir)})
        for color in item_colors(item):
            safe_color = safe_path_segment(color, "Color")
            product_dir = product_root / safe_color
            if product_dir.exists():
                for old_file in product_dir.iterdir():
                    if old_file.is_file() and old_file.suffix.lower() in RPA_IMAGE_EXTENSIONS:
                        old_file.unlink(missing_ok=True)
            urls = color_upload_images(item, color)
            written = 0
            if detail_urls:
                target = product_dir / "0.jpg"
                if materialize_image_for_rpa(detail_urls[0], target):
                    written += 1
            for index, image_url in enumerate(urls, start=1):
                target = product_dir / f"{index}.jpg"
                if materialize_image_for_rpa(image_url, target):
                    written += 1
            if written >= 3:
                prepared.append({"skc": item.skc, "color": color, "count": written, "dir": str(product_dir)})
            else:
                missing.append({"skc": item.skc, "color": color, "reason": f"??? {written} ? RPA ????"})
    return {
        "root": str(target_root),
        "prepared_count": len(prepared),
        "missing_count": len(missing),
        "detail_prepared_count": len(prepared_details),
        "prepared": prepared,
        "prepared_details": prepared_details,
        "missing": missing,
    }


def miaoshou_rows(items: list[ProcessingItem], image_url_resolver: Callable[[ProcessingItem, str, str | None, int], str] | None = None) -> list[list[object]]:
    resolve_image_url = image_url_resolver or (lambda item, source_url, _color, _index: miaoshou_image_url(source_url))
    rows: list[list[object]] = []
    for item in items:
        sizes = item_supported_upload_sizes(item)
        if not sizes:
            continue
        colors = item_colors(item)
        suggested_price = get_setting_value("miaoshou_default_suggested_price_usd", "60").strip() or "60"
        detail_urls = [resolve_image_url(item, url, "detail", index) for index, url in enumerate(detail_upload_images(item)[:10], start=1)]
        detail_urls = [url for url in detail_urls if url]
        detail_description = miaoshou_detail_description(item, detail_urls)
        first_row_for_item = True
        for color in colors:
            color_urls = [resolve_image_url(item, url, color, index) for index, url in enumerate(color_upload_images(item, color)[:10], start=1)]
            color_urls = [url for url in color_urls if url]
            carousel_images = "\n".join(color_urls)
            material_image = next((url for url in color_urls if str(url or "").startswith("http")), "")
            for size in sizes:
                color_code = color_english_label(color)
                sku_code = f"{item.skc}-{color_code}-{safe_path_segment(size, 'SIZE')}"
                rows.append([
                    item.title,
                    item.english_title,
                    detail_description if first_row_for_item else "",
                    item.skc,
                    "颜色",
                    color,
                    "尺码",
                    size,
                    "",
                    f"{float(item.declared_price):.2f}",
                    sku_code,
                    str(item.length_cm),
                    str(item.width_cm),
                    str(item.height_cm),
                    str(item.weight_g),
                    "",
                    "",
                    item.source_url,
                    carousel_images,
                    material_image,
                    "",
                    "",
                    "",
                    suggested_price,
                    str(item.stock),
                    str(item.ship_days),
                ])
                first_row_for_item = False
    return rows


def upload_operation_settings_snapshot() -> dict[str, object]:
    return {
        "executor": {
            "mode": get_setting_value("executor_mode", "local_python"),
            "server_url": get_setting_value("executor_server_url", "http://127.0.0.1:8000"),
            "download_url": get_setting_value("executor_download_url", ""),
            "version": get_setting_value("executor_version", "0.1.0"),
            "poll_seconds": get_setting_value("executor_poll_seconds", "5"),
            "heartbeat_timeout": get_setting_value("executor_heartbeat_timeout", "60"),
            "task_scope": get_setting_value("executor_task_scope", "manual"),
        },
        "temu": {
            "shop_account": get_setting_value("temu_shop_account", ""),
            "site": get_setting_value("temu_site", "美国站"),
            "product_template": get_setting_value("temu_product_template", ""),
            "size_category": get_setting_value("temu_size_category", ""),
            "size_template": get_setting_value("temu_size_template", ""),
            "warehouse_template": get_setting_value("temu_warehouse_template", ""),
            "logistics_template": get_setting_value("temu_logistics_template", ""),
            "ship_days": get_setting_value("temu_ship_days", "9"),
            "declare_markup": get_setting_value("temu_declare_markup", "239.0"),
            "default_weight_g": get_setting_value("temu_default_weight_g", "350"),
            "default_stock": get_setting_value("temu_default_stock", "999"),
            "package_size_cm": {
                "length": get_setting_value("temu_package_length_cm", "10"),
                "width": get_setting_value("temu_package_width_cm", "5"),
                "height": get_setting_value("temu_package_height_cm", "1"),
            },
            "excel_path": get_setting_value("temu_1688_excel_path", ""),
            "batch_limit": get_setting_value("temu_batch_limit", ""),
            "start_skc": get_setting_value("temu_start_skc", ""),
            "append_sku_suffix": get_setting_value("temu_append_sku_suffix", "true"),
            "add_model_info": get_setting_value("temu_add_model_info", "false"),
            "model_index": get_setting_value("temu_model_index", "2"),
        },
        "flow": {
            "fill_skc": get_setting_value("upload_fill_skc", "true"),
            "skc_missing_policy": get_setting_value("upload_skc_missing_policy", "pause"),
            "auto_submit": get_setting_value("upload_auto_submit", "false"),
            "error_policy": get_setting_value("upload_error_policy", "skip"),
            "save_screenshots": get_setting_value("upload_save_screenshots", "false"),
            "save_html": get_setting_value("upload_save_html", "false"),
            "trace": get_setting_value("upload_trace", "off"),
            "step_delay_ms": get_setting_value("upload_step_delay_ms", "500"),
        },
    }


def upload_operation_settings_summary(settings: dict[str, object]) -> str:
    flow = settings.get("flow", {}) if isinstance(settings.get("flow"), dict) else {}
    executor = settings.get("executor", {}) if isinstance(settings.get("executor"), dict) else {}
    temu = settings.get("temu", {}) if isinstance(settings.get("temu"), dict) else {}
    return (
        f"执行器={executor.get('mode', '-')}; "
        f"店铺={temu.get('shop_account', '-') or '-'}; 起始SKC={temu.get('start_skc', '-') or '-'}; "
        f"SKC={'自动填写' if flow.get('fill_skc') == 'true' else '不填写'}; "
        f"缺SKC策略={flow.get('skc_missing_policy', '-')}; "
        f"自动提交={'开启' if flow.get('auto_submit') == 'true' else '关闭'}; "
        f"遇错={flow.get('error_policy', '-')}; Trace={flow.get('trace', '-')}"
    )


def write_task_manifest(task_id: int, task_type: str, status: str, export_path: str, items: list[ProcessingItem], run_log: str = "") -> str:
    manifest_dir = DATA_DIR / "logs"
    manifest_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = manifest_dir / f"upload_task_{task_id}.json"
    settings_snapshot = upload_operation_settings_snapshot()
    payload = {
        "task_id": task_id,
        "task_type": task_type,
        "status": status,
        "export_path": export_path,
        "run_log": run_log,
        "created_at": now_text(),
        "settings": settings_snapshot,
        "settings_summary": upload_operation_settings_summary(settings_snapshot),
        "preflight": upload_preflight_summary(items),
        "rpa_images": prepare_rpa_sku_images([item for item in items if not processing_item_issues(item)]),
        "items": [item.model_dump() for item in items],
    }
    manifest_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(manifest_path)


def upload_task_from_row(row: sqlite3.Row) -> UploadTask:
    return UploadTask(**dict(row))


def summarize_upload_failure(stdout_text: str, run_log: str = "") -> tuple[str, str, str]:
    combined = "\n".join(part for part in [run_log, stdout_text] if part).strip()
    if not combined:
        return "executor", "执行器未返回失败详情", ""
    stage_patterns = [
        ("import", ["导入", "import"]),
        ("ready-check", ["就绪", "搜索", "ready"]),
        ("open-edit", ["编辑", "open-edit"]),
        ("apply-template", ["模板", "template"]),
        ("upload-color-images", ["上传", "图片", "file chooser", "image"]),
        ("prune-empty-colors", ["剪枝", "无图", "prune"]),
        ("fill-size-chart", ["尺码", "size"]),
        ("select-warehouse", ["仓库", "warehouse"]),
        ("fill-batch-inventory", ["库存", "inventory"]),
        ("save", ["保存", "错误", "save"]),
    ]
    stage = "rpa"
    lower_text = combined.lower()
    for candidate, keywords in stage_patterns:
        if any(keyword.lower() in lower_text for keyword in keywords):
            stage = candidate
            break
    lines = [line.strip() for line in combined.splitlines() if line.strip()]
    reason = ""
    for line in reversed(lines):
        if any(keyword in line for keyword in ["失败", "错误", "Exception", "Traceback", "Timeout", "ERROR", "fail", "Failed"]):
            reason = line
            break
    if not reason and lines:
        reason = lines[-1]
    evidence_path = ""
    for match in re.finditer(r"[A-Za-z]:[^\r\n<>|?*]+?\.(?:png|jpg|jpeg|html|log|txt)", combined, re.I):
        evidence_path = match.group(0).strip().strip("'\")")
    return stage, reason[:500], evidence_path


def publish_record_from_row(row: sqlite3.Row) -> PublishRecord:
    return PublishRecord(**dict(row))


def build_processing_items(
    limit: int | None = None,
    offset: int = 0,
    keyword: str = "",
    status: str = "",
    exception_type: str = "",
) -> list[ProcessingItem]:
    clauses, params = processing_item_filter_sql(keyword, status, exception_type)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    limit_sql = ""
    if limit is not None:
        limit_sql = " LIMIT ? OFFSET ?"
        params.extend([max(1, int(limit)), max(0, int(offset or 0))])
    with connect() as db:
        rows = db.execute(
            f"""
            SELECT p.*, o.english_title AS override_english_title, o.color AS override_color,
                   o.size AS override_size, o.sku_code AS override_sku_code,
                   o.declared_price AS override_declared_price, o.weight_g AS override_weight_g,
                   o.length_cm AS override_length_cm, o.width_cm AS override_width_cm,
                   o.height_cm AS override_height_cm, o.source_url AS override_source_url,
                   o.stock AS override_stock, o.ship_days AS override_ship_days,
                   d.status AS detail_status, d.note AS detail_note
            FROM products p
            LEFT JOIN processing_overrides o ON o.product_id = p.id
            LEFT JOIN product_detail_snapshots d ON d.product_id = p.id
            {where_sql}
            ORDER BY p.id DESC
            {limit_sql}
            """,
            params,
        ).fetchall()
    items: list[ProcessingItem] = []
    for row in rows:
        title = row["title"]
        skc = row["skc"]
        main_image = str(row["main_image"] or "").strip()
        image_options = build_product_image_options(row["id"], main_image, row["override_source_url"] or "")
        color_assignments = list_color_image_assignments(row["id"])
        resolved_override_color = clean_color_label(row["override_color"] or "")
        fallback_color = normalize_assignment_color(row["override_color"] or "")
        if resolved_override_color and color_assignments:
            real_color_assignments = [assignment for assignment in color_assignments if clean_color_label(assignment.color) == resolved_override_color]
            fallback_assignments = [assignment for assignment in color_assignments if clean_color_label(assignment.color) in {"", "混合色"}]
            if not real_color_assignments and fallback_assignments:
                first_assignment = fallback_assignments[0]
                color_assignments = [
                    ColorImageAssignment(
                        color=resolved_override_color,
                        image_url=first_assignment.image_url,
                        preview_url=first_assignment.preview_url,
                        source=first_assignment.source,
                        sort_order=0,
                        is_auto=first_assignment.is_auto,
                        confidence=first_assignment.confidence,
                    )
                ]
        if main_image and not color_assignments:
            color_assignments = [
                ColorImageAssignment(
                    color=fallback_color,
                    image_url=main_image,
                    preview_url=image_preview_url(main_image),
                    source="main_image",
                    sort_order=0,
                    is_auto=True,
                    confidence=0.5,
                )
            ]
        has_image = bool(main_image) or any(option.kind == "main_image" and option.url for option in image_options)
        default_declared_price = round(float(row["total_cost"]) + 239, 2)
        source_url = row["override_source_url"] or ""
        items.append(
            ProcessingItem(
                product_id=row["id"],
                title=title,
                skc=skc,
                english_title=row["override_english_title"] if valid_english_title(row["override_english_title"] or "") else english_title_fallback(title),
                color=clean_color_label(row["override_color"] or "") or (row["override_color"] or ""),
                size=row["override_size"] or "pending",
                sku_code=row["override_sku_code"] or f"{skc}-COLOR-SIZE",
                declared_price=float(row["override_declared_price"] or default_declared_price),
                weight_g=int(row["override_weight_g"] or 350),
                length_cm=int(row["override_length_cm"] or 15),
                width_cm=int(row["override_width_cm"] or 10),
                height_cm=int(row["override_height_cm"] or 2),
                source_url=source_url,
                main_image=main_image,
                image_options=image_options,
                color_image_assignments=color_assignments,
                detail_image_assignments=list_detail_image_assignments(row["id"]),
                detail_status=row["detail_status"] or "",
                detail_note=row["detail_note"] or "",
                image_status="has_main_image" if has_image else "missing_main_image",
                stock=int(row["override_stock"] or 999),
                ship_days=int(row["override_ship_days"] or 9),
                status="ready_to_export" if has_image else "needs_image",
            )
        )
    return items


def count_processing_products(keyword: str = "", status: str = "", exception_type: str = "") -> int:
    clauses, params = processing_item_filter_sql(keyword, status, exception_type)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as db:
        return int(db.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM products p
            LEFT JOIN processing_overrides o ON o.product_id = p.id
            {where_sql}
            """,
            params,
        ).fetchone()["total"])


def processing_item_filter_sql(keyword: str = "", status: str = "", exception_type: str = "") -> tuple[list[str], list[object]]:
    clauses: list[str] = ["COALESCE(p.processing_hidden, 0) = 0"]
    params: list[object] = []
    clean_keyword = str(keyword or "").strip()
    if clean_keyword:
        clauses.append(
            """
            (p.title LIKE ? OR p.skc LIKE ? OR p.sku_summary LIKE ? OR
             o.english_title LIKE ? OR o.color LIKE ? OR o.size LIKE ? OR
             o.sku_code LIKE ? OR o.source_url LIKE ?)
            """
        )
        like_keyword = f"%{clean_keyword}%"
        params.extend([like_keyword] * 8)
    clean_status = str(status or "").strip()
    if clean_status == "ready_to_export":
        clauses.append("NULLIF(TRIM(COALESCE(p.main_image, '')), '') IS NOT NULL")
    elif clean_status == "needs_image":
        clauses.append("NULLIF(TRIM(COALESCE(p.main_image, '')), '') IS NULL")
    elif clean_status == "exception_pool":
        clauses.append("EXISTS (SELECT 1 FROM processing_exceptions pe WHERE pe.product_id = p.id AND pe.status = 'open')")
    elif clean_status:
        clauses.append("p.status = ?")
        params.append(clean_status)
    exception_patterns = {
        "title": ["标题", "SKU Code"],
        "spec": ["颜色", "尺码", "标准色"],
        "image": ["主图", "颜色图", "图片"],
        "field": ["申报价", "重量", "尺寸", "库存", "发货"],
        "link": ["链接"],
    }.get(str(exception_type or "").strip(), [])
    if exception_patterns:
        exception_text_sql = "COALESCE(pe.issues_json, '') || COALESCE(pe.warnings_json, '') || COALESCE(pe.note, '')"
        clauses.append(
            "EXISTS ("
            "SELECT 1 FROM processing_exceptions pe "
            "WHERE pe.product_id = p.id AND pe.status = 'open' AND "
            f"({' OR '.join([exception_text_sql + ' LIKE ?' for _ in exception_patterns])})"
            ")"
        )
        params.extend([f"%{pattern}%" for pattern in exception_patterns])
    return clauses, params


def get_processing_item(product_id: int) -> ProcessingItem:
    for item in build_processing_items():
        if item.product_id == product_id:
            return item
    raise HTTPException(status_code=404, detail="商品不存在")


def increment_skc_value(start_skc: str, offset: int) -> str:
    clean_skc = str(start_skc or "").strip()
    match = re.search(r"(\d+)$", clean_skc)
    if not match:
        return clean_skc if offset == 0 else f"{clean_skc}-{offset + 1}"
    number_text = match.group(1)
    prefix = clean_skc[:match.start(1)]
    return f"{prefix}{int(number_text) + offset:0{len(number_text)}d}"


def ensure_unique_skc(db: sqlite3.Connection, desired_skc: str, product_id: int) -> str:
    clean_skc = str(desired_skc or "").strip()
    if not clean_skc:
        return clean_skc
    candidate = clean_skc
    offset = 0
    while db.execute("SELECT 1 FROM products WHERE skc = ? AND id != ?", (candidate, product_id)).fetchone():
        offset += 1
        candidate = increment_skc_value(clean_skc, offset)
    return candidate


def save_processing_override(product_id: int, payload: ProcessingItemPayload) -> ProcessingItem:
    timestamp = now_text()
    with connect() as db:
        product = db.execute("SELECT id FROM products WHERE id = ?", (product_id,)).fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="商品不存在")
        db.execute(
            """
            INSERT INTO processing_overrides (
                product_id, english_title, color, size, sku_code, declared_price,
                weight_g, length_cm, width_cm, height_cm, source_url, stock, ship_days, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(product_id) DO UPDATE SET
                english_title = excluded.english_title,
                color = excluded.color,
                size = excluded.size,
                sku_code = excluded.sku_code,
                declared_price = excluded.declared_price,
                weight_g = excluded.weight_g,
                length_cm = excluded.length_cm,
                width_cm = excluded.width_cm,
                height_cm = excluded.height_cm,
                source_url = excluded.source_url,
                stock = excluded.stock,
                ship_days = excluded.ship_days,
                updated_at = excluded.updated_at
            """,
            (
                product_id,
                payload.english_title,
                payload.color,
                payload.size,
                payload.sku_code,
                payload.declared_price,
                payload.weight_g,
                payload.length_cm,
                payload.width_cm,
                payload.height_cm,
                payload.source_url,
                payload.stock,
                payload.ship_days,
                timestamp,
            ),
        )
        sync_product_weight_from_processing(db, product_id, payload.weight_g)
    return get_processing_item(product_id)



def update_product_title(product_id: int, title: str) -> None:
    clean_title = clean_generated_title(title, 180)
    if not clean_title:
        return
    with connect() as db:
        db.execute(
            "UPDATE products SET title = ?, updated_at = ? WHERE id = ?",
            (clean_title, now_text(), product_id),
        )


def extract_title_pair_from_model_response(response_text: str, fallback_chinese_title: str) -> tuple[str, str]:
    text = str(response_text or "").strip().strip("` ")
    if text.lower().startswith("json"):
        text = text[4:].strip()
    chinese_title = ""
    english_title = ""
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            chinese_title = str(
                parsed.get("chinese_title")
                or parsed.get("title_zh")
                or parsed.get("title_cn")
                or parsed.get("cn_title")
                or parsed.get("zh_title")
                or parsed.get("title")
                or ""
            )
            english_title = str(
                parsed.get("english_title")
                or parsed.get("title_en")
                or parsed.get("en_title")
                or ""
            )
    except Exception:
        pass
    if not chinese_title or not english_title:
        lines = [clean_generated_title(line, 180) for line in re.split(r"[\r\n]+", text) if clean_generated_title(line, 180)]
        for line in lines:
            label_removed = re.sub(r"^(\u4e2d\u6587\u6807\u9898|\u4e2d\u6587|Chinese title|CN)[:\uff1a]\s*", "", line, flags=re.I).strip()
            if label_removed != line or (contains_cjk(line) and not chinese_title):
                chinese_title = label_removed
                continue
            label_removed = re.sub(r"^(\u82f1\u6587\u6807\u9898|\u82f1\u6587|English title|EN)[:\uff1a]\s*", "", line, flags=re.I).strip()
            if label_removed != line or (not contains_cjk(line) and not english_title):
                english_title = label_removed
    chinese_title = clean_generated_title(chinese_title or fallback_chinese_title, 180)
    english_title = clean_generated_title(english_title, 120)
    if not valid_english_title(english_title):
        english_title = english_title_fallback(chinese_title)
    return chinese_title, english_title

def chinese_title_fallback(item: ProcessingItem) -> str:
    text = clean_generated_title(item.title, 180)
    remove_words = [
        "厂家直批", "厂家", "直批", "批发", "现货", "包邮", "网红", "爆款",
        "2024年", "2025年", "2026年", "2024", "2025", "2026",
        "新款", "热卖", "ins", "INS",
    ]
    for word in remove_words:
        text = text.replace(word, "")
    color = item_colors(item)[0] if item_colors(item) else clean_color_label(item.color)
    parts: list[str] = []
    if color:
        parts.append(color)
    if "高腰" in text:
        parts.append("高腰")
    if "宽松" in text:
        parts.append("宽松")
    if "紧身" in text or "包臀" in text:
        parts.append("修身")
    if "卷边" in text:
        parts.append("卷边")
    if "破洞" in text:
        parts.append("破洞")
    if "五分" in text or "5分" in text:
        parts.append("五分")
    if "牛仔" in text:
        parts.append("牛仔")
    if "短裤" in text or "热裤" in text:
        parts.append("短裤")
    if "女" in text:
        parts.append("女")
    if "夏" in text:
        parts.append("夏季")
    result = "".join(dict.fromkeys(part for part in parts if part))
    if len(result) < 8:
        result = text[:40]
    if result == clean_generated_title(item.title, 180):
        result = f"{result[:32]} 上架款"
    return clean_generated_title(result, 80)


def fallback_title_pair_for_processing_item(item: ProcessingItem, cache_key: str, reason: str = "") -> tuple[str, str]:
    chinese_title = chinese_title_fallback(item)
    english_title = english_title_fallback(chinese_title)
    save_title_cache(cache_key, item.title, chinese_title, english_title, "fallback", reason[:80])
    return chinese_title, english_title


def generate_title_pair_for_processing_item(item: ProcessingItem) -> tuple[str, str]:
    title_template = active_title_prompt_template()
    cache_key = title_cache_key(item, title_template)
    with connect() as db:
        cached = db.execute("SELECT chinese_title, english_title FROM ai_title_cache WHERE cache_key = ?", (cache_key,)).fetchone()
    if cached and cached["chinese_title"] and valid_english_title(cached["english_title"]):
        return cached["chinese_title"], cached["english_title"]
    with connect() as db:
        deepseek = db.execute("SELECT * FROM api_configs WHERE key = 'deepseek'").fetchone()
    if not deepseek or not deepseek["enabled"] or not deepseek["api_key"]:
        return fallback_title_pair_for_processing_item(item, cache_key, "missing_api_config")
    system_prompt = "You are a clothing ecommerce title assistant. Return strict JSON only, no markdown, no explanation."
    if title_template:
        user_prompt = render_title_prompt(title_template, item)
    else:
        user_prompt = (
            "\u8bf7\u57fa\u4e8e\u539f\u59cb\u5546\u54c1\u6807\u9898\u751f\u6210\u4e00\u4e2a\u9002\u5408\u4e0a\u67b6\u7684\u4e2d\u6587\u5546\u54c1\u6807\u9898\uff0c\u7136\u540e\u628a\u8fd9\u4e2a\u4e2d\u6587\u6807\u9898\u9010\u6761\u7b49\u4e49\u7ffb\u8bd1\u6210\u82f1\u6587\u3002\n"
            "\u91cd\u70b9\u662f\u4e2d\u6587\u6807\u9898\u8d28\u91cf\uff1b\u82f1\u6587\u5fc5\u987b\u548c\u4e2d\u6587\u6807\u9898 1:1 \u5bf9\u5e94\uff0c\u4e0d\u8981\u989d\u5916\u6269\u5199\u3002\n"
            "\u8981\u6c42\uff1a\u4e2d\u6587\u6807\u9898\u81ea\u7136\u3001\u6e05\u695a\u3001\u4e0d\u8981\u54c1\u724c\u8bcd\u3001\u4e0d\u8981\u5938\u5927\u529f\u6548\uff1b\u82f1\u6587\u53ea\u5141\u8bb8\u82f1\u6587\uff0c120\u5b57\u7b26\u4ee5\u5185\u3002\n"
            "\u8fd4\u56de JSON\uff1a{\"chinese_title\":\"...\",\"english_title\":\"...\"}\n\n"
            f"\u539f\u59cb\u6807\u9898\uff1a{item.title}\n"
            f"SKC\uff1a{item.skc}\n"
            f"\u989c\u8272\uff1a{item.color}\n"
            f"\u5c3a\u7801\uff1a{item.size}\n"
        )
    try:
        response_text = call_openai_compatible_chat(
            deepseek["base_url"] or "https://api.deepseek.com",
            deepseek["api_key"],
            deepseek["model"] or "deepseek-chat",
            [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            timeout=15,
        )
        chinese_title, english_title = extract_title_pair_from_model_response(response_text, item.title)
        validate_title_pair(chinese_title, english_title)
        save_title_cache(cache_key, item.title, chinese_title, english_title, "deepseek", deepseek["model"] or "deepseek-chat")
        return chinese_title, english_title
    except Exception as exc:
        return fallback_title_pair_for_processing_item(item, cache_key, f"ai_error:{exc}")


def title_cache_key(item: ProcessingItem, title_template: str = "") -> str:
    raw = "|".join([item.title or "", item.skc or "", item.color or "", item.size or "", title_template or active_title_prompt_template()])
    return hashlib.sha256(raw.encode("utf-8", errors="ignore")).hexdigest()


def save_title_cache(cache_key: str, source_title: str, chinese_title: str, english_title: str, provider: str, model: str) -> None:
    timestamp = now_text()
    with connect() as db:
        db.execute(
            """
            INSERT INTO ai_title_cache (cache_key, source_title, chinese_title, english_title, provider, model, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(cache_key) DO UPDATE SET
                chinese_title = excluded.chinese_title,
                english_title = excluded.english_title,
                provider = excluded.provider,
                model = excluded.model,
                updated_at = excluded.updated_at
            """,
            (cache_key, source_title, chinese_title, english_title, provider, model, timestamp, timestamp),
        )


def title_cache_exists(item: ProcessingItem) -> bool:
    with connect() as db:
        row = db.execute("SELECT 1 FROM ai_title_cache WHERE cache_key = ?", (title_cache_key(item),)).fetchone()
    return bool(row)


def validate_title_pair(chinese_title: str, english_title: str) -> None:
    if not clean_generated_title(chinese_title, 180):
        raise ValueError("中文标题为空")
    if not valid_english_title(english_title):
        raise ValueError("英文标题校验失败")

def generate_title_for_processing_item(product_id: int) -> ProcessingItem:
    item = get_processing_item(product_id)
    with connect() as db:
        deepseek = db.execute("SELECT * FROM api_configs WHERE key = 'deepseek'").fetchone()
    if deepseek and deepseek["enabled"] and deepseek["api_key"]:
        title_template = active_prompt_template("标题提示词", "标题生成") or active_prompt_template("标题提示词")
        if title_template:
            system_prompt = "You are a product title copywriting assistant. Follow the user's prompt exactly. Return only the requested result."
            user_prompt = render_title_prompt(title_template, item)
        else:
            system_prompt = "You are a Temu ecommerce listing assistant. Generate one concise, natural English product title. Return only the title, no quotes, no explanation."
            user_prompt = (
                f"Source title: {item.title}\n"
                f"SKC: {item.skc}\n"
                f"Color: {item.color}\n"
                f"Size: {item.size}\n"
                "Requirements: English only, under 120 characters, no brand names unless present, no exaggerated claims."
            )
        generated_title = call_openai_compatible_chat(
            deepseek["base_url"] or "https://api.deepseek.com",
            deepseek["api_key"],
            deepseek["model"] or "deepseek-chat",
            [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            timeout=15,
        )
        generated_title = extract_title_from_model_response(generated_title, item.title)
        if not valid_english_title(generated_title):
            generated_title = english_title_fallback(item.title)
    else:
        generated_title = english_title_fallback(item.title)
    updated = save_processing_override(
        product_id,
        ProcessingItemPayload(
            english_title=generated_title,
            color=item.color,
            size=item.size,
            sku_code=item.sku_code,
            declared_price=item.declared_price,
            weight_g=item.weight_g,
            length_cm=item.length_cm,
            width_cm=item.width_cm,
            height_cm=item.height_cm,
            source_url=item.source_url,
            stock=item.stock,
            ship_days=item.ship_days,
        ),
    )
    updated.english_title = f"{updated.english_title}"
    return updated




def api_config_from_row(row: sqlite3.Row) -> ApiConfig:
    data = dict(row)
    data["enabled"] = bool(data["enabled"])
    return ApiConfig(**data)


def setting_from_row(row: sqlite3.Row) -> AppSetting:
    return AppSetting(**dict(row))


def prompt_from_row(row: sqlite3.Row) -> PromptTemplate:
    return PromptTemplate(**dict(row))


def is_default_prompt_template_row(row: sqlite3.Row) -> bool:
    return (row["name"], row["category"], row["prompt_type"], row["usage"]) in DEFAULT_PROMPT_TEMPLATE_KEYS


@app.get("/api/config/apis", response_model=list[ApiConfig])
def list_api_configs() -> list[ApiConfig]:
    with connect() as db:
        rows = db.execute("SELECT * FROM api_configs ORDER BY key").fetchall()
    return [api_config_from_row(row) for row in rows]


@app.put("/api/config/apis/{config_key}", response_model=ApiConfig)
def save_api_config(config_key: str, payload: ApiConfig) -> ApiConfig:
    timestamp = now_text()
    with connect() as db:
        db.execute(
            """
            INSERT INTO api_configs (key, name, enabled, base_url, model, api_key, usage, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET name=excluded.name, enabled=excluded.enabled, base_url=excluded.base_url,
            model=excluded.model, api_key=excluded.api_key, usage=excluded.usage, updated_at=excluded.updated_at
            """,
            (config_key, payload.name, int(payload.enabled), payload.base_url, payload.model, payload.api_key, payload.usage, timestamp),
        )
        row = db.execute("SELECT * FROM api_configs WHERE key = ?", (config_key,)).fetchone()
    return api_config_from_row(row)


@app.get("/api/settings", response_model=list[AppSetting])
def list_settings() -> list[AppSetting]:
    with connect() as db:
        rows = db.execute("SELECT * FROM app_settings ORDER BY key").fetchall()
    return [setting_from_row(row) for row in rows]


@app.put("/api/settings", response_model=list[AppSetting])
def save_settings(payload: list[AppSetting]) -> list[AppSetting]:
    timestamp = now_text()
    with connect() as db:
        for item in payload:
            db.execute(
                """
                INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, ?)
                ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
                """,
                (item.key, item.value, timestamp),
            )
        rows = db.execute("SELECT * FROM app_settings ORDER BY key").fetchall()
    return [setting_from_row(row) for row in rows]


@app.get("/api/prompts", response_model=list[PromptTemplate])
def list_prompts() -> list[PromptTemplate]:
    with connect() as db:
        rows = db.execute("SELECT * FROM prompt_templates ORDER BY id DESC").fetchall()
    return [prompt_from_row(row) for row in rows]


@app.post("/api/prompts", response_model=PromptTemplate)
def create_prompt(payload: PromptTemplatePayload) -> PromptTemplate:
    timestamp = now_text()
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO prompt_templates (name, category, prompt_type, usage, content, status, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (payload.name, payload.category, payload.prompt_type, payload.usage, payload.content, payload.status, timestamp),
        )
        row = db.execute("SELECT * FROM prompt_templates WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return prompt_from_row(row)


@app.put("/api/prompts/{prompt_id}", response_model=PromptTemplate)
def update_prompt(prompt_id: int, payload: PromptTemplatePayload) -> PromptTemplate:
    with connect() as db:
        result = db.execute(
            """
            UPDATE prompt_templates SET name=?, category=?, prompt_type=?, usage=?, content=?, status=?, updated_at=? WHERE id=?
            """,
            (payload.name, payload.category, payload.prompt_type, payload.usage, payload.content, payload.status, now_text(), prompt_id),
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="prompt not found")
        row = db.execute("SELECT * FROM prompt_templates WHERE id = ?", (prompt_id,)).fetchone()
    return prompt_from_row(row)


@app.delete("/api/prompts/{prompt_id}")
def delete_prompt(prompt_id: int) -> dict[str, bool]:
    with connect() as db:
        row = db.execute("SELECT * FROM prompt_templates WHERE id = ?", (prompt_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="prompt not found")
        if is_default_prompt_template_row(row):
            raise HTTPException(status_code=400, detail="默认提示词不允许删除")
        result = db.execute("DELETE FROM prompt_templates WHERE id = ?", (prompt_id,))
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="prompt not found")
    return {"ok": True}


REQUIRED_TEMU_UPLOAD_SETTINGS = [
    ("temu_shop_account", "店铺账号"),
    ("temu_site", "经营站点"),
    ("temu_product_template", "产品模板"),
    ("temu_size_category", "尺码分类"),
    ("temu_size_template", "尺码模板"),
    ("temu_warehouse_template", "仓库模板"),
    ("temu_logistics_template", "物流模板"),
]


def validate_required_temu_upload_settings() -> None:
    missing = [label for key, label in REQUIRED_TEMU_UPLOAD_SETTINGS if not get_setting_value(key, "").strip()]
    if missing:
        raise HTTPException(status_code=400, detail=f"请先填写店铺与模板必填项：{'、'.join(missing)}")


def export_miaoshou_for_items(items: list[ProcessingItem], prefix: str = "miaoshou_upload") -> dict[str, str]:
    export_dir = DATA_DIR / "export"
    export_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    target = export_dir / filename
    image_result = prepare_rpa_sku_images(items)
    cos_sync_result = {"uploaded_count": 0, "failed_count": 0}
    if upload_image_source() == "cos":
        cos_sync_result = sync_rpa_images_to_cos(items)
    image_url_resolver = cos_export_url_resolver(items) if upload_image_source() == "cos" else None
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Miaoshou Upload"
    sheet.append(MIAOSHOU_COLUMNS)
    for row in miaoshou_rows(items, image_url_resolver=image_url_resolver):
        sheet.append(row)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 36)
    workbook.save(target)
    return {
        "path": str(target),
        "filename": filename,
        "download_url": f"/api/export/miaoshou/{filename}",
        "rpa_image_dir": str(image_result["root"]),
        "cos_uploaded_count": str(cos_sync_result["uploaded_count"]),
    }


@app.post("/api/export/miaoshou")
def export_miaoshou() -> dict[str, str]:
    return export_miaoshou_for_items(build_processing_items())


@app.get("/api/export/miaoshou/{filename}")
def download_miaoshou_export(filename: str) -> FileResponse:
    export_dir = (DATA_DIR / "export").resolve()
    target = (export_dir / filename).resolve()
    if target.parent != export_dir or not target.name.startswith(("miaoshou_upload_", "miaoshou_retry_")) or target.suffix != ".xlsx":
        raise HTTPException(status_code=400, detail="Invalid export filename")
    if not target.exists():
        raise HTTPException(status_code=404, detail="Export file not found")
    return FileResponse(target, filename=target.name)


@app.get("/api/upload/preflight")
def upload_preflight() -> dict[str, object]:
    scripts = ["export_miaoshou_xlsx.py", "test_sku_feishu.py", "rpa_upload.py", "pipeline.py"]
    script_dir = configured_script_dir()
    script_status = {name: (script_dir / name).exists() for name in scripts}
    latest_exports = sorted((DATA_DIR / "export").glob("miaoshou_upload_*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True) if (DATA_DIR / "export").exists() else []
    image_root = DATA_DIR / "images"
    items = build_processing_items()
    item_summary = upload_preflight_summary(items)
    rpa_image_result = {"root": str(configured_rpa_sku_image_dir()), "prepared_count": 0, "missing_count": 0, "prepared": [], "missing": []}
    cos = cos_preflight()
    infra_ready = image_root.exists()
    data_ready = item_summary["total_count"] > 0 and item_summary["blocked_count"] == 0
    return {
        "script_dir": str(script_dir),
        "script_dir_exists": script_dir.exists(),
        "scripts": script_status,
        "latest_export": str(latest_exports[0]) if latest_exports else "",
        "image_dir": str(image_root),
        "image_dir_exists": image_root.exists(),
        "rpa_image_dir": str(rpa_image_result["root"]),
        "rpa_images": rpa_image_result,
        "cos": cos,
        "infra_ready": infra_ready,
        "items": item_summary,
        "command_preview": display_upload_rpa_command(str(latest_exports[0]) if latest_exports else "<export_path>"),
        "ready": infra_ready and data_ready,
    }


@app.get("/api/upload-tasks", response_model=list[UploadTask])
def list_upload_tasks() -> list[UploadTask]:
    with connect() as db:
        rows = db.execute("SELECT * FROM upload_tasks ORDER BY id DESC").fetchall()
    return [upload_task_from_row(row) for row in rows]


@app.delete("/api/upload-tasks/{task_id}")
def delete_upload_task(task_id: int) -> dict[str, bool]:
    with connect() as db:
        result = db.execute("DELETE FROM upload_tasks WHERE id = ?", (task_id,))
    log_path = DATA_DIR / "logs" / f"upload_task_{task_id}.log"
    manifest_path = DATA_DIR / "logs" / f"upload_task_{task_id}.json"
    log_path.unlink(missing_ok=True)
    manifest_path.unlink(missing_ok=True)
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="任务不存在")
    return {"ok": True}


@app.post("/api/upload-tasks/clear")
def clear_upload_tasks() -> dict[str, int]:
    with connect() as db:
        count = db.execute("SELECT COUNT(*) AS total FROM upload_tasks").fetchone()["total"]
        db.execute("DELETE FROM upload_tasks")
        db.execute("DELETE FROM publish_records")
    return {"deleted_count": count}


@app.post("/api/upload-tasks", response_model=UploadTask)
def create_upload_task() -> UploadTask:
    validate_required_temu_upload_settings()
    export = export_miaoshou()
    items = build_processing_items()
    item_summary = upload_preflight_summary(items)
    success_count = int(item_summary["ready_count"])
    failed_count = int(item_summary["blocked_count"])
    status = "export_ready" if failed_count == 0 else "needs_review"
    timestamp = now_text()
    settings_snapshot = upload_operation_settings_snapshot()
    settings_summary = upload_operation_settings_summary(settings_snapshot)
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO upload_tasks (name, status, total_count, success_count, failed_count, export_path, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (f"\u4e0a\u8d27\u4efb\u52a1 {timestamp}", status, len(items), success_count, failed_count, export["path"], timestamp, timestamp),
        )
        db.execute("DELETE FROM publish_records")
        for item in items:
            issues = processing_item_issues(item)
            if not issues:
                db.execute(
                    "INSERT INTO publish_records (result, skc, title, reason, created_at) VALUES (?, ?, ?, ?, ?)",
                    ("\u6210\u529f", item.skc, item.title, "", timestamp),
                )
            else:
                db.execute(
                    "INSERT INTO publish_records (result, skc, title, reason, created_at) VALUES (?, ?, ?, ?, ?)",
                    ("\u5931\u8d25", item.skc, item.title, "；".join(issues), timestamp),
                )
        manifest_path = write_task_manifest(cursor.lastrowid, "create", status, export["path"], items)
        db.execute("UPDATE upload_tasks SET run_log = ? WHERE id = ?", (f"设置：{settings_summary}；Manifest: {manifest_path}", cursor.lastrowid))
        row = db.execute("SELECT * FROM upload_tasks WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return upload_task_from_row(row)




@app.post("/api/upload-tasks/run", response_model=UploadTask)
def run_upload_task() -> UploadTask:
    validate_required_temu_upload_settings()
    mode = "real"
    preflight = upload_preflight()
    export = export_miaoshou()
    items = build_processing_items()
    diagnostics = upload_item_diagnostics(items)
    blocked_items = [item for item in diagnostics if not item["ready"]]
    timestamp = now_text()
    status = "queued_for_executor"
    if blocked_items:
        status = "needs_review"
        run_log = f"Real RPA blocked: {len(blocked_items)} products need review. SKC: {', '.join(str(item['skc']) for item in blocked_items)}"
    elif not preflight["ready"]:
        status = "needs_review"
        run_log = "Real RPA blocked: upload data preflight failed."
    else:
        run_log = "Real RPA queued for Windows executor. Open the upload software on the Windows machine to claim and run this task."
    failed_count = len(blocked_items)
    success_count = len(items) - failed_count
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO upload_tasks (name, status, total_count, success_count, failed_count, export_path, run_log, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (f"正式上货 {timestamp}", status, len(items), success_count, failed_count, export["path"], run_log, timestamp, timestamp),
        )
        task_id = cursor.lastrowid
        log_path = write_run_log(task_id, run_log)
        manifest_path = write_task_manifest(task_id, mode, status, export["path"], items, run_log)
        db.execute("UPDATE upload_tasks SET run_log = ? WHERE id = ?", (run_log + f" Log: {log_path} Manifest: {manifest_path}", task_id))
        db.execute("DELETE FROM publish_records")
        for item in items:
            issues = processing_item_issues(item)
            if issues:
                db.execute(
                    "INSERT INTO publish_records (result, skc, title, reason, created_at) VALUES (?, ?, ?, ?, ?)",
                    ("Failed", item.skc, item.title, "; ".join(issues), timestamp),
                )
            else:
                db.execute(
                    "INSERT INTO publish_records (result, skc, title, reason, created_at) VALUES (?, ?, ?, ?, ?)",
                    ("Queued", item.skc, item.title, "Waiting for Windows executor", timestamp),
                )
        row = db.execute("SELECT * FROM upload_tasks WHERE id = ?", (task_id,)).fetchone()
    return upload_task_from_row(row)


@app.post("/api/upload-tasks/retry-failed", response_model=UploadTask)
def retry_failed_upload_task(payload: RetryFailedUploadPayload = Body(default_factory=RetryFailedUploadPayload)) -> UploadTask:
    with connect() as db:
        product_ids = product_ids_for_failed_publish_records(db, payload.ids)
    if not product_ids:
        raise HTTPException(status_code=400, detail="没有可重试的失败商品")
    items = build_processing_items_for_ids(product_ids)
    if not items:
        raise HTTPException(status_code=400, detail="失败商品已不存在，无法重试")
    timestamp = now_text()
    export = export_miaoshou_for_items(items, prefix="miaoshou_retry")
    preflight_summary = upload_preflight_summary(items)
    blocked_items = preflight_summary["blocked_items"]
    status = "needs_review" if blocked_items else "queued_for_executor"
    run_log = "Retry upload queued for Windows executor." if not blocked_items else "Retry upload blocked: failed products still need review."
    failed_count = len(blocked_items)
    success_count = len(items) - failed_count
    with connect() as db:
        previous = db.execute("SELECT MAX(retry_count) AS retry_count FROM upload_tasks").fetchone()
        retry_count = int(previous["retry_count"] or 0) + 1 if previous else 1
        cursor = db.execute(
            """
            INSERT INTO upload_tasks (name, status, total_count, success_count, failed_count, export_path, run_log, retry_count, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (f"失败重试 {timestamp}", status, len(items), success_count, failed_count, export["path"], run_log, retry_count, timestamp, timestamp),
        )
        task_id = cursor.lastrowid
        log_path = write_run_log(task_id, run_log)
        manifest_path = write_task_manifest(task_id, "retry_failed", status, export["path"], items, run_log)
        db.execute("UPDATE upload_tasks SET run_log = ? WHERE id = ?", (run_log + f" Log: {log_path} Manifest: {manifest_path}", task_id))
        db.execute("DELETE FROM publish_records")
        for item in items:
            issues = processing_item_issues(item)
            db.execute(
                "INSERT INTO publish_records (result, skc, title, reason, created_at) VALUES (?, ?, ?, ?, ?)",
                ("Failed" if issues else "Queued", item.skc, item.title, "; ".join(issues) if issues else "Retry queued for Windows executor", timestamp),
            )
        row = db.execute("SELECT * FROM upload_tasks WHERE id = ?", (task_id,)).fetchone()
    return upload_task_from_row(row)






@app.post("/api/products/images/batch")
def batch_upload_product_images(images: list[UploadFile] = File(...)) -> dict[str, object]:
    updated: list[str] = []
    unmatched: list[str] = []
    with connect() as db:
        products = db.execute("SELECT id, skc FROM products").fetchall()
        product_by_skc = {row["skc"].lower(): row for row in products}
        for image in images:
            filename = image.filename or ""
            stem = Path(filename).stem.lower()
            matched = None
            for skc, row in product_by_skc.items():
                if skc in stem:
                    matched = row
                    break
            if not matched:
                unmatched.append(filename)
                continue
            suffix = Path(filename).suffix.lower() or ".jpg"
            product_dir = IMAGE_DIR / str(matched["id"])
            product_dir.mkdir(parents=True, exist_ok=True)
            target = product_dir / f"main{suffix}"
            for old_file in product_dir.glob("main.*"):
                old_file.unlink(missing_ok=True)
            with target.open("wb") as file:
                shutil.copyfileobj(image.file, file)
            image_path = f"/images/{matched['id']}/{target.name}"
            db.execute(
                "UPDATE products SET main_image = ?, status = ?, updated_at = ? WHERE id = ?",
                (image_path, "ok", now_text(), matched["id"]),
            )
            updated.append(matched["skc"])
    return {"updated": updated, "unmatched": unmatched, "updated_count": len(updated), "unmatched_count": len(unmatched)}


@app.get("/api/upload-tasks/{task_id}/log")
def get_upload_task_log(task_id: int) -> dict[str, str]:
    log_path = DATA_DIR / "logs" / f"upload_task_{task_id}.log"
    if not log_path.exists():
        return {"path": str(log_path), "content": ""}
    return {"path": str(log_path), "content": log_path.read_text(encoding="utf-8", errors="replace")}


@app.get("/api/upload-tasks/{task_id}/manifest")
def get_upload_task_manifest(task_id: int) -> dict[str, object]:
    manifest_path = DATA_DIR / "logs" / f"upload_task_{task_id}.json"
    if not manifest_path.exists():
        return {"path": str(manifest_path), "content": {}}
    return {"path": str(manifest_path), "content": json.loads(manifest_path.read_text(encoding="utf-8"))}


def _executor_task_payload(row: sqlite3.Row) -> dict[str, object]:
    manifest_path = DATA_DIR / "logs" / f"upload_task_{row['id']}.json"
    manifest: dict[str, object] = {}
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    export_path = Path(row["export_path"] or "")
    filename = export_path.name if export_path.name else ""
    return {
        "task": upload_task_from_row(row).model_dump(),
        "manifest": manifest,
        "export_download_url": f"/api/executor/tasks/{row['id']}/export" if filename else "",
        "export_filename": filename,
        "manifest_download_url": f"/api/upload-tasks/{row['id']}/manifest",
        "settings": upload_operation_settings_snapshot(),
    }


@app.post("/api/executor/tasks/claim")
def executor_claim_task(payload: ExecutorClaimPayload) -> dict[str, object]:
    executor_id = payload.executor_id.strip() or "default-windows-executor"
    timestamp = now_text()
    with connect() as db:
        row = db.execute(
            """
            SELECT * FROM upload_tasks
            WHERE status IN ('export_ready', 'queued_for_executor')
            ORDER BY id ASC
            LIMIT 1
            """
        ).fetchone()
        if not row:
            return {"task": None}
        db.execute(
            """
            UPDATE upload_tasks
            SET status = 'claimed_by_executor', executor_id = ?, claimed_at = ?, heartbeat_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (executor_id, timestamp, timestamp, timestamp, row["id"]),
        )
        row = db.execute("SELECT * FROM upload_tasks WHERE id = ?", (row["id"],)).fetchone()
    return _executor_task_payload(row)


@app.get("/api/executor/status", response_model=ExecutorStatus)
def executor_status() -> ExecutorStatus:
    work_dir = Path(get_setting_value("executor_work_dir", str(DATA_DIR / "executor")))
    status_path = work_dir / "status.json"
    pending_dir = work_dir / "pending_results"
    pending_count = len(list(pending_dir.glob("task_*.json"))) if pending_dir.exists() else 0
    payload: dict[str, object] = {}
    if status_path.exists():
        try:
            payload = json.loads(status_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            payload = {"status": "invalid_status_file", "message": "执行器状态文件无法读取"}
    updated_at = str(payload.get("updated_at") or "")
    online = False
    if updated_at:
        try:
            heartbeat_time = datetime.strptime(updated_at, "%Y-%m-%d %H:%M:%S")
            timeout_seconds = int_setting("executor_heartbeat_timeout", 60, 10, 3600)
            online = (datetime.now() - heartbeat_time).total_seconds() <= timeout_seconds
        except ValueError:
            online = False
    return ExecutorStatus(
        online=online,
        executor_id=str(payload.get("executor_id") or ""),
        status=str(payload.get("status") or ("online" if online else "offline")),
        message=str(payload.get("message") or ""),
        task_id=int(payload.get("task_id") or 0),
        updated_at=updated_at,
        pending_result_count=pending_count,
        status_path=str(status_path),
    )


@app.post("/api/executor/tasks/{task_id}/heartbeat")
def executor_heartbeat(task_id: int, payload: ExecutorHeartbeatPayload) -> dict[str, object]:
    timestamp = now_text()
    with connect() as db:
        row = db.execute("SELECT * FROM upload_tasks WHERE id = ?", (task_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="上货任务不存在")
        if row["executor_id"] and payload.executor_id and row["executor_id"] != payload.executor_id:
            raise HTTPException(status_code=409, detail="任务已被其他执行器领取")
        db.execute("UPDATE upload_tasks SET heartbeat_at = ?, updated_at = ? WHERE id = ?", (timestamp, timestamp, task_id))
    return {"ok": True, "heartbeat_at": timestamp}


@app.get("/api/executor/tasks/{task_id}/export")
def executor_download_export(task_id: int) -> FileResponse:
    with connect() as db:
        row = db.execute("SELECT * FROM upload_tasks WHERE id = ?", (task_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="上货任务不存在")
    target = Path(row["export_path"] or "").resolve()
    export_dir = (DATA_DIR / "export").resolve()
    if target.parent != export_dir or target.suffix != ".xlsx" or not target.exists():
        raise HTTPException(status_code=404, detail="导出表不存在")
    return FileResponse(target, filename=target.name)


@app.post("/api/executor/tasks/{task_id}/report", response_model=UploadTask)
def executor_report_task(task_id: int, payload: ExecutorReportPayload) -> UploadTask:
    allowed_status = {"rpa_success", "rpa_failed", "needs_review", "blocked"}
    status = payload.status if payload.status in allowed_status else "rpa_failed"
    timestamp = now_text()
    log_text = payload.run_log.strip()
    if payload.stdout.strip():
        log_text = (log_text + "\n\n" if log_text else "") + payload.stdout.strip()
    failure_stage = payload.failure_stage.strip()
    failure_reason = payload.failure_reason.strip()
    evidence_path = payload.evidence_path.strip()
    if status != "rpa_success" and (not failure_stage or not failure_reason):
        inferred_stage, inferred_reason, inferred_evidence = summarize_upload_failure(payload.stdout, payload.run_log)
        failure_stage = failure_stage or inferred_stage
        failure_reason = failure_reason or inferred_reason
        evidence_path = evidence_path or inferred_evidence
    if status == "rpa_success":
        failure_stage = ""
        failure_reason = ""
        evidence_path = ""
    with connect() as db:
        row = db.execute("SELECT * FROM upload_tasks WHERE id = ?", (task_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="上货任务不存在")
        if row["executor_id"] and payload.executor_id and row["executor_id"] != payload.executor_id:
            raise HTTPException(status_code=409, detail="任务已被其他执行器领取")
        if log_text:
            log_path = write_run_log(task_id, log_text)
            run_log = f"Executor {payload.executor_id or row['executor_id']} reported {status}. Log: {log_path}"
        else:
            run_log = f"Executor {payload.executor_id or row['executor_id']} reported {status}."
        db.execute(
            """
            UPDATE upload_tasks
            SET status = ?, success_count = ?, failed_count = ?, run_log = ?, failure_stage = ?,
                failure_reason = ?, evidence_path = ?, heartbeat_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, payload.success_count, payload.failed_count, run_log, failure_stage, failure_reason, evidence_path, timestamp, timestamp, task_id),
        )
        db.execute("DELETE FROM publish_records")
        manifest_path = DATA_DIR / "logs" / f"upload_task_{task_id}.json"
        manifest: dict[str, object] = {}
        if manifest_path.exists():
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        for item in manifest.get("items", []) if isinstance(manifest.get("items"), list) else []:
            result = "Success" if status == "rpa_success" else "Failed"
            reason = "Executor reported success" if status == "rpa_success" else (failure_reason or run_log)
            skc = str(item.get("skc", ""))
            db.execute(
                "INSERT INTO publish_records (result, skc, title, reason, created_at) VALUES (?, ?, ?, ?, ?)",
                (result, skc, str(item.get("title", "")), reason, timestamp),
            )
            if status == "rpa_success" and skc:
                db.execute("UPDATE products SET status = ?, updated_at = ? WHERE skc = ?", ("已上货", timestamp, skc))
        row = db.execute("SELECT * FROM upload_tasks WHERE id = ?", (task_id,)).fetchone()
    return upload_task_from_row(row)


@app.get("/api/products/missing-images", response_model=list[Product])
def list_missing_image_products() -> list[Product]:
    with connect() as db:
        rows = db.execute("SELECT * FROM products WHERE main_image IS NULL OR main_image = '' ORDER BY id DESC").fetchall()
    return [product_from_row(row) for row in rows]


@app.get("/api/publish-records", response_model=list[PublishRecord])
def list_publish_records(result: str = "", q: str = "") -> list[PublishRecord]:
    clauses: list[str] = []
    params: list[object] = []
    if result.strip():
        clauses.append("result = ?")
        params.append(result.strip())
    if q.strip():
        keyword = f"%{q.strip()}%"
        clauses.append("(skc LIKE ? OR title LIKE ? OR reason LIKE ?)")
        params.extend([keyword, keyword, keyword])
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as db:
        rows = db.execute(f"SELECT id, result, skc, title, reason, created_at FROM publish_records {where_sql} ORDER BY id DESC", params).fetchall()
    return [publish_record_from_row(row) for row in rows]


@app.get("/api/processing-items", response_model=ProcessingItemPage)
def list_processing_items(page: int = 1, page_size: int = 50, q: str = "", status: str = "", exception_type: str = "") -> ProcessingItemPage:
    safe_page_size = max(10, min(int(page_size or 50), 200))
    total = count_processing_products(q, status, exception_type)
    total_pages = max(1, (total + safe_page_size - 1) // safe_page_size)
    safe_page = min(max(1, int(page or 1)), total_pages)
    offset = (safe_page - 1) * safe_page_size
    items = build_processing_items(limit=safe_page_size, offset=offset, keyword=q, status=status, exception_type=exception_type)
    return ProcessingItemPage(items=items, total=total, page=safe_page, page_size=safe_page_size, total_pages=total_pages)


@app.get("/api/processing-items/by-id/{product_id}", response_model=ProcessingItem)
def read_processing_item(product_id: int) -> ProcessingItem:
    return get_processing_item(product_id)


@app.post("/api/processing-items/preflight")
def run_processing_preflight() -> dict[str, object]:
    items = build_processing_items()
    summary = sync_processing_exceptions(items)
    summary["exception_pool_synced"] = True
    return summary


@app.get("/api/processing-items/exceptions", response_model=list[ProcessingExceptionItem])
def list_processing_exceptions(status: str = "open") -> list[ProcessingExceptionItem]:
    clauses: list[str] = []
    params: list[object] = []
    if status:
        clauses.append("status = ?")
        params.append(status)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as db:
        rows = db.execute(
            f"SELECT * FROM processing_exceptions {where_sql} ORDER BY updated_at DESC, product_id DESC",
            params,
        ).fetchall()
    return [processing_exception_from_row(row) for row in rows]


@app.put("/api/processing-items/bulk", response_model=list[ProcessingItem])
def update_processing_items_bulk(payload: ProcessingBulkPayload) -> list[ProcessingItem]:
    if not payload.ids:
        raise HTTPException(status_code=400, detail="请选择要批量编辑的商品")
    ordered_ids = [int(product_id) for product_id in payload.ids]
    start_skc = payload.start_skc.strip()
    if start_skc:
        timestamp = now_text()
        with connect() as db:
            for index, product_id in enumerate(ordered_ids):
                desired_skc = increment_skc_value(start_skc, index)
                skc = ensure_unique_skc(db, desired_skc, product_id)
                db.execute("UPDATE products SET skc = ?, updated_at = ? WHERE id = ?", (skc, timestamp, product_id))
    updated_items: list[ProcessingItem] = []
    for product_id in ordered_ids:
        current = get_processing_item(product_id)
        merged_fields = ProcessingItemPayload(
            english_title=current.english_title,
            color=current.color,
            size=current.size,
            sku_code=current.sku_code,
            declared_price=payload.fields.declared_price,
            weight_g=payload.fields.weight_g,
            length_cm=payload.fields.length_cm,
            width_cm=payload.fields.width_cm,
            height_cm=payload.fields.height_cm,
            source_url=current.source_url,
            stock=payload.fields.stock,
            ship_days=payload.fields.ship_days,
        )
        updated_items.append(save_processing_override(product_id, merged_fields))
    return updated_items


@app.post("/api/processing-items/field-tasks", response_model=ProcessingFieldTask)
def create_processing_field_task(payload: ProcessingBulkPayload) -> ProcessingFieldTask:
    ordered_ids = []
    seen_ids: set[int] = set()
    for product_id in payload.ids:
        product_id = int(product_id)
        if product_id > 0 and product_id not in seen_ids:
            ordered_ids.append(product_id)
            seen_ids.add(product_id)
    if not ordered_ids:
        raise HTTPException(status_code=400, detail="请选择要批量编辑的商品")
    timestamp = now_text()
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO processing_field_tasks (
                status, ids_json, fields_json, start_skc, total_count, processed_count, success_count, failed_count, note, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "queued",
                json.dumps(ordered_ids),
                json.dumps(payload.fields.model_dump(), ensure_ascii=False),
                payload.start_skc.strip(),
                len(ordered_ids),
                0,
                0,
                0,
                "已加入批量字段处理队列",
                timestamp,
                timestamp,
            ),
        )
        task_id = cursor.lastrowid
        row = db.execute("SELECT * FROM processing_field_tasks WHERE id = ?", (task_id,)).fetchone()
    schedule_processing_field_task_execution()
    return processing_field_task_from_row(row)


@app.get("/api/processing-items/field-tasks/{task_id}", response_model=ProcessingFieldTask)
def get_processing_field_task(task_id: int) -> ProcessingFieldTask:
    with connect() as db:
        row = db.execute("SELECT * FROM processing_field_tasks WHERE id = ?", (task_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="批量字段处理任务不存在")
    return processing_field_task_from_row(row)


def processing_task_history_from_row(row: sqlite3.Row, task_type: str) -> ProcessingTaskHistoryItem:
    failed_ids = json.loads(row["failed_ids_json"] or "[]") if "failed_ids_json" in row.keys() else []
    return ProcessingTaskHistoryItem(
        id=int(row["id"]),
        task_type=task_type,
        task_label="批量标题" if task_type == "title" else "批量字段",
        status=row["status"] or "",
        total_count=int(row["total_count"] or 0),
        processed_count=int(row["processed_count"] or 0),
        success_count=int(row["success_count"] or 0),
        failed_count=int(row["failed_count"] or 0),
        cache_hit_count=int(row["cache_hit_count"] or 0) if "cache_hit_count" in row.keys() else 0,
        retryable_count=len(failed_ids),
        note=row["note"] or "",
        created_at=row["created_at"] or "",
        updated_at=row["updated_at"] or "",
    )


@app.get("/api/processing-items/tasks", response_model=ProcessingTaskHistoryPage)
def list_processing_tasks(task_type: str = "", status: str = "", page: int = 1, page_size: int = 20, summary: bool = False) -> ProcessingTaskHistoryPage:
    safe_page = max(1, int(page or 1))
    safe_page_size = max(1, min(int(page_size or 20), 100))
    types = [task_type.strip().lower()] if task_type.strip().lower() in {"title", "field"} else ["title", "field"]
    status_filter = status.strip().lower()
    with connect() as db:
        tasks: list[ProcessingTaskHistoryItem] = []
        if "title" in types:
            title_rows = db.execute("SELECT * FROM processing_title_tasks ORDER BY id DESC LIMIT 500").fetchall()
            tasks.extend(processing_task_history_from_row(row, "title") for row in title_rows)
        if "field" in types:
            field_rows = db.execute("SELECT * FROM processing_field_tasks ORDER BY id DESC LIMIT 500").fetchall()
            tasks.extend(processing_task_history_from_row(row, "field") for row in field_rows)
    if status_filter:
        if status_filter == "retryable":
            tasks = [task for task in tasks if task.retryable_count > 0]
        else:
            tasks = [task for task in tasks if task.status == status_filter]
    tasks = sorted(tasks, key=lambda task: (task.updated_at, task.id), reverse=True)
    if summary:
        priority_tasks = [task for task in tasks if task.status in {"queued", "running"} or task.retryable_count > 0]
        recent_tasks = [task for task in tasks if task not in priority_tasks]
        tasks = [*priority_tasks, *recent_tasks][:5]
        return ProcessingTaskHistoryPage(items=tasks, total=len(tasks), page=1, page_size=5, total_pages=1)
    total = len(tasks)
    total_pages = max(1, (total + safe_page_size - 1) // safe_page_size)
    safe_page = min(safe_page, total_pages)
    start = (safe_page - 1) * safe_page_size
    return ProcessingTaskHistoryPage(
        items=tasks[start:start + safe_page_size],
        total=total,
        page=safe_page,
        page_size=safe_page_size,
        total_pages=total_pages,
    )


@app.post("/api/processing-items/auto-assign-color-images", response_model=AutoAssignColorImagesResult)
def auto_assign_processing_color_images(payload: AutoAssignColorImagesPayload) -> AutoAssignColorImagesResult:
    return auto_assign_color_images(payload)


@app.post("/api/processing-items/color-image-assignment", response_model=list[ColorImageAssignment])
def manual_assign_processing_color_image(payload: ManualColorImageAssignmentPayload) -> list[ColorImageAssignment]:
    return save_manual_color_image_assignment(payload)


@app.post("/api/processing-items/detail-image-assignment", response_model=list[DetailImageAssignment])
def manual_assign_processing_detail_image(payload: ManualDetailImageAssignmentPayload) -> list[DetailImageAssignment]:
    return save_manual_detail_image_assignment(payload)


@app.post("/api/processing-items/dedupe-images", response_model=DedupeProcessingImagesResult)
def dedupe_processing_item_images(payload: DedupeProcessingImagesPayload) -> DedupeProcessingImagesResult:
    return dedupe_processing_images(payload.ids)


@app.delete("/api/processing-items/{product_id}/detail-images/{slot_index}", response_model=list[DetailImageAssignment])
def clear_processing_detail_image(product_id: int, slot_index: int) -> list[DetailImageAssignment]:
    return save_detail_image_slot(product_id, slot_index, "")


@app.post("/api/processing-items/generate-title", response_model=ProcessingItem)
def generate_processing_title(payload: GenerateTitlePayload) -> ProcessingItem:
    return generate_title_for_processing_item(payload.product_id)


@app.post("/api/processing-items/generate-titles", response_model=BulkGenerateTitleResult)
def generate_processing_titles(payload: BulkGenerateTitlePayload) -> BulkGenerateTitleResult:
    seen_ids: set[int] = set()
    ordered_ids: list[int] = []
    for product_id in payload.ids:
        if product_id > 0 and product_id not in seen_ids:
            ordered_ids.append(product_id)
            seen_ids.add(product_id)
    if not ordered_ids:
        raise HTTPException(status_code=400, detail="请选择要批量生成标题的商品")

    results: list[BulkGenerateTitleItemResult] = []
    for product_id in ordered_ids:
        try:
            item = get_processing_item(product_id)
            chinese_title, english_title = generate_title_pair_for_processing_item(item)
            update_product_title(product_id, chinese_title)
            updated = save_processing_override(
                product_id,
                ProcessingItemPayload(
                    english_title=english_title,
                    color=item.color,
                    size=item.size,
                    sku_code=item.sku_code,
                    declared_price=item.declared_price,
                    weight_g=item.weight_g,
                    length_cm=item.length_cm,
                    width_cm=item.width_cm,
                    height_cm=item.height_cm,
                    source_url=item.source_url,
                    stock=item.stock,
                    ship_days=item.ship_days,
                ),
            )
            results.append(
                BulkGenerateTitleItemResult(
                    product_id=product_id,
                    ok=True,
                    chinese_title=updated.title,
                    english_title=updated.english_title,
                )
            )
        except Exception as exc:
            results.append(BulkGenerateTitleItemResult(product_id=product_id, ok=False, error=str(exc)))
    success_count = sum(1 for result in results if result.ok)
    return BulkGenerateTitleResult(success_count=success_count, failed_count=len(results) - success_count, results=results)


@app.post("/api/processing-items/title-tasks", response_model=ProcessingTitleTask)
def create_processing_title_task(payload: BulkGenerateTitlePayload) -> ProcessingTitleTask:
    seen_ids: set[int] = set()
    ordered_ids: list[int] = []
    for product_id in payload.ids:
        product_id = int(product_id)
        if product_id > 0 and product_id not in seen_ids:
            ordered_ids.append(product_id)
            seen_ids.add(product_id)
    if not ordered_ids:
        raise HTTPException(status_code=400, detail="请选择要批量生成标题的商品")
    timestamp = now_text()
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO processing_title_tasks (
                status, ids_json, total_count, processed_count, success_count, failed_count, cache_hit_count, note, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("queued", json.dumps(ordered_ids), len(ordered_ids), 0, 0, 0, 0, "已加入标题处理队列", timestamp, timestamp),
        )
        task_id = cursor.lastrowid
        row = db.execute("SELECT * FROM processing_title_tasks WHERE id = ?", (task_id,)).fetchone()
    schedule_processing_title_task_execution()
    return processing_title_task_from_row(row)


@app.get("/api/processing-items/title-tasks/{task_id}", response_model=ProcessingTitleTask)
def get_processing_title_task(task_id: int) -> ProcessingTitleTask:
    with connect() as db:
        row = db.execute("SELECT * FROM processing_title_tasks WHERE id = ?", (task_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="标题处理任务不存在")
    return processing_title_task_from_row(row)


@app.post("/api/processing-items/tasks/{task_type}/{task_id}/retry-failed", response_model=ProcessingTaskHistoryItem)
def retry_failed_processing_task(task_type: str, task_id: int) -> ProcessingTaskHistoryItem:
    task_type = task_type.strip().lower()
    timestamp = now_text()
    with connect() as db:
        if task_type == "title":
            source = db.execute("SELECT * FROM processing_title_tasks WHERE id = ?", (task_id,)).fetchone()
            if not source:
                raise HTTPException(status_code=404, detail="标题处理任务不存在")
            failed_ids = [int(product_id) for product_id in json.loads(source["failed_ids_json"] or "[]") if int(product_id) > 0]
            if not failed_ids:
                raise HTTPException(status_code=400, detail="没有可重试的失败商品")
            cursor = db.execute(
                """
                INSERT INTO processing_title_tasks (
                    status, ids_json, failed_ids_json, total_count, processed_count, success_count, failed_count, cache_hit_count, note, created_at, updated_at
                ) VALUES (?, ?, ?, ?, 0, 0, 0, 0, ?, ?, ?)
                """,
                ("queued", json.dumps(failed_ids), "[]", len(failed_ids), f"重试标题任务 #{task_id} 的失败商品", timestamp, timestamp),
            )
            new_row = db.execute("SELECT * FROM processing_title_tasks WHERE id = ?", (cursor.lastrowid,)).fetchone()
            schedule_processing_title_task_execution()
            return processing_task_history_from_row(new_row, "title")
        if task_type == "field":
            source = db.execute("SELECT * FROM processing_field_tasks WHERE id = ?", (task_id,)).fetchone()
            if not source:
                raise HTTPException(status_code=404, detail="批量字段处理任务不存在")
            failed_ids = [int(product_id) for product_id in json.loads(source["failed_ids_json"] or "[]") if int(product_id) > 0]
            if not failed_ids:
                raise HTTPException(status_code=400, detail="没有可重试的失败商品")
            cursor = db.execute(
                """
                INSERT INTO processing_field_tasks (
                    status, ids_json, failed_ids_json, fields_json, start_skc, total_count, processed_count, success_count, failed_count, note, created_at, updated_at
                ) VALUES (?, ?, ?, ?, '', ?, 0, 0, 0, ?, ?, ?)
                """,
                ("queued", json.dumps(failed_ids), "[]", source["fields_json"] or "{}", len(failed_ids), f"重试字段任务 #{task_id} 的失败商品", timestamp, timestamp),
            )
            new_row = db.execute("SELECT * FROM processing_field_tasks WHERE id = ?", (cursor.lastrowid,)).fetchone()
            schedule_processing_field_task_execution()
            return processing_task_history_from_row(new_row, "field")
    raise HTTPException(status_code=400, detail="任务类型必须是 title 或 field")


@app.put("/api/processing-items/{product_id}", response_model=ProcessingItem)
def update_processing_item(product_id: int, payload: ProcessingItemPayload) -> ProcessingItem:
    return save_processing_override(product_id, payload)


@app.delete("/api/processing-items/{product_id}")
def hide_processing_item(product_id: int) -> dict[str, bool]:
    with connect() as db:
        result = db.execute(
            "UPDATE products SET processing_hidden = 1, updated_at = ? WHERE id = ?",
            (now_text(), product_id),
        )
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="商品不存在")
    return {"ok": True}


@app.post("/api/processing-items/bulk/hide")
def hide_processing_items_bulk(payload: CollectionBulkPayload) -> dict[str, int]:
    ids = [int(product_id) for product_id in payload.ids if int(product_id) > 0]
    if not ids:
        return {"hidden_count": 0}
    hidden_count = 0
    with connect() as db:
        for id_batch in chunked_ids(ids):
            placeholders = ",".join("?" for _ in id_batch)
            result = db.execute(
                f"UPDATE products SET processing_hidden = 1, updated_at = ? WHERE id IN ({placeholders})",
                [now_text(), *id_batch],
            )
            hidden_count += result.rowcount
    return {"hidden_count": hidden_count}


@app.get("/api/collection-items", response_model=CollectionItemPage)
def list_collection_items(
    q: str = "",
    source: str = "",
    status: str = "",
    min_price: float = 0,
    max_price: float = 0,
    sort: str = "newest",
    page: int = 1,
    page_size: int = 50,
) -> CollectionItemPage:
    clauses: list[str] = []
    params: list[object] = []
    if q.strip():
        keyword = f"%{q.strip()}%"
        clauses.append("(title LIKE ? OR source LIKE ? OR source_url LIKE ?)")
        params.extend([keyword, keyword, keyword])
    if source.strip():
        clauses.append("source = ?")
        params.append(source.strip())
    if status.strip():
        clauses.append("status = ?")
        params.append(status.strip())
    if min_price:
        clauses.append("price >= ?")
        params.append(min_price)
    if max_price:
        clauses.append("price <= ?")
        params.append(max_price)
    order_by = {
        "sales_desc": "sales DESC, id DESC",
        "price_asc": "price ASC, id DESC",
        "price_desc": "price DESC, id DESC",
        "oldest": "id ASC",
        "newest": "id DESC",
    }.get(sort, "id DESC")
    clean_page = max(1, int(page or 1))
    clean_page_size = max(10, min(int(page_size or 50), 200))
    offset = (clean_page - 1) * clean_page_size
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as db:
        total = int(db.execute(f"SELECT COUNT(*) AS total FROM collection_items {where_sql}", params).fetchone()["total"])
        rows = db.execute(
            f"SELECT * FROM collection_items {where_sql} ORDER BY {order_by} LIMIT ? OFFSET ?",
            [*params, clean_page_size, offset],
        ).fetchall()
    total_pages = max(1, (total + clean_page_size - 1) // clean_page_size)
    return CollectionItemPage(
        items=[collection_item_from_row(row) for row in rows],
        total=total,
        page=clean_page,
        page_size=clean_page_size,
        total_pages=total_pages,
    )


@app.get("/api/collection-items/ids")
def list_collection_item_ids(
    q: str = "",
    source: str = "",
    status: str = "",
    min_price: float = 0,
    max_price: float = 0,
    sort: str = "newest",
    limit: int = 5000,
) -> dict[str, object]:
    clauses: list[str] = []
    params: list[object] = []
    if q.strip():
        keyword = f"%{q.strip()}%"
        clauses.append("(title LIKE ? OR source LIKE ? OR source_url LIKE ?)")
        params.extend([keyword, keyword, keyword])
    if source.strip():
        clauses.append("source = ?")
        params.append(source.strip())
    if status.strip():
        clauses.append("status = ?")
        params.append(status.strip())
    if min_price:
        clauses.append("price >= ?")
        params.append(min_price)
    if max_price:
        clauses.append("price <= ?")
        params.append(max_price)
    order_by = {
        "sales_desc": "sales DESC, id DESC",
        "price_asc": "price ASC, id DESC",
        "price_desc": "price DESC, id DESC",
        "oldest": "id ASC",
        "newest": "id DESC",
    }.get(sort, "id DESC")
    clean_limit = max(1, min(int(limit or 5000), 20000))
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    with connect() as db:
        total = int(db.execute(f"SELECT COUNT(*) AS total FROM collection_items {where_sql}", params).fetchone()["total"])
        rows = db.execute(
            f"SELECT id FROM collection_items {where_sql} ORDER BY {order_by} LIMIT ?",
            [*params, clean_limit],
        ).fetchall()
    return {"ids": [int(row["id"]) for row in rows], "total": total, "limit": clean_limit}


@app.get("/api/collection-tasks", response_model=list[CollectionTask])
def list_collection_tasks() -> list[CollectionTask]:
    with connect() as db:
        rows = db.execute("SELECT * FROM collection_tasks ORDER BY id DESC").fetchall()
    return [collection_task_from_row(row) for row in rows]


@app.get("/api/collection-tasks/queue")
def collection_task_queue_status() -> dict[str, object]:
    with connect() as db:
        queued = int(db.execute("SELECT COUNT(*) AS total FROM collection_tasks WHERE status = 'queued'").fetchone()["total"])
        running = int(db.execute("SELECT COUNT(*) AS total FROM collection_tasks WHERE status = 'running'").fetchone()["total"])
        next_row = db.execute("SELECT id, keyword FROM collection_tasks WHERE status = 'queued' ORDER BY id ASC LIMIT 1").fetchone()
    return {
        "active": COLLECTION_TASK_QUEUE_ACTIVE,
        "queued_count": queued,
        "running_count": running,
        "next_task": dict(next_row) if next_row else None,
    }


@app.get("/api/collection-tasks/{task_id}/events", response_model=list[CollectionTaskEvent])
def list_collection_task_events(task_id: int) -> list[CollectionTaskEvent]:
    with connect() as db:
        task = db.execute("SELECT id FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
        if not task:
            raise HTTPException(status_code=404, detail="Collection task not found")
        rows = db.execute(
            "SELECT * FROM collection_task_events WHERE task_id = ? ORDER BY id ASC",
            (task_id,),
        ).fetchall()
    return [collection_task_event_from_row(row) for row in rows]


@app.post("/api/collection-tasks/clear")
def clear_collection_tasks() -> dict[str, int]:
    with connect() as db:
        count = db.execute("SELECT COUNT(*) AS total FROM collection_tasks").fetchone()["total"]
        db.execute("DELETE FROM collection_tasks")
    return {"deleted_count": count}


@app.get("/api/collection/preflight")
def get_collection_preflight() -> dict[str, object]:
    return collection_preflight()


@app.get("/api/collection-tasks/{task_id}/request")
def get_collection_task_request(task_id: int) -> dict[str, object]:
    with connect() as db:
        row = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
        if row and row["mode"] == "1688":
            request_path = ensure_collection_request(row)
            db.execute("UPDATE collection_tasks SET request_path = ?, updated_at = ? WHERE id = ?", (request_path, now_text(), task_id))
            row = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="采集任务不存在")
    request_path = row["request_path"]
    if not request_path:
        return {"path": "", "content": {}}
    target = Path(request_path)
    if not target.exists():
        return {"path": request_path, "content": {}}
    return {"path": request_path, "content": json.loads(target.read_text(encoding="utf-8"))}


@app.post("/api/collection-tasks/{task_id}/execute", response_model=CollectionTask)
def execute_collection_task(task_id: int) -> CollectionTask:
    with connect() as db:
        row = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Collection task not found")
        if row["mode"] != "1688":
            raise HTTPException(status_code=400, detail="Only 1688 collection tasks can be executed")
        if row["status"] in {"running", "queued"}:
            return collection_task_from_row(row)
        if row["status"] not in {"external_pending", "blocked", "executed_stub", "empty", "failed", "cancelled"}:
            raise HTTPException(status_code=400, detail="Current status cannot be executed again")
        db.execute("UPDATE collection_tasks SET status = ?, updated_at = ? WHERE id = ?", ("queued", now_text(), task_id))
        add_collection_task_event(db, task_id, "queued", "info", "\u4efb\u52a1\u5df2\u91cd\u65b0\u52a0\u5165\u91c7\u96c6\u961f\u5217")
        updated = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
    schedule_collection_task_execution(task_id)
    return collection_task_from_row(updated)


@app.post("/api/collection-tasks/{task_id}/cancel", response_model=CollectionTask)
def cancel_collection_task(task_id: int) -> CollectionTask:
    with connect() as db:
        row = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Collection task not found")
        if row["mode"] != "1688":
            raise HTTPException(status_code=400, detail="Only 1688 collection tasks can be cancelled")
        status = row["status"]
        if status == "queued":
            note = "采集任务已取消，未开始执行"
            db.execute("UPDATE collection_tasks SET status = ?, note = ?, updated_at = ? WHERE id = ?", ("cancelled", note, now_text(), task_id))
            add_collection_task_event(db, task_id, "cancelled", "info", note)
        elif status == "running":
            note = "已收到取消请求，任务会在当前采集步骤结束后停止"
            db.execute("UPDATE collection_tasks SET status = ?, note = ?, updated_at = ? WHERE id = ?", ("cancel_requested", note, now_text(), task_id))
            add_collection_task_event(db, task_id, "cancel_requested", "info", note)
        elif status == "cancel_requested":
            pass
        else:
            raise HTTPException(status_code=400, detail="Only queued or running collection tasks can be cancelled")
        updated = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
    return collection_task_from_row(updated)


@app.post("/api/collection-tasks", response_model=CollectionTask)
def create_collection_task(payload: CollectionTaskPayload) -> CollectionTask:
    keyword = payload.keyword.strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="请输入采集关键词")
    requested_count = max(1, min(int(payload.target_count or 1), COLLECTION_BATCH_TARGET_MAX))
    batches = collection_task_batches(requested_count)
    if len(batches) > 1:
        return create_collection_task_batch(payload, keyword, requested_count, batches)
    return create_single_collection_task(payload, keyword, batches[0])


def create_single_collection_task(
    payload: CollectionTaskPayload,
    keyword: str,
    target_count: int,
    parent_task_id: int = 0,
    batch_index: int = 1,
    batch_total: int = 1,
) -> CollectionTask:
    min_price = max(float(payload.min_price or 0), 0)
    max_price = max(float(payload.max_price or 0), 0)
    if max_price and min_price > max_price:
        raise HTTPException(status_code=400, detail="最低价不能高于最高价")
    timestamp = now_text()
    mode, collector, note = resolve_collector(payload)
    batch_note = f"批次 {batch_index}/{batch_total}，目标 {target_count} 条。" if batch_total > 1 else ""
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO collection_tasks (
                keyword, source, mode, collector, target_count, min_price, max_price, blacklist,
                status, note, request_path, result_count, parent_task_id, batch_index, batch_total, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                keyword,
                payload.source,
                mode,
                collector,
                target_count,
                min_price,
                max_price,
                payload.blacklist,
                "queued",
                f"{batch_note}{note}",
                "",
                0,
                parent_task_id,
                batch_index,
                batch_total,
                timestamp,
                timestamp,
            ),
        )
        task_id = cursor.lastrowid
        add_collection_task_event(db, task_id, "created", "running", f"采集任务已创建，目标 {target_count} 条")
        result = run_collector(CollectionTaskPayload(**{**payload.model_dump(), "target_count": target_count}), task_id)
        request_path = write_collection_request(task_id, CollectionTaskPayload(**{**payload.model_dump(), "target_count": target_count}), result.mode, result.collector) if result.mode == "1688" else ""
        created, skipped = insert_collection_candidates_with_db(db, result.candidates)
        add_collection_task_event(db, task_id, "parsed", "info", f"采集器返回 {len(result.candidates)} 条，新增 {created} 条，跳过 {skipped} 条", parsed_count=len(result.candidates), imported_count=created, skipped_count=skipped)
        status = "external_pending" if request_path else ("completed" if created else "empty")
        final_note = result.note if request_path else f"{result.note} 写入 {created} 条，跳过重复 {skipped} 条。".strip()
        if batch_note:
            final_note = f"{batch_note}{final_note}"
        add_collection_task_event(db, task_id, "finished", "success" if created else "info", final_note, imported_count=created, skipped_count=skipped)
        db.execute(
            """
            UPDATE collection_tasks
            SET mode = ?, collector = ?, status = ?, note = ?, request_path = ?, result_count = ?, updated_at = ?
            WHERE id = ?
            """,
            (result.mode, result.collector, status, final_note, request_path, created, now_text(), task_id),
        )
        row = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
    task = collection_task_from_row(row)
    if task.mode == "1688" and task.status == "external_pending":
        return execute_collection_task(task.id)
    return task


def create_collection_task_batch(payload: CollectionTaskPayload, keyword: str, requested_count: int, batches: list[int]) -> CollectionTask:
    min_price = max(float(payload.min_price or 0), 0)
    max_price = max(float(payload.max_price or 0), 0)
    if max_price and min_price > max_price:
        raise HTTPException(status_code=400, detail="最低价不能高于最高价")
    timestamp = now_text()
    batch_total = len(batches)
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO collection_tasks (
                keyword, source, mode, collector, target_count, min_price, max_price, blacklist,
                status, note, request_path, result_count, parent_task_id, batch_index, batch_total, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                keyword,
                payload.source,
                payload.mode,
                "batch_parent",
                requested_count,
                min_price,
                max_price,
                payload.blacklist,
                "completed",
                f"已拆分为 {batch_total} 个采集子任务，每批最多 {COLLECTION_TARGET_MAX} 条，子任务会按队列串行执行。",
                "",
                0,
                0,
                0,
                batch_total,
                timestamp,
                timestamp,
            ),
        )
        parent_task_id = cursor.lastrowid
        add_collection_task_event(db, parent_task_id, "split", "info", f"目标 {requested_count} 条已拆分为 {batch_total} 个子任务")
        first_child_row: sqlite3.Row | None = None
        for index, batch_count in enumerate(batches, start=1):
            child_cursor = db.execute(
                """
                INSERT INTO collection_tasks (
                    keyword, source, mode, collector, target_count, min_price, max_price, blacklist,
                    status, note, request_path, result_count, parent_task_id, batch_index, batch_total, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    keyword,
                    payload.source,
                    "1688",
                    payload.collector,
                    batch_count,
                    min_price,
                    max_price,
                    payload.blacklist,
                    "queued",
                    f"批次 {index}/{batch_total}，目标 {batch_count} 条，等待队列执行。",
                    "",
                    0,
                    parent_task_id,
                    index,
                    batch_total,
                    timestamp,
                    timestamp,
                ),
            )
            child_id = child_cursor.lastrowid
            add_collection_task_event(db, child_id, "created", "running", f"批次 {index}/{batch_total} 已创建，目标 {batch_count} 条")
            request_payload = CollectionTaskPayload(**{**payload.model_dump(), "target_count": batch_count})
            request_path = write_collection_request(child_id, request_payload, "1688", payload.collector)
            db.execute("UPDATE collection_tasks SET request_path = ?, updated_at = ? WHERE id = ?", (request_path, now_text(), child_id))
            if first_child_row is None:
                first_child_row = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (child_id,)).fetchone()
        row = first_child_row or db.execute("SELECT * FROM collection_tasks WHERE id = ?", (parent_task_id,)).fetchone()
    if row and int(row["parent_task_id"] or 0):
        schedule_collection_task_execution(int(row["id"]))
    return collection_task_from_row(row)

@app.post("/api/collection-items", response_model=CollectionItem)
def create_collection_item(payload: CollectionItemPayload) -> CollectionItem:
    with connect() as db:
        cursor = db.execute(
            """
            INSERT INTO collection_items (title, source, source_url, image_url, price, sales, image_count, status, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', ?)
            """,
            (payload.title, payload.source, payload.source_url, payload.image_url, payload.price, payload.sales, payload.image_count, now_text()),
        )
        enqueue_collection_image_download(db, cursor.lastrowid, payload.image_url)
        row = db.execute("SELECT * FROM collection_items WHERE id = ?", (cursor.lastrowid,)).fetchone()
    schedule_collection_image_task_execution()
    return collection_item_from_row(row)


@app.get("/api/collection-image-tasks/status", response_model=CollectionImageTaskStatus)
def collection_image_task_status() -> CollectionImageTaskStatus:
    with connect() as db:
        counts = {
            row["status"]: int(row["total"])
            for row in db.execute("SELECT status, COUNT(*) AS total FROM collection_image_tasks GROUP BY status").fetchall()
        }
    return CollectionImageTaskStatus(
        queued_count=counts.get("queued", 0),
        running_count=counts.get("running", 0),
        completed_count=counts.get("completed", 0),
        failed_count=counts.get("failed", 0),
        active=COLLECTION_IMAGE_TASK_QUEUE_ACTIVE,
    )


@app.post("/api/collection-items/localize-images")
def localize_collection_images(payload: CollectionBulkPayload) -> dict[str, int]:
    if not payload.ids:
        return {"localized_count": 0, "failed_count": 0}
    localized_count = 0
    failed_count = 0
    placeholders = ",".join("?" for _ in payload.ids)
    with connect() as db:
        rows = db.execute(f"SELECT id, image_url FROM collection_items WHERE id IN ({placeholders})", payload.ids).fetchall()
        for row in rows:
            current = str(row["image_url"] or "").strip()
            if not current or current.startswith("/images/"):
                continue
            local_image = save_collection_image_from_url(row["id"], current)
            if local_image:
                db.execute("UPDATE collection_items SET image_url = ? WHERE id = ?", (local_image, row["id"]))
                localized_count += 1
            else:
                failed_count += 1
    return {"localized_count": localized_count, "failed_count": failed_count}


@app.post("/api/collection-items/import-file", response_model=CollectionImportResult)
def import_collection_file(source: str = Form("外部文件"), file: UploadFile = File(...)) -> CollectionImportResult:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xlsm", ".json"}:
        raise HTTPException(status_code=400, detail="仅支持 CSV / XLSX / JSON 文件")
    import_dir = DATA_DIR / "imports"
    import_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"collection_{datetime.now().strftime('%Y%m%d_%H%M%S')}{suffix}"
    target = import_dir / safe_name
    with target.open("wb") as output:
        shutil.copyfileobj(file.file, output)
    rows = parse_collection_rows(target)
    imported, skipped = insert_collection_rows(rows, source)
    return CollectionImportResult(imported_count=imported, skipped_count=skipped, source=source, filename=file.filename or safe_name)


@app.post("/api/collection-tasks/{task_id}/import-result", response_model=CollectionTask)
def import_collection_task_result(task_id: int, file: UploadFile = File(...)) -> CollectionTask:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xlsm", ".json"}:
        raise HTTPException(status_code=400, detail="Only CSV / XLSX / JSON files are supported")
    with connect() as db:
        task = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
    if not task:
        raise HTTPException(status_code=404, detail="Collection task not found")
    import_dir = DATA_DIR / "collection_results"
    import_dir.mkdir(parents=True, exist_ok=True)
    target = import_dir / f"collection_task_{task_id}_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}{suffix}"
    with target.open("wb") as output:
        shutil.copyfileobj(file.file, output)
    rows = parse_collection_rows(target)
    with connect() as db:
        add_collection_task_event(db, task_id, "import_result", "running", f"\u5f00\u59cb\u56de\u586b\u91c7\u96c6\u7ed3\u679c\u6587\u4ef6\uff1a{file.filename or target.name}")
        imported, skipped = insert_collection_rows_with_db(db, rows, task["source"])
        status = "completed" if imported else "empty"
        note = f"\u56de\u586b\u5b8c\u6210\uff1a\u65b0\u589e {imported} \u6761\uff0c\u8df3\u8fc7 {skipped} \u6761\uff0c\u6587\u4ef6={target}"
        add_collection_task_event(db, task_id, "write", "success" if imported else "info", note, parsed_count=len(rows), imported_count=imported, skipped_count=skipped)
        db.execute(
            "UPDATE collection_tasks SET status = ?, note = ?, result_count = ?, updated_at = ? WHERE id = ?",
            (status, note, imported, now_text(), task_id),
        )
        updated = db.execute("SELECT * FROM collection_tasks WHERE id = ?", (task_id,)).fetchone()
    return collection_task_from_row(updated)


@app.get("/api/collection-items/import-template")
def download_collection_import_template() -> FileResponse:
    template_dir = DATA_DIR / "templates"
    template_dir.mkdir(parents=True, exist_ok=True)
    target = template_dir / "collection_import_template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采集候选导入"
    sheet.append(["title", "source", "url", "image_url", "price", "sales", "image_count"])
    sheet.append(["示例 牛仔短裤 高腰显瘦", "1688", "https://example.com/item/1", "https://example.com/image.jpg", 29.8, 1200, 12])
    workbook.save(target)
    return FileResponse(target, filename=target.name)


@app.post("/api/collection-items/import", response_model=list[Product])
def import_collection_items(payload: ImportCollectionPayload) -> list[Product]:
    if not payload.ids:
        raise HTTPException(status_code=400, detail="select collection items first")
    imported: list[Product] = []
    requested_ids = [int(item_id) for item_id in payload.ids if int(item_id) > 0]
    with connect() as db:
        rows: list[sqlite3.Row] = []
        for id_batch in chunked_ids(requested_ids):
            placeholders = ",".join("?" for _ in id_batch)
            rows.extend(db.execute(f"SELECT * FROM collection_items WHERE id IN ({placeholders})", id_batch).fetchall())
        row_by_id = {int(row["id"]): row for row in rows}
        ordered_rows = [row_by_id[item_id] for item_id in requested_ids if item_id in row_by_id]
        if not ordered_rows:
            raise HTTPException(status_code=404, detail="collection item not found")
        for row in ordered_rows:
            if row["status"] == "imported":
                continue
            imported.append(import_collection_row_as_product(db, row))
    return imported


@app.post("/api/collection-items/import-tasks", response_model=CollectionImportTask)
def create_collection_import_task(payload: ImportCollectionPayload) -> CollectionImportTask:
    requested_ids = [int(item_id) for item_id in payload.ids if int(item_id) > 0]
    if not requested_ids:
        raise HTTPException(status_code=400, detail="select collection items first")
    timestamp = now_text()
    with connect() as db:
        existing_count = 0
        for id_batch in chunked_ids(requested_ids):
            placeholders = ",".join("?" for _ in id_batch)
            existing_count += int(db.execute(f"SELECT COUNT(*) AS total FROM collection_items WHERE id IN ({placeholders})", id_batch).fetchone()["total"])
        if not existing_count:
            raise HTTPException(status_code=404, detail="collection item not found")
        cursor = db.execute(
            """
            INSERT INTO collection_import_tasks (
                status, ids_json, total_count, processed_count, imported_count, skipped_count, failed_count, note, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("queued", json.dumps(requested_ids), len(requested_ids), 0, 0, 0, 0, "已加入商品库入库队列", timestamp, timestamp),
        )
        task_id = cursor.lastrowid
        row = db.execute("SELECT * FROM collection_import_tasks WHERE id = ?", (task_id,)).fetchone()
    schedule_collection_import_task_execution(task_id)
    return collection_import_task_from_row(row)


@app.get("/api/collection-items/import-tasks/{task_id}", response_model=CollectionImportTask)
def get_collection_import_task(task_id: int) -> CollectionImportTask:
    with connect() as db:
        row = db.execute("SELECT * FROM collection_import_tasks WHERE id = ?", (task_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="collection import task not found")
    return collection_import_task_from_row(row)

@app.post("/api/collection-items/skip")
def skip_collection_items(payload: CollectionBulkPayload) -> dict[str, int]:
    if not payload.ids:
        return {"updated_count": 0}
    updated_count = 0
    ids = [int(item_id) for item_id in payload.ids if int(item_id) > 0]
    with connect() as db:
        for id_batch in chunked_ids(ids):
            placeholders = ",".join("?" for _ in id_batch)
            result = db.execute(f"UPDATE collection_items SET status = 'skipped' WHERE id IN ({placeholders}) AND status != 'imported'", id_batch)
            updated_count += result.rowcount
    return {"updated_count": updated_count}


@app.post("/api/collection-items/delete")
def delete_collection_items(payload: CollectionBulkPayload) -> dict[str, int]:
    if not payload.ids:
        return {"deleted_count": 0}
    deleted_count = 0
    ids = [int(item_id) for item_id in payload.ids if int(item_id) > 0]
    with connect() as db:
        for id_batch in chunked_ids(ids):
            placeholders = ",".join("?" for _ in id_batch)
            result = db.execute(f"DELETE FROM collection_items WHERE id IN ({placeholders})", id_batch)
            deleted_count += result.rowcount
    return {"deleted_count": deleted_count}


@app.delete("/api/products/{product_id}")
def delete_product(product_id: int) -> dict[str, bool]:
    with connect() as db:
        db.execute("DELETE FROM processing_overrides WHERE product_id = ?", (product_id,))
        result = db.execute("DELETE FROM products WHERE id = ?", (product_id,))
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="商品不存在")
    product_dir = IMAGE_DIR / str(product_id)
    if product_dir.exists():
        shutil.rmtree(product_dir)
    return {"ok": True}
